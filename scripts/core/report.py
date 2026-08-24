"""Deterministic event feedback and final reports from one canonical result model."""

from __future__ import annotations

import csv
import json
import os
from collections import defaultdict
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill

from value_semantics import parse_iso_timestamp

from .constants import (
    DOMAIN_LABELS,
    DOMAINS,
    OPERATION_COUNTERS,
    compact_reason,
    status_label,
    worst_status,
)
from .correlate import action_windows
from .judge import judge_run
from .state import RunPaths, StateError, load_plan, read_stream


def formula_safe_text(value: Any) -> str:
    text = str(value if value is not None else "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def set_safe_cell(cell: Cell, value: Any) -> Cell:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str):
        value = formula_safe_text(value)
    cell.value = value
    return cell


def markdown_safe(value: Any) -> str:
    text = (
        json.dumps(value, ensure_ascii=False, sort_keys=True)
        if isinstance(value, (dict, list))
        else str(value if value is not None else "")
    )
    return text.replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def _append_row(sheet: Any, values: Iterable[Any]) -> None:
    row_values = list(values)
    sheet.append([None] * len(row_values))
    row_index = sheet.max_row
    for column, value in enumerate(row_values, start=1):
        set_safe_cell(sheet.cell(row=row_index, column=column), value)


def _style_sheet(sheet: Any) -> None:
    if sheet.max_row:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [len(str(cell.value or "")) for cell in column[:200]]
        sheet.column_dimensions[column[0].column_letter].width = min(max(values or [8]) + 2, 60)
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _status_text(status: str) -> str:
    return f"{status_label(status)} ({status})" if status in {"PASS", "FAIL"} else status


MANDATORY_LAYER_ORDER = (
    "Page/action reality",
    "Data Layer API Call",
    "GTM Tags",
    "Browser request",
    "Surrounding behavior",
)


def _operational_layer(row: dict[str, Any]) -> str:
    check = str(row.get("target", {}).get("check") or "")
    if not check and row.get("domain") == "source":
        return "Data Layer API Call"
    mapping = {
        "binding": "Page/action reality",
        "valid_outcome": "Page/action reality",
        "page_value": "Page/action reality",
        "relationship": "Page/action reality",
        "settlement": "Page/action reality",
        "acquisition": "Page/action reality",
        "event_occurrence": "Data Layer API Call",
        "data_layer_state": "Data Layer state diagnostic",
        "resolved_variable": "GTM Variables diagnostic",
        "event_match": "GTM Tags",
        "tag_inventory": "GTM Tags",
        "in_scope_tag_discovery": "GTM Tags",
        "tag_configuration": "GTM Tags",
        "effective_mapping": "GTM Tags",
        "tag_firing": "GTM Tags",
        "runtime_parameter": "GTM Tags",
        "tag_consent": "Consent diagnostic",
        "event_time_consent": "Consent diagnostic",
        "request_parameter": "Browser request",
        "destination_request": "Browser request",
        "tag_request": "Browser request",
        "shared_send": "Browser request",
        "transport_failure": "Browser request",
        "surrounding_behavior": "Surrounding behavior",
        "execution_protocol": "Surrounding behavior",
        "semantic_finding": "Surrounding behavior",
        "sensitive_data": "Data safety",
    }
    domain_layers = {
        "reality": "Page/action reality",
        "source": "Data Layer API Call",
        "gtm": "GTM Tags",
        "delivery": "Browser request",
        "behavior": "Surrounding behavior",
        "safety": "Data safety",
    }
    return mapping.get(
        check,
        domain_layers.get(str(row.get("domain")), str(row.get("domain") or "Other")),
    )


def _layer_component(row: dict[str, Any]) -> str:
    check = str(row.get("target", {}).get("check") or "")
    if not check and row.get("domain") == "source":
        return "fields"
    groups = {
        "binding": "identity",
        "valid_outcome": "outcome",
        "page_value": "business reality",
        "relationship": "business reality",
        "settlement": "settlement",
        "acquisition": "acquisition",
        "event_occurrence": "occurrence",
        "event_match": "inventory",
        "tag_inventory": "inventory",
        "in_scope_tag_discovery": "inventory",
        "tag_configuration": "mapping",
        "effective_mapping": "mapping",
        "tag_firing": "firing",
        "runtime_parameter": "runtime",
        "request_parameter": "fields",
        "destination_request": "routing",
        "tag_request": "routing",
        "shared_send": "routing",
        "transport_failure": "transport",
        "surrounding_behavior": "chronology",
        "execution_protocol": "action protocol",
        "semantic_finding": "analyst finding",
        "sensitive_data": "safety",
    }
    return groups.get(check, check.replace("_", " ") or "checks")


def _layer_detail(rows: list[dict[str, Any]]) -> str:
    components: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        components[_layer_component(row)].append(row)
    component_text = []
    for component, values in components.items():
        applicable = [row for row in values if row.get("status") != "NOT_APPLICABLE"]
        status = worst_status(
            (str(row.get("status")) for row in applicable), default="NOT_APPLICABLE"
        )
        component_text.append(
            f"{component} {status} "
            f"{sum(row.get('status') == 'PASS' for row in applicable)}/{len(applicable)} passed"
        )
    exceptions: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row.get("status") in {"PASS", "NOT_APPLICABLE"}:
            continue
        key = (
            str(row.get("scenario_id") or "ordinary"),
            str(row.get("status") or "BLOCKED"),
            str(row.get("reason") or "No reason supplied."),
        )
        exceptions[key].append(row)
    reasons = []
    for (scenario_id, status, reason), affected in exceptions.items():
        paths = list(
            dict.fromkeys(
                str(row.get("target", {}).get("path") or row.get("inspection_target") or "check")
                for row in affected
            )
        )
        path_text = ", ".join(paths[:8])
        if len(paths) > 8:
            path_text += f", +{len(paths) - 8} more"
        value_text = ""
        if len(affected) == 1:
            observed = compact_reason(affected[0].get("observed"), 90)
            expected = compact_reason(affected[0].get("expected"), 90)
            if observed or expected:
                value_text = f" Observed: {observed or '-'}; expected: {expected or '-'} ."
        reasons.append(
            f"{len(affected)} {status} in {scenario_id} ({path_text}): "
            f"{compact_reason(reason, 180)}{value_text}"
        )
    return "; ".join(component_text) + ((" | " + " | ".join(reasons)) if reasons else "")


def operational_layer_summary(result: dict[str, Any]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    expected_layers: set[str] = set()
    for expected in result.get("expected_checks", []):
        if not isinstance(expected, dict):
            continue
        layer = _operational_layer(
            {
                "domain": expected.get("domain"),
                "target": {"check": expected.get("check")},
            }
        )
        expected_layers.add(layer)
    for row in result.get("inspections", []):
        layer = _operational_layer(row)
        if layer not in grouped:
            grouped[layer] = []
        grouped[layer].append(row)
    optional = [
        layer
        for layer in (
            "Data Layer state diagnostic",
            "GTM Variables diagnostic",
            "Consent diagnostic",
            "Data safety",
        )
        if layer in grouped or layer in expected_layers
    ]
    order = [*MANDATORY_LAYER_ORDER, *optional]
    output = []
    for layer in order:
        rows = grouped.get(layer, [])
        if not rows:
            expected = layer in expected_layers
            status = (
                "PENDING"
                if expected and not result.get("scenarios")
                else "BLOCKED"
                if expected
                else "NOT_APPLICABLE"
            )
            output.append(
                {
                    "layer": layer,
                    "status": status,
                    "checks": 0,
                    "passed": 0,
                    "detail": (
                        "No judgement row was produced for this expected layer."
                        if expected
                        else "The tracking plan creates no obligation on this layer."
                    ),
                    "check_next": ["Capture and judge the planned layer once"] if expected else [],
                    "evidence": [],
                }
            )
            continue
        status = worst_status((str(row.get("status")) for row in rows), default="NOT_APPLICABLE")
        if layer == "Data safety" and status in {"PASS", "NOT_APPLICABLE"}:
            continue
        output.append(
            {
                "layer": layer,
                "status": status,
                "checks": len(rows),
                "passed": sum(row.get("status") == "PASS" for row in rows),
                "detail": _layer_detail(rows),
                "check_next": list(
                    dict.fromkeys(
                        str(row.get("check_next")) for row in rows if row.get("check_next")
                    )
                )[:5],
                "evidence": list(
                    dict.fromkeys(
                        str(reference)
                        for row in rows
                        for reference in row.get("evidence", [])
                        if reference
                    )
                )[:12],
            }
        )
    for layer, gate in (
        ("Evidence confidence", result.get("gates", {}).get("evidence_confidence", {})),
        ("Scenario coverage", result.get("gates", {}).get("scenario_completeness", {})),
    ):
        if gate.get("status") in {"PASS", "NOT_APPLICABLE"}:
            continue
        output.append(
            {
                "layer": layer,
                "status": gate.get("status", "PENDING"),
                "checks": 1,
                "passed": int(gate.get("status") == "PASS"),
                "detail": gate.get("reason") or gate.get("rationale") or gate.get("errors"),
                "check_next": [],
                "evidence": [gate["record_id"]] if gate.get("record_id") else [],
            }
        )
    return output


def render_event_feedback(result: dict[str, Any], *, compact: bool = False) -> str:
    lines = [
        f"Event: {result.get('event_name') or result.get('event_id')} - {_status_text(result['status'])}",
        f"Final: {'yes' if result.get('final') else 'no'}",
        f"Why: {result.get('reason')}",
        (
            "Coverage: "
            + str(result.get("coverage", {}).get("mode") or "pending")
            + (" - complete" if result.get("coverage", {}).get("complete") else " - incomplete")
        ),
        "",
    ]
    scenarios = result.get("scenarios", [])
    if len(scenarios) > 1:
        lines.extend(
            [
                "| Scenario | Status | Actions |",
                "|---|---|---|",
            ]
        )
        for scenario in scenarios:
            lines.append(
                "| "
                + " | ".join(
                    [
                        markdown_safe(scenario.get("label")),
                        str(scenario.get("status")),
                        markdown_safe(scenario.get("action_ids", [])),
                    ]
                )
                + " |"
            )
        lines.append("")
    lines.extend(
        [
            "| Layer | Status | Checks | Detail | Check next | Evidence |",
            "|---|---|---|---|---|---|",
        ]
    )
    for row in operational_layer_summary(result):
        lines.append(
            "| "
            + " | ".join(
                markdown_safe(value)
                for value in (
                    row.get("layer"),
                    row.get("status"),
                    f"{row.get('passed', 0)}/{row.get('checks', 0)} passed",
                    row.get("detail"),
                    "; ".join(row.get("check_next", [])) or "-",
                    ", ".join(row.get("evidence", [])) or "-",
                )
            )
            + " |"
        )
    return "\n".join(lines).rstrip() + "\n"


def render_conclusion(result: dict[str, Any]) -> str:
    lines = [
        "# GTM Client Recette Conclusion",
        "",
        f"Run: `{result.get('run_id')}`",
        f"Overall status: **{_status_text(result.get('status', 'PENDING'))}**",
        "",
        "| Event | Reality | API Call | GTM Tags | Browser request | Surrounding behavior | Status | Why |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for event in result.get("events", []):
        layers = {row["layer"]: row for row in operational_layer_summary(event)}
        lines.append(
            "| "
            + " | ".join(
                markdown_safe(value)
                for value in (
                    event.get("event_name") or event.get("event_id"),
                    *[
                        layers.get(layer, {}).get("status", "NOT_APPLICABLE")
                        for layer in MANDATORY_LAYER_ORDER
                    ],
                    event.get("status"),
                    event.get("reason"),
                )
            )
            + " |"
        )
    lines.extend(["", "## Event details", ""])
    for event in result.get("events", []):
        lines.append(render_event_feedback(event).rstrip())
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def defect_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for event in result.get("events", []):
        for inspection in event.get("inspections", []):
            if inspection.get("status") in {"PASS", "NOT_APPLICABLE"}:
                continue
            rows.append(
                {
                    "event_id": event.get("event_id"),
                    "event_name": event.get("event_name"),
                    "scenario_id": inspection.get("scenario_id"),
                    "action_id": inspection.get("action_id"),
                    "domain": inspection.get("domain"),
                    "inspection_target": inspection.get("inspection_target"),
                    "status": inspection.get("status"),
                    "reason_code": inspection.get("reason_code"),
                    "reason": inspection.get("reason"),
                    "check_next": inspection.get("check_next"),
                    "evidence": inspection.get("evidence", []),
                }
            )
    return rows


def telemetry_view(plan: dict[str, Any], records: list[dict[str, Any]]) -> dict[str, Any]:
    kinds = defaultdict(int)
    for record in records:
        kinds[str(record.get("kind"))] += 1
    created = parse_iso_timestamp(plan.get("created_at"))
    first_action = next(
        (
            parse_iso_timestamp(record.get("recorded_at"))
            for record in records
            if record.get("kind") == "ACTION_BEGIN"
        ),
        None,
    )
    first_feedback = next(
        (
            parse_iso_timestamp(record.get("recorded_at"))
            for record in records
            if record.get("kind") == "EVENT_FEEDBACK_ISSUED"
        ),
        None,
    )

    def elapsed(end: Any) -> float | None:
        return round((end - created).total_seconds(), 3) if created and end else None

    operations: dict[str, int | None] = {key: None for key in OPERATION_COUNTERS}
    runtime: dict[str, Any] = {}
    milestones: dict[str, Any] = {}
    preview_events_observed = 0
    network_requests_observed = 0
    preview_deep_read_captures_observed = 0
    unique_preview: set[tuple[str, str]] = set()
    unique_network: set[tuple[str, str]] = set()
    for record in records:
        summary = record.get("data", {}).get("summary", {})
        if record.get("kind") == "CAPTURE_CAPABILITY" and isinstance(summary, dict):
            if isinstance(summary.get("runtime"), dict):
                runtime = summary["runtime"]
            if isinstance(summary.get("milestones"), dict):
                milestones = summary["milestones"]
        if record.get("kind") == "CAPTURE_PREVIEW" and isinstance(summary, dict):
            reference = record.get("data", {}).get("evidence_ref", {})
            identity = (str(reference.get("path") or ""), str(reference.get("sha256") or ""))
            if identity in unique_preview:
                continue
            unique_preview.add(identity)
            preview_events_observed += int(summary.get("event_count") or 0)
            events = summary.get("events", [])
            if isinstance(events, list) and any(
                isinstance(event, dict)
                and (
                    event.get("has_runtime_extract") is True
                    or event.get("has_data_layer_state") is True
                    or event.get("has_api_call") is True
                )
                for event in events
            ):
                preview_deep_read_captures_observed += 1
        if record.get("kind") == "CAPTURE_NETWORK" and isinstance(summary, dict):
            reference = record.get("data", {}).get("evidence_ref", {})
            identity = (str(reference.get("path") or ""), str(reference.get("sha256") or ""))
            if identity in unique_network:
                continue
            unique_network.add(identity)
            network_requests_observed += int(summary.get("request_count") or 0)
        if record.get("kind") != "CAPTURE_HEALTH":
            continue
        observed = summary.get("operations", {}) if isinstance(summary, dict) else {}
        if not isinstance(observed, dict):
            continue
        for key in OPERATION_COUNTERS:
            value = observed.get(key)
            if isinstance(value, int) and not isinstance(value, bool):
                operations[key] = value if operations[key] is None else max(operations[key], value)

    return {
        "first_action_seconds": elapsed(first_action),
        "first_feedback_seconds": elapsed(first_feedback),
        "browser_ready_seconds": elapsed(parse_iso_timestamp(milestones.get("browser_ready_at"))),
        "preview_ready_seconds": elapsed(parse_iso_timestamp(milestones.get("preview_ready_at"))),
        "runtime_provider": runtime.get("provider"),
        "runtime_version": runtime.get("mcp_version"),
        "browser_channel": runtime.get("browser_channel"),
        "actions": kinds["ACTION_BEGIN"],
        "commits": kinds["ACTION_COMMIT"],
        "preview_syncs": kinds["PREVIEW_SYNC"],
        "preview_captures": len(unique_preview),
        "network_captures": len(unique_network),
        "capability_probes": kinds["CAPTURE_CAPABILITY"],
        "feedbacks": kinds["EVENT_FEEDBACK_ISSUED"],
        "stream_records": len(records),
        "operation_counters_instrumented": any(value is not None for value in operations.values()),
        "preview_events_observed": preview_events_observed,
        "preview_deep_read_captures_observed": preview_deep_read_captures_observed,
        "network_requests_observed": network_requests_observed,
        **operations,
    }


def compact_status_view(value: dict[str, Any]) -> dict[str, Any]:
    """Return the user-facing event/layer checkpoint without canonical payload bulk."""
    if "event_id" in value:
        return {
            "event_id": value.get("event_id"),
            "event_name": value.get("event_name"),
            "status": value.get("status"),
            "final": value.get("final"),
            "reason": value.get("reason"),
            "domains": {
                key: {
                    "status": row.get("status"),
                    "reason": row.get("reason"),
                    "checks": row.get("checks"),
                }
                for key, row in value.get("domains", {}).items()
            },
            "layers": operational_layer_summary(value),
            "gates": {
                key: {
                    "status": row.get("status"),
                    "reason": row.get("reason", row.get("rationale")),
                }
                for key, row in value.get("gates", {}).items()
                if isinstance(row, dict)
            },
        }
    return {
        "schema_version": value.get("schema_version"),
        "run_id": value.get("run_id"),
        "status": value.get("status"),
        "finished": value.get("finished"),
        "counts": value.get("counts", {}),
        "events": [compact_status_view(event) for event in value.get("events", [])],
        "open_actions": value.get("open_actions", []),
        "warnings": value.get("warnings", []),
        "telemetry": value.get("telemetry", {}),
    }


def status_view(run_dir: Path | str) -> dict[str, Any]:
    plan = load_plan(run_dir)
    records, warnings = read_stream(run_dir)
    result = judge_run(run_dir, plan, records)
    return {
        **result,
        "warnings": warnings,
        "open_actions": [
            action["action_id"] for action in action_windows(records) if action["status"] == "OPEN"
        ],
        "telemetry": telemetry_view(plan, records),
    }


def _workbook(
    plan: dict[str, Any], result: dict[str, Any], records: list[dict[str, Any]]
) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append_row(summary, ["Metric", "Value"])
    for key, value in (
        ("Run ID", result.get("run_id")),
        ("Status", result.get("status")),
        ("Events", len(result.get("events", []))),
        ("Claims", plan.get("claim_count")),
        ("Source kind", plan.get("source", {}).get("kind")),
        ("Source SHA-256", plan.get("source", {}).get("sha256")),
    ):
        _append_row(summary, [key, value])

    events_sheet = workbook.create_sheet("Events")
    _append_row(
        events_sheet,
        [
            "Order",
            "Event ID",
            "Event",
            *MANDATORY_LAYER_ORDER,
            "Confidence",
            "Coverage",
            "Status",
            "Final",
            "Why",
        ],
    )
    order = {event["event_id"]: event.get("plan_order") for event in plan.get("events", [])}
    for event in result.get("events", []):
        layers = {row["layer"]: row for row in operational_layer_summary(event)}
        _append_row(
            events_sheet,
            [
                order.get(event.get("event_id")),
                event.get("event_id"),
                event.get("event_name"),
                *[
                    layers.get(layer, {}).get("status", "NOT_APPLICABLE")
                    for layer in MANDATORY_LAYER_ORDER
                ],
                event.get("gates", {}).get("evidence_confidence", {}).get("status"),
                event.get("coverage", {}).get("status"),
                event.get("status"),
                event.get("final"),
                event.get("reason"),
            ],
        )

    scenarios = workbook.create_sheet("Scenarios")
    _append_row(
        scenarios,
        [
            "Event ID",
            "Scenario ID",
            "Scenario",
            "Values",
            "Signature",
            *[DOMAIN_LABELS[d] for d in DOMAINS],
            "Status",
            "Actions",
        ],
    )
    for event in result.get("events", []):
        for scenario in event.get("scenarios", []):
            _append_row(
                scenarios,
                [
                    event.get("event_id"),
                    scenario.get("scenario_id"),
                    scenario.get("label"),
                    scenario.get("values"),
                    scenario.get("behavior_signature"),
                    *[
                        scenario.get("domains", {}).get(domain, {}).get("status")
                        for domain in DOMAINS
                    ],
                    scenario.get("status"),
                    scenario.get("action_ids"),
                ],
            )

    inspections = workbook.create_sheet("Inspections")
    _append_row(
        inspections,
        [
            "Event ID",
            "Scenario ID",
            "Action ID",
            "Domain",
            "Inspection target",
            "Status",
            "Observed",
            "Expected",
            "Reason code",
            "Reason",
            "Check next",
            "Evidence",
        ],
    )
    for event in result.get("events", []):
        for row in event.get("inspections", []):
            _append_row(
                inspections,
                [
                    event.get("event_id"),
                    row.get("scenario_id"),
                    row.get("action_id"),
                    row.get("domain"),
                    row.get("inspection_target"),
                    row.get("status"),
                    row.get("observed"),
                    row.get("expected"),
                    row.get("reason_code"),
                    row.get("reason"),
                    row.get("check_next"),
                    row.get("evidence"),
                ],
            )

    defects = workbook.create_sheet("Defects and limits")
    _append_row(
        defects,
        [
            "Event ID",
            "Scenario",
            "Action ID",
            "Domain",
            "Inspection target",
            "Status",
            "Reason code",
            "Reason",
            "Check next",
            "Evidence",
        ],
    )
    for row in defect_rows(result):
        _append_row(
            defects,
            [
                row.get("event_id"),
                row.get("scenario_id"),
                row.get("action_id"),
                row.get("domain"),
                row.get("inspection_target"),
                row.get("status"),
                row.get("reason_code"),
                row.get("reason"),
                row.get("check_next"),
                row.get("evidence"),
            ],
        )

    evidence = workbook.create_sheet("Evidence index")
    _append_row(
        evidence, ["Record ID", "Sequence", "Kind", "Provenance", "Recorded at", "Evidence path"]
    )
    for record in records:
        reference = record.get("data", {}).get("evidence_ref", {})
        _append_row(
            evidence,
            [
                record.get("record_id"),
                record.get("seq"),
                record.get("kind"),
                record.get("provenance"),
                record.get("recorded_at"),
                reference.get("path") if isinstance(reference, dict) else None,
            ],
        )

    telemetry = workbook.create_sheet("Telemetry")
    _append_row(telemetry, ["Metric", "Value"])
    for key, value in telemetry_view(plan, records).items():
        _append_row(telemetry, [key, value])
    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    return workbook


def validate_workbook(path: Path, expected_event_count: int) -> list[str]:
    errors = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as error:  # pragma: no cover - openpyxl supplies detailed errors
        return [f"Workbook cannot be opened: {error}"]
    try:
        required = {
            "Summary",
            "Events",
            "Scenarios",
            "Inspections",
            "Defects and limits",
            "Evidence index",
            "Telemetry",
        }
        missing = required - set(workbook.sheetnames)
        if missing:
            errors.append("Missing sheets: " + ", ".join(sorted(missing)))
        if (
            "Events" in workbook.sheetnames
            and workbook["Events"].max_row - 1 != expected_event_count
        ):
            errors.append("Events sheet row count differs from the plan.")
    finally:
        workbook.close()
    return errors


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = list(rows[0]) if rows else ["event_id", "status"]
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: formula_safe_text(
                        json.dumps(value, ensure_ascii=False, sort_keys=True)
                        if isinstance(value, (dict, list))
                        else value
                    )
                    for key, value in row.items()
                }
            )
    os.replace(temporary, path)


def _write_text_atomic(path: Path, value: str, *, encoding: str = "utf-8") -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(value, encoding=encoding, newline="\n")
    os.replace(temporary, path)


def build_reports(run_dir: Path | str) -> dict[str, str]:
    paths = RunPaths.at(run_dir)
    plan = load_plan(run_dir)
    records, warnings = read_stream(run_dir)
    if warnings:
        raise StateError("Cannot build reports from an incomplete stream tail.")
    if not any(record.get("kind") == "RUN_FINISHED" for record in records) or any(
        record.get("kind") == "RUN_REOPENED"
        for record in records[
            1
            + max(
                (
                    index
                    for index, record in enumerate(records)
                    if record.get("kind") == "RUN_FINISHED"
                ),
                default=-1,
            ) :
        ]
    ):
        raise StateError("Final reports can only be built from a currently finished run.")
    result = judge_run(run_dir, plan, records)
    paths.reports.mkdir(exist_ok=True)
    outputs = {
        "json": paths.reports / "results.json",
        "markdown": paths.reports / "conclusion.md",
        "xlsx": paths.reports / "results.xlsx",
        "defects_csv": paths.reports / "defects.csv",
    }
    _write_text_atomic(outputs["json"], json.dumps(result, ensure_ascii=False, indent=2) + "\n")
    _write_text_atomic(outputs["markdown"], render_conclusion(result))
    _write_csv(outputs["defects_csv"], defect_rows(result))
    workbook = _workbook(plan, result, records)
    temporary = outputs["xlsx"].with_name(outputs["xlsx"].stem + ".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, outputs["xlsx"])
    errors = validate_workbook(outputs["xlsx"], plan.get("event_count", 0))
    if errors:
        raise StateError("Generated workbook failed validation: " + " | ".join(errors))
    return {key: str(path) for key, path in outputs.items()}
