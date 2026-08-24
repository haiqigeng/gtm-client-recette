"""Staged, source-preserving tracking-plan compiler.

The compiler emits typed proof obligations and records event-local errors instead of
building run-wide cases or allowing later malformed rows to block the first valid event.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import OrderedDict
from collections.abc import Callable
from pathlib import Path
from typing import Any
from uuid import uuid4

import yaml
from openpyxl import load_workbook

from client_side_rules import valid_path

from .constants import (
    ARCHETYPE_DOMAIN,
    CLAIM_ARCHETYPES,
    SCHEMA_VERSION,
    utc_now,
)
from .predicates import PredicateError, compile_predicate
from .state import StateError, canonical_json, write_immutable_plan

HEADER_ALIASES = {
    "event_name": {"event", "event_name", "nom_evenement", "nom_événement"},
    "event_id": {"event_id", "event_group_id", "groupe_evenement", "groupe_événement"},
    "event_label": {"event_label", "label", "description"},
    "requirement_id": {"requirement_id", "id", "requirement", "exigence"},
    "claim_type": {"claim_type", "archetype", "type_requete", "type_exigence"},
    "field_path": {
        "field_path",
        "data_layer_path",
        "datalayer_path",
        "parameter",
        "parametre",
        "paramètre",
    },
    "expected_value": {"expected_value", "value", "valeur_attendue"},
    "allowed_values": {"allowed_values", "enum", "valeurs_autorisees", "valeurs_autorisées"},
    "expected_type": {"expected_type", "json_type", "type_json"},
    "match_rule": {"match_rule", "operator", "rule", "regle", "règle"},
    "occurrence": {"occurrence", "expected_occurrence", "count", "nombre"},
    "action": {"action", "trigger", "interaction", "declencheur", "déclencheur"},
    "url": {"url", "page", "route"},
    "locale": {"locale", "language", "langue"},
    "scenario": {"scenario", "case", "cas"},
    "tag": {"tag", "tag_name", "balise"},
    "destination": {"destination", "measurement_id", "pixel_id"},
    "source_mechanism": {"source_mechanism", "source", "mecanisme_source"},
    "resolved_path": {"resolved_path", "gtm_variable", "variable_gtm"},
    "runtime_path": {"runtime_path", "tag_parameter", "parametre_balise"},
    "request_path": {"request_path", "request_parameter", "parametre_requete"},
    "configuration_path": {"configuration_path", "tag_configuration_path"},
    "negative": {"negative", "expected_absence", "absence_attendue"},
    "condition": {"condition", "applicability", "applicabilite", "applicabilité"},
    "mode": {"mode", "event_mode", "measurement_mode"},
    "source_event_name": {"source_event_name", "source_event", "raw_event_name"},
    "delivery_event_name": {
        "delivery_event_name",
        "destination_event_name",
        "ga4_event_name",
        "send_event_name",
    },
    "source_only": {"source_only", "source_seulement"},
    "forwarding_required": {"forwarding_required", "send_required", "delivery_required"},
    "notes": {"notes", "comment", "commentaire", "example", "exemple"},
}

EVENT_METADATA_FIELDS = {
    "mode",
    "source_event_name",
    "delivery_event_name",
    "source_only",
    "forwarding_required",
}
TABULAR_REQUIREMENT_FIELDS = frozenset(HEADER_ALIASES) - {
    "event_name",
    "event_id",
    "event_label",
    "notes",
    *EVENT_METADATA_FIELDS,
}

SECTION_EVENT_HEADERS = {
    "name_of_the_event",
    "event_name",
    "nom_de_l_evenement",
    "nom_evenement",
    "nature",
}
SECTION_FIELD_HEADERS = {
    "variable",
    "variables",
    "parameter",
    "parameters",
    "dimension",
    "dimensions",
}
SECTION_TYPE_HEADERS = {"type", "json_type", "format"}
SECTION_STATUS_HEADERS = {"status", "statut", "required", "requiredness"}
SECTION_VALUE_HEADERS = {"value", "values", "valeur", "valeurs"}
SECTION_SUMMARY_HEADERS = {
    "summary",
    "description",
    "definition",
    "définition",
    "example",
    "examples",
}
SECTION_STOP_MARKERS = {"code", "images", "image", "screenshots", "screenshot"}
SECTION_EXACT_SINGLETON_FIELDS = {"event", "event_name", "action", "checkout_step"}
GA4_ECOMMERCE_EVENTS = {
    "view_promotion",
    "select_promotion",
    "view_item_list",
    "select_item",
    "view_item",
    "add_to_wishlist",
    "add_to_cart",
    "remove_from_cart",
    "view_cart",
    "begin_checkout",
    "add_shipping_info",
    "add_payment_info",
    "purchase",
    "refund",
}
GA4_ITEM_FIELDS = {
    "item_id",
    "item_name",
    "item_brand",
    "item_category",
    "item_category2",
    "item_category3",
    "item_category4",
    "item_category5",
    "item_variant",
    "item_list_id",
    "item_list_name",
    "index",
    "price",
    "quantity",
    "discount",
    "coupon",
    "affiliation",
    "location_id",
    "creative_name",
    "creative_slot",
    "promotion_id",
    "promotion_name",
}

ARCHETYPE_ALIASES = {
    "page": "reality",
    "business": "reality",
    "business_state": "reality",
    "source_state": "source",
    "source_event": "source",
    "tag": "gtm",
    "gtm_tag": "gtm",
    "request": "delivery",
    "destination": "delivery",
    "anomaly": "sequence",
    "behavior": "sequence",
    "behaviour": "sequence",
    "privacy": "safety",
}


def _slug(value: Any, fallback: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(value or "").strip()).strip("-")
    return text[:80] or fallback


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _list_value(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if value in (None, ""):
        return []
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("["):
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError as error:
                raise StateError("allowed_values contains invalid JSON array syntax.") from error
            if not isinstance(parsed, list):
                raise StateError("allowed_values JSON must be an array.")
            return parsed
        separator = "|" if "|" in stripped else ","
        return [item.strip() for item in stripped.split(separator) if item.strip()]
    return [value]


def _header_key(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", "_", ascii_value.casefold()).strip("_")


def _headers(row: list[Any]) -> dict[int, str]:
    aliases = {
        _header_key(alias): target for target, values in HEADER_ALIASES.items() for alias in values
    }
    return {
        index: aliases[key]
        for index, value in enumerate(row)
        if (key := _header_key(value)) in aliases
    }


def _header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]] | None:
    for index, row in enumerate(rows[:60]):
        mapping = _headers(list(row))
        if "event_name" in mapping.values() and (
            "field_path" in mapping.values()
            or "claim_type" in mapping.values()
            or "requirement_id" in mapping.values()
        ):
            return index, mapping
    return None


def _has_cell_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _tabular_expected_value(value: Any, expected_type: Any) -> Any:
    """Preserve spreadsheet strings unless their declared JSON type requires parsing."""
    normalized_type = str(expected_type or "").strip().casefold()
    if not isinstance(value, str):
        if normalized_type == "string" and value is not None:
            return str(value)
        return value
    stripped = value.strip()
    if normalized_type in {"boolean", "number", "integer", "null", "array", "object"}:
        try:
            parsed = json.loads(stripped)
        except json.JSONDecodeError as error:
            raise StateError(
                f"Expected value {value!r} is not valid JSON for declared type {normalized_type}."
            ) from error
        actual_type = (
            "boolean"
            if isinstance(parsed, bool)
            else "integer"
            if isinstance(parsed, int) and not isinstance(parsed, bool)
            else "number"
            if isinstance(parsed, float)
            else "null"
            if parsed is None
            else "array"
            if isinstance(parsed, list)
            else "object"
            if isinstance(parsed, dict)
            else "string"
        )
        compatible = actual_type == normalized_type or (
            normalized_type == "number" and actual_type == "integer"
        )
        if not compatible:
            raise StateError(
                f"Expected value {value!r} does not match declared type {normalized_type}."
            )
        return parsed
    return value


def _tabular_boolean(value: Any, *, field: str, source: str) -> bool | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().casefold()
    if normalized in {"1", "true", "yes", "oui"}:
        return True
    if normalized in {"0", "false", "no", "non"}:
        return False
    raise StateError(f"{source}: {field} must be an explicit true/false value.")


def _canonical_event_name(value: Any) -> str:
    text = " ".join(str(value or "").split())
    if not text:
        return ""
    if re.fullmatch(r"[A-Za-z0-9_.-]+", text):
        return text
    return re.sub(r"[^a-z0-9]+", "_", text.casefold()).strip("_")


def _canonical_field_name(value: Any) -> str:
    original = " ".join(str(value or "").split())
    if valid_path(original):
        return original
    text = unicodedata.normalize("NFKD", original)
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-zA-Z0-9_]+", "_", text).strip("_").casefold()


def _section_field_path(event_name: str, field_name: str) -> str:
    if "." in field_name or "[" in field_name:
        return field_name
    if event_name in GA4_ECOMMERCE_EVENTS:
        if field_name in GA4_ITEM_FIELDS:
            return f"ecommerce.items[].{field_name}"
        if field_name != "event":
            return f"ecommerce.{field_name}"
    return field_name


def _finite_declared_values(value: Any) -> list[str]:
    """Accept only clearly exhaustive, small pipe-separated enums from prose workbooks."""
    if not isinstance(value, str) or "|" not in value:
        return []
    candidates = [item.strip() for item in value.split("|") if item.strip()]
    if not 2 <= len(candidates) <= 20:
        return []
    if any(
        len(item) > 80
        or "..." in item
        or item.casefold() in {"xxx", "n/a", "na", "etc"}
        or "{{" in item
        for item in candidates
    ):
        return []
    return candidates


def _section_declared_predicate(
    value: Any, expected_type: str | None, field_name: str, event_name: str
) -> dict[str, Any]:
    """Interpret only unambiguous constants/enums; leave examples and identities dynamic."""
    allowed_values = _finite_declared_values(value)
    if allowed_values:
        return {"match_rule": "one_of", "allowed_values": allowed_values}

    if field_name == "event" and event_name:
        return {"match_rule": "equals", "expected_value": event_name}
    if not _has_cell_value(value):
        return {"match_rule": "present"}

    text = value.strip() if isinstance(value, str) else ""
    dynamic = bool(
        text
        and re.search(
            r"(?:\.\.\.|…|\b(?:xxx|example|sample|dynamic|variable|n/?a|etc)\b|\{\{|<[^>]+>)",
            text,
            re.IGNORECASE,
        )
    )
    ambiguous_list = bool(text and ("|" in text or re.search(r"\s(?:-|/)\s", text) or "," in text))
    if dynamic or ambiguous_list:
        return {"match_rule": "present"}

    normalized_type = str(expected_type or "").strip().casefold()
    if field_name not in SECTION_EXACT_SINGLETON_FIELDS and normalized_type not in {
        "boolean",
        "null",
    }:
        # A generic "Values" column often contains one illustrative member, not a
        # global oracle for products, amounts, labels, methods, or other contextual
        # values. Flat plans can declare expected_value explicitly; finite pipes remain
        # strict enums above.
        return {"match_rule": "present"}

    parsed = _tabular_expected_value(value, expected_type)
    if normalized_type in {"boolean", "number", "integer", "null"}:
        return {"match_rule": "equals", "expected_value": parsed}
    if not isinstance(parsed, str):
        return {"match_rule": "present"}

    if len(text) <= 120:
        return {"match_rule": "equals", "expected_value": parsed.strip()}
    return {"match_rule": "present"}


def _sectioned_rows(
    rows: list[tuple[Any, ...]], sheet_title: str
) -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    """Parse the common two-block plan sheet: event metadata, then a variable table."""
    event_header_index = None
    event_column = None
    for row_index, row in enumerate(rows[:40]):
        for column, value in enumerate(row):
            if _header_key(value) in SECTION_EVENT_HEADERS:
                event_header_index, event_column = row_index, column
                break
        if event_header_index is not None:
            break
    if event_header_index is None or event_column is None:
        return None

    event_row_index = next(
        (
            index
            for index in range(event_header_index + 1, min(len(rows), event_header_index + 8))
            if event_column < len(rows[index]) and _has_cell_value(rows[index][event_column])
        ),
        None,
    )
    if event_row_index is None:
        return None
    event_label = " ".join(str(rows[event_row_index][event_column]).split())
    state_only = _header_key(event_label) in {
        "core_datalayer",
        "core_data_layer",
        "core_state",
    } or _header_key(sheet_title) in {"core_datalayer", "core_data_layer", "core_state"}
    event_name = "" if state_only else _canonical_event_name(event_label)

    variable_header_index = None
    column_map: dict[str, int] = {}
    for row_index in range(event_row_index + 1, min(len(rows), event_row_index + 30)):
        normalized = [_header_key(value) for value in rows[row_index]]
        field_columns = [
            index for index, key in enumerate(normalized) if key in SECTION_FIELD_HEADERS
        ]
        if not field_columns or not any(key in SECTION_TYPE_HEADERS for key in normalized):
            continue
        variable_header_index = row_index
        column_map["field_path"] = field_columns[0]
        for index, key in enumerate(normalized):
            if key in SECTION_TYPE_HEADERS:
                column_map.setdefault("expected_type", index)
            elif key in SECTION_STATUS_HEADERS:
                column_map.setdefault("requiredness", index)
            elif key in SECTION_VALUE_HEADERS:
                column_map.setdefault("declared_values", index)
            elif key in SECTION_SUMMARY_HEADERS:
                column_map.setdefault("summary", index)
        break
    if variable_header_index is None:
        return None

    output: list[dict[str, Any]] = []
    ignored: list[dict[str, str]] = []
    for row_index in range(variable_header_index + 1, len(rows)):
        row = rows[row_index]
        field_column = column_map["field_path"]
        raw_field = row[field_column] if field_column < len(row) else None
        if not _has_cell_value(raw_field):
            if any(_has_cell_value(value) for value in row):
                source = f"{sheet_title}!{row_index + 1}"
                reason = (
                    "code_example"
                    if any(
                        re.search(
                            r"(?:data\s*layer\s*\.\s*push|window\s*\.\s*data\s*layer|<script|```)",
                            str(value),
                            re.I,
                        )
                        for value in row
                        if _has_cell_value(value)
                    )
                    else "non_requirement_row"
                )
                ignored.append({"source": source, "reason": reason})
            continue
        normalized_field = _header_key(raw_field)
        if normalized_field in SECTION_STOP_MARKERS:
            break
        source = f"{sheet_title}!{row_index + 1}"
        if any(
            re.search(
                r"(?:data\s*layer\s*\.\s*push|window\s*\.\s*data\s*layer|<script|```)",
                str(value),
                re.I,
            )
            for value in row
            if _has_cell_value(value)
        ):
            ignored.append({"source": source, "reason": "code_example"})
            continue
        field_name = _canonical_field_name(raw_field)
        if not field_name:
            ignored.append({"source": source, "reason": "empty_field_name"})
            continue
        expected_type = None
        if "expected_type" in column_map and column_map["expected_type"] < len(row):
            expected_type = str(row[column_map["expected_type"]] or "").strip().casefold()
            if expected_type in {"numer", "numeric", "float", "double"}:
                expected_type = "number"
            if expected_type in {"str", "text", "texte"}:
                expected_type = "string"
        requiredness = (
            row[column_map["requiredness"]]
            if "requiredness" in column_map and column_map["requiredness"] < len(row)
            else None
        )
        declared_values = (
            row[column_map["declared_values"]]
            if "declared_values" in column_map and column_map["declared_values"] < len(row)
            else None
        )
        summary = (
            row[column_map["summary"]]
            if "summary" in column_map and column_map["summary"] < len(row)
            else None
        )
        declared_predicate = _section_declared_predicate(
            declared_values,
            expected_type,
            field_name,
            event_name,
        )
        output.append(
            {
                "event_name": event_name or event_label,
                "event_id": _slug(sheet_title, f"event-{event_row_index + 1}"),
                "event_label": event_label,
                "mode": "state_only" if state_only else "named_event",
                "delivery_event_name": "page_view" if state_only else None,
                "forwarding_required": True if state_only else None,
                "field_path": _section_field_path(event_name, field_name),
                **declared_predicate,
                "expected_type": expected_type or None,
                "action": (
                    rows[event_row_index][event_column + 2]
                    if event_column + 2 < len(rows[event_row_index])
                    else None
                ),
                "notes": {
                    "requiredness": requiredness,
                    "summary": summary,
                    "declared_values": declared_values,
                },
                "_source": source,
            }
        )
    if not output:
        return None
    return output, {
        "rows_seen": len(output) + len(ignored),
        "requirements_compiled": len(output),
        "carried_event_rows": max(0, len(output) - 1),
        "rows_ignored": len(ignored),
        "ignored": ignored,
        "layout": "event_metadata_plus_variable_table",
    }


def _index_event_names(rows: list[tuple[Any, ...]]) -> list[str]:
    """Read only a simple event-name index; never interpret prose as requirements."""
    for row_index, row in enumerate(rows[:60]):
        mapping = _headers(list(row))
        event_columns = [index for index, name in mapping.items() if name == "event_name"]
        if not event_columns:
            continue
        if any(name in TABULAR_REQUIREMENT_FIELDS for name in mapping.values()):
            return []
        column = event_columns[0]
        output: list[str] = []
        for candidate in rows[row_index + 1 :]:
            value = candidate[column] if column < len(candidate) else None
            if _has_cell_value(value):
                normalized = " ".join(str(value).split())
                if normalized not in output:
                    output.append(normalized)
        return output
    return []


def _sheet_kind(rows: list[tuple[Any, ...]], title: str) -> tuple[str, str, list[str]]:
    nonempty = [row for row in rows if any(_has_cell_value(value) for value in row)]
    if not nonempty:
        return "empty", "sheet_has_no_values", []
    index_events = _index_event_names(rows)
    if index_events:
        return "index", "simple_event_index", index_events
    searchable = " ".join(str(value) for row in nonempty[:80] for value in row if value)
    title_key = _header_key(title)
    if re.search(r"(?:data\s*layer\s*\.\s*push|<script|```)", searchable, re.I) or any(
        token in title_key for token in ("code", "example", "exemple")
    ):
        return "example", "code_or_example_content", []
    if any(
        token in title_key
        for token in ("reference", "referentiel", "catalog", "dimension", "category", "values")
    ):
        return "reference", "reference_content_not_requirements", []
    return "ignored", "unrecognized_nonempty_layout", []


def _identity_key(value: Any) -> str:
    return _header_key(value)


def _tabular_rows(
    rows: list[tuple[Any, ...]],
    header_index: int,
    mapping: dict[int, str],
    reference_for_row: Callable[[int], str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Read one recognized table without silently dropping continuation rows.

    Tracking workbooks commonly merge or visually fill down event identity cells. A
    non-empty requirement row may inherit the immediately preceding event identity,
    but never across a blank separator. Ambiguous orphan rows fail at intake instead of
    producing an incomplete plan that could later pass falsely.
    """

    output: list[dict[str, Any]] = []
    ignored: list[dict[str, str]] = []
    current_event_name: str | None = None
    current_event_id: str | None = None
    carried_event_rows = 0
    rows_seen = 0

    for source_row, raw in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row = {name: raw[index] for index, name in mapping.items() if index < len(raw)}
        if not any(_has_cell_value(value) for value in row.values()):
            current_event_name = None
            current_event_id = None
            continue

        rows_seen += 1
        reference = reference_for_row(source_row)
        event_name = str(row.get("event_name") or "").strip()
        event_id = str(row.get("event_id") or "").strip()
        has_requirement = any(
            _has_cell_value(row.get(field)) for field in TABULAR_REQUIREMENT_FIELDS
        )

        if event_name:
            current_event_name = event_name
            current_event_id = event_id or None
        elif has_requirement:
            if current_event_name is None:
                raise StateError(
                    f"{reference} contains requirement data before an event name; "
                    "the row cannot be assigned safely."
                )
            if event_id and event_id != current_event_id:
                raise StateError(
                    f"{reference} introduces event group {event_id!r} without an event name; "
                    "the row cannot be assigned safely."
                )
            row["event_name"] = current_event_name
            if current_event_id:
                row["event_id"] = current_event_id
            carried_event_rows += 1
        elif _has_cell_value(row.get("notes")):
            ignored.append({"source": reference, "reason": "notes_only"})
            continue
        else:
            raise StateError(
                f"{reference} contains tabular data but no event name or requirement fields."
            )

        row["_source"] = reference
        output.append(row)

    return output, {
        "rows_seen": rows_seen,
        "requirements_compiled": len(output),
        "carried_event_rows": carried_event_rows,
        "rows_ignored": len(ignored),
        "ignored": ignored,
    }


def _rows_from_xlsx(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: list[dict[str, Any]] = []
    tables: list[str] = []
    ignored_sheets: list[str] = []
    sheet_manifest: list[dict[str, Any]] = []
    indexed_events: list[str] = []
    diagnostics = {
        "rows_seen": 0,
        "requirements_compiled": 0,
        "carried_event_rows": 0,
        "rows_ignored": 0,
        "ignored": [],
    }
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header = _header_row(rows)
            if header is None:
                sectioned = _sectioned_rows(rows, sheet.title)
                if sectioned is None:
                    classification, reason, events = _sheet_kind(rows, sheet.title)
                    ignored_sheets.append(sheet.title)
                    indexed_events.extend(event for event in events if event not in indexed_events)
                    sheet_manifest.append(
                        {
                            "sheet": sheet.title,
                            "classification": classification,
                            "reason": reason,
                            "nonempty_rows": sum(
                                any(_has_cell_value(value) for value in row) for row in rows
                            ),
                            "indexed_events": events,
                        }
                    )
                    continue
                table_rows, table_diagnostics = sectioned
            else:
                header_index, mapping = header
                table_rows, table_diagnostics = _tabular_rows(
                    rows,
                    header_index,
                    mapping,
                    lambda excel_row, title=sheet.title: f"{title}!{excel_row}",
                )
            tables.append(sheet.title)
            sheet_manifest.append(
                {
                    "sheet": sheet.title,
                    "classification": "requirements",
                    "reason": table_diagnostics.get("layout", "recognized_requirement_table"),
                    "nonempty_rows": sum(
                        any(_has_cell_value(value) for value in row) for row in rows
                    ),
                    "requirements_compiled": len(table_rows),
                }
            )
            output.extend(table_rows)
            for key in (
                "rows_seen",
                "requirements_compiled",
                "carried_event_rows",
                "rows_ignored",
            ):
                diagnostics[key] += table_diagnostics[key]
            diagnostics["ignored"].extend(table_diagnostics["ignored"])
    finally:
        workbook.close()
    if not output:
        raise StateError(
            "No worksheet contained an event column plus a requirement, field, or claim column."
        )
    compiled_identities = {
        _identity_key(value)
        for row in output
        for value in (row.get("event_name"), row.get("event_label"), row.get("event_id"))
        if _has_cell_value(value)
    }
    indexed_by_key = {_identity_key(value): value for value in indexed_events}
    index_only = [value for key, value in indexed_by_key.items() if key not in compiled_identities]
    detail_only = sorted(
        {
            str(row.get("event_label") or row.get("event_name") or row.get("event_id"))
            for row in output
            if indexed_events
            and not any(
                _identity_key(value) in indexed_by_key
                for value in (row.get("event_name"), row.get("event_label"), row.get("event_id"))
                if _has_cell_value(value)
            )
        }
    )
    return output, {
        "format": "xlsx",
        "tables": tables,
        "ignored_sheets": ignored_sheets,
        "sheet_manifest": sheet_manifest,
        "reconciliation": {
            "indexed_events": indexed_events,
            "index_only_events": index_only,
            "detail_only_events": detail_only,
            "status": "REVIEW" if index_only or detail_only else "PASS",
        },
        **diagnostics,
    }


def _rows_from_delimited(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sample = path.read_text(encoding="utf-8-sig")
    try:
        dialect = csv.Sniffer().sniff(sample[:8192], delimiters=",;\t")
    except csv.Error as error:
        raise StateError(f"Cannot detect the tracking-plan delimiter: {error}") from error
    values = list(csv.reader(sample.splitlines(), dialect))
    header = _header_row([tuple(row) for row in values])
    if header is None:
        raise StateError("No event column plus a requirement, field, or claim column was found.")
    header_index, mapping = header
    output, diagnostics = _tabular_rows(
        [tuple(row) for row in values],
        header_index,
        mapping,
        lambda line: f"line {line}",
    )
    if not output:
        raise StateError("The tracking-plan table contains no compilable requirement rows.")
    return output, {"format": "delimited", "tables": ["file"], **diagnostics}


def _requirements_from_rows(
    rows: list[dict[str, Any]], diagnostics: dict[str, Any]
) -> dict[str, Any]:
    requirements = []
    for index, row in enumerate(rows, start=1):
        event_name = str(row.get("event_name") or "").strip()
        source = str(row.get("_source") or f"requirement {index}")
        expected = _tabular_expected_value(row.get("expected_value"), row.get("expected_type"))
        allowed = _list_value(row.get("allowed_values"))
        has_expected = _has_cell_value(row.get("expected_value"))
        rule = str(
            row.get("match_rule")
            or ("one_of" if allowed else "equals" if has_expected else "present")
        )
        expectation = {
            "event_name": event_name,
            "field_path": str(row.get("field_path") or "event").strip(),
            "match_rule": rule,
            "expected_value": expected,
            "expected_type": row.get("expected_type"),
            "expected_occurrence": row.get("occurrence") or "once_per_action",
            "source_mechanism": row.get("source_mechanism") or "data_layer_push",
        }
        for key in (
            "claim_type",
            "resolved_path",
            "runtime_path",
            "request_path",
            "configuration_path",
            "condition",
            "delivery_event_name",
            "forwarding_required",
            "source_only",
        ):
            if row.get(key) not in (None, ""):
                expectation[key] = row[key]
        if allowed:
            expectation["allowed_values"] = allowed
        negative = _tabular_boolean(row.get("negative"), field="negative", source=source)
        source_only = _tabular_boolean(row.get("source_only"), field="source_only", source=source)
        forwarding_required = _tabular_boolean(
            row.get("forwarding_required"), field="forwarding_required", source=source
        )
        if source_only is not None:
            expectation["source_only"] = source_only
        if forwarding_required is not None:
            expectation["forwarding_required"] = forwarding_required
        requirements.append(
            {
                "requirement_id": str(row.get("requirement_id") or f"REQ-{index:04d}"),
                "event_group_id": str(row.get("event_id") or event_name),
                "event_label": row.get("event_label"),
                "source": {"reference": row.get("_source"), "plan_order": index},
                "journey": {
                    "action": row.get("action"),
                    "url": row.get("url"),
                    "locale": row.get("locale"),
                },
                "expectation": expectation,
                "scenario": row.get("scenario"),
                "tag": row.get("tag"),
                "destination": row.get("destination"),
                "negative": negative is True,
                "mode": row.get("mode"),
                "source_event_name": row.get("source_event_name"),
                "delivery_event_name": row.get("delivery_event_name"),
                "source_only": source_only,
                "forwarding_required": forwarding_required,
            }
        )
    return {"requirements": requirements, "_normalization": diagnostics}


def _load_source(source: Path) -> tuple[dict[str, Any], str, str]:
    if source.is_dir():
        from import_ga4_tracking_plan_handoff import interpreted_requirements, verify_delivery

        handoff, plan, expected = verify_delivery(source)
        value = interpreted_requirements(handoff, plan, expected)
        return value, "ga4_tracking_plan_handoff", str(value.get("source_plan_sha256"))
    suffix = source.suffix.lower()
    if suffix == ".json":
        try:
            value = json.loads(source.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError) as error:
            raise StateError(f"Cannot read tracking plan: {error}") from error
        kind = "json"
    elif suffix in {".yaml", ".yml"}:
        try:
            value = yaml.safe_load(source.read_text(encoding="utf-8-sig"))
        except (OSError, yaml.YAMLError) as error:
            raise StateError(f"Cannot read tracking plan: {error}") from error
        kind = "yaml"
    elif suffix in {".xlsx", ".xlsm"}:
        rows, diagnostics = _rows_from_xlsx(source)
        return _requirements_from_rows(rows, diagnostics), "xlsx", _sha256(source)
    elif suffix in {".csv", ".tsv"}:
        rows, diagnostics = _rows_from_delimited(source)
        return _requirements_from_rows(rows, diagnostics), "delimited", _sha256(source)
    else:
        raise StateError("Tracking plan must be JSON, YAML, XLSX, CSV, TSV, or a handoff folder.")
    if not isinstance(value, dict):
        raise StateError("Tracking-plan root must be an object.")
    return value, kind, _sha256(source)


def _normalize_scope(scope: dict[str, Any] | None) -> dict[str, Any]:
    value = dict(scope or {})
    if value.get("approved") is not True:
        raise StateError(
            "Scope must contain approved=true after the user accepts the test boundary."
        )
    origins = value.get("origins", [])
    if isinstance(origins, str):
        origins = [origins]
    if not isinstance(origins, list):
        raise StateError("Scope origins must be an array when supplied.")
    value["origins"] = [str(item).strip() for item in origins if str(item).strip()]
    value["origin_mode"] = "explicit" if value["origins"] else "prepared_runtime"
    for key in ("expected_container", "destination", "tag_scope"):
        item = value.get(key, [])
        if isinstance(item, str):
            item = [item]
        value[key] = [str(entry).strip() for entry in item if str(entry).strip()]

    scope_aliases = {
        "ga4": "GA4",
        "ga4only": "GA4",
        "ga4tag": "GA4",
        "ga4tags": "GA4",
        "ga4tagonly": "GA4",
        "ga4tagsonly": "GA4",
        "googleanalytics": "GA4",
        "googleanalyticstag": "GA4",
        "googleanalyticstags": "GA4",
        "googleanalytics4": "GA4",
        "googleanalytics4only": "GA4",
        "googleanalytics4tag": "GA4",
        "googleanalytics4tags": "GA4",
        "googleanalytics4tagsonly": "GA4",
        "googletag": "GA4",
        "googletags": "GA4",
        "googleads": "Google Ads",
        "googleadsonly": "Google Ads",
        "googleadstag": "Google Ads",
        "googleadstags": "Google Ads",
        "googleadstagsonly": "Google Ads",
        "adwords": "Google Ads",
    }

    def plan_declared(item: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", item.casefold())
        return compact in {"plan", "trackingplan", "plandeclared", "allplanned"} or (
            "plan" in compact
            and any(token in compact for token in ("all", "every", "specified", "declared"))
        )

    declared_tag_scope = False
    normalized_tag_scope: list[str] = []
    for item in value["tag_scope"]:
        if plan_declared(item):
            declared_tag_scope = True
            continue
        key = re.sub(r"[^a-z0-9]+", "", item.casefold())
        canonical = scope_aliases.get(key)
        if canonical is None:
            # Preserve concise media/vendor categories such as "Meta Pixel". Reject
            # sentence-like scope prose because it cannot identify a tag deterministically.
            words = item.split()
            if len(words) > 4 or any(mark in item for mark in (";", ".", ":")):
                raise StateError(
                    "Tag scope must be a concise vendor category or the plan-declared "
                    "scope, not free-form prose."
                )
            canonical = item
        if canonical not in normalized_tag_scope:
            normalized_tag_scope.append(canonical)
    value["tag_scope"] = normalized_tag_scope
    value["tag_scope_mode"] = "explicit" if normalized_tag_scope else "plan_declared"
    value["tag_scope_declared"] = declared_tag_scope

    declared_destinations = False
    destination_categories: list[str] = []
    normalized_destinations: list[str] = []
    for item in value["destination"]:
        if plan_declared(item):
            declared_destinations = True
            continue
        compact = re.sub(r"[^a-z0-9]+", "", item.casefold())
        category = scope_aliases.get(compact)
        if category is not None:
            if category not in normalized_tag_scope:
                normalized_tag_scope.append(category)
            if category not in destination_categories:
                destination_categories.append(category)
            continue
        if any(character.isspace() for character in item):
            raise StateError(
                "Destination scope must contain exact destination IDs or the plan-declared "
                "scope, not free-form prose."
            )
        if item not in normalized_destinations:
            normalized_destinations.append(item)
    value["destination"] = normalized_destinations
    value["tag_scope"] = normalized_tag_scope
    if destination_categories and value.get("tag_scope_mode") == "plan_declared":
        value["tag_scope_mode"] = "explicit"
    value["destination_categories"] = destination_categories
    value["destination_mode"] = (
        "explicit"
        if normalized_destinations
        else "runtime_discovered"
        if destination_categories or normalized_tag_scope
        else "plan_declared"
    )
    value["destination_scope_declared"] = declared_destinations
    if not normalized_tag_scope:
        inferred = []
        if any(destination.upper().startswith("G-") for destination in normalized_destinations):
            inferred.append("GA4")
        if any(destination.upper().startswith("AW-") for destination in normalized_destinations):
            inferred.append("Google Ads")
        if inferred:
            value["tag_scope"] = inferred
            value["tag_scope_mode"] = "inferred_from_destination"
    if declared_tag_scope or declared_destinations:
        value["scope_resolution"] = "Generic plan wording was resolved to plan-declared identities."
    bad_containers = [item for item in value["expected_container"] if not item.startswith("GTM-")]
    bad_destinations = [item for item in value["destination"] if item.startswith("GTM-")]
    if bad_containers:
        raise StateError("Expected container IDs must use the GTM- identity type.")
    if bad_destinations:
        raise StateError("Destination IDs cannot use the GTM container identity type.")
    value.setdefault("certify_tags", True)
    value.setdefault("browser_send_required", bool(value["certify_tags"]))
    value.setdefault("delivery_mode", "gtm_browser" if value["certify_tags"] else "source_only")
    runtime = str(value.get("browser_runtime") or "playwright_mcp").strip().casefold()
    runtime_aliases = {
        "playwright": "playwright_mcp",
        "playwrightmcp": "playwright_mcp",
        "managed": "playwright_mcp",
        "existing": "existing_chromium",
        "existingbrowser": "existing_chromium",
        "existingchromium": "existing_chromium",
    }
    runtime = runtime_aliases.get(re.sub(r"[^a-z0-9]+", "", runtime), runtime)
    if runtime not in {"playwright_mcp", "existing_chromium"}:
        raise StateError("browser_runtime must be playwright_mcp or existing_chromium.")
    value["browser_runtime"] = runtime
    value.setdefault("browser_channel", "msedge" if runtime == "playwright_mcp" else "chromium")
    value.setdefault("browser_profile_mode", "persistent")
    value.setdefault("browser_headed", True)
    return value


def _normalize_tag(value: Any, index: int) -> dict[str, Any]:
    row = value if isinstance(value, dict) else {"tag_name": str(value)}
    identity = str(row.get("tag_id") or row.get("tag_name") or row.get("name") or "").strip()
    if not identity:
        identity = f"tag-{index}"
    expected = str(row.get("expected") or row.get("expectation") or "fire").casefold()
    if expected not in {"fire", "not_fire", "optional"}:
        expected = "fire"
    configuration = row.get("configuration", row.get("expected_configuration"))
    configured_destination = None
    if isinstance(configuration, dict):
        configured_destination = configuration.get(
            "measurement_id", configuration.get("destination")
        )
    return {
        "tag_id": identity,
        "tag_name": str(row.get("tag_name") or row.get("name") or identity),
        "category": str(row.get("category") or "").strip() or None,
        "expected": expected,
        "destination": row.get("destination") or configured_destination,
        "configuration": configuration,
        "consent_requirements": row.get("consent_requirements", []),
        "browser_send_required": row.get("browser_send_required"),
    }


def _tag_in_scope(tag: dict[str, Any], scope: dict[str, Any]) -> bool:
    """Apply an explicit tag-category scope without guessing unknown vendors."""
    requested = {
        re.sub(r"[^a-z0-9]+", "", str(value).casefold())
        for value in scope.get("tag_scope", [])
        if str(value).strip()
    }
    if not requested:
        return True
    searchable = re.sub(
        r"[^a-z0-9]+",
        "",
        " ".join(
            str(value or "")
            for value in (tag.get("category"), tag.get("tag_name"), tag.get("tag_id"))
        ).casefold(),
    )
    aliases = {
        "ga4": {
            "ga4",
            "googleanalytics",
            "googleanalytics4",
            "googleanalyticsevent",
            "googletag",
        },
        "googleads": {"googleads", "ads", "adwords"},
    }
    for item in requested:
        candidates = aliases.get(item, {item})
        if any(candidate and candidate in searchable for candidate in candidates):
            return True
    return False


def _tag_protocol(tag: dict[str, Any]) -> str | None:
    searchable = " ".join(
        str(value or "") for value in (tag.get("category"), tag.get("tag_name"), tag.get("tag_id"))
    ).casefold()
    if "ga4" in searchable or "analytics" in searchable:
        return "ga4"
    if "google ads" in searchable or "adwords" in searchable:
        return "google_ads"
    return None


def _scenario(value: Any, event_id: str, index: int, *, negative: bool = False) -> dict[str, Any]:
    row = value if isinstance(value, dict) else {"label": str(value)}
    is_negative = negative or row.get("negative") is True
    return {
        "scenario_id": str(row.get("scenario_id") or f"{_slug(event_id, 'event')}-S{index:02d}"),
        "label": str(row.get("label") or row.get("name") or f"Scenario {index}"),
        "values": row.get("values", {}),
        "explicit": True,
        "negative": is_negative,
        "role": "NEGATIVE" if is_negative else row.get("role"),
    }


class _ClaimBuilder:
    def __init__(self, event_id: str) -> None:
        self.event_id = event_id
        self.claims: list[dict[str, Any]] = []
        self._keys: dict[str, str] = {}
        self.warnings: list[str] = []

    def add(
        self,
        archetype: str,
        target: dict[str, Any],
        predicate: dict[str, Any],
        evidence: list[str],
        source: dict[str, Any],
        *,
        applicability: dict[str, Any] | None = None,
        label: str | None = None,
    ) -> None:
        if archetype not in CLAIM_ARCHETYPES:
            raise PredicateError(f"Unsupported claim archetype '{archetype}'.")
        core = {
            "archetype": archetype,
            "domain": ARCHETYPE_DOMAIN[archetype],
            "target": target,
            "predicate": predicate,
            "evidence": evidence,
            "applicability": applicability or {},
        }
        key = canonical_json(core)
        if key in self._keys:
            self.warnings.append(
                f"Duplicate claim at {source.get('reference') or 'unknown source'} matches {self._keys[key]}."
            )
            return
        claim_id = f"{_slug(self.event_id, 'event')}::C{len(self.claims) + 1:03d}"
        self._keys[key] = claim_id
        self.claims.append(
            {"claim_id": claim_id, **core, "source": source, "label": label or target.get("label")}
        )


def _occurrence_predicate(
    value: Any, *, negative: bool = False, location: str = "event"
) -> dict[str, Any]:
    if negative:
        return {"operator": "count", "exact": 0}
    if isinstance(value, int) and not isinstance(value, bool):
        if value < 0:
            raise PredicateError(f"{location}: expected occurrence cannot be negative.")
        return {"operator": "count", "exact": value}
    normalized = str(value or "once_per_action").strip().casefold()
    if normalized.isdigit():
        return {"operator": "count", "exact": int(normalized)}
    if normalized in {"none", "never", "absent", "zero"}:
        return {"operator": "count", "exact": 0}
    if normalized in {"one_or_more", "at_least_once"}:
        return {"operator": "count", "minimum": 1}
    if normalized in {"once", "once_per_action", "one", "exactly_once"}:
        return {"operator": "count", "exact": 1}
    raise PredicateError(f"{location}: unsupported expected occurrence '{value}'.")


def _requirement_source(row: dict[str, Any], index: int) -> dict[str, Any]:
    source = row.get("source") if isinstance(row.get("source"), dict) else {}
    return {
        **source,
        "reference": source.get("reference") or row.get("_source") or f"requirement {index}",
        "plan_order": source.get("plan_order", index),
        "requirement_id": str(row.get("requirement_id") or f"REQ-{index:04d}"),
    }


def _safe_path(value: Any, location: str) -> str:
    path = str(value or "").strip()
    if not path:
        raise PredicateError(f"{location}: field path is empty.")
    looks_like_example = bool(
        re.search(
            r"(?:^\s*(?:ex(?:ample|emple)?|code)\s*:|data\s*layer\s*\.\s*push\s*\(|<script|=>)",
            path,
            re.IGNORECASE,
        )
    )
    if (
        len(path) > 512
        or "\n" in path
        or "```" in path
        or looks_like_example
        or not valid_path(path)
    ):
        raise PredicateError(f"{location}: invalid or implausible field path '{path[:80]}'.")
    return path


def _event_tags_and_destinations(
    raw: dict[str, Any], requirements: list[dict[str, Any]], scope: dict[str, Any]
) -> tuple[list[dict[str, Any]], list[str]]:
    tag_values = list(raw.get("tags", [])) if isinstance(raw.get("tags"), list) else []
    destination_values = (
        list(raw.get("destinations", [])) if isinstance(raw.get("destinations"), list) else []
    )
    for row in requirements:
        expectation = row.get("expectation") if isinstance(row.get("expectation"), dict) else row
        if row.get("tag") or expectation.get("tag_name"):
            tag_values.append(row.get("tag") or expectation.get("tag_name"))
        if row.get("destination") or expectation.get("destination"):
            destination_values.append(row.get("destination") or expectation.get("destination"))

    tags: list[dict[str, Any]] = []
    seen_tags: set[str] = set()
    for index, value in enumerate(tag_values, start=1):
        tag = _normalize_tag(value, index)
        if tag["tag_id"] not in seen_tags and _tag_in_scope(tag, scope):
            tags.append(tag)
            seen_tags.add(tag["tag_id"])

    destination_values.extend(
        tag.get("destination") for tag in tags if tag.get("destination") not in (None, "")
    )

    destinations: list[str] = []
    for value in [*destination_values, *scope.get("destination", [])]:
        identity = str(value or "").strip()
        if identity and identity not in destinations:
            destinations.append(identity)
    if len(destinations) == 1:
        for tag in tags:
            if tag.get("destination") in (None, ""):
                tag["destination"] = destinations[0]
    return tags, destinations


def _claim_target(
    archetype: str,
    path: str,
    source_event_name: str | None,
    expectation: dict[str, Any],
) -> dict[str, Any]:
    target = {
        "surface": "source",
        "path": path,
        "event_name": source_event_name,
        "label": f"Source - {path}",
    }
    variants = {
        "gtm": ("preview", "resolved_variable", f"Resolved variable - {path}"),
        "delivery": ("network", "request_parameter", f"Request parameter - {path}"),
        "reality": ("page", "page_value", f"Page/business value - {path}"),
        "sequence": ("stream", "order", f"Sequence - {path}"),
    }
    if archetype in variants:
        surface, check, label = variants[archetype]
        target.update({"surface": surface, "check": check, "label": label})
    if archetype == "delivery":
        target["destination"] = expectation.get("destination")
    return target


def _add_explicit_mirrors(
    builder: _ClaimBuilder,
    expectation: dict[str, Any],
    predicate: dict[str, Any],
    source: dict[str, Any],
    applicability: dict[str, Any],
    source_event_name: str | None,
    inferred_tag: Any,
    inferred_destination: Any,
    location: str,
) -> None:
    mirrors = (
        ("state_path", "gtm", "data_layer_state", "preview"),
        ("resolved_path", "gtm", "resolved_variable", "preview"),
        ("configuration_path", "gtm", "tag_configuration", "preview"),
        ("runtime_path", "delivery", "runtime_parameter", "preview"),
        ("request_path", "delivery", "request_parameter", "network"),
    )
    for key, mirror_type, check, evidence in mirrors:
        if expectation.get(key) in (None, ""):
            continue
        mirror_path = _safe_path(expectation[key], location)
        builder.add(
            mirror_type,
            {
                "surface": evidence,
                "check": check,
                "path": mirror_path,
                "event_name": expectation.get(
                    "resolved_event_name", expectation.get("forward_event_name", source_event_name)
                ),
                "tag_id": inferred_tag,
                "destination": inferred_destination,
                "label": f"{check.replace('_', ' ').title()} - {mirror_path}",
            },
            predicate,
            [evidence],
            source,
            applicability=applicability,
        )


def _add_ga4_forwarding_claims(
    builder: _ClaimBuilder,
    expectation: dict[str, Any],
    predicate: dict[str, Any],
    source: dict[str, Any],
    applicability: dict[str, Any],
    path: str,
    delivery_event_name: str | None,
    tag_event_name: str | None,
    ga4_tags: list[dict[str, Any]],
    destinations: list[str],
    scope: dict[str, Any],
) -> None:
    dynamic_target = {
        "tag_scope": ["GA4"],
        "destination_allowlist": destinations,
    }
    targets = ga4_tags or [None]
    for tag in targets:
        tag_id = tag["tag_id"] if tag else None
        destination = tag.get("destination") if tag else None
        tag_name = tag["tag_name"] if tag else "runtime-discovered GA4 tag"
        common = {
            "tag_id": tag_id,
            "destination": destination,
            **({} if tag else dynamic_target),
        }
        builder.add(
            "gtm",
            {
                "surface": "preview",
                "check": "effective_mapping",
                "path": path,
                "event_name": tag_event_name,
                **common,
                "label": f"Effective tag mapping - {tag_name} - {path}",
            },
            {"operator": "present"},
            ["preview"],
            source,
            applicability=applicability,
        )
        if expectation.get("runtime_path") in (None, ""):
            builder.add(
                "delivery",
                {
                    "surface": "preview",
                    "check": "runtime_parameter",
                    "path": path,
                    "event_name": tag_event_name,
                    **common,
                    "label": f"Tag runtime - {tag_name} - {path}",
                },
                predicate,
                ["preview"],
                source,
                applicability=applicability,
            )
        send_required = tag.get("browser_send_required") if tag else None
        if send_required is None:
            send_required = scope.get("browser_send_required", True)
        if send_required and expectation.get("request_path") in (None, ""):
            builder.add(
                "delivery",
                {
                    "surface": "network",
                    "check": "request_parameter",
                    "path": path,
                    "event_name": delivery_event_name,
                    **common,
                    "protocol": "ga4",
                    "label": f"Browser parameter - {tag_name} - {path}",
                },
                predicate,
                ["network"],
                source,
                applicability=applicability,
            )


def _compile_requirement(
    builder: _ClaimBuilder,
    row: dict[str, Any],
    index: int,
    *,
    tags: list[dict[str, Any]],
    destinations: list[str],
    scope: dict[str, Any],
    source_event_name: str | None,
    delivery_event_name: str | None,
    source_only: bool,
    event_negative: bool,
) -> None:
    expectation = row.get("expectation") if isinstance(row.get("expectation"), dict) else row
    source = _requirement_source(row, index)
    location = str(source["reference"])
    explicit = str(expectation.get("claim_type") or row.get("claim_type") or "source").casefold()
    archetype = ARCHETYPE_ALIASES.get(explicit, explicit)
    if archetype not in CLAIM_ARCHETYPES:
        raise PredicateError(f"{location}: unsupported claim type '{explicit}'.")
    path = _safe_path(expectation.get("field_path", "event"), location)
    predicate = compile_predicate(expectation, location=location)
    condition = expectation.get("condition")
    applicability = condition if isinstance(condition, dict) else {}
    if not (event_negative and archetype == "source"):
        builder.add(
            archetype,
            _claim_target(archetype, path, source_event_name, expectation),
            predicate,
            ["datalayer", "source"] if archetype == "source" else [archetype],
            source,
            applicability=applicability,
        )

    inferred_tag = expectation.get("tag_name")
    if inferred_tag in (None, "") and len(tags) == 1:
        inferred_tag = tags[0]["tag_id"]
    inferred_destination = expectation.get("destination")
    if inferred_destination in (None, "") and len(destinations) == 1:
        inferred_destination = destinations[0]
    _add_explicit_mirrors(
        builder,
        expectation,
        predicate,
        source,
        applicability,
        source_event_name,
        inferred_tag,
        inferred_destination,
        location,
    )

    ga4_tags = [
        tag for tag in tags if _tag_protocol(tag) == "ga4" and tag.get("expected") == "fire"
    ]
    forwarding_required = (
        expectation.get("source_only") is not True
        and expectation.get("forwarding_required") is not False
    )
    ga4_scope = any(str(item).casefold() == "ga4" for item in scope.get("tag_scope", []))
    if (
        archetype == "source"
        and forwarding_required
        and not source_only
        and delivery_event_name is not None
        and not event_negative
        and (ga4_tags or (not tags and ga4_scope))
    ):
        _add_ga4_forwarding_claims(
            builder,
            expectation,
            predicate,
            source,
            applicability,
            path,
            delivery_event_name,
            delivery_event_name,
            ga4_tags,
            destinations,
            scope,
        )
    for business_rule in expectation.get("business_rules", []):
        builder.add(
            "reality",
            {"surface": "business", "check": "relationship", "label": "Business relationship"},
            compile_predicate(
                {"operator": "relationship", "relationship": business_rule}, location=location
            ),
            ["page", "datalayer"],
            source,
            applicability=applicability,
        )


def _add_event_tag_claims(
    builder: _ClaimBuilder,
    raw: dict[str, Any],
    event_name: str | None,
    tags: list[dict[str, Any]],
    destinations: list[str],
    scope: dict[str, Any],
    source: dict[str, Any],
) -> None:
    occurrence = _occurrence_predicate(
        raw.get("expected_occurrence"),
        negative=raw.get("negative") is True,
        location=str(source.get("reference") or "event"),
    )
    builder.add(
        "gtm",
        {
            "surface": "preview",
            "check": "event_match",
            "event_name": event_name,
            "label": "GTM Preview event match",
        },
        occurrence,
        ["preview"],
        source,
    )
    builder.add(
        "gtm",
        {
            "surface": "preview",
            "check": "tag_inventory",
            "event_name": event_name,
            "label": "GTM fired/non-fired inventory",
        },
        {"operator": "present"},
        ["preview"],
        source,
    )
    if not tags and scope.get("tag_scope"):
        builder.add(
            "gtm",
            {
                "surface": "preview",
                "check": "in_scope_tag_discovery",
                "event_name": event_name,
                "tag_scope": scope.get("tag_scope", []),
                "label": "Runtime in-scope tag discovery",
            },
            {"operator": "present"},
            ["preview"],
            source,
        )
        dynamic_tag = {
            "tag_id": None,
            "tag_name": "Runtime-discovered in-scope tag",
            "category": None,
            "expected": "not_fire" if raw.get("negative") is True else "fire",
            "destination": None,
            "configuration": None,
            "consent_requirements": [],
            "browser_send_required": scope.get("browser_send_required", True),
        }
        for check, label in (
            ("tag_configuration", "Concerned tag effective configuration"),
            ("tag_firing", "Concerned tag firing"),
        ):
            builder.add(
                "gtm",
                {
                    "surface": "preview",
                    "check": check,
                    "event_name": event_name,
                    "tag_id": None,
                    "tag": dynamic_tag,
                    "tag_scope": scope.get("tag_scope", []),
                    "destination_allowlist": destinations,
                    "label": label,
                },
                {"operator": "present"}
                if check == "tag_configuration"
                else {"operator": "count", "exact": 0}
                if raw.get("negative") is True
                else {"operator": "count", "minimum": 1},
                ["preview"],
                source,
            )
    for tag in tags:
        _add_one_tag_claims(builder, tag, event_name, destinations, scope, source)
    if not tags and scope.get("browser_send_required"):
        for destination in destinations:
            builder.add(
                "delivery",
                {
                    "surface": "network",
                    "check": "destination_request",
                    "destination": destination,
                    "event_name": event_name,
                    "label": f"Destination routing - {destination}",
                },
                occurrence,
                ["network"],
                source,
            )
        if not destinations and "GA4" in scope.get("tag_scope", []):
            builder.add(
                "delivery",
                {
                    "surface": "network",
                    "check": "destination_request",
                    "destination": None,
                    "event_name": event_name,
                    "protocol": "ga4",
                    "label": "Runtime-discovered GA4 destination routing",
                },
                occurrence,
                ["network"],
                source,
            )


def _add_state_delivery_claims(
    builder: _ClaimBuilder,
    raw: dict[str, Any],
    delivery_event_name: str,
    destinations: list[str],
    scope: dict[str, Any],
    source: dict[str, Any],
) -> None:
    """Inspect a global/core state against its GA4 page send without inventing a source event."""
    builder.add(
        "gtm",
        {
            "surface": "preview",
            "check": "tag_inventory",
            "event_name": delivery_event_name,
            "label": "GTM fired/non-fired inventory",
        },
        {"operator": "present"},
        ["preview"],
        source,
    )
    dynamic_tag = {
        "tag_id": None,
        "tag_name": "Runtime-discovered in-scope tag",
        "category": None,
        "expected": "fire",
        "destination": None,
        "configuration": None,
        "consent_requirements": [],
        "browser_send_required": scope.get("browser_send_required", True),
    }
    for check, label, predicate in (
        ("tag_configuration", "Concerned tag effective configuration", {"operator": "present"}),
        ("tag_firing", "Concerned tag firing", {"operator": "count", "minimum": 1}),
    ):
        builder.add(
            "gtm",
            {
                "surface": "preview",
                "check": check,
                "event_name": delivery_event_name,
                "tag_id": None,
                "tag": dynamic_tag,
                "tag_scope": scope.get("tag_scope", []),
                "destination_allowlist": destinations,
                "label": label,
            },
            predicate,
            ["preview"],
            source,
        )
    if scope.get("browser_send_required"):
        for destination in destinations:
            builder.add(
                "delivery",
                {
                    "surface": "network",
                    "check": "destination_request",
                    "destination": destination,
                    "event_name": delivery_event_name,
                    "protocol": "ga4",
                    "label": f"Destination routing - {destination}",
                },
                _occurrence_predicate(
                    raw.get("expected_occurrence"),
                    negative=False,
                    location=str(source.get("reference") or "state event"),
                ),
                ["network"],
                source,
            )
        if not destinations and "GA4" in scope.get("tag_scope", []):
            builder.add(
                "delivery",
                {
                    "surface": "network",
                    "check": "destination_request",
                    "destination": None,
                    "event_name": delivery_event_name,
                    "protocol": "ga4",
                    "label": "Runtime-discovered GA4 page destination routing",
                },
                _occurrence_predicate(
                    raw.get("expected_occurrence"),
                    negative=False,
                    location=str(source.get("reference") or "state event"),
                ),
                ["network"],
                source,
            )


def _add_one_tag_claims(
    builder: _ClaimBuilder,
    tag: dict[str, Any],
    event_name: str | None,
    destinations: list[str],
    scope: dict[str, Any],
    source: dict[str, Any],
) -> None:
    builder.add(
        "gtm",
        {
            "surface": "preview",
            "check": "tag_configuration",
            "event_name": event_name,
            "tag_id": tag["tag_id"],
            "tag": tag,
            "destination_allowlist": destinations,
            "label": f"Tag configuration - {tag['tag_name']}",
        },
        {"operator": "present"},
        ["preview"],
        source,
    )
    expected = tag["expected"]
    if expected == "not_fire":
        firing_predicate = {"operator": "count", "exact": 0}
    elif expected == "optional":
        firing_predicate = {"operator": "count", "minimum": 0, "maximum": 1}
    else:
        firing_predicate = {"operator": "count", "exact": 1}
    builder.add(
        "gtm",
        {
            "surface": "preview",
            "check": "tag_firing",
            "event_name": event_name,
            "tag_id": tag["tag_id"],
            "tag": tag,
            "destination_allowlist": destinations,
            "label": f"Tag firing - {tag['tag_name']}",
        },
        firing_predicate,
        ["preview"],
        source,
    )
    send_required = tag.get("browser_send_required")
    if send_required is None:
        send_required = scope.get("browser_send_required", True)
    if not send_required or expected == "optional":
        return
    destination = tag.get("destination")
    if destination in (None, "") and len(destinations) == 1:
        destination = destinations[0]
    builder.add(
        "delivery",
        {
            "surface": "network",
            "check": "tag_request",
            "tag_id": tag["tag_id"],
            "tag": tag,
            "destination": destination,
            "destination_allowlist": destinations,
            "event_name": event_name if _tag_protocol(tag) == "ga4" else None,
            "protocol": _tag_protocol(tag),
            "label": f"Browser request - {tag['tag_name']}",
        },
        firing_predicate,
        ["preview", "network"],
        source,
    )


def _compile_event(
    raw: dict[str, Any], requirements: list[dict[str, Any]], order: int, scope: dict[str, Any]
) -> dict[str, Any]:
    event_id = str(
        raw.get("event_id")
        or raw.get("event_group_id")
        or raw.get("event_name")
        or raw.get("name")
        or f"event-{order}"
    )
    event_name_value = raw.get("event_name", raw.get("name"))
    event_name = str(event_name_value).strip() if event_name_value not in (None, "") else None
    state_only = raw.get("state_only") is True or str(raw.get("mode") or "").casefold() in {
        "state",
        "state_only",
        "core_state",
    }
    event_label_name = event_name
    if state_only:
        event_name = None
    explicit_source_event = raw.get("source_event_name")
    source_event_name = (
        str(explicit_source_event).strip()
        if explicit_source_event not in (None, "")
        else None
        if state_only
        else event_name
    )
    explicit_delivery_event = raw.get("delivery_event_name")
    delivery_event_name = (
        str(explicit_delivery_event).strip()
        if explicit_delivery_event not in (None, "")
        else None
        if state_only
        else event_name
    )
    source_only = raw.get("source_only") is True or not scope.get("certify_tags", True)
    errors: list[str] = [str(value) for value in raw.get("_input_errors", [])]
    builder = _ClaimBuilder(event_id)
    builder.warnings.extend(str(value) for value in raw.get("_input_warnings", []))
    journey = (
        raw.get("journey")
        if isinstance(raw.get("journey"), dict)
        else {
            "action": raw.get("trigger", raw.get("action")),
            "url": raw.get("url"),
            "locale": raw.get("locale"),
        }
    )

    tags, destinations = _event_tags_and_destinations(raw, requirements, scope)
    if not source_only and scope.get("tag_scope_declared") is True and not tags:
        errors.append(
            "Plan-declared tag scope has no exact event tag identity; normalize the planned "
            "tag name/category before browser inspection."
        )
    if (
        not source_only
        and scope.get("browser_send_required") is True
        and scope.get("destination_scope_declared") is True
        and not destinations
        and not scope.get("tag_scope")
    ):
        errors.append(
            "Plan-declared destination scope has no exact event destination; normalize the "
            "planned measurement/conversion destination before browser inspection."
        )

    default_source = {"reference": f"event {event_id}", "plan_order": order}
    builder.add(
        "reality",
        {"surface": "page", "check": "valid_outcome", "label": "Page/API outcome"},
        {"operator": "present"},
        ["page"],
        default_source,
    )
    if source_event_name:
        try:
            source_occurrence = _occurrence_predicate(
                raw.get("expected_occurrence"),
                negative=raw.get("negative") is True,
                location=str(default_source["reference"]),
            )
            builder.add(
                "source",
                {
                    "surface": "source",
                    "check": "event_occurrence",
                    "event_name": source_event_name,
                    "label": f"dataLayer/direct source - {source_event_name}",
                },
                source_occurrence,
                ["datalayer", "source"],
                default_source,
            )
        except PredicateError as error:
            errors.append(str(error))

    for index, row in enumerate(requirements, start=1):
        try:
            _compile_requirement(
                builder,
                row,
                index,
                tags=tags,
                destinations=destinations,
                scope=scope,
                source_event_name=source_event_name,
                delivery_event_name=delivery_event_name,
                source_only=source_only,
                event_negative=raw.get("negative") is True,
            )
        except PredicateError as error:
            errors.append(str(error))

    if not source_only and delivery_event_name is not None:
        try:
            if state_only:
                _add_state_delivery_claims(
                    builder,
                    raw,
                    delivery_event_name,
                    destinations,
                    scope,
                    default_source,
                )
            else:
                _add_event_tag_claims(
                    builder,
                    raw,
                    delivery_event_name,
                    tags,
                    destinations,
                    scope,
                    default_source,
                )
        except PredicateError as error:
            errors.append(str(error))

    builder.add(
        "sequence",
        {"surface": "stream", "check": "surrounding_behavior", "label": "Sequence/anomaly stream"},
        {"operator": "present"},
        ["datalayer", "preview", "network", "lifecycle", "page"],
        default_source,
    )
    builder.add(
        "safety",
        {
            "surface": "all",
            "check": "sensitive_data",
            "event_name": delivery_event_name or source_event_name,
            "tag_ids": [tag["tag_id"] for tag in tags],
            "destinations": destinations,
            "label": "Data safety",
        },
        {"operator": "present"},
        ["page", "datalayer", "source", "preview", "network"],
        default_source,
    )

    explicit = raw.get("scenarios") if isinstance(raw.get("scenarios"), list) else []
    negative = (
        raw.get("negative_scenarios") if isinstance(raw.get("negative_scenarios"), list) else []
    )
    scenarios = [
        _scenario(value, event_id, index, negative=value in negative)
        for index, value in enumerate([*explicit, *negative], start=1)
    ]
    known_dimensions = [
        dict(value) for value in raw.get("known_dimensions", []) if isinstance(value, dict)
    ]
    known_names = {str(value.get("name")) for value in known_dimensions}
    for row in requirements:
        expectation = row.get("expectation") if isinstance(row.get("expectation"), dict) else row
        allowed = expectation.get("allowed_values")
        path = str(expectation.get("field_path") or "")
        if not path or not isinstance(allowed, list) or not allowed or path in known_names:
            continue
        known_dimensions.append(
            {
                "name": path,
                "kind": "manageable_finite",
                "material": True,
                "values": [{"value": value, "source": "plan"} for value in allowed],
            }
        )
        known_names.add(path)
    return {
        "event_id": event_id,
        "event_name": event_name,
        "source_event_name": source_event_name,
        "delivery_event_name": delivery_event_name,
        "label": str(raw.get("label") or raw.get("event_label") or event_label_name or event_id),
        "plan_order": order,
        "mode": "state_only" if state_only else "named_event",
        "journey": journey,
        "claims": builder.claims,
        "claim_count": len(builder.claims),
        "tags": tags,
        "destinations": destinations,
        "explicit_scenarios": scenarios,
        "known_dimensions": known_dimensions,
        "allowed_companions": raw.get("allowed_companions", []),
        "required_consent_signals": raw.get("required_consent_signals", []),
        "compile_errors": errors,
        "compile_warnings": builder.warnings,
        "executable": not errors,
    }


def _event_inputs(value: dict[str, Any]) -> list[tuple[dict[str, Any], list[dict[str, Any]]]]:
    if isinstance(value.get("events"), list):
        output = []
        for event in value["events"]:
            if not isinstance(event, dict):
                continue
            requirements = event.get("requirements")
            if not isinstance(requirements, list):
                requirements = event.get("claims") if isinstance(event.get("claims"), list) else []
            output.append((event, [item for item in requirements if isinstance(item, dict)]))
        return output
    if not isinstance(value.get("requirements"), list):
        raise StateError("Tracking plan needs an events or requirements array.")
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for index, row in enumerate(value["requirements"], start=1):
        if not isinstance(row, dict):
            continue
        expectation = row.get("expectation") if isinstance(row.get("expectation"), dict) else row
        name = str(
            expectation.get("event_name")
            or row.get("event_name")
            or row.get("event_group_id")
            or ""
        ).strip()
        if not name:
            raise StateError(f"Requirement {index} has no event group/name.")
        group = str(row.get("event_group_id") or name)
        source = row.get("source") if isinstance(row.get("source"), dict) else {}
        normalized = {**row, "source": {**source, "plan_order": source.get("plan_order", index)}}
        grouped.setdefault(group, []).append(normalized)
    output = []
    for group, rows in grouped.items():
        first_expectation = rows[0].get("expectation", rows[0])
        occurrence_values = {
            str(expectation.get("expected_occurrence"))
            for row in rows
            if (
                expectation := row.get("expectation")
                if isinstance(row.get("expectation"), dict)
                else row
            ).get("expected_occurrence")
            not in (None, "")
        }
        negative_values = {row.get("negative") is True for row in rows}
        input_errors = []
        if len(occurrence_values) > 1:
            input_errors.append(
                "Conflicting expected_occurrence values exist inside one event group."
            )
        if len(negative_values) > 1:
            input_errors.append("Negative and positive requirements are mixed in one event group.")
        metadata = {
            key: next((row.get(key) for row in rows if row.get(key) not in (None, "")), None)
            for key in EVENT_METADATA_FIELDS
        }
        raw = {
            "event_id": group,
            "event_name": first_expectation.get("event_name", group),
            "event_label": rows[0].get("event_label"),
            "journey": rows[0].get("journey", {}),
            "negative": True in negative_values,
            "expected_occurrence": next(iter(occurrence_values), "once_per_action"),
            "_input_errors": input_errors,
            **{key: item for key, item in metadata.items() if item not in (None, "")},
        }
        output.append((raw, rows))
    return output


def normalize_plan(
    source: Path | str,
    *,
    run_id: str | None = None,
    scope: dict[str, Any] | None = None,
) -> dict[str, Any]:
    path = Path(source).expanduser().resolve()
    value, source_kind, digest = _load_source(path)
    normalized_scope = _normalize_scope(scope)
    inputs = _event_inputs(value)
    normalization = value.get("_normalization", {})
    reconciliation = (
        normalization.get("reconciliation", {}) if isinstance(normalization, dict) else {}
    )
    detail_only = {
        _identity_key(item)
        for item in reconciliation.get("detail_only_events", [])
        if str(item).strip()
    }
    reconciled_inputs: list[tuple[dict[str, Any], list[dict[str, Any]]]] = []
    for event, requirements in inputs:
        raw = dict(event)
        identities = {
            _identity_key(value)
            for value in (
                raw.get("event_id"),
                raw.get("event_group_id"),
                raw.get("event_name"),
                raw.get("event_label"),
            )
            if str(value or "").strip()
        }
        if detail_only.intersection(identities):
            raw.setdefault("_input_warnings", []).append(
                "The event has a requirement sheet/table but is absent from the workbook event index."
            )
        reconciled_inputs.append((raw, requirements))
    for name in reconciliation.get("index_only_events", []):
        label = " ".join(str(name).split())
        if not label:
            continue
        state_only = _identity_key(label) in {
            "core_datalayer",
            "core_data_layer",
            "core_state",
        }
        reconciled_inputs.append(
            (
                {
                    "event_id": label,
                    "event_name": label,
                    "event_label": label,
                    "mode": "state_only" if state_only else "named_event",
                    "_input_errors": [
                        "The event is listed in the workbook index but has no requirement sheet/table."
                    ],
                },
                [],
            )
        )
    inputs = reconciled_inputs
    seen_ids: dict[str, int] = {}
    prepared_inputs = []
    for event, requirements in inputs:
        raw = dict(event)
        base_id = str(
            raw.get("event_id")
            or raw.get("event_group_id")
            or raw.get("event_name")
            or raw.get("name")
            or "event"
        )
        seen_ids[base_id] = seen_ids.get(base_id, 0) + 1
        if seen_ids[base_id] > 1:
            explicit_id = raw.get("event_id") not in (None, "") or raw.get(
                "event_group_id"
            ) not in (
                None,
                "",
            )
            if explicit_id:
                raise StateError(f"Duplicate explicit event identity: {base_id}")
            raw["event_id"] = f"{base_id}--{seen_ids[base_id]}"
        prepared_inputs.append((raw, requirements))
    events = [
        _compile_event(event, requirements, order, normalized_scope)
        for order, (event, requirements) in enumerate(prepared_inputs, start=1)
    ]
    if not events:
        raise StateError("Tracking plan contains no event groups.")
    if (
        normalized_scope.get("certify_tags") is True
        and normalized_scope.get("tag_scope_declared") is True
        and not any(event.get("tags") for event in events)
    ):
        raise StateError(
            "The broad plan-declared tag scope resolves to no tag identity. Supply a "
            "concise accepted vendor category such as GA4/Google Ads, or exact event tags."
        )
    if (
        normalized_scope.get("browser_send_required") is True
        and normalized_scope.get("destination_scope_declared") is True
        and not any(event.get("destinations") for event in events)
        and not normalized_scope.get("tag_scope")
    ):
        raise StateError(
            "The broad plan-declared destination scope resolves to no exact destination. "
            "Supply accepted measurement/conversion destinations or omit destination "
            "certification until they are confirmed."
        )
    if not normalized_scope.get("origins"):
        inferred_origins = []
        for event in events:
            candidate = str(event.get("journey", {}).get("url") or "").strip()
            if not candidate:
                continue
            match = re.match(r"^(https?://[^/]+)", candidate, re.I)
            if match and match.group(1) not in inferred_origins:
                inferred_origins.append(match.group(1))
        if inferred_origins:
            normalized_scope["origins"] = inferred_origins
            normalized_scope["origin_mode"] = "plan_inferred"

    claim_count = sum(event["claim_count"] for event in events)
    requirement_count = sum(len(requirements) for _, requirements in inputs)
    source_details = {"kind": source_kind, "path": str(path), "sha256": digest}
    if isinstance(value.get("_normalization"), dict):
        source_details["normalization"] = value["_normalization"]
    plan = {
        "schema_version": SCHEMA_VERSION,
        "run_id": run_id or f"GTM-RECETTE-{uuid4().hex[:12].upper()}",
        "created_at": utc_now(),
        "source": source_details,
        "scope": normalized_scope,
        "event_count": len(events),
        "requirement_count": requirement_count,
        "claim_count": claim_count,
        "events": events,
    }
    plan["compile_digest"] = hashlib.sha256(canonical_json(events).encode("utf-8")).hexdigest()
    return plan


def initialize_run(
    source: Path | str,
    run_dir: Path | str,
    *,
    run_id: str | None = None,
    scope: dict[str, Any] | None = None,
) -> dict[str, Any]:
    plan = normalize_plan(source, run_id=run_id, scope=scope)
    write_immutable_plan(run_dir, plan)
    return plan
