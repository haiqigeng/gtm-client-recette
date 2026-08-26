#!/usr/bin/env python3
"""Render the exact one-sheet, four-column final XLSX."""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

LAYER_ORDER = (
    "Page/action reality",
    "Data Layer API Call",
    "GTM Tags",
    "Browser request",
    "Surrounding behavior",
)
HEADERS = ("event_name", "layer_name", "status", "details")
STATUS_FILL = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "BLOCKED": "F4B183",
    "REVIEW": "FFEB9C",
    "NOT_APPLICABLE": "E7E6E6",
}
PROBLEM_STATUSES = {"FAIL", "BLOCKED", "REVIEW"}


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _safe_text(value: Any) -> str:
    text = str(value if value is not None else "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _relevant_checks(layer: dict[str, Any]) -> list[dict[str, Any]]:
    checks = [check for check in layer.get("checks", []) if isinstance(check, dict)]
    problems = [check for check in checks if check.get("status") in PROBLEM_STATUSES]
    return problems or checks


def _comparison(checks: list[dict[str, Any]], key: str) -> Any:
    values = [
        {
            "check": check.get("check"),
            "path": check.get("path"),
            "value": check.get(key),
        }
        for check in checks
        if check.get(key) is not None
    ]
    return values or None


def layer_details(result: dict[str, Any], layer: dict[str, Any]) -> str:
    """Render the one fixed details grammar for immediate and XLSX feedback."""
    checks = _relevant_checks(layer)
    reasons = list(
        dict.fromkeys(
            str(check.get("reason") or "Unspecified evidence result.") for check in checks
        )
    )
    evidence_file = str(result.get("evidence_file") or f"evidence-{result['event_id']}.json")
    evidence = [f"{evidence_file}#{check.get('check') or 'layer'}" for check in checks] or [
        evidence_file
    ]
    return (
        f"reason={' | '.join(reasons)}; "
        f"expected={_json(_comparison(checks, 'expected'))}; "
        f"observed={_json(_comparison(checks, 'observed'))}; "
        f"evidence={','.join(dict.fromkeys(evidence))}"
    )


def feedback_rows(result: dict[str, Any]) -> list[dict[str, str]]:
    """Return exactly five fixed-schema rows for one completed event."""
    layers = {layer.get("layer"): layer for layer in result.get("layers", [])}
    if set(layers) != set(LAYER_ORDER) or len(result.get("layers", [])) != len(LAYER_ORDER):
        raise ValueError("Completed event must contain each evidence layer exactly once.")
    return [
        {
            "event_name": _safe_text(result["event_name"]),
            "layer_name": name,
            "status": _safe_text(layers[name]["status"]),
            "details": _safe_text(layer_details(result, layers[name])),
        }
        for name in LAYER_ORDER
    ]


def _validate_plan_results(
    plan: dict[str, Any], results: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    if len(results) != len(plan.get("events", [])):
        raise ValueError("Final report requires exactly one result per tracking-plan event.")
    by_event: dict[str, dict[str, Any]] = {}
    for result in results:
        event_id = str(result.get("event_id") or "")
        if event_id in by_event:
            raise ValueError(f"Duplicate final result for {event_id}.")
        by_event[event_id] = result
    ordered = []
    for event in plan["events"]:
        result = by_event.get(event["event_id"])
        if result is None:
            raise ValueError(f"Missing final result for {event['event_name']}.")
        if result.get("event_name") != event["event_name"]:
            raise ValueError(f"Final event identity mismatch for {event['event_id']}.")
        ordered.append(result)
    return ordered


def _validate_workbook(path: Path, ordered: list[dict[str, Any]]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != ["Event feedback"]:
            raise ValueError("Final workbook must contain only Event feedback.")
        sheet = workbook["Event feedback"]
        headers = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        if headers != HEADERS:
            raise ValueError("Final workbook column contract is invalid.")
        expected_rows = [row for result in ordered for row in feedback_rows(result)]
        actual_rows = [
            dict(zip(HEADERS, (cell.value for cell in row), strict=True))
            for row in sheet.iter_rows(min_row=2)
        ]
        if actual_rows != expected_rows:
            raise ValueError("Final workbook rows or order do not match committed evidence.")
    finally:
        workbook.close()


def render_report(plan: dict[str, Any], results: list[dict[str, Any]], output: Path | str) -> None:
    """Atomically write and reopen-validate the sole supported final workbook."""
    target = Path(output).resolve()
    if target.exists():
        raise ValueError(f"Final workbook already exists: {target}")
    ordered = _validate_plan_results(plan, results)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(target.name + ".tmp.xlsx")
    if temporary.exists():
        raise ValueError(f"Temporary report path already exists: {temporary}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Event feedback"
    sheet.append(list(HEADERS))
    for result in ordered:
        for row in feedback_rows(result):
            sheet.append([row[header] for header in HEADERS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 100
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[2].fill = PatternFill("solid", fgColor=STATUS_FILL.get(row[2].value, "FFFFFF"))
    try:
        workbook.save(temporary)
        workbook.close()
        _validate_workbook(temporary, ordered)
        os.replace(temporary, target)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()
