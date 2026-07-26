#!/usr/bin/env python3
"""Inspect XLSX or delimited tracking plans while preserving source coordinates."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Tracking-plan .xlsx, .csv, or .tsv file.")
    parser.add_argument("output", type=Path, help="Destination inspection JSON.")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=0,
        help="Optional per-sheet row limit; 0 reads all populated rows.",
    )
    parser.add_argument(
        "--assets-dir",
        type=Path,
        help=(
            "Directory for embedded workbook images. Defaults to "
            "<output-stem>-assets when the workbook contains images."
        ),
    )
    return parser.parse_args()


def value_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "formula" if value.startswith("=") else "string"
    return type(value).__name__


def _safe_name(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9._-]+", "-", value.strip()).strip("-")
    return normalized or "sheet"


def _hyperlink(cell: Any) -> dict[str, Any] | None:
    link = getattr(cell, "hyperlink", None)
    if link is None:
        return None
    target = getattr(link, "target", None)
    location = getattr(link, "location", None)
    if not target and not location:
        return None
    return {
        "target": target,
        "location": location,
        "tooltip": getattr(link, "tooltip", None),
        "display": getattr(link, "display", None),
    }


def _comment(cell: Any) -> dict[str, Any] | None:
    comment = getattr(cell, "comment", None)
    if comment is None:
        return None
    return {
        "text": comment.text,
        "author": comment.author,
    }


def _image_anchor(image: Any) -> dict[str, Any]:
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str):
        return {"from_cell": anchor, "to_cell": anchor}
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)

    def marker_cell(marker: Any) -> str | None:
        if marker is None:
            return None
        row = getattr(marker, "row", None)
        column = getattr(marker, "col", None)
        if not isinstance(row, int) or not isinstance(column, int):
            return None
        return f"{get_column_letter(column + 1)}{row + 1}"

    return {
        "from_cell": marker_cell(start),
        "to_cell": marker_cell(end) or marker_cell(start),
    }


def _extract_images(
    worksheet: Any,
    assets_dir: Path | None,
) -> list[dict[str, Any]]:
    images = []
    for index, image in enumerate(getattr(worksheet, "_images", []), start=1):
        image_format = str(getattr(image, "format", "") or "png").lower()
        if image_format == "jpeg":
            image_format = "jpg"
        anchor = _image_anchor(image)
        row = {
            "index": index,
            "anchor": anchor,
            "width": getattr(image, "width", None),
            "height": getattr(image, "height", None),
            "format": image_format,
            "extracted_file": None,
        }
        if assets_dir is not None:
            try:
                payload = image._data()
            except Exception as exc:  # pragma: no cover - depends on workbook media
                row["extraction_error"] = type(exc).__name__
            else:
                digest = hashlib.sha256(payload).hexdigest()[:12]
                filename = (
                    f"{_safe_name(worksheet.title)}-"
                    f"{anchor.get('from_cell') or index}-{digest}.{image_format}"
                )
                assets_dir.mkdir(parents=True, exist_ok=True)
                destination = assets_dir / filename
                destination.write_bytes(payload)
                row["extracted_file"] = str(destination.resolve())
                row["sha256"] = hashlib.sha256(payload).hexdigest()
        images.append(row)
    return images


def inspect_xlsx(
    path: Path,
    max_rows: int,
    assets_dir: Path | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    sheets = []
    try:
        for worksheet in workbook.worksheets:
            rows = []
            populated = 0
            for row_number, row in enumerate(worksheet.iter_rows(), start=1):
                cells = []
                for column_number, cell in enumerate(row, start=1):
                    if isinstance(cell, MergedCell):
                        continue
                    link = _hyperlink(cell)
                    comment = _comment(cell)
                    if cell.value is None and link is None and comment is None:
                        continue
                    column = get_column_letter(column_number)
                    item = {
                        "cell": f"{column}{row_number}",
                        "column": column,
                        "value": cell.value,
                        "value_type": value_type(cell.value),
                    }
                    if link is not None:
                        item["hyperlink"] = link
                    if comment is not None:
                        item["comment"] = comment
                    if worksheet.column_dimensions[column].hidden:
                        item["column_hidden"] = True
                    cells.append(item)
                if not cells:
                    continue
                row_item = {"row": row_number, "cells": cells}
                if worksheet.row_dimensions[row_number].hidden:
                    row_item["hidden"] = True
                rows.append(row_item)
                populated += 1
                if max_rows and populated >= max_rows:
                    break
            images = _extract_images(worksheet, assets_dir)
            sheets.append(
                {
                    "sheet": worksheet.title,
                    "sheet_state": worksheet.sheet_state,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "merged_ranges": [
                        str(cell_range) for cell_range in worksheet.merged_cells.ranges
                    ],
                    "images": images,
                    "populated_rows": rows,
                    "truncated": bool(max_rows and populated >= max_rows),
                }
            )
    finally:
        workbook.close()
    return {"source": str(path.resolve()), "format": "xlsx", "sheets": sheets}


def inspect_delimited(path: Path, max_rows: int) -> dict[str, Any]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row_number, row in enumerate(reader, start=1):
            cells = [
                {
                    "cell": f"{get_column_letter(column_number)}{row_number}",
                    "column": get_column_letter(column_number),
                    "value": value,
                    "value_type": "string",
                }
                for column_number, value in enumerate(row, start=1)
                if value != ""
            ]
            if not cells:
                continue
            rows.append({"row": row_number, "cells": cells})
            if max_rows and len(rows) >= max_rows:
                break
    max_column = max((len(row["cells"]) for row in rows), default=0)
    return {
        "source": str(path.resolve()),
        "format": path.suffix.lower().lstrip("."),
        "sheets": [
            {
                "sheet": path.stem,
                "max_row": len(rows),
                "max_column": max_column,
                "populated_rows": rows,
                "truncated": bool(max_rows and len(rows) >= max_rows),
            }
        ],
    }


def main() -> int:
    args = parse_args()
    if args.max_rows < 0:
        raise SystemExit("--max-rows must be zero or positive")
    suffix = args.input.suffix.lower()
    if suffix == ".xlsx":
        assets_dir = args.assets_dir or args.output.with_name(
            f"{args.output.stem}-assets"
        )
        result = inspect_xlsx(args.input, args.max_rows, assets_dir)
    elif suffix in {".csv", ".tsv"}:
        result = inspect_delimited(args.input, args.max_rows)
    else:
        raise SystemExit("Unsupported input format; use .xlsx, .csv, or .tsv")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"Created {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
