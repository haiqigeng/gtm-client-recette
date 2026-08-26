#!/usr/bin/env python3
"""Extract one variable XLSX and validate its sole adaptive GA4 mapping."""

from __future__ import annotations

import base64
import hashlib
import json
import re
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from tag_assistant import parse_api_call

SCHEMA_VERSION = "3.0.0"
SCOPE = "GA4_ONLY"
SUPPORTED_TYPES = {"string", "number", "integer", "boolean", "array", "object", "null"}
EXACT_RULE_HEADERS = re.compile(
    r"\b(?:expected|fixed|constant|required value|must equal|valeur attendue|valeur fixe)\b",
    re.I,
)
ENUM_RULE_HEADERS = re.compile(
    r"\b(?:allowed|accepted|permitted|one of|enum|valeurs autoris(?:e|é)es|valeurs possibles)\b",
    re.I,
)
DYNAMIC_VALUE = re.compile(r"(?:%[^%]+%|\$\{[^}]+\}|<[^>]+>|\.\.\.|…)")


class PlanError(ValueError):
    """Raised when XLSX evidence or its adaptive mapping violates the contract."""


def _reject_unknown(value: dict[str, Any], allowed: set[str], context: str) -> None:
    unknown = sorted(set(value) - allowed)
    if unknown:
        raise PlanError(f"{context}: unknown fields: {', '.join(unknown)}")


def _text(value: Any) -> str:
    return " ".join(str(value or "").split())


def _portable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _image_anchor(image: Any) -> str:
    if isinstance(image.anchor, str):
        return image.anchor
    marker = getattr(image.anchor, "_from", None)
    if marker is None:
        raise PlanError("Embedded image has no readable worksheet anchor.")
    return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"


def _partial_selector_from_code(text: str) -> dict[str, Any] | None:
    """Recover only quoted event identity when the full code payload is malformed."""
    gtag = re.search(r"gtag\s*\(\s*(['\"])event\1\s*,\s*(['\"])([^'\"]+)\2", text, re.I)
    if gtag:
        return {"event": gtag.group(3)}
    event = re.search(r"(?:['\"]?event['\"]?)\s*:\s*(['\"])([^'\"]+)\1", text, re.I)
    if not event:
        return None
    selector: dict[str, Any] = {"event": event.group(2)}
    semantic = re.search(r"(?:['\"]?event_name['\"]?)\s*:\s*(['\"])([^'\"]+)\1", text, re.I)
    if selector["event"].casefold() == "gtm.custom_event" and semantic:
        selector["event_name"] = semantic.group(2)
    return selector


def extract_workbook(path: Path | str) -> dict[str, Any]:
    """Extract source-addressed XLSX evidence without interpreting requirements."""
    source = Path(path).resolve()
    if source.suffix.casefold() != ".xlsx":
        raise PlanError("Tracking plan must be exactly one .xlsx file.")
    if not source.is_file():
        raise PlanError(f"Tracking plan does not exist: {source}")
    try:
        workbook = load_workbook(source, read_only=False, data_only=False, keep_links=False)
        displayed = load_workbook(source, read_only=False, data_only=True, keep_links=False)
    except (OSError, ValueError, BadZipFile, InvalidFileException) as error:
        raise PlanError(f"Cannot open XLSX tracking plan: {error}") from error
    sheets: list[dict[str, Any]] = []
    references: list[str] = []
    code_calls: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            displayed_sheet = displayed[worksheet.title]
            cells = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None and cell.hyperlink is None:
                        continue
                    reference = f"{worksheet.title}!{cell.coordinate}"
                    references.append(reference)
                    raw_text = str(cell.value or "")
                    if re.search(r"(?:dataLayer\s*\.\s*push|gtag)\s*\(", raw_text, re.I):
                        parsed_call = parse_api_call(raw_text)
                        code_calls.append(
                            {
                                "ref": reference,
                                "sheet": worksheet.title,
                                "complete": parsed_call.get("complete") is True,
                                "payload": parsed_call.get("payload"),
                                "partial_selector": _partial_selector_from_code(raw_text),
                                "arguments": parsed_call.get("arguments", []),
                                "reason": parsed_call.get("reason"),
                            }
                        )
                    cells.append(
                        {
                            "ref": reference,
                            "address": cell.coordinate,
                            "value": _portable(cell.value),
                            "displayed_value": _portable(displayed_sheet[cell.coordinate].value),
                            "formula": str(cell.value) if cell.data_type == "f" else None,
                            "data_type": cell.data_type,
                            "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
                        }
                    )
            images = []
            for index, image in enumerate(worksheet._images, 1):
                try:
                    content = image._data()
                except (OSError, ValueError, TypeError) as error:
                    raise PlanError(
                        f"{worksheet.title}: cannot extract embedded image {index}: {error}"
                    ) from error
                reference = f"image:{worksheet.title}:{index}"
                references.append(reference)
                images.append(
                    {
                        "ref": reference,
                        "anchor": _image_anchor(image),
                        "format": str(image.format or "").casefold(),
                        "sha256": hashlib.sha256(content).hexdigest(),
                        "base64": base64.b64encode(content).decode("ascii"),
                    }
                )
            sheets.append(
                {
                    "name": worksheet.title,
                    "visibility": worksheet.sheet_state,
                    "used_range": worksheet.calculate_dimension(),
                    "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                    "tables": [
                        {"name": table.name, "ref": table.ref}
                        for table in worksheet.tables.values()
                    ],
                    "cells": cells,
                    "images": images,
                }
            )
    finally:
        workbook.close()
        displayed.close()
    if not any(sheet["cells"] or sheet["images"] for sheet in sheets):
        raise PlanError("Tracking plan workbook is empty.")
    return {
        "contract": "gtm-client-recette/workbook-evidence/1",
        "source": {
            "path": str(source),
            "filename": source.name,
            "sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        },
        "sheets": sheets,
        "source_refs": references,
        "code_calls": code_calls,
    }


def _source_refs(value: Any, valid: set[str], context: str) -> list[str]:
    if not isinstance(value, list) or not value or not all(isinstance(item, str) for item in value):
        raise PlanError(f"{context}: source_refs must be a non-empty string list.")
    invalid = [item for item in value if item not in valid]
    if invalid:
        raise PlanError(f"{context}: unknown source reference {invalid[0]!r}.")
    return list(dict.fromkeys(value))


def _rule(value: Any, context: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise PlanError(f"{context}: expected_value_or_rule must be an object.")
    rule = value.get("rule")
    if rule == "present":
        _reject_unknown(value, {"rule"}, context)
        return {"rule": "present"}
    if rule == "equals":
        _reject_unknown(value, {"rule", "expected"}, context)
        if "expected" not in value:
            raise PlanError(f"{context}: equals rule requires expected.")
        return {"rule": "equals", "expected": value["expected"]}
    if rule == "one_of":
        _reject_unknown(value, {"rule", "allowed_values"}, context)
        allowed = value.get("allowed_values")
        if not isinstance(allowed, list) or not allowed:
            raise PlanError(f"{context}: one_of requires non-empty allowed_values.")
        return {"rule": "one_of", "allowed_values": allowed}
    raise PlanError(f"{context}: unsupported expected-value rule {rule!r}.")


def _cell_records(evidence: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(cell.get("ref")): cell
        for sheet in evidence.get("sheets", [])
        if isinstance(sheet, dict)
        for cell in sheet.get("cells", [])
        if isinstance(cell, dict) and cell.get("ref")
    }


def _rule_is_explicit(
    evidence: dict[str, Any], source_refs: list[str], compiled_rule: dict[str, Any]
) -> bool:
    """Accept value assertions only when workbook semantics explicitly declare them."""
    rule = compiled_rule["rule"]
    if rule == "present":
        return True
    values = (
        [compiled_rule.get("expected")]
        if rule == "equals"
        else list(compiled_rule.get("allowed_values", []))
    )
    if any(isinstance(value, str) and DYNAMIC_VALUE.search(value) for value in values):
        return False
    cells = _cell_records(evidence)
    header_pattern = EXACT_RULE_HEADERS if rule == "equals" else ENUM_RULE_HEADERS
    cited_text = "\n".join(_text(cells[ref].get("value")) for ref in source_refs if ref in cells)
    if header_pattern.search(cited_text):
        return True
    for reference in source_refs:
        cell = cells.get(reference)
        if not cell:
            continue
        match = re.fullmatch(r"([A-Z]+)(\d+)", str(cell.get("address") or ""), re.I)
        if not match:
            continue
        column, row_text = match.groups()
        row = int(row_text)
        sheet = reference.split("!", 1)[0]
        for preceding in range(max(1, row - 12), row):
            header = cells.get(f"{sheet}!{column.upper()}{preceding}")
            if header and header_pattern.search(_text(header.get("value"))):
                return True
    return False


def _selector_from_payload(payload: Any) -> dict[str, Any] | None:
    if not isinstance(payload, dict):
        return None
    event = payload.get("event")
    if not isinstance(event, str) or not event.strip():
        return None
    selector: dict[str, Any] = {"event": event.strip()}
    semantic = payload.get("event_name")
    if event.casefold() == "gtm.custom_event" and isinstance(semantic, str) and semantic.strip():
        selector["event_name"] = semantic.strip()
    return selector


def _selector_from_call(call: dict[str, Any]) -> dict[str, Any] | None:
    return _selector_from_payload(call.get("payload")) or call.get("partial_selector")


def _semantic_event_name(selector: dict[str, Any]) -> str:
    if str(selector.get("event") or "").casefold() == "gtm.custom_event":
        return _text(selector.get("event_name"))
    return _text(selector.get("event"))


def _same_selector(left: Any, right: Any) -> bool:
    return isinstance(left, dict) and isinstance(right, dict) and left == right


def _code_authority(
    evidence: dict[str, Any], source_refs: list[str], mapped_selector: Any
) -> tuple[dict[str, Any] | None, dict[str, Any] | None, list[str], bool]:
    """Resolve one unambiguous cited or same-sheet technical code authority."""
    calls = [
        call
        for call in evidence.get("code_calls", [])
        if isinstance(call, dict) and _selector_from_call(call)
    ]
    cited = [call for call in calls if call.get("ref") in source_refs]
    sheets = {ref.split("!", 1)[0] for ref in source_refs if "!" in ref}
    candidates = cited or [call for call in calls if call.get("sheet") in sheets]
    if not candidates:
        return None, None, [], False
    if isinstance(mapped_selector, dict):
        exact = [
            call
            for call in candidates
            if _same_selector(_selector_from_call(call), mapped_selector)
        ]
        if len(exact) == 1:
            call = exact[0]
            return (
                _selector_from_call(call),
                call.get("payload") if call.get("complete") is True else None,
                [str(call["ref"])],
                False,
            )
    selectors = {json.dumps(_selector_from_call(call), sort_keys=True): call for call in candidates}
    if len(selectors) == 1:
        call = next(iter(selectors.values()))
        return (
            _selector_from_call(call),
            call.get("payload") if call.get("complete") is True else None,
            [str(call["ref"])],
            False,
        )
    return None, None, [str(call["ref"]) for call in candidates], True


def _unparsed_code(evidence: dict[str, Any], source_refs: list[str]) -> list[dict[str, Any]]:
    sheets = {ref.split("!", 1)[0] for ref in source_refs if "!" in ref}
    return [
        call
        for call in evidence.get("code_calls", [])
        if isinstance(call, dict)
        and call.get("complete") is not True
        and (call.get("ref") in source_refs or call.get("sheet") in sheets)
    ]


def _json_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return "string"


def _technical_fields(
    payload: dict[str, Any] | None, source_refs: list[str]
) -> list[dict[str, Any]]:
    """Flatten an authoritative code payload into present/type field expectations."""
    output: list[dict[str, Any]] = []

    def walk(value: Any, path: str) -> None:
        if isinstance(value, dict):
            if not value and path:
                output.append({"path": path, "type": "object"})
            for key, child in value.items():
                if not path and key in {"event", "event_name"}:
                    continue
                walk(child, f"{path}.{key}" if path else str(key))
            return
        if isinstance(value, list):
            if not value:
                output.append({"path": path, "type": "array"})
                return
            item_path = f"{path}[]"
            if all(isinstance(item, dict) for item in value):
                for key in dict.fromkeys(key for item in value for key in item):
                    sample = next(item[key] for item in value if key in item)
                    walk(sample, f"{item_path}.{key}")
            else:
                types = {_json_type(item) for item in value}
                output.append(
                    {"path": item_path, "type": next(iter(types)) if len(types) == 1 else "array"}
                )
            return
        if path:
            output.append({"path": path, "type": _json_type(value)})

    if payload:
        walk(payload, "")
    return [
        {
            **field,
            "ga4_parameter_name": field["path"].rsplit(".", 1)[-1].replace("[]", ""),
            "required": True,
            "rule": "present",
            "source_refs": source_refs,
        }
        for field in output
    ]


def _notice(
    code: str,
    field: str,
    message: str,
    source_refs: list[str],
    *,
    expected: Any = None,
    observed: Any = None,
) -> dict[str, Any]:
    return {
        "code": code,
        "field": field,
        "message": message,
        "source_refs": list(dict.fromkeys(source_refs)),
        "expected": expected,
        "observed": observed,
    }


def _compile_mapping(evidence: dict[str, Any], mapping: dict[str, Any]) -> dict[str, Any]:
    """Compile the fixed one-scenario canonical plan from one cited adaptive mapping."""
    if not isinstance(mapping, dict):
        raise PlanError("Canonical mapping must be a JSON object.")
    _reject_unknown(mapping, {"schema_version", "scope", "events"}, "mapping")
    if mapping.get("schema_version") != SCHEMA_VERSION:
        raise PlanError(f"mapping: schema_version must be {SCHEMA_VERSION}.")
    if mapping.get("scope") != SCOPE:
        raise PlanError(f"mapping: scope must be {SCOPE}.")
    raw_events = mapping.get("events")
    if not isinstance(raw_events, list) or not raw_events:
        raise PlanError("mapping: events must be a non-empty list.")
    valid_refs = set(evidence.get("source_refs", []))
    events = []
    names: set[str] = set()
    event_fields = {
        "event_name",
        "data_layer_selector",
        "definition",
        "trigger_description",
        "entry_url",
        "expected_destination_id",
        "expected_parameters",
        "source_refs",
    }
    parameter_fields = {
        "data_layer_path",
        "json_type",
        "required",
        "expected_value_or_rule",
        "ga4_parameter_name",
        "source_refs",
    }
    for index, raw_event in enumerate(raw_events, 1):
        context = f"events[{index}]"
        if not isinstance(raw_event, dict):
            raise PlanError(f"{context}: event must be an object.")
        _reject_unknown(raw_event, event_fields, context)
        refs = _source_refs(raw_event.get("source_refs"), valid_refs, context)
        mapped_selector = raw_event.get("data_layer_selector")
        code_selector, code_payload, technical_refs, ambiguous_code = _code_authority(
            evidence, refs, mapped_selector
        )
        notices: list[dict[str, Any]] = []
        for unparsed in _unparsed_code(evidence, refs):
            notices.append(
                _notice(
                    "UNPARSED_CODE",
                    "data_layer_code",
                    str(unparsed.get("reason") or "Technical code could not be parsed."),
                    [str(unparsed.get("ref"))],
                )
            )
        if ambiguous_code:
            notices.append(
                _notice(
                    "AMBIGUOUS_CODE_AUTHORITY",
                    "data_layer_selector",
                    "Multiple technical code calls apply; the cited mapped selector was retained.",
                    technical_refs,
                    observed=mapped_selector,
                )
            )
        selector = code_selector or mapped_selector
        if not isinstance(selector, dict) or not selector or "event" not in selector:
            raise PlanError(
                f"{context}: one identifiable event selector is required from code or plan evidence."
            )
        for path, expected in selector.items():
            if not isinstance(path, str) or not re.fullmatch(
                r"[A-Za-z_$][A-Za-z0-9_$]*(?:\.[A-Za-z_$][A-Za-z0-9_$]*)*", path
            ):
                raise PlanError(f"{context}: invalid dataLayer selector path {path!r}.")
            if not isinstance(expected, (str, int, float, bool)) or expected == "":
                raise PlanError(f"{context}: selector {path!r} requires one scalar value.")
        mapped_name = _text(raw_event.get("event_name"))
        event_name = _semantic_event_name(selector) or mapped_name
        if not re.fullmatch(r"[A-Za-z][A-Za-z0-9_]{0,39}", event_name):
            raise PlanError(f"{context}: invalid GA4 event_name {event_name!r}.")
        if event_name.casefold() in names:
            raise PlanError(f"{context}: duplicate canonical event {event_name!r}.")
        names.add(event_name.casefold())
        if code_selector and mapped_selector and code_selector != mapped_selector:
            notices.append(
                _notice(
                    "SOURCE_CONFLICT",
                    "data_layer_selector",
                    "Technical code overrides a conflicting mapped selector.",
                    technical_refs,
                    expected=code_selector,
                    observed=mapped_selector,
                )
            )
        if mapped_name and mapped_name != event_name:
            notices.append(
                _notice(
                    "SOURCE_CONFLICT",
                    "event_name",
                    "Technical code overrides a conflicting mapped event name.",
                    [*technical_refs, *refs],
                    expected=event_name,
                    observed=mapped_name,
                )
            )
        definition = _text(raw_event.get("definition")) or None
        trigger = _text(raw_event.get("trigger_description")) or None
        if definition is None:
            notices.append(
                _notice(
                    "MISSING_PLAN_FIELD",
                    "definition",
                    "The tracking plan supplies no event definition.",
                    refs,
                )
            )
        if trigger is None:
            notices.append(
                _notice(
                    "MISSING_PLAN_FIELD",
                    "trigger_description",
                    "The tracking plan supplies no trigger description.",
                    refs,
                )
            )
        entry_url = _text(raw_event.get("entry_url")) or None
        if entry_url is not None:
            parsed = urlparse(entry_url)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                raise PlanError(f"{context}: supplied entry_url must be one exact HTTP(S) URL.")
        else:
            notices.append(
                _notice(
                    "MISSING_PLAN_FIELD",
                    "entry_url",
                    "The tracking plan supplies no exact event entry URL; resolve it live.",
                    refs,
                )
            )
        destination = raw_event.get("expected_destination_id")
        if destination is not None:
            destination = _text(destination)
            if not re.fullmatch(r"[A-Za-z0-9-]+", destination):
                raise PlanError(f"{context}: invalid expected_destination_id.")
        raw_parameters = raw_event.get("expected_parameters", [])
        if not isinstance(raw_parameters, list):
            raise PlanError(f"{context}: expected_parameters must be a list.")
        fields = []
        parameter_names: set[str] = set()
        for parameter_index, parameter in enumerate(raw_parameters, 1):
            parameter_context = f"{context}.expected_parameters[{parameter_index}]"
            if not isinstance(parameter, dict):
                raise PlanError(f"{parameter_context}: parameter must be an object.")
            _reject_unknown(parameter, parameter_fields, parameter_context)
            path = _text(parameter.get("data_layer_path"))
            ga4_name = _text(parameter.get("ga4_parameter_name"))
            value_type = _text(parameter.get("json_type")).casefold()
            if not path or not ga4_name:
                raise PlanError(f"{parameter_context}: both parameter names are required.")
            if ga4_name.casefold() in parameter_names:
                raise PlanError(f"{parameter_context}: duplicate GA4 parameter {ga4_name!r}.")
            parameter_names.add(ga4_name.casefold())
            if value_type not in SUPPORTED_TYPES:
                raise PlanError(f"{parameter_context}: unsupported JSON type {value_type!r}.")
            if not isinstance(parameter.get("required"), bool):
                raise PlanError(f"{parameter_context}: required must be boolean.")
            parameter_refs = _source_refs(
                parameter.get("source_refs"), valid_refs, parameter_context
            )
            compiled_rule = _rule(parameter.get("expected_value_or_rule"), parameter_context)
            if not _rule_is_explicit(evidence, parameter_refs, compiled_rule):
                notices.append(
                    _notice(
                        "AMBIGUOUS_VALUE_RULE",
                        path,
                        "The workbook does not explicitly declare this example/value as a "
                        "runtime equality rule; presence and JSON type are retained.",
                        parameter_refs,
                        expected={"rule": "present"},
                        observed=compiled_rule,
                    )
                )
                compiled_rule = {"rule": "present"}
            fields.append(
                {
                    "path": path,
                    "ga4_parameter_name": ga4_name,
                    "type": value_type,
                    "required": parameter["required"],
                    **compiled_rule,
                    "source_refs": parameter_refs,
                }
            )
        by_path = {field["path"]: field for field in fields}
        for technical_field in _technical_fields(code_payload, technical_refs):
            existing = by_path.get(technical_field["path"])
            if existing is None:
                fields.append(technical_field)
                by_path[technical_field["path"]] = technical_field
                continue
            if existing["type"] != technical_field["type"]:
                notices.append(
                    _notice(
                        "SOURCE_CONFLICT",
                        technical_field["path"],
                        "Technical code overrides a conflicting declared JSON type.",
                        [*technical_refs, *existing["source_refs"]],
                        expected=technical_field["type"],
                        observed=existing["type"],
                    )
                )
                existing["type"] = technical_field["type"]
            existing["source_refs"] = list(
                dict.fromkeys([*existing["source_refs"], *technical_refs])
            )
        events.append(
            {
                "event_id": f"E-{index:04d}",
                "event_name": event_name,
                "definition": definition,
                "trigger_description": trigger,
                "entry_url": entry_url,
                "expected_destination_id": destination,
                "source_refs": refs,
                "technical_source_refs": technical_refs,
                "mapping_notices": notices,
                "fields": fields,
                "selector": dict(selector),
                "scenario": {"id": event_name},
                "plan_order": index,
            }
        )
    return {
        "schema_version": SCHEMA_VERSION,
        "scope": SCOPE,
        "created_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "source": evidence["source"],
        "event_count": len(events),
        "field_count": sum(len(event["fields"]) for event in events),
        "events": events,
    }


PARAMETER_SEMANTICS = {"FIXED", "EXAMPLE", "DYNAMIC"}
INTERPRETATION_EVENT_FIELDS = {
    "event_name",
    "parameters",
    "data_layer_payload",
    "definition",
    "trigger_description",
    "entry_url",
    "expected_destination_id",
    "source_refs",
}
INTERPRETATION_PARAMETER_FIELDS = {
    "data_layer_path",
    "ga4_parameter_name",
    "value",
    "value_semantics",
    "json_type",
    "required",
    "source_refs",
}


def _value_fields(payload: dict[str, Any], source_refs: list[str]) -> list[dict[str, Any]]:
    """Flatten one Data Layer payload while retaining representative values."""
    output: list[dict[str, Any]] = []

    def walk(value: Any, path: str) -> None:
        if isinstance(value, dict):
            for key, child in value.items():
                if not path and key in {"event", "event_name"}:
                    continue
                walk(child, f"{path}.{key}" if path else str(key))
            return
        if isinstance(value, list):
            if not value:
                output.append((path, [], "array"))
                return
            if all(isinstance(item, dict) for item in value):
                for key in dict.fromkeys(key for item in value for key in item):
                    sample = next(item[key] for item in value if key in item)
                    walk(sample, f"{path}[].{key}")
                return
            sample = value[0]
            output.append((f"{path}[]", sample, _json_type(sample)))
            return
        output.append((path, value, _json_type(value)))

    walk(payload, "")
    return [
        {
            "data_layer_path": path,
            "ga4_parameter_name": path.rsplit(".", 1)[-1].replace("[]", ""),
            "value": value,
            "value_semantics": "DYNAMIC" if DYNAMIC_VALUE.search(str(value)) else "EXAMPLE",
            "json_type": value_type,
            "required": None,
            "source_refs": list(source_refs),
        }
        for path, value, value_type in output
        if path
    ]


def _assign_payload_path(payload: dict[str, Any], path: str, value: Any) -> None:
    """Assign one dotted path, where [] denotes the representative first item."""
    segments = path.split(".")
    current: Any = payload
    for index, segment in enumerate(segments):
        array = segment.endswith("[]")
        key = segment[:-2] if array else segment
        if not re.fullmatch(r"[A-Za-z_$][A-Za-z0-9_$]*", key):
            raise PlanError(f"Invalid dataLayer parameter path {path!r}.")
        final = index == len(segments) - 1
        if array:
            if final:
                current[key] = [value]
                continue
            current.setdefault(key, [{}])
            if (
                not isinstance(current[key], list)
                or not current[key]
                or not isinstance(current[key][0], dict)
            ):
                raise PlanError(f"Conflicting dataLayer parameter paths at {path!r}.")
            current = current[key][0]
        elif final:
            current[key] = value
        else:
            current.setdefault(key, {})
            if not isinstance(current[key], dict):
                raise PlanError(f"Conflicting dataLayer parameter paths at {path!r}.")
            current = current[key]


def _canonical_snippet(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, indent=2)
    return "window.dataLayer = window.dataLayer || [];\nwindow.dataLayer.push(" + encoded + ");"


def _parameter_map(parameters: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {parameter["data_layer_path"]: parameter for parameter in parameters}


def validate_inspection_plan(
    evidence: dict[str, Any], interpretation: dict[str, Any]
) -> dict[str, Any]:
    """Reconcile irregular workbook interpretation into one executable inspection plan."""
    if not isinstance(interpretation, dict):
        raise PlanError("Workbook interpretation must be a JSON object.")
    _reject_unknown(interpretation, {"schema_version", "events"}, "interpretation")
    missing_root = {"schema_version", "events"} - set(interpretation)
    if missing_root:
        raise PlanError(f"interpretation: missing fields: {', '.join(sorted(missing_root))}")
    if interpretation.get("schema_version") != SCHEMA_VERSION:
        raise PlanError(f"interpretation: schema_version must be {SCHEMA_VERSION}.")
    raw_events = interpretation.get("events")
    if not isinstance(raw_events, list) or not raw_events:
        raise PlanError("interpretation: events must be a non-empty list.")
    valid_refs = set(evidence.get("source_refs", []))
    mapping_events: list[dict[str, Any]] = []
    canonical_parts: list[dict[str, Any]] = []
    for index, raw_event in enumerate(raw_events, 1):
        context = f"events[{index}]"
        if not isinstance(raw_event, dict):
            raise PlanError(f"{context}: event must be an object.")
        _reject_unknown(raw_event, INTERPRETATION_EVENT_FIELDS, context)
        missing_event = INTERPRETATION_EVENT_FIELDS - set(raw_event)
        if missing_event:
            raise PlanError(f"{context}: missing fields: {', '.join(sorted(missing_event))}")
        refs = _source_refs(raw_event.get("source_refs"), valid_refs, context)
        event_name = _text(raw_event.get("event_name"))
        payload = raw_event.get("data_layer_payload")
        if payload is not None and not isinstance(payload, dict):
            raise PlanError(f"{context}: data_layer_payload must be an object or null.")
        cited_calls = [
            call
            for call in evidence.get("code_calls", [])
            if call.get("complete") is True and call.get("ref") in refs
        ]
        if len(cited_calls) > 1:
            raise PlanError(f"{context}: multiple cited complete Data Layer calls are ambiguous.")
        code_payload = cited_calls[0].get("payload") if cited_calls else None
        if code_payload is not None and payload is not None and code_payload != payload:
            raise PlanError(
                f"{context}: interpreted parameters contradict the cited Data Layer payload."
            )
        payload = code_payload or payload
        raw_parameters = raw_event.get("parameters")
        if not isinstance(raw_parameters, list):
            raise PlanError(f"{context}: parameters must be a list.")
        parameters: list[dict[str, Any]] = []
        for parameter_index, parameter in enumerate(raw_parameters, 1):
            parameter_context = f"{context}.parameters[{parameter_index}]"
            if not isinstance(parameter, dict):
                raise PlanError(f"{parameter_context}: parameter must be an object.")
            _reject_unknown(parameter, INTERPRETATION_PARAMETER_FIELDS, parameter_context)
            missing_parameter = INTERPRETATION_PARAMETER_FIELDS - set(parameter)
            if missing_parameter:
                raise PlanError(
                    f"{parameter_context}: missing fields: {', '.join(sorted(missing_parameter))}"
                )
            parameter_refs = _source_refs(
                parameter.get("source_refs"), valid_refs, parameter_context
            )
            path = _text(parameter.get("data_layer_path"))
            ga4_name = _text(parameter.get("ga4_parameter_name"))
            semantics = parameter.get("value_semantics")
            value_type = parameter.get("json_type")
            required = parameter.get("required")
            if not path or not ga4_name:
                raise PlanError(f"{parameter_context}: both parameter names are required.")
            if semantics not in PARAMETER_SEMANTICS:
                raise PlanError(f"{parameter_context}: value_semantics is invalid.")
            if value_type is None:
                value_type = _json_type(parameter.get("value"))
            if value_type not in SUPPORTED_TYPES:
                raise PlanError(f"{parameter_context}: unsupported JSON type {value_type!r}.")
            if required is not None and not isinstance(required, bool):
                raise PlanError(f"{parameter_context}: required must be boolean or null.")
            parameters.append(
                {
                    "data_layer_path": path,
                    "ga4_parameter_name": ga4_name,
                    "value": parameter.get("value"),
                    "value_semantics": semantics,
                    "json_type": value_type,
                    "required": required,
                    "source_refs": parameter_refs,
                }
            )
        if payload is None:
            if not event_name or not parameters:
                raise PlanError(
                    f"{context}: event identity plus parameter values or one complete Data Layer "
                    "payload is required."
                )
            payload = {"event": event_name}
            for parameter in parameters:
                _assign_payload_path(payload, parameter["data_layer_path"], parameter["value"])
        selector = _selector_from_payload(payload)
        if selector is None:
            raise PlanError(f"{context}: Data Layer payload has no identifiable event.")
        payload_event_name = _semantic_event_name(selector)
        if event_name and event_name != payload_event_name:
            raise PlanError(f"{context}: event_name contradicts the Data Layer payload.")
        event_name = payload_event_name
        payload_parameters = _value_fields(payload, refs)
        if not parameters:
            parameters = payload_parameters
        elif not payload_parameters:
            raise PlanError(f"{context}: Data Layer payload contains no event parameters.")
        else:
            declared = _parameter_map(parameters)
            observed = _parameter_map(payload_parameters)
            if set(declared) != set(observed):
                raise PlanError(f"{context}: parameter paths contradict the Data Layer payload.")
            for path, parameter in declared.items():
                actual = observed[path]
                if parameter["json_type"] != actual["json_type"]:
                    raise PlanError(f"{context}: JSON type contradicts payload at {path!r}.")
                if (
                    parameter["value_semantics"] != "DYNAMIC"
                    and parameter["value"] != actual["value"]
                ):
                    raise PlanError(f"{context}: parameter value contradicts payload at {path!r}.")
        if not parameters:
            raise PlanError(f"{context}: at least one event parameter and value is required.")
        expected_parameters = []
        for parameter in parameters:
            rule = (
                {"rule": "equals", "expected": parameter["value"]}
                if parameter["value_semantics"] == "FIXED"
                else {"rule": "present"}
            )
            expected_parameters.append(
                {
                    "data_layer_path": parameter["data_layer_path"],
                    "json_type": parameter["json_type"],
                    "required": True,
                    "expected_value_or_rule": rule,
                    "ga4_parameter_name": parameter["ga4_parameter_name"],
                    "source_refs": parameter["source_refs"],
                }
            )
        mapping_events.append(
            {
                "event_name": event_name,
                "data_layer_selector": selector,
                "definition": raw_event.get("definition"),
                "trigger_description": raw_event.get("trigger_description"),
                "entry_url": raw_event.get("entry_url"),
                "expected_destination_id": raw_event.get("expected_destination_id"),
                "expected_parameters": expected_parameters,
                "source_refs": refs,
            }
        )
        canonical_parts.append(
            {
                "parameters": parameters,
                "data_layer_payload": payload,
                "data_layer_snippet": _canonical_snippet(payload),
            }
        )
    compiled = _compile_mapping(
        evidence,
        {"schema_version": SCHEMA_VERSION, "scope": SCOPE, "events": mapping_events},
    )
    for event, canonical in zip(compiled["events"], canonical_parts, strict=True):
        event.update(canonical)
    return compiled
