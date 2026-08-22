#!/usr/bin/env python3
"""Build and semantically validate a schema-v3 GTM Client Recette workbook."""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
from collections import Counter
from collections.abc import Iterable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from acceptance_contract import status_of, worst_status
from client_side_rules import evaluate_report_business_rules
from event_feedback import event_feedback, final_conclusion
from execution_contract import (
    business_push_rows,
    case_action_rows,
    validate_session,
)
from layer_contract import CANONICAL_LAYERS, TAG_RESULT_LAYERS
from recette_schema import (
    ReportValidationError,
    as_rows,
    dumps_structured,
    validate,
)
from state_io import atomic_write_bytes, recover_file_pair

STATUS_FILLS = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "BLOCKED": "F4B183",
    "REVIEW": "FFF2CC",
    "NOT_TESTED": "D9E1F2",
    "NOT_APPLICABLE": "E7E6E6",
}
UNSAFE_EVIDENCE_MARKERS = (
    "unallowlisted sensitive content",
    "provenance contains sensitive content",
    "must not retain an unredacted value",
    "session contains unredacted sensitive content",
)


def refuse_unsafe_evidence(errors: Iterable[str]) -> None:
    if any(marker in error for error in errors for marker in UNSAFE_EVIDENCE_MARKERS):
        raise ReportValidationError(
            "Workbook generation refused because normalized evidence contains "
            "unsafe sensitive content. Use the redacted scanner output, quarantine "
            "the source evidence, and rebuild only from a safe normalized record."
        )


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(bottom=Side(style="thin", color="B7C9D6"))
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
EXCEL_CELL_LIMIT = 32767
EXCEL_CHUNK_PAYLOAD = 32000
EXCEL_CONTINUATION_PREFIX = "[part {part}/{total}] "

REQUIRED_SHEETS = [
    "Client Summary",
    "Defect Register",
    "Requirement Matrix",
    "Journey Coverage",
    "Interaction Cases",
    "Layer Verdicts",
    "Event Evidence",
    "Observed Push Stream",
    "Tag Evidence",
    "Destination Evidence",
    "Trigger & Sequence",
    "Consent",
    "Business Rules",
    "Sensitive Data",
    "Client Checks",
    "Regression",
    "Container Context",
    "Unexpected Events-Tags",
    "Blockers",
    "Evidence Catalogue",
    "Run Context",
]
V2_SHEETS = [
    "Coverage Decisions",
    "Scenario Classes",
    "Semantic Checks",
    "Journey State",
    "Stream Segments",
    "Protected Handoffs",
    "Gated Flows",
    "Final Conclusion",
]


def required_sheets(session: dict[str, Any] | None) -> list[str]:
    """Return the exact workbook contract for legacy or operator-v2 runs."""
    if isinstance(session, dict) and session.get("operator_contract_version") == 2:
        return [*REQUIRED_SHEETS[:-1], *V2_SHEETS, REQUIRED_SHEETS[-1]]
    return list(REQUIRED_SHEETS)


LEGACY_OUTPUT_CONTRACT_VERSION = 2
OUTPUT_CONTRACT_VERSION = 3


def output_contract_version(session: dict[str, Any] | None) -> int:
    """Keep legacy workbook columns stable; operator-v2 uses the expanded contract."""
    return (
        OUTPUT_CONTRACT_VERSION
        if isinstance(session, dict) and session.get("operator_contract_version") == 2
        else LEGACY_OUTPUT_CONTRACT_VERSION
    )


DEFECT_HEADERS = [
    "output_contract_version",
    "defect_id",
    "plan_order",
    "event_group_id",
    "event_name",
    "status",
    "primary_outcome",
    "anomaly_flags",
    "case_id",
    "requirement_id",
    "tag_name",
    "placement",
    "material_variant",
    "failed_layer",
    "expected_value",
    "expected_type",
    "observed_value",
    "observed_type",
    "concise_reason",
    "exact_retest",
    "evidence_ids",
    "retest_status",
]

REQUIREMENT_HEADERS = [
    "requirement_id",
    "event_group_id",
    "plan_order",
    "source_reference",
    "source_file",
    "source_sheet",
    "source_row",
    "source_cells",
    "source_section",
    "journey_id",
    "step_id",
    "action",
    "action_value",
    "action_value_type",
    "action_value_source",
    "url",
    "selector_or_element",
    "inferred",
    "inference_source",
    "browser_context_id",
    "container_id",
    "scenario_id",
    "scenario_kind",
    "event_name",
    "source_mechanism",
    "field_path",
    "match_rule",
    "expected_value",
    "expected_type",
    "expected_occurrence",
    "occurrence_verdict",
    "source_signal_verdict",
    "raw_state",
    "raw_value",
    "raw_type",
    "resolved_state",
    "resolved_value",
    "resolved_type",
    "gtm_variable",
    "gtm_variable_state",
    "gtm_variable_value",
    "gtm_variable_type",
    "tag_name",
    "tag_delivery",
    "vendor_family",
    "destination_id",
    "tag_relevance",
    "expected_firing",
    "actual_firing",
    "fire_count",
    "tag_configuration_field",
    "expected_tag_configuration",
    "configured_value",
    "runtime_state",
    "runtime_value",
    "runtime_type",
    "request_behavior",
    "request_count",
    "request_id",
    "destination_parameter_path",
    "destination_parameter_value",
    "consent_scenario",
    "consent_source",
    "consent_state",
    "raw_verdict",
    "resolved_verdict",
    "variable_verdict",
    "tag_configuration_verdict",
    "tag_firing_verdict",
    "tag_parameter_verdict",
    "destination_request_verdict",
    "destination_parameter_verdict",
    "trigger_logic_verdict",
    "tag_sequence_verdict",
    "consent_verdict",
    "business_rule_verdict",
    "sensitive_data_verdict",
    "client_checks_verdict",
    "regression_verdict",
    "overall_status",
    "failure_layer",
    "mismatch_or_reason",
    "evidence_ids",
    "notes",
]

JOURNEY_HEADERS = [
    "plan_order",
    "event_group_id",
    "event_name",
    "requirement_id",
    "journey_id",
    "step_id",
    "action",
    "url",
    "selector_or_element",
    "inferred",
    "inference_source",
    "confidence",
    "attempted_routes",
    "execution_status",
    "blocker_id",
    "overall_status",
    "evidence_ids",
    "notes",
]

CASE_HEADERS = [
    "event_group_id",
    "case_id",
    "scope_status",
    "execution_status",
    "url",
    "element",
    "placement",
    "action",
    "material_variant",
    "discovered_from",
    "coverage_decision_id",
    "scenario_class_id",
    "sample_role",
    "selection_rationale",
    "population_member_id",
    "acquisition_context",
    "gated_flow_kind",
    "tag_scope",
    "tag_inventory_status",
    "tag_inventory",
    "applicability_status",
    "layer_applicability",
    "applicable_layers",
    "blocker_id",
    "case_reason",
    "final_action_id",
    "action_id",
    "attempt_number",
    "retry_of_action_id",
    "readiness_check_id",
    "settlement_check_id",
    "last_event_before",
    "first_event_after",
    "settled_final_event",
    "network_request_cursor_before",
    "network_request_cursor_after",
    "datalayer_call_index_before",
    "datalayer_call_index_after",
    "interaction_outcome",
    "completion_signal",
    "stream_settled",
    "settlement_reason",
    "observed_business_push_count",
    "layer_results",
    "tag_layer_results",
]
V2_CASE_ONLY_HEADERS = {
    "coverage_decision_id",
    "scenario_class_id",
    "sample_role",
    "selection_rationale",
    "population_member_id",
    "acquisition_context",
    "gated_flow_kind",
    "datalayer_call_index_before",
    "datalayer_call_index_after",
}
LEGACY_CASE_HEADERS = [header for header in CASE_HEADERS if header not in V2_CASE_ONLY_HEADERS]

LAYER_VERDICT_HEADERS = [
    "plan_order",
    "event_group_id",
    "event_name",
    "case_id",
    "action_id",
    "tag_id",
    "tag_name",
    "tag_scope_status",
    "layer",
    "status",
    "reason",
    "predicate_result",
    "details",
    "evidence_ids",
    "exact_retest",
]

EVENT_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "event_observed",
    "actual_occurrence_count",
    "occurrence_event_indexes",
    "anchor_event_name",
    "anchor_event_index",
    "occurrence_evidence_id",
    "source_mechanism",
    "source_signal_capture_source",
    "source_signal_observed",
    "source_signal_evidence_id",
    "capture_source",
    "event_index",
    "event_timestamp",
    "raw_api_call_payload",
    "raw_field_path",
    "raw_field_state",
    "raw_field_value",
    "raw_field_type",
    "resolved_data_layer_snapshot",
    "resolved_field_state",
    "resolved_field_value",
    "resolved_field_type",
    "action_id",
    "retry_of_action_id",
    "readiness_check_id",
    "settlement_check_id",
    "preview_connected_before",
    "target_ready_before",
    "last_event_before",
    "network_request_cursor_before",
    "action_timestamp",
    "interaction_outcome",
    "completion_signal",
    "first_event_after",
    "settled_final_event",
    "network_request_cursor_after",
    "quiet_window_ms",
    "timeout_ms",
    "stream_settled",
    "settlement_reason",
    "raw_evidence_id",
    "resolved_evidence_id",
    "status",
    "notes",
]

PUSH_HEADERS = [
    "stream_id",
    "connection_epoch",
    "event_index",
    "preview_event_index",
    "datalayer_call_index",
    "segment_id",
    "captured_at",
    "push_id",
    "action_id",
    "case_id",
    "event_group_id",
    "event_name",
    "url",
    "page_state",
    "classification",
    "classification_reason",
    "container_id",
    "evidence_id",
]
V2_PUSH_ONLY_HEADERS = {"preview_event_index", "datalayer_call_index", "segment_id"}
LEGACY_PUSH_HEADERS = [header for header in PUSH_HEADERS if header not in V2_PUSH_ONLY_HEADERS]


def case_headers(session: dict[str, Any] | None) -> list[str]:
    return (
        CASE_HEADERS
        if isinstance(session, dict) and session.get("operator_contract_version") == 2
        else LEGACY_CASE_HEADERS
    )


def push_headers(session: dict[str, Any] | None) -> list[str]:
    return (
        PUSH_HEADERS
        if isinstance(session, dict) and session.get("operator_contract_version") == 2
        else LEGACY_PUSH_HEADERS
    )


TAG_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "container_id",
    "vendor_family",
    "destination_id",
    "template_type",
    "tag_name",
    "tag_id",
    "tag_category",
    "tag_delivery",
    "scope_status",
    "scope_reason",
    "inventory_status",
    "relevance",
    "expected_firing",
    "actual_firing",
    "fire_count",
    "configuration_field",
    "configured_value",
    "runtime_state",
    "runtime_value",
    "runtime_type",
    "firing_status",
    "parameter_status",
    "non_firing_reason",
    "reason_source",
    "configuration_evidence_id",
    "runtime_evidence_id",
    "execution_error",
    "error_evidence_id",
    "notes",
    "layer_results",
]

CONSENT_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "applicable",
    "scenario_id",
    "scenario",
    "source",
    "expected_state",
    "state_at_event",
    "transport_mode",
    "ads_data_redaction",
    "url_passthrough",
    "transition",
    "tag_consent_checks",
    "override_approved",
    "override_method",
    "override_scope",
    "native_cmp_status",
    "native_cmp_acceptance_in_scope",
    "production_exception_approved",
    "production_approval_evidence_id",
    "restoration_confirmed",
    "blocker_id",
    "approval_evidence_id",
    "evidence_id",
    "status",
    "notes",
]

DESTINATION_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "container_id",
    "vendor_family",
    "destination_id",
    "destination_id_parameter_path",
    "expected_destination_event_name",
    "actual_destination_event_name",
    "destination_event_parameter_path",
    "expected_request_behavior",
    "request_behavior",
    "request_count",
    "request_id",
    "request_method",
    "request_url",
    "expected_endpoint_pattern",
    "expected_parameter_path",
    "actual_parameter_path",
    "field_state",
    "field_value",
    "field_type",
    "request_status",
    "parameter_status",
    "capture_source",
    "vendor_helper_status",
    "evidence_id",
    "vendor_helper_evidence_id",
    "notes",
]

TRIGGER_SEQUENCE_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "tag_name",
    "trigger_mode",
    "expected_trigger_result",
    "actual_trigger_result",
    "conditions",
    "blocking_exceptions",
    "trigger_status",
    "trigger_evidence_id",
    "expected_sequence",
    "actual_sequence",
    "sequence_status",
    "sequence_evidence_id",
    "notes",
]

BUSINESS_RULE_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "rule_id",
    "operator",
    "evaluation_source",
    "declared_rule",
    "actual",
    "expected",
    "status",
    "reason",
    "evidence_id",
]

SENSITIVE_DATA_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "scan_status",
    "scanned_targets",
    "path",
    "category",
    "confidence",
    "basis",
    "allowlisted",
    "finding_status",
    "redacted_value",
    "value_fingerprint",
    "evidence_id",
]

CLIENT_CHECK_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "check_id",
    "category",
    "browser_context_id",
    "comparison",
    "expected",
    "actual",
    "status",
    "limit_source",
    "evidence_id",
    "notes",
]

REGRESSION_HEADERS = [
    "plan_order",
    "event_group_id",
    "requirement_id",
    "event_name",
    "baseline_run_id",
    "baseline_status",
    "current_status",
    "change",
    "regression_status",
    "evidence_id",
    "notes",
]

CONTAINER_HEADERS = [
    "container_id",
    "workspace",
    "role",
    "container_type",
    "preview_environment",
    "version",
    "evidence_id",
    "notes",
]

UNEXPECTED_HEADERS = [
    "unexpected_id",
    "observed_push_id",
    "event_group_id",
    "action_id",
    "case_id",
    "event_index",
    "url",
    "page_state",
    "kind",
    "classification",
    "classification_reason",
    "event_name",
    "tag_name",
    "actual",
    "status",
    "review_basis",
    "review_question",
    "evidence_ids",
    "notes",
]

BLOCKER_HEADERS = [
    "blocker_id",
    "type",
    "checkpoint",
    "description",
    "requirement_ids",
    "attempted_methods",
    "analyst_intervention_required",
    "analyst_help_requested",
    "analyst_response",
    "outcome",
    "status",
    "evidence_ids",
    "notes",
]

EVIDENCE_HEADERS = [
    "evidence_id",
    "kind",
    "source",
    "capture_mode",
    "action_id",
    "event_index",
    "container_id",
    "tag_id",
    "request_id",
    "tag_name",
    "configuration_field",
    "source_detail",
    "path_or_url",
    "captured_at",
    "description",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="Schema-v3 normalized JSON path, or - for stdin.")
    parser.add_argument("output", nargs="?", help="Destination .xlsx path.")
    parser.add_argument("--strict", action="store_true", help="Reject every semantic error.")
    parser.add_argument(
        "--session-ledger",
        type=Path,
        help="Case/action/push ledger required for strict final certification.",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate normalized data without writing a workbook.",
    )
    parser.add_argument("--defects-csv", type=Path)
    parser.add_argument("--defects-md", type=Path)
    parser.add_argument("--stakeholder-summary", type=Path)
    return parser.parse_args()


def load_data(source: str) -> dict[str, Any]:
    raw = sys.stdin.read() if source == "-" else Path(source).read_text(encoding="utf-8")
    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ReportValidationError("Top-level JSON value must be an object.")
    return data


def serialize(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return dumps_structured(value)
    if isinstance(value, bool):
        return "true" if value else "false"
    return value


def set_safe_cell(cell: Cell, value: Any) -> Cell:
    """Write a workbook value without allowing untrusted text to become a formula."""
    serialized = serialize(value)
    if isinstance(serialized, str) and len(serialized) > EXCEL_CELL_LIMIT:
        raise ReportValidationError(
            "A non-tabular workbook cell exceeds Excel's 32,767-character limit; "
            "move the structured evidence to a table where it can be split safely."
        )
    cell.value = serialized
    if isinstance(serialized, str):
        cell.data_type = "s"
    return cell


def _excel_parts(value: Any) -> list[Any]:
    serialized = serialize(value)
    if not isinstance(serialized, str) or len(serialized) <= EXCEL_CELL_LIMIT:
        return [serialized]
    raw_parts = [
        serialized[index : index + EXCEL_CHUNK_PAYLOAD]
        for index in range(0, len(serialized), EXCEL_CHUNK_PAYLOAD)
    ]
    total = len(raw_parts)
    return [
        EXCEL_CONTINUATION_PREFIX.format(part=index, total=total) + part
        for index, part in enumerate(raw_parts, start=1)
    ]


def expanded_row_count(values: Iterable[Any]) -> int:
    """Return how many physical Excel rows one logical row requires."""
    return max((len(_excel_parts(value)) for value in values), default=1)


def append_safe_row(ws, values: Iterable[Any]) -> int:
    """Append one logical row, splitting oversized cells with explicit part markers."""
    serialized = [serialize(value) for value in values]
    parts = [_excel_parts(value) for value in serialized]
    physical_rows = max((len(value_parts) for value_parts in parts), default=1)
    for index in range(physical_rows):
        row_values = [
            (
                value_parts[index]
                if len(value_parts) > 1 and index < len(value_parts)
                else ""
                if len(value_parts) > 1
                else value_parts[0]
            )
            for value_parts in parts
        ]
        ws.append(row_values)
        for cell, value in zip(ws[ws.max_row], row_values, strict=False):
            if isinstance(value, str):
                cell.data_type = "s"
    return physical_rows


def apply_status_fill(cell: Cell) -> None:
    status = status_of(cell.value)
    colour = STATUS_FILLS.get(status)
    if colour:
        cell.fill = PatternFill("solid", fgColor=colour)
        cell.font = Font(bold=True)


def _column_cap(header: str) -> int:
    if any(
        token in header
        for token in (
            "payload",
            "snapshot",
            "attempted_routes",
            "state",
            "notes",
            "reason",
            "description",
            "value",
        )
    ):
        return 80
    return 45


def style_table(
    ws,
    status_headers: tuple[str, ...] = (
        "status",
        "overall_status",
        "occurrence_verdict",
        "raw_verdict",
        "resolved_verdict",
        "variable_verdict",
        "tag_configuration_verdict",
        "tag_firing_verdict",
        "tag_parameter_verdict",
        "destination_request_verdict",
        "destination_parameter_verdict",
        "trigger_logic_verdict",
        "tag_sequence_verdict",
        "consent_verdict",
        "business_rule_verdict",
        "sensitive_data_verdict",
        "client_checks_verdict",
        "regression_verdict",
        "firing_status",
        "parameter_status",
        "request_status",
        "trigger_status",
        "sequence_status",
        "finding_status",
        "scan_status",
        "regression_status",
    ),
) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    header_map = {str(cell.value): cell.column for cell in ws[1]}
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER
        for header in status_headers:
            column = header_map.get(header)
            if column:
                apply_status_fill(row[column - 1])
    for column_index, cells in enumerate(ws.columns, start=1):
        values = [str(cell.value or "") for cell in cells]
        longest = max((len(value) for value in values), default=0)
        header = str(cells[0].value or "")
        ws.column_dimensions[get_column_letter(column_index)].width = min(
            max(longest + 2, 12), _column_cap(header)
        )


def add_table_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: Iterable[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title)
    append_safe_row(ws, headers)
    for row in rows:
        append_safe_row(ws, [row.get(header) for header in headers])
    style_table(ws)


def table_sheet_row_count(headers: list[str], rows: Iterable[dict[str, Any]]) -> int:
    """Return the exact physical row count after safe continuation splitting."""
    return 1 + sum(expanded_row_count(row.get(header) for header in headers) for row in rows)


def _observation_value(observation: Any, field: str) -> Any:
    if not isinstance(observation, dict):
        return ""
    if field not in observation:
        return ""
    return observation[field]


def requirement_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        source = req.get("source", {})
        journey = req.get("journey", {})
        expectation = req.get("expectation", {})
        raw = req.get("raw_api_call") or {}
        resolved = req.get("resolved_data_layer") or {}
        variable = req.get("gtm_variable") or {}
        tag = req.get("tag") or {}
        destination = req.get("destination_request") or {}
        consent = req.get("consent") or {}
        scenario = req.get("scenario") or {}
        verdict = req.get("verdict", {})
        output.append(
            {
                "requirement_id": req.get("requirement_id"),
                "event_group_id": req.get("event_group_id"),
                "plan_order": source.get("plan_order"),
                "source_reference": source.get("reference"),
                "source_file": source.get("file"),
                "source_sheet": source.get("sheet"),
                "source_row": source.get("row"),
                "source_cells": source.get("cells"),
                "source_section": source.get("section"),
                "journey_id": journey.get("journey_id"),
                "step_id": journey.get("step_id"),
                "action": journey.get("action"),
                "action_value": journey.get("action_value"),
                "action_value_type": journey.get("action_value_type"),
                "action_value_source": journey.get("action_value_source"),
                "url": journey.get("url"),
                "selector_or_element": journey.get("selector_or_element"),
                "inferred": journey.get("inferred"),
                "inference_source": journey.get("inference_source"),
                "browser_context_id": req.get("browser_context_id"),
                "container_id": req.get("container_id")
                or tag.get("container_id")
                or data.get("run", {}).get("container_id"),
                "scenario_id": scenario.get("scenario_id"),
                "scenario_kind": scenario.get("kind"),
                "event_name": expectation.get("event_name"),
                "source_mechanism": expectation.get("source_mechanism", "data_layer_push"),
                "field_path": expectation.get("field_path"),
                "match_rule": expectation.get("match_rule"),
                "expected_value": expectation.get("expected_value"),
                "expected_type": expectation.get("expected_type"),
                "expected_occurrence": expectation.get("expected_occurrence"),
                "occurrence_verdict": verdict.get("event_occurrence"),
                "source_signal_verdict": verdict.get("source_signal"),
                "raw_state": raw.get("field_state"),
                "raw_value": _observation_value(raw, "field_value"),
                "raw_type": raw.get("field_type"),
                "resolved_state": resolved.get("field_state"),
                "resolved_value": _observation_value(resolved, "field_value"),
                "resolved_type": resolved.get("field_type"),
                "gtm_variable": variable.get("name"),
                "gtm_variable_state": variable.get("field_state"),
                "gtm_variable_value": _observation_value(variable, "field_value"),
                "gtm_variable_type": variable.get("field_type"),
                "tag_name": tag.get("name"),
                "tag_delivery": expectation.get("tag_delivery"),
                "vendor_family": expectation.get("vendor_family") or tag.get("vendor_family"),
                "destination_id": expectation.get("destination_id") or tag.get("destination_id"),
                "tag_relevance": tag.get("relevance"),
                "expected_firing": tag.get("expected_firing"),
                "actual_firing": tag.get("actual_firing"),
                "fire_count": tag.get("fire_count"),
                "tag_configuration_field": tag.get("configuration_field"),
                "expected_tag_configuration": expectation.get("expected_tag_configuration"),
                "configured_value": tag.get("configured_value"),
                "runtime_state": tag.get("runtime_state"),
                "runtime_value": _observation_value(tag, "runtime_value"),
                "runtime_type": tag.get("runtime_type"),
                "request_behavior": destination.get("request_behavior"),
                "request_count": destination.get("request_count"),
                "request_id": destination.get("request_id"),
                "destination_parameter_path": destination.get("parameter_path"),
                "destination_parameter_value": _observation_value(destination, "field_value"),
                "consent_scenario": consent.get("scenario"),
                "consent_source": consent.get("source"),
                "consent_state": consent.get("state_at_event"),
                "raw_verdict": verdict.get("raw_payload"),
                "resolved_verdict": verdict.get("resolved_data_layer"),
                "variable_verdict": verdict.get("gtm_variable"),
                "tag_configuration_verdict": verdict.get("tag_configuration"),
                "tag_firing_verdict": verdict.get("tag_firing"),
                "tag_parameter_verdict": verdict.get("tag_parameter"),
                "destination_request_verdict": verdict.get("destination_request"),
                "destination_parameter_verdict": verdict.get("destination_parameter"),
                "trigger_logic_verdict": verdict.get("trigger_logic"),
                "tag_sequence_verdict": verdict.get("tag_sequence"),
                "consent_verdict": verdict.get("consent"),
                "business_rule_verdict": verdict.get("business_rule"),
                "sensitive_data_verdict": verdict.get("sensitive_data"),
                "client_checks_verdict": verdict.get("client_checks"),
                "regression_verdict": verdict.get("regression"),
                "overall_status": verdict.get("overall"),
                "failure_layer": verdict.get("failure_layer"),
                "mismatch_or_reason": verdict.get("mismatch"),
                "evidence_ids": req.get("evidence_ids"),
                "notes": req.get("notes"),
            }
        )
    return output


def _defect_observation(requirement: dict[str, Any], layer: str) -> tuple[Any, Any]:
    expectation = requirement.get("expectation", {})
    raw = requirement.get("raw_api_call") or {}
    resolved = requirement.get("resolved_data_layer") or {}
    variable = requirement.get("gtm_variable") or {}
    tag = requirement.get("tag") or {}
    destination = requirement.get("destination_request") or {}
    occurrence = requirement.get("occurrence_evidence") or {}
    mapping = {
        "event_occurrence": (occurrence.get("actual_count"), "number"),
        "source_signal": (
            (requirement.get("source_signal") or {}).get("value"),
            (requirement.get("source_signal") or {}).get("value_type"),
        ),
        "raw_payload": (raw.get("field_value"), raw.get("field_type")),
        "resolved_data_layer": (resolved.get("field_value"), resolved.get("field_type")),
        "gtm_variable": (variable.get("field_value"), variable.get("field_type")),
        "tag_configuration": (tag.get("configured_value"), None),
        "tag_firing": (tag.get("actual_firing"), "string"),
        "tag_parameter": (tag.get("runtime_value"), tag.get("runtime_type")),
        "destination_request": (destination.get("request_behavior"), "string"),
        "destination_parameter": (
            destination.get("field_value"),
            destination.get("field_type"),
        ),
        "consent": ((requirement.get("consent") or {}).get("state_at_event"), "object"),
        "trigger_logic": (requirement.get("trigger_evaluation"), "object"),
        "tag_sequence": (requirement.get("tag_sequence"), "object"),
        "business_rule": (requirement.get("business_rule_results"), "array"),
        "sensitive_data": (requirement.get("sensitive_data_scan"), "object"),
        "client_checks": (requirement.get("client_checks"), "array"),
        "regression": (requirement.get("regression"), "object"),
    }
    return mapping.get(layer, (None, expectation.get("expected_type")))


def _defect_expectation(requirement: dict[str, Any], layer: str) -> tuple[Any, Any]:
    expectation = requirement.get("expectation", {})
    if layer == "event_occurrence":
        return expectation.get("expected_occurrence"), "occurrence_rule"
    if layer == "tag_configuration":
        return expectation.get("expected_tag_configuration"), None
    if layer == "tag_firing":
        return expectation.get("expected_firing"), "string"
    if layer == "destination_request":
        return expectation.get("expected_request_behavior"), "string"
    if layer == "consent":
        return expectation.get("expected_consent_state"), "object"
    if layer == "trigger_logic":
        return expectation.get("trigger_contract"), "object"
    if layer == "tag_sequence":
        return expectation.get("sequence_contract"), "object"
    if layer == "business_rule":
        return expectation.get("business_rules"), "array"
    if layer == "sensitive_data":
        return expectation.get("sensitive_data_policy"), "object"
    if layer == "client_checks":
        return [
            {
                "check_id": row.get("check_id"),
                "comparison": row.get("comparison"),
                "expected": row.get("expected"),
            }
            for row in requirement.get("client_checks", [])
            if isinstance(row, dict)
        ], "array"
    if layer == "regression":
        return "no acceptance regression", "regression_state"
    return expectation.get("expected_value"), expectation.get("expected_type")


def defect_rows(
    data: dict[str, Any],
    session: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Build the actionable non-PASS register used by XLSX and sidecars."""
    run_inventory = {
        str(row.get("event_group_id")): row
        for row in data.get("run", {}).get("event_inventory", [])
        if isinstance(row, dict)
    }
    feedback = {str(row.get("event_group_id")): row for row in event_feedback(data, session)}
    cases = {
        str(row.get("case_id")): row
        for row in (session or {}).get("cases", [])
        if isinstance(row, dict)
    }
    actions = {
        str(row.get("action_id")): row
        for row in (session or {}).get("actions", [])
        if isinstance(row, dict)
    }
    output: list[dict[str, Any]] = []
    for requirement in as_rows(data.get("requirements"), "requirements"):
        verdict = requirement.get("verdict") or {}
        defect_status = status_of(verdict.get("overall"))
        if defect_status == "PASS":
            continue
        group_id = str(requirement.get("event_group_id", ""))
        inventory = run_inventory.get(group_id, {})
        layer = str(verdict.get("failure_layer") or "event_occurrence")
        expected_value, expected_type = _defect_expectation(requirement, layer)
        observed_value, observed_type = _defect_observation(requirement, layer)
        boundary = requirement.get("action_boundary") or {}
        action = actions.get(str(boundary.get("action_id", "")), {})
        case = cases.get(str(action.get("case_id", "")), {})
        reason = str(
            verdict.get("mismatch")
            or verdict.get("review_question")
            or requirement.get("notes")
            or feedback.get(group_id, {}).get("reason")
            or "Non-PASS requirement; inspect the cited layer evidence."
        ).strip()
        output.append(
            {
                "defect_id": f"DEF-REQ-{requirement.get('requirement_id')}",
                "plan_order": inventory.get("plan_order")
                or (requirement.get("source") or {}).get("plan_order"),
                "event_group_id": group_id,
                "event_name": inventory.get("event_name")
                or (requirement.get("expectation") or {}).get("event_name"),
                "status": defect_status,
                "case_id": case.get("case_id"),
                "requirement_id": requirement.get("requirement_id"),
                "placement": case.get("placement"),
                "material_variant": case.get("material_variant"),
                "failed_layer": layer,
                "expected_value": expected_value,
                "expected_type": expected_type,
                "observed_value": observed_value,
                "observed_type": observed_type,
                "concise_reason": reason,
                "exact_retest": feedback.get(group_id, {}).get("retest"),
                "evidence_ids": requirement.get("evidence_ids"),
                "retest_status": "PENDING",
            }
        )
    for unexpected in as_rows(data.get("unexpected"), "unexpected"):
        defect_status = status_of(unexpected.get("status"))
        if defect_status == "PASS":
            continue
        group_id = str(unexpected.get("event_group_id", ""))
        inventory = run_inventory.get(group_id, {})
        case = cases.get(str(unexpected.get("case_id", "")), {})
        output.append(
            {
                "defect_id": f"DEF-UNX-{unexpected.get('unexpected_id')}",
                "plan_order": inventory.get("plan_order"),
                "event_group_id": group_id,
                "event_name": unexpected.get("event_name") or inventory.get("event_name"),
                "status": defect_status,
                "case_id": unexpected.get("case_id"),
                "requirement_id": None,
                "placement": case.get("placement"),
                "material_variant": case.get("material_variant"),
                "failed_layer": "unexpected_business_push",
                "expected_value": "absent in this action window",
                "expected_type": "occurrence_rule",
                "observed_value": unexpected.get("classification"),
                "observed_type": "string",
                "concise_reason": unexpected.get("classification_reason")
                or unexpected.get("review_question")
                or unexpected.get("notes"),
                "exact_retest": feedback.get(group_id, {}).get("retest"),
                "evidence_ids": unexpected.get("evidence_ids"),
                "retest_status": "PENDING",
            }
        )
    for case in cases.values():
        group_id = str(case.get("event_group_id", ""))
        inventory = run_inventory.get(group_id, {})
        case_status = str(case.get("execution_status", "")).upper()
        final_action = actions.get(str(case.get("final_action_id", "")), {})
        layer_results = [
            row
            for row in final_action.get("layer_results", [])
            if isinstance(row, dict)
            and status_of(row.get("status")) not in {"PASS", "NOT_APPLICABLE"}
        ]
        if case_status == "EXECUTED" and not layer_results:
            continue
        if not layer_results:
            layer_results = [
                {
                    "layer": "interaction_case",
                    "status": case_status,
                    "reason": case.get("reason") or "Case did not reach an executed result.",
                    "evidence_ids": [],
                }
            ]
        variant = case.get("material_variant") or {}
        retest = (
            f"{case.get('url')} — {case.get('action')} {case.get('element')} "
            f"at {case.get('placement')}"
            + (f" with {dumps_structured(variant)}" if variant else "")
        )
        for layer_result in layer_results:
            layer = str(layer_result.get("layer", "interaction_case"))
            output.append(
                {
                    "defect_id": f"DEF-CASE-{case.get('case_id')}-{layer}",
                    "plan_order": inventory.get("plan_order"),
                    "event_group_id": group_id,
                    "event_name": inventory.get("event_name"),
                    "status": status_of(layer_result.get("status")) or case_status,
                    "case_id": case.get("case_id"),
                    "requirement_id": None,
                    "tag_name": None,
                    "placement": case.get("placement"),
                    "material_variant": variant,
                    "failed_layer": layer,
                    "expected_value": "applicable layer accepted",
                    "expected_type": "layer_status",
                    "observed_value": layer_result.get("status") or case_status,
                    "observed_type": "status",
                    "concise_reason": layer_result.get("reason") or case.get("reason"),
                    "exact_retest": retest,
                    "evidence_ids": layer_result.get("evidence_ids") or [],
                    "retest_status": "PENDING",
                }
            )
        inventory_by_id = {
            str(row.get("tag_id", "")): row
            for row in case.get("tag_inventory", [])
            if isinstance(row, dict)
        }
        for tag_result in final_action.get("tag_layer_results", []) or []:
            if not isinstance(tag_result, dict) or status_of(tag_result.get("status")) in {
                "PASS",
                "NOT_APPLICABLE",
            }:
                continue
            tag = inventory_by_id.get(str(tag_result.get("tag_id", "")), {})
            layer = str(tag_result.get("layer", "tag_layer"))
            output.append(
                {
                    "defect_id": (
                        f"DEF-TAG-{case.get('case_id')}-{tag_result.get('tag_id')}-{layer}"
                    ),
                    "plan_order": inventory.get("plan_order"),
                    "event_group_id": group_id,
                    "event_name": inventory.get("event_name"),
                    "status": status_of(tag_result.get("status")),
                    "case_id": case.get("case_id"),
                    "requirement_id": None,
                    "tag_name": tag.get("tag_name"),
                    "placement": case.get("placement"),
                    "material_variant": variant,
                    "failed_layer": layer,
                    "expected_value": "per-tag layer accepted",
                    "expected_type": "layer_status",
                    "observed_value": tag_result.get("details"),
                    "observed_type": "object",
                    "concise_reason": tag_result.get("reason"),
                    "exact_retest": retest,
                    "evidence_ids": tag_result.get("evidence_ids") or [],
                    "retest_status": "PENDING",
                }
            )
    for row in output:
        group_feedback = feedback.get(str(row.get("event_group_id", "")), {})
        row["output_contract_version"] = output_contract_version(session)
        row["primary_outcome"] = group_feedback.get("primary_outcome")
        row["anomaly_flags"] = group_feedback.get("anomaly_flags")
    return sorted(
        output,
        key=lambda row: (
            float(row.get("plan_order"))
            if isinstance(row.get("plan_order"), (int, float))
            else 1e18,
            str(row.get("defect_id", "")),
        ),
    )


def journey_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        source = req.get("source", {})
        journey = req.get("journey", {})
        expectation = req.get("expectation", {})
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "event_name": expectation.get("event_name"),
                "requirement_id": req.get("requirement_id"),
                "journey_id": journey.get("journey_id"),
                "step_id": journey.get("step_id"),
                "action": journey.get("action"),
                "url": journey.get("url"),
                "selector_or_element": journey.get("selector_or_element"),
                "inferred": journey.get("inferred"),
                "inference_source": journey.get("inference_source"),
                "confidence": journey.get("confidence"),
                "attempted_routes": journey.get("attempted_routes"),
                "execution_status": journey.get("execution_status"),
                "blocker_id": req.get("blocker_id"),
                "overall_status": req.get("verdict", {}).get("overall"),
                "evidence_ids": req.get("evidence_ids"),
                "notes": req.get("notes"),
            }
        )
    return output


def event_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        raw = req.get("raw_api_call") or {}
        resolved = req.get("resolved_data_layer") or {}
        signal = req.get("source_signal") or {}
        occurrence = req.get("occurrence_evidence") or {}
        boundary = req.get("action_boundary") or {}
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "requirement_id": req.get("requirement_id"),
                "event_name": expectation.get("event_name"),
                "event_observed": req.get("event_observed"),
                "actual_occurrence_count": occurrence.get("actual_count"),
                "occurrence_event_indexes": occurrence.get("event_indexes"),
                "anchor_event_name": occurrence.get("anchor_event_name"),
                "anchor_event_index": occurrence.get("anchor_event_index"),
                "occurrence_evidence_id": occurrence.get("evidence_id"),
                "source_mechanism": expectation.get("source_mechanism", "data_layer_push"),
                "source_signal_capture_source": signal.get("capture_source"),
                "source_signal_observed": signal.get("observed"),
                "source_signal_evidence_id": signal.get("evidence_id"),
                "capture_source": raw.get("capture_source"),
                "event_index": raw.get("event_index"),
                "event_timestamp": raw.get("timestamp"),
                "raw_api_call_payload": raw.get("payload"),
                "raw_field_path": expectation.get("field_path"),
                "raw_field_state": raw.get("field_state"),
                "raw_field_value": _observation_value(raw, "field_value"),
                "raw_field_type": raw.get("field_type"),
                "resolved_data_layer_snapshot": resolved.get("snapshot"),
                "resolved_field_state": resolved.get("field_state"),
                "resolved_field_value": _observation_value(resolved, "field_value"),
                "resolved_field_type": resolved.get("field_type"),
                "action_id": boundary.get("action_id"),
                "retry_of_action_id": boundary.get("retry_of_action_id"),
                "readiness_check_id": boundary.get("readiness_check_id"),
                "settlement_check_id": boundary.get("settlement_check_id"),
                "preview_connected_before": boundary.get("preview_connected_before"),
                "target_ready_before": boundary.get("target_ready_before"),
                "last_event_before": boundary.get("last_event_before"),
                "network_request_cursor_before": boundary.get("network_request_cursor_before"),
                "action_timestamp": boundary.get("action_timestamp"),
                "interaction_outcome": boundary.get("interaction_outcome"),
                "completion_signal": boundary.get("completion_signal"),
                "first_event_after": boundary.get("first_event_after"),
                "settled_final_event": boundary.get("settled_final_event"),
                "network_request_cursor_after": boundary.get("network_request_cursor_after"),
                "quiet_window_ms": boundary.get("quiet_window_ms"),
                "timeout_ms": boundary.get("timeout_ms"),
                "stream_settled": boundary.get("stream_settled"),
                "settlement_reason": boundary.get("settlement_reason"),
                "raw_evidence_id": raw.get("evidence_id"),
                "resolved_evidence_id": resolved.get("evidence_id"),
                "status": req.get("verdict", {}).get("raw_payload"),
                "notes": req.get("notes"),
            }
        )
    return output


def tag_rows(
    data: dict[str, Any],
    session: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    if isinstance(session, dict):
        inventory_by_group = {
            str(row.get("event_group_id", "")): row
            for row in data.get("run", {}).get("event_inventory", [])
            if isinstance(row, dict)
        }
        actions = {
            str(row.get("action_id", "")): row
            for row in session.get("actions", [])
            if isinstance(row, dict)
        }
        output: list[dict[str, Any]] = []
        for case in session.get("cases", []):
            if not isinstance(case, dict):
                continue
            event = inventory_by_group.get(str(case.get("event_group_id", "")), {})
            action = actions.get(str(case.get("final_action_id", "")), {})
            results_by_tag: dict[str, list[dict[str, Any]]] = {}
            for result in action.get("tag_layer_results", []) or []:
                if isinstance(result, dict):
                    results_by_tag.setdefault(str(result.get("tag_id", "")), []).append(result)
            for tag in case.get("tag_inventory", []) or []:
                if not isinstance(tag, dict):
                    continue
                tag_id = str(tag.get("tag_id", ""))
                layers = results_by_tag.get(tag_id, [])
                layer_by_name = {
                    str(row.get("layer", "")): row for row in layers if isinstance(row, dict)
                }
                firing = (layer_by_name.get("tag_firing") or {}).get("details") or {}
                configuration = (layer_by_name.get("tag_configuration") or {}).get("details") or {}
                parameters = (layer_by_name.get("tag_parameter") or {}).get("details") or {}
                output.append(
                    {
                        "plan_order": event.get("plan_order"),
                        "event_group_id": case.get("event_group_id"),
                        "requirement_id": None,
                        "event_name": event.get("event_name"),
                        "container_id": tag.get("container_id"),
                        "vendor_family": tag.get("vendor_family"),
                        "destination_id": tag.get("destination_id"),
                        "template_type": tag.get("template_type"),
                        "tag_name": tag.get("tag_name"),
                        "tag_id": tag_id,
                        "tag_category": tag.get("tag_category"),
                        "tag_delivery": tag.get("tag_delivery"),
                        "scope_status": tag.get("scope_status"),
                        "scope_reason": tag.get("scope_reason"),
                        "inventory_status": case.get("tag_inventory_status"),
                        "relevance": "runtime_discovered",
                        "expected_firing": firing.get("expected_firing"),
                        "actual_firing": firing.get("actual_firing"),
                        "fire_count": firing.get("fire_count"),
                        "configuration_field": configuration.get("configuration_field"),
                        "configured_value": configuration.get("configuration"),
                        "runtime_state": parameters.get("runtime_state"),
                        "runtime_value": parameters.get("parameters"),
                        "runtime_type": parameters.get("runtime_type"),
                        "firing_status": (layer_by_name.get("tag_firing") or {}).get("status"),
                        "parameter_status": (layer_by_name.get("tag_parameter") or {}).get(
                            "status"
                        ),
                        "non_firing_reason": (layer_by_name.get("tag_firing") or {}).get("reason"),
                        "reason_source": "direct runtime evidence" if layers else None,
                        "configuration_evidence_id": (
                            layer_by_name.get("tag_configuration") or {}
                        ).get("evidence_ids"),
                        "runtime_evidence_id": (layer_by_name.get("tag_parameter") or {}).get(
                            "evidence_ids"
                        ),
                        "execution_error": firing.get("execution_error"),
                        "error_evidence_id": firing.get("error_evidence_id"),
                        "notes": case.get("tag_inventory_reason"),
                        "layer_results": layers,
                    }
                )
        return output
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        tag = req.get("tag")
        if not isinstance(tag, dict) or tag.get("applicable") is not True:
            continue
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        verdict = req.get("verdict", {})
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "requirement_id": req.get("requirement_id"),
                "event_name": expectation.get("event_name"),
                "container_id": tag.get("container_id")
                or req.get("container_id")
                or data.get("run", {}).get("container_id"),
                "vendor_family": tag.get("vendor_family") or expectation.get("vendor_family"),
                "destination_id": tag.get("destination_id") or expectation.get("destination_id"),
                "template_type": tag.get("template_type"),
                "tag_name": tag.get("name"),
                "tag_id": None,
                "tag_category": "analytics" if expectation.get("vendor_family") == "ga4" else None,
                "tag_delivery": expectation.get("tag_delivery"),
                "scope_status": "IN_SCOPE",
                "scope_reason": "Plan-declared tag; no session inventory supplied.",
                "inventory_status": None,
                "relevance": tag.get("relevance"),
                "expected_firing": tag.get("expected_firing"),
                "actual_firing": tag.get("actual_firing"),
                "fire_count": tag.get("fire_count"),
                "configuration_field": tag.get("configuration_field"),
                "configured_value": tag.get("configured_value"),
                "runtime_state": tag.get("runtime_state"),
                "runtime_value": _observation_value(tag, "runtime_value"),
                "runtime_type": tag.get("runtime_type"),
                "firing_status": verdict.get("tag_firing"),
                "parameter_status": verdict.get("tag_parameter"),
                "non_firing_reason": tag.get("non_firing_reason"),
                "reason_source": tag.get("reason_source"),
                "configuration_evidence_id": tag.get("configuration_evidence_id"),
                "runtime_evidence_id": tag.get("runtime_evidence_id"),
                "execution_error": tag.get("execution_error"),
                "error_evidence_id": tag.get("error_evidence_id"),
                "notes": req.get("notes"),
                "layer_results": None,
            }
        )
    return output


def layer_verdict_rows(
    data: dict[str, Any],
    session: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    """Return one visible row per event layer and one subrow per in-scope tag layer."""
    if not isinstance(session, dict):
        return []
    inventory = {
        str(row.get("event_group_id", "")): row
        for row in data.get("run", {}).get("event_inventory", [])
        if isinstance(row, dict)
    }
    feedback_by_group = {
        str(row.get("event_group_id", "")): row for row in event_feedback(data, session)
    }
    output: list[dict[str, Any]] = []
    for group_id, feedback in feedback_by_group.items():
        event = inventory.get(group_id, {})
        for row in feedback.get("layer_feedback", []):
            output.append(
                {
                    "plan_order": event.get("plan_order"),
                    "event_group_id": group_id,
                    "event_name": event.get("event_name"),
                    "case_id": row.get("case_id"),
                    "action_id": row.get("action_id"),
                    "tag_id": None,
                    "tag_name": None,
                    "tag_scope_status": None,
                    "layer": row.get("layer"),
                    "status": row.get("status"),
                    "reason": row.get("reason"),
                    "predicate_result": row.get("predicate_result"),
                    "details": None,
                    "evidence_ids": row.get("evidence_ids"),
                    "exact_retest": feedback.get("retest"),
                }
            )
        for tag in feedback.get("tag_feedback", []):
            if tag.get("scope_status") == "OUT_OF_SCOPE":
                output.append(
                    {
                        "plan_order": event.get("plan_order"),
                        "event_group_id": group_id,
                        "event_name": event.get("event_name"),
                        "case_id": tag.get("case_id"),
                        "action_id": tag.get("action_id"),
                        "tag_id": tag.get("tag_id"),
                        "tag_name": tag.get("tag_name"),
                        "tag_scope_status": "OUT_OF_SCOPE",
                        "layer": "concerned_tag_inventory",
                        "status": "NOT_APPLICABLE",
                        "reason": tag.get("scope_reason"),
                        "predicate_result": False,
                        "details": {
                            "tag_category": tag.get("tag_category"),
                            "tag_delivery": tag.get("tag_delivery"),
                        },
                        "evidence_ids": tag.get("evidence_ids"),
                        "exact_retest": feedback.get("retest"),
                    }
                )
                continue
            for layer in tag.get("layers", []):
                output.append(
                    {
                        "plan_order": event.get("plan_order"),
                        "event_group_id": group_id,
                        "event_name": event.get("event_name"),
                        "case_id": tag.get("case_id"),
                        "action_id": tag.get("action_id"),
                        "tag_id": tag.get("tag_id"),
                        "tag_name": tag.get("tag_name"),
                        "tag_scope_status": tag.get("scope_status"),
                        "layer": layer.get("layer"),
                        "status": layer.get("status"),
                        "reason": layer.get("reason"),
                        "predicate_result": None,
                        "details": layer.get("details"),
                        "evidence_ids": layer.get("evidence_ids"),
                        "exact_retest": feedback.get("retest"),
                    }
                )
    canonical_order = {layer: index for index, layer in enumerate(CANONICAL_LAYERS)}
    tag_order = {layer: index for index, layer in enumerate(TAG_RESULT_LAYERS)}
    return sorted(
        output,
        key=lambda row: (
            float(row.get("plan_order"))
            if isinstance(row.get("plan_order"), (int, float))
            else 1e18,
            str(row.get("case_id", "")),
            0 if row.get("tag_id") in (None, "") else 1,
            str(row.get("tag_id") or ""),
            (
                canonical_order.get(str(row.get("layer", "")), len(canonical_order))
                if row.get("tag_id") in (None, "")
                else tag_order.get(str(row.get("layer", "")), len(tag_order))
            ),
        ),
    )


def consent_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        consent = req.get("consent")
        if not isinstance(consent, dict):
            continue
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "requirement_id": req.get("requirement_id"),
                "event_name": expectation.get("event_name"),
                "applicable": consent.get("applicable"),
                "scenario_id": consent.get("scenario_id"),
                "scenario": consent.get("scenario"),
                "source": consent.get("source"),
                "expected_state": expectation.get("expected_consent_state"),
                "state_at_event": consent.get("state_at_event"),
                "transport_mode": consent.get("transport_mode"),
                "ads_data_redaction": consent.get("ads_data_redaction"),
                "url_passthrough": consent.get("url_passthrough"),
                "transition": consent.get("transition"),
                "tag_consent_checks": consent.get("tag_consent_checks"),
                "override_approved": consent.get("override_approved"),
                "override_method": consent.get("override_method"),
                "override_scope": consent.get("override_scope"),
                "native_cmp_status": consent.get("native_cmp_status"),
                "native_cmp_acceptance_in_scope": consent.get("native_cmp_acceptance_in_scope"),
                "production_exception_approved": consent.get("production_exception_approved"),
                "production_approval_evidence_id": consent.get("production_approval_evidence_id"),
                "restoration_confirmed": consent.get("restoration_confirmed"),
                "blocker_id": consent.get("blocker_id"),
                "approval_evidence_id": consent.get("approval_evidence_id"),
                "evidence_id": consent.get("evidence_id"),
                "status": req.get("verdict", {}).get("consent"),
                "notes": req.get("notes"),
            }
        )
    return output


def destination_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        destination = req.get("destination_request")
        if not isinstance(destination, dict) or destination.get("applicable") is not True:
            continue
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        verdict = req.get("verdict", {})
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "requirement_id": req.get("requirement_id"),
                "event_name": expectation.get("event_name"),
                "container_id": destination.get("container_id")
                or req.get("container_id")
                or data.get("run", {}).get("container_id"),
                "vendor_family": destination.get("vendor_family"),
                "destination_id": destination.get("destination_id"),
                "destination_id_parameter_path": expectation.get("destination_id_parameter_path"),
                "expected_destination_event_name": expectation.get("destination_event_name"),
                "actual_destination_event_name": destination.get("event_name"),
                "destination_event_parameter_path": expectation.get(
                    "destination_event_parameter_path"
                ),
                "expected_request_behavior": expectation.get("expected_request_behavior"),
                "request_behavior": destination.get("request_behavior"),
                "request_count": destination.get("request_count"),
                "request_id": destination.get("request_id"),
                "request_method": destination.get("method"),
                "request_url": destination.get("request_url"),
                "expected_endpoint_pattern": expectation.get("expected_endpoint_pattern"),
                "expected_parameter_path": expectation.get("destination_parameter_path"),
                "actual_parameter_path": destination.get("parameter_path"),
                "field_state": destination.get("field_state"),
                "field_value": _observation_value(destination, "field_value"),
                "field_type": destination.get("field_type"),
                "request_status": verdict.get("destination_request"),
                "parameter_status": verdict.get("destination_parameter"),
                "capture_source": destination.get("capture_source"),
                "vendor_helper_status": destination.get("vendor_helper_status"),
                "evidence_id": destination.get("evidence_id"),
                "vendor_helper_evidence_id": destination.get("vendor_helper_evidence_id"),
                "notes": req.get("notes"),
            }
        )
    return output


def trigger_sequence_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        trigger = req.get("trigger_evaluation")
        sequence = req.get("tag_sequence")
        if not isinstance(trigger, dict) and not isinstance(sequence, dict):
            continue
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        trigger_contract = expectation.get("trigger_contract") or {}
        sequence_contract = expectation.get("sequence_contract") or {}
        tag = req.get("tag") or {}
        verdict = req.get("verdict", {})
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "requirement_id": req.get("requirement_id"),
                "event_name": expectation.get("event_name"),
                "tag_name": tag.get("name"),
                "trigger_mode": (trigger or {}).get("mode"),
                "expected_trigger_result": trigger_contract.get("expected_result"),
                "actual_trigger_result": (trigger or {}).get("actual_result"),
                "conditions": (trigger or {}).get("conditions"),
                "blocking_exceptions": (trigger or {}).get("blocking_exceptions"),
                "trigger_status": verdict.get("trigger_logic"),
                "trigger_evidence_id": (trigger or {}).get("evidence_id"),
                "expected_sequence": sequence_contract.get("expected_order"),
                "actual_sequence": (sequence or {}).get("actual_order"),
                "sequence_status": verdict.get("tag_sequence"),
                "sequence_evidence_id": (sequence or {}).get("evidence_id"),
                "notes": req.get("notes"),
            }
        )
    return output


def business_rule_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    computed = {
        (str(item.get("requirement_id")), str(item.get("rule_id"))): item
        for item in evaluate_report_business_rules(data)
    }
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        rules = expectation.get("business_rules")
        if not isinstance(rules, list):
            continue
        stored = {
            str(item.get("rule_id")): item
            for item in req.get("business_rule_results", [])
            if isinstance(item, dict)
        }
        for rule in rules:
            if not isinstance(rule, dict):
                continue
            rule_id = str(rule.get("rule_id", ""))
            result = stored.get(rule_id, {})
            evaluated = computed.get((str(req.get("requirement_id")), rule_id), {})
            output.append(
                {
                    "plan_order": source.get("plan_order"),
                    "event_group_id": req.get("event_group_id"),
                    "requirement_id": req.get("requirement_id"),
                    "event_name": expectation.get("event_name"),
                    "rule_id": rule_id,
                    "operator": rule.get("operator"),
                    "evaluation_source": result.get(
                        "evaluation_source",
                        evaluated.get("evaluation_source"),
                    ),
                    "declared_rule": rule,
                    "actual": result.get("actual", evaluated.get("actual")),
                    "expected": result.get("expected", evaluated.get("expected")),
                    "status": result.get("status", evaluated.get("status")),
                    "reason": result.get("reason", evaluated.get("reason")),
                    "evidence_id": result.get("evidence_id"),
                }
            )
    return output


def sensitive_data_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        scan = req.get("sensitive_data_scan")
        if not isinstance(scan, dict) or scan.get("applicable") is not True:
            continue
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        findings = scan.get("findings")
        rows = findings if isinstance(findings, list) and findings else [{}]
        for finding in rows:
            if not isinstance(finding, dict):
                continue
            output.append(
                {
                    "plan_order": source.get("plan_order"),
                    "event_group_id": req.get("event_group_id"),
                    "requirement_id": req.get("requirement_id"),
                    "event_name": expectation.get("event_name"),
                    "scan_status": scan.get("status"),
                    "scanned_targets": scan.get("scanned_targets"),
                    "path": finding.get("path"),
                    "category": finding.get("category"),
                    "confidence": finding.get("confidence"),
                    "basis": finding.get("basis"),
                    "allowlisted": finding.get("allowlisted"),
                    "finding_status": finding.get("status"),
                    "redacted_value": finding.get("redacted_value"),
                    "value_fingerprint": finding.get("value_fingerprint"),
                    "evidence_id": scan.get("evidence_id"),
                }
            )
    return output


def client_check_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        for check in req.get("client_checks", []):
            if not isinstance(check, dict):
                continue
            output.append(
                {
                    "plan_order": source.get("plan_order"),
                    "event_group_id": req.get("event_group_id"),
                    "requirement_id": req.get("requirement_id"),
                    "event_name": expectation.get("event_name"),
                    "check_id": check.get("check_id"),
                    "category": check.get("category"),
                    "browser_context_id": check.get("context_id") or req.get("browser_context_id"),
                    "comparison": check.get("comparison"),
                    "expected": check.get("expected"),
                    "actual": check.get("actual"),
                    "status": check.get("status"),
                    "limit_source": check.get("limit_source"),
                    "evidence_id": check.get("evidence_id"),
                    "notes": check.get("notes"),
                }
            )
    return output


def regression_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    output = []
    for req in as_rows(data.get("requirements"), "requirements"):
        regression = req.get("regression")
        if not isinstance(regression, dict) or regression.get("applicable") is not True:
            continue
        source = req.get("source", {})
        expectation = req.get("expectation", {})
        output.append(
            {
                "plan_order": source.get("plan_order"),
                "event_group_id": req.get("event_group_id"),
                "requirement_id": req.get("requirement_id"),
                "event_name": expectation.get("event_name"),
                "baseline_run_id": regression.get("baseline_run_id"),
                "baseline_status": regression.get("baseline_status"),
                "current_status": regression.get("current_status"),
                "change": regression.get("change"),
                "regression_status": req.get("verdict", {}).get("regression"),
                "evidence_id": regression.get("evidence_id"),
                "notes": regression.get("notes"),
            }
        )
    return output


def container_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    run = data.get("run", {})
    containers = run.get("containers")
    if not isinstance(containers, list):
        containers = [
            {
                "container_id": run.get("container_id"),
                "workspace": run.get("workspace"),
                "role": "primary",
                "container_type": "web",
            }
        ]
    return [
        {header: container.get(header) for header in CONTAINER_HEADERS}
        for container in containers
        if isinstance(container, dict)
    ]


def client_category(req: dict[str, Any]) -> str:
    status = status_of(req.get("verdict", {}).get("overall"))
    status_categories = {
        "PASS": "tested and correct",
        "BLOCKED": "blocked journey or unavailable element",
        "REVIEW": "review or outside scope",
        "NOT_TESTED": "review or outside scope",
    }
    if status in status_categories:
        return status_categories[status]
    if req.get("event_observed") is False:
        return "expected event not triggered"
    verdict = req.get("verdict", {})
    layer = str(verdict.get("failure_layer", "")).lower()
    raw = req.get("raw_api_call") or {}
    tag = req.get("tag") or {}
    if raw.get("field_state") == "absent":
        return "raw dataLayer field missing"
    if tag.get("fire_count", 0) > 1 or layer in {"duplicate_tag", "unexpected_tag"}:
        return "tag fired unexpectedly or duplicated"
    category_by_layer = {
        "raw_payload": "wrong raw value/type",
        "raw_value": "wrong raw value/type",
        "raw_type": "wrong raw value/type",
        "resolved_data_layer": "resolved Data Layer or GTM variable mismatch",
        "gtm_variable": "resolved Data Layer or GTM variable mismatch",
        "tag_firing": "tag not fired",
        "tag_not_fired": "tag not fired",
        "tag_configuration": "tag configuration mismatch",
        "configuration": "tag configuration mismatch",
        "tag_parameter": "runtime tag parameter mismatch",
        "runtime_parameter": "runtime tag parameter mismatch",
        "destination_request": "client-side destination request mismatch",
        "destination_parameter": "client-side destination request mismatch",
        "trigger_logic": "trigger or tag-sequencing mismatch",
        "tag_sequence": "trigger or tag-sequencing mismatch",
        "business_rule": "cross-field business-rule mismatch",
        "cross_field": "cross-field business-rule mismatch",
        "sensitive_data": "sensitive-data exposure or review",
        "pii": "sensitive-data exposure or review",
        "client_checks": "client-side browser-context mismatch",
        "spa": "client-side browser-context mismatch",
        "cross_domain": "client-side browser-context mismatch",
        "responsive": "client-side browser-context mismatch",
        "regression": "regression from a previously passing requirement",
        "consent": "consent/timing issue",
        "timing": "consent/timing issue",
    }
    if layer in category_by_layer:
        return category_by_layer[layer]
    return "other confirmed mismatch"


def add_client_summary(
    wb: Workbook,
    data: dict[str, Any],
    warnings: list[str],
    session: dict[str, Any] | None = None,
) -> None:
    ws = wb.active
    ws.title = "Client Summary"
    run = data.get("run", {})
    rollup = event_feedback(data, session)
    overall = worst_status(
        [item["status"] for item in rollup]
        + [status_of(item) for item in as_rows(data.get("unexpected"), "unexpected")]
    )
    counts = Counter(item["status"] for item in rollup)

    set_safe_cell(ws["A1"], run.get("report_title", "GTM Client Recette"))
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    summary_rows = [
        ("Overall status", overall),
        ("Output contract", output_contract_version(session)),
        ("Acceptance scope", run.get("acceptance_scope")),
        ("Client", run.get("client")),
        ("Environment", run.get("environment")),
        ("Target URL", run.get("site_url")),
        (
            "Container(s) / workspace(s)",
            run.get("containers") or f"{run.get('container_id')} / {run.get('workspace')}",
        ),
        ("Tracking plan", run.get("tracking_plan_source")),
        ("Generated at", datetime.now(UTC).isoformat(timespec="seconds")),
        ("Validation warnings", len(warnings)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        set_safe_cell(ws.cell(row=row_index, column=1), label)
        set_safe_cell(ws.cell(row=row_index, column=2), value)
    apply_status_fill(ws["B3"])

    event_start = 14
    set_safe_cell(ws.cell(row=event_start, column=1), "Plan-ordered event status")
    ws.cell(row=event_start, column=1).fill = SECTION_FILL
    ws.cell(row=event_start, column=1).font = Font(bold=True)
    event_headers = [
        "plan_order",
        "event_name",
        "status",
        "primary_outcome",
        "anomaly_flags",
        "requirement_count",
        "case_counts",
        "verified_layers",
        "reason",
        "retest",
        "evidence_ids",
    ]
    for column, header in enumerate(event_headers, start=1):
        cell = set_safe_cell(ws.cell(row=event_start + 1, column=column), header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_index, item in enumerate(rollup, start=event_start + 2):
        for column, header in enumerate(event_headers, start=1):
            set_safe_cell(ws.cell(row=row_index, column=column), item.get(header))
        apply_status_fill(ws.cell(row=row_index, column=3))

    counts_start = event_start + len(rollup) + 4
    set_safe_cell(ws.cell(row=counts_start, column=1), "Event status totals")
    ws.cell(row=counts_start, column=1).fill = SECTION_FILL
    ws.cell(row=counts_start, column=1).font = Font(bold=True)
    for offset, status in enumerate(("PASS", "FAIL", "BLOCKED", "REVIEW", "NOT_TESTED"), start=1):
        set_safe_cell(ws.cell(row=counts_start + offset, column=1), status)
        set_safe_cell(ws.cell(row=counts_start + offset, column=2), counts.get(status, 0))
        apply_status_fill(ws.cell(row=counts_start + offset, column=1))

    category_start = counts_start + 7
    set_safe_cell(
        ws.cell(row=category_start, column=1),
        "Requirement outcome categories",
    )
    ws.cell(row=category_start, column=1).fill = SECTION_FILL
    ws.cell(row=category_start, column=1).font = Font(bold=True)
    category_counts = Counter(
        client_category(req) for req in as_rows(data.get("requirements"), "requirements")
    )
    for offset, (category, count) in enumerate(sorted(category_counts.items()), start=1):
        set_safe_cell(ws.cell(row=category_start + offset, column=1), category)
        set_safe_cell(ws.cell(row=category_start + offset, column=2), count)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 80
    ws.column_dimensions["J"].width = 80
    ws.column_dimensions["K"].width = 50
    ws.freeze_panes = "A3"


def add_run_context(wb: Workbook, run: dict[str, Any]) -> None:
    ws = wb.create_sheet("Run Context")
    append_safe_row(ws, ["field", "value"])
    for key, value in run.items():
        append_safe_row(ws, [key, value])
    style_table(ws)


def _generic_rows(rows: list[dict[str, Any]], headers: list[str]) -> list[dict[str, Any]]:
    return [{header: row.get(header) for header in headers} for row in rows]


COVERAGE_HEADERS = [
    "coverage_decision_id",
    "event_group_id",
    "revision",
    "status",
    "population_scope",
    "population_complete",
    "discovery_sources",
    "dimensions",
    "limitations",
    "recorded_at",
    "frozen_at",
]
SCENARIO_HEADERS = [
    "event_group_id",
    "coverage_decision_id",
    "scenario_class_id",
    "name",
    "selection_mode",
    "population_source",
    "population_estimate",
    "selection_method",
    "required_sample_roles",
    "dimension_values",
    "behavior_signature",
    "case_ids",
    "limitations",
]
SEMANTIC_HEADERS = [
    "check_id",
    "event_group_id",
    "case_id",
    "action_id",
    "requirement_id",
    "kind",
    "subject",
    "authority",
    "comparison",
    "anchor_state",
    "anchor_value",
    "observed_value",
    "status",
    "reason",
    "evidence_ids",
]
JOURNEY_STATE_HEADERS = [
    "state_id",
    "event_group_id",
    "case_id",
    "action_id",
    "phase",
    "values",
    "summary",
    "captured_at",
    "evidence_ids",
]
STREAM_HEADERS = [
    "segment_id",
    "kind",
    "status",
    "connection_epoch",
    "action_id",
    "previous_segment_id",
    "start_preview_event_index",
    "end_preview_event_index",
    "start_datalayer_call_index",
    "end_datalayer_call_index",
    "observed_push_ids",
    "started_at",
    "ended_at",
    "evidence_ids",
]
HANDOFF_HEADERS = [
    "handoff_id",
    "gate_type",
    "status",
    "event_group_id",
    "case_id",
    "action_id",
    "reason",
    "requested_at",
    "resumed_at",
    "before_binding",
    "after_binding",
    "evidence_ids",
]
GATED_FLOW_HEADERS = [
    "flow_id",
    "kind",
    "status",
    "event_group_id",
    "case_id",
    "action_id",
    "states",
    "consent_outcome",
    "captcha_outcome",
    "synthetic_data_used",
    "safe_environment_confirmed",
    "handoff_id",
    "reason",
    "evidence_ids",
]
CONCLUSION_HEADERS = [
    "plan_order",
    "event_group_id",
    "event_name",
    "status",
    "status_label",
    "technical_status",
    "semantic_status",
    "stream_status",
    "coverage_status",
    "layers_inspected",
    "why",
]


def scenario_class_rows(session: dict[str, Any]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for decision in session.get("coverage_decisions", []):
        if not isinstance(decision, dict):
            continue
        for scenario in decision.get("scenario_classes", []):
            if isinstance(scenario, dict):
                output.append(
                    {
                        "event_group_id": decision.get("event_group_id"),
                        "coverage_decision_id": decision.get("coverage_decision_id"),
                        **scenario,
                    }
                )
    return output


def conclusion_rows(data: dict[str, Any], session: dict[str, Any]) -> list[dict[str, Any]]:
    return list(final_conclusion(data, session).get("events", []))


def build_workbook(
    data: dict[str, Any],
    output: Path,
    warnings: list[str] | None = None,
    session: dict[str, Any] | None = None,
) -> None:
    warnings = warnings or []
    refuse_unsafe_evidence(validate(data, strict=False))
    if session is not None:
        execution_errors = validate_session(session, results=data, final=True)
        if execution_errors:
            raise ReportValidationError("\n".join(execution_errors))
    wb = Workbook()
    selected_case_headers = case_headers(session)
    selected_push_headers = push_headers(session)
    add_client_summary(wb, data, warnings, session)
    add_table_sheet(wb, "Defect Register", DEFECT_HEADERS, defect_rows(data, session))
    add_table_sheet(wb, "Requirement Matrix", REQUIREMENT_HEADERS, requirement_rows(data))
    add_table_sheet(wb, "Journey Coverage", JOURNEY_HEADERS, journey_rows(data))
    add_table_sheet(
        wb,
        "Interaction Cases",
        selected_case_headers,
        _generic_rows(case_action_rows(session or {}), selected_case_headers),
    )
    add_table_sheet(
        wb,
        "Layer Verdicts",
        LAYER_VERDICT_HEADERS,
        layer_verdict_rows(data, session),
    )
    add_table_sheet(wb, "Event Evidence", EVENT_HEADERS, event_rows(data))
    add_table_sheet(
        wb,
        "Observed Push Stream",
        selected_push_headers,
        _generic_rows(business_push_rows(session or {}), selected_push_headers),
    )
    add_table_sheet(wb, "Tag Evidence", TAG_HEADERS, tag_rows(data, session))
    add_table_sheet(
        wb,
        "Destination Evidence",
        DESTINATION_HEADERS,
        destination_rows(data),
    )
    add_table_sheet(
        wb,
        "Trigger & Sequence",
        TRIGGER_SEQUENCE_HEADERS,
        trigger_sequence_rows(data),
    )
    add_table_sheet(wb, "Consent", CONSENT_HEADERS, consent_rows(data))
    add_table_sheet(
        wb,
        "Business Rules",
        BUSINESS_RULE_HEADERS,
        business_rule_rows(data),
    )
    add_table_sheet(
        wb,
        "Sensitive Data",
        SENSITIVE_DATA_HEADERS,
        sensitive_data_rows(data),
    )
    add_table_sheet(
        wb,
        "Client Checks",
        CLIENT_CHECK_HEADERS,
        client_check_rows(data),
    )
    add_table_sheet(
        wb,
        "Regression",
        REGRESSION_HEADERS,
        regression_rows(data),
    )
    add_table_sheet(
        wb,
        "Container Context",
        CONTAINER_HEADERS,
        container_rows(data),
    )
    add_table_sheet(
        wb,
        "Unexpected Events-Tags",
        UNEXPECTED_HEADERS,
        _generic_rows(as_rows(data.get("unexpected"), "unexpected"), UNEXPECTED_HEADERS),
    )
    add_table_sheet(
        wb,
        "Blockers",
        BLOCKER_HEADERS,
        _generic_rows(as_rows(data.get("blockers"), "blockers"), BLOCKER_HEADERS),
    )
    add_table_sheet(
        wb,
        "Evidence Catalogue",
        EVIDENCE_HEADERS,
        _generic_rows(as_rows(data.get("evidence"), "evidence"), EVIDENCE_HEADERS),
    )
    if isinstance(session, dict) and session.get("operator_contract_version") == 2:
        add_table_sheet(
            wb,
            "Coverage Decisions",
            COVERAGE_HEADERS,
            _generic_rows(
                [row for row in session.get("coverage_decisions", []) if isinstance(row, dict)],
                COVERAGE_HEADERS,
            ),
        )
        add_table_sheet(
            wb,
            "Scenario Classes",
            SCENARIO_HEADERS,
            _generic_rows(scenario_class_rows(session), SCENARIO_HEADERS),
        )
        add_table_sheet(
            wb,
            "Semantic Checks",
            SEMANTIC_HEADERS,
            _generic_rows(
                [row for row in session.get("semantic_checks", []) if isinstance(row, dict)],
                SEMANTIC_HEADERS,
            ),
        )
        add_table_sheet(
            wb,
            "Journey State",
            JOURNEY_STATE_HEADERS,
            _generic_rows(
                [row for row in session.get("journey_states", []) if isinstance(row, dict)],
                JOURNEY_STATE_HEADERS,
            ),
        )
        add_table_sheet(
            wb,
            "Stream Segments",
            STREAM_HEADERS,
            _generic_rows(
                [row for row in session.get("stream_segments", []) if isinstance(row, dict)],
                STREAM_HEADERS,
            ),
        )
        add_table_sheet(
            wb,
            "Protected Handoffs",
            HANDOFF_HEADERS,
            _generic_rows(
                [row for row in session.get("protected_handoffs", []) if isinstance(row, dict)],
                HANDOFF_HEADERS,
            ),
        )
        add_table_sheet(
            wb,
            "Gated Flows",
            GATED_FLOW_HEADERS,
            _generic_rows(
                [row for row in session.get("gated_flows", []) if isinstance(row, dict)],
                GATED_FLOW_HEADERS,
            ),
        )
        add_table_sheet(
            wb,
            "Final Conclusion",
            CONCLUSION_HEADERS,
            _generic_rows(conclusion_rows(data, session), CONCLUSION_HEADERS),
        )
    add_run_context(wb, data.get("run", {}))

    evidence_ws = wb["Evidence Catalogue"]
    path_column = EVIDENCE_HEADERS.index("path_or_url") + 1
    for row_index in range(2, evidence_ws.max_row + 1):
        cell = evidence_ws.cell(row=row_index, column=path_column)
        target = str(cell.value or "").strip()
        if target and (
            target.startswith(("https://", "http://"))
            or Path(target).is_absolute()
            or Path(target).exists()
        ):
            cell.hyperlink = target
            cell.style = "Hyperlink"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    validate_workbook(output, data, session)


def validate_workbook(
    path: Path,
    data: dict[str, Any],
    session: dict[str, Any] | None = None,
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        selected_case_headers = case_headers(session)
        selected_push_headers = push_headers(session)
        case_rows_for_validation = _generic_rows(
            case_action_rows(session or {}), selected_case_headers
        )
        push_rows_for_validation = _generic_rows(
            business_push_rows(session or {}), selected_push_headers
        )
        layer_rows_for_validation = layer_verdict_rows(data, session)
        evidence_rows_for_validation = _generic_rows(
            as_rows(data.get("evidence"), "evidence"), EVIDENCE_HEADERS
        )
        expected_sheets = required_sheets(session)
        missing = [title for title in expected_sheets if title not in workbook.sheetnames]
        if missing:
            raise ReportValidationError(
                "Generated workbook is missing sheets: " + ", ".join(missing)
            )
        if workbook.sheetnames != expected_sheets:
            raise ReportValidationError("Generated workbook sheets are not in the required order.")
        expected_rows = {
            "Defect Register": table_sheet_row_count(DEFECT_HEADERS, defect_rows(data, session)),
            "Requirement Matrix": table_sheet_row_count(
                REQUIREMENT_HEADERS, requirement_rows(data)
            ),
            "Journey Coverage": table_sheet_row_count(JOURNEY_HEADERS, journey_rows(data)),
            "Interaction Cases": table_sheet_row_count(
                selected_case_headers,
                case_rows_for_validation,
            ),
            "Layer Verdicts": table_sheet_row_count(
                LAYER_VERDICT_HEADERS, layer_rows_for_validation
            ),
            "Event Evidence": table_sheet_row_count(EVENT_HEADERS, event_rows(data)),
            "Observed Push Stream": table_sheet_row_count(
                selected_push_headers,
                push_rows_for_validation,
            ),
            "Tag Evidence": table_sheet_row_count(TAG_HEADERS, tag_rows(data, session)),
            "Destination Evidence": table_sheet_row_count(
                DESTINATION_HEADERS, destination_rows(data)
            ),
            "Trigger & Sequence": table_sheet_row_count(
                TRIGGER_SEQUENCE_HEADERS, trigger_sequence_rows(data)
            ),
            "Consent": table_sheet_row_count(CONSENT_HEADERS, consent_rows(data)),
            "Business Rules": table_sheet_row_count(
                BUSINESS_RULE_HEADERS, business_rule_rows(data)
            ),
            "Sensitive Data": table_sheet_row_count(
                SENSITIVE_DATA_HEADERS, sensitive_data_rows(data)
            ),
            "Client Checks": table_sheet_row_count(CLIENT_CHECK_HEADERS, client_check_rows(data)),
            "Regression": table_sheet_row_count(REGRESSION_HEADERS, regression_rows(data)),
            "Container Context": table_sheet_row_count(CONTAINER_HEADERS, container_rows(data)),
            "Unexpected Events-Tags": table_sheet_row_count(
                UNEXPECTED_HEADERS,
                _generic_rows(as_rows(data.get("unexpected"), "unexpected"), UNEXPECTED_HEADERS),
            ),
            "Blockers": table_sheet_row_count(
                BLOCKER_HEADERS,
                _generic_rows(as_rows(data.get("blockers"), "blockers"), BLOCKER_HEADERS),
            ),
            "Evidence Catalogue": table_sheet_row_count(
                EVIDENCE_HEADERS,
                evidence_rows_for_validation,
            ),
        }
        expected_headers = {
            "Defect Register": DEFECT_HEADERS,
            "Requirement Matrix": REQUIREMENT_HEADERS,
            "Journey Coverage": JOURNEY_HEADERS,
            "Interaction Cases": selected_case_headers,
            "Layer Verdicts": LAYER_VERDICT_HEADERS,
            "Event Evidence": EVENT_HEADERS,
            "Observed Push Stream": selected_push_headers,
            "Tag Evidence": TAG_HEADERS,
            "Destination Evidence": DESTINATION_HEADERS,
            "Trigger & Sequence": TRIGGER_SEQUENCE_HEADERS,
            "Consent": CONSENT_HEADERS,
            "Business Rules": BUSINESS_RULE_HEADERS,
            "Sensitive Data": SENSITIVE_DATA_HEADERS,
            "Client Checks": CLIENT_CHECK_HEADERS,
            "Regression": REGRESSION_HEADERS,
            "Container Context": CONTAINER_HEADERS,
            "Unexpected Events-Tags": UNEXPECTED_HEADERS,
            "Blockers": BLOCKER_HEADERS,
            "Evidence Catalogue": EVIDENCE_HEADERS,
        }
        conclusion_rows_for_validation: list[dict[str, Any]] = []
        if isinstance(session, dict) and session.get("operator_contract_version") == 2:
            conclusion_rows_for_validation = _generic_rows(
                conclusion_rows(data, session), CONCLUSION_HEADERS
            )
            expected_rows.update(
                {
                    "Coverage Decisions": table_sheet_row_count(
                        COVERAGE_HEADERS,
                        _generic_rows(
                            [
                                row
                                for row in session.get("coverage_decisions", [])
                                if isinstance(row, dict)
                            ],
                            COVERAGE_HEADERS,
                        ),
                    ),
                    "Scenario Classes": table_sheet_row_count(
                        SCENARIO_HEADERS,
                        _generic_rows(scenario_class_rows(session), SCENARIO_HEADERS),
                    ),
                    "Semantic Checks": table_sheet_row_count(
                        SEMANTIC_HEADERS,
                        _generic_rows(
                            [
                                row
                                for row in session.get("semantic_checks", [])
                                if isinstance(row, dict)
                            ],
                            SEMANTIC_HEADERS,
                        ),
                    ),
                    "Journey State": table_sheet_row_count(
                        JOURNEY_STATE_HEADERS,
                        _generic_rows(
                            [
                                row
                                for row in session.get("journey_states", [])
                                if isinstance(row, dict)
                            ],
                            JOURNEY_STATE_HEADERS,
                        ),
                    ),
                    "Stream Segments": table_sheet_row_count(
                        STREAM_HEADERS,
                        _generic_rows(
                            [
                                row
                                for row in session.get("stream_segments", [])
                                if isinstance(row, dict)
                            ],
                            STREAM_HEADERS,
                        ),
                    ),
                    "Protected Handoffs": table_sheet_row_count(
                        HANDOFF_HEADERS,
                        _generic_rows(
                            [
                                row
                                for row in session.get("protected_handoffs", [])
                                if isinstance(row, dict)
                            ],
                            HANDOFF_HEADERS,
                        ),
                    ),
                    "Gated Flows": table_sheet_row_count(
                        GATED_FLOW_HEADERS,
                        _generic_rows(
                            [
                                row
                                for row in session.get("gated_flows", [])
                                if isinstance(row, dict)
                            ],
                            GATED_FLOW_HEADERS,
                        ),
                    ),
                    "Final Conclusion": table_sheet_row_count(
                        CONCLUSION_HEADERS,
                        conclusion_rows_for_validation,
                    ),
                }
            )
            expected_headers.update(
                {
                    "Coverage Decisions": COVERAGE_HEADERS,
                    "Scenario Classes": SCENARIO_HEADERS,
                    "Semantic Checks": SEMANTIC_HEADERS,
                    "Journey State": JOURNEY_STATE_HEADERS,
                    "Stream Segments": STREAM_HEADERS,
                    "Protected Handoffs": HANDOFF_HEADERS,
                    "Gated Flows": GATED_FLOW_HEADERS,
                    "Final Conclusion": CONCLUSION_HEADERS,
                }
            )
        for title, count in expected_rows.items():
            sheet = workbook[title]
            actual_headers = [
                sheet.cell(row=1, column=index).value
                for index in range(1, len(expected_headers[title]) + 1)
            ]
            if actual_headers != expected_headers[title] or sheet.max_column != len(
                expected_headers[title]
            ):
                raise ReportValidationError(
                    f"Generated workbook sheet '{title}' has an invalid column contract."
                )
            if sheet.max_row != count:
                raise ReportValidationError(
                    f"Generated workbook sheet '{title}' has {sheet.max_row} rows; expected {count}."
                )
            if not sheet.auto_filter.ref:
                raise ReportValidationError(
                    f"Generated workbook sheet '{title}' is missing its filter."
                )
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and len(cell.value) > EXCEL_CELL_LIMIT:
                        raise ReportValidationError(
                            f"Generated workbook sheet '{title}' contains an oversized cell."
                        )

        def assert_projection(
            title: str,
            headers: list[str],
            rows: list[dict[str, Any]],
            keys: tuple[str, ...],
        ) -> None:
            sheet = workbook[title]
            positions = {header: headers.index(header) + 1 for header in keys}
            excel_row = 2
            for logical_row in rows:
                for key, column in positions.items():
                    if sheet.cell(row=excel_row, column=column).value != serialize(
                        logical_row.get(key)
                    ):
                        raise ReportValidationError(
                            f"Generated workbook sheet '{title}' changed decision field '{key}'."
                        )
                excel_row += expanded_row_count(logical_row.get(header) for header in headers)

        assert_projection(
            "Interaction Cases",
            selected_case_headers,
            case_rows_for_validation,
            ("event_group_id", "case_id", "execution_status"),
        )
        assert_projection(
            "Layer Verdicts",
            LAYER_VERDICT_HEADERS,
            layer_rows_for_validation,
            ("event_group_id", "case_id", "layer", "status"),
        )
        assert_projection(
            "Evidence Catalogue",
            EVIDENCE_HEADERS,
            evidence_rows_for_validation,
            ("evidence_id", "kind", "path_or_url"),
        )
        if conclusion_rows_for_validation:
            assert_projection(
                "Final Conclusion",
                CONCLUSION_HEADERS,
                conclusion_rows_for_validation,
                (
                    "event_group_id",
                    "status",
                    "technical_status",
                    "semantic_status",
                    "stream_status",
                    "coverage_status",
                    "why",
                ),
            )

        evidence_sheet = workbook["Evidence Catalogue"]
        path_column = EVIDENCE_HEADERS.index("path_or_url") + 1
        excel_row = 2
        for logical_row in evidence_rows_for_validation:
            target = str(logical_row.get("path_or_url") or "").strip()
            hyperlink = evidence_sheet.cell(row=excel_row, column=path_column).hyperlink
            should_link = bool(
                target
                and (
                    target.startswith(("https://", "http://"))
                    or Path(target).is_absolute()
                    or Path(target).exists()
                )
            )
            if should_link and (hyperlink is None or hyperlink.target != target):
                raise ReportValidationError(
                    "Generated workbook evidence hyperlink does not match its catalog path."
                )
            excel_row += expanded_row_count(logical_row.get(header) for header in EVIDENCE_HEADERS)
    finally:
        workbook.close()


def _sidecar_text(value: Any) -> str:
    return str(serialize(value)).replace("\r", " ").replace("\n", " ").strip()


def _csv_safe_text(value: Any) -> str:
    """Prevent spreadsheet programs from interpreting exported text as a formula."""
    serialized = serialize(value)
    text = str(serialized).replace("\r", " ").replace("\n", " ").strip()
    return (
        "'" + text
        if isinstance(serialized, str) and text.startswith(("=", "+", "-", "@"))
        else text
    )


def write_defects_csv(path: Path, rows_to_write: list[dict[str, Any]]) -> None:
    handle = io.StringIO(newline="")
    writer = csv.DictWriter(handle, fieldnames=DEFECT_HEADERS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(
        {header: _csv_safe_text(row.get(header)) for header in DEFECT_HEADERS}
        for row in rows_to_write
    )
    atomic_write_bytes(path, ("\ufeff" + handle.getvalue()).encode("utf-8"))


def write_defects_markdown(
    path: Path,
    rows_to_write: list[dict[str, Any]],
    session: dict[str, Any] | None = None,
) -> None:
    columns = [
        "plan_order",
        "event_name",
        "status",
        "primary_outcome",
        "anomaly_flags",
        "failed_layer",
        "concise_reason",
        "exact_retest",
    ]
    lines = [
        "# GTM recette defects",
        "",
        f"Output contract: {output_contract_version(session)}",
        "",
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows_to_write:
        values = [_sidecar_text(row.get(column)).replace("|", "\\|") for column in columns]
        lines.append("| " + " | ".join(values) + " |")
    atomic_write_bytes(path, ("\n".join(lines) + "\n").encode("utf-8"))


def write_stakeholder_summary(
    path: Path,
    data: dict[str, Any],
    session: dict[str, Any] | None,
) -> None:
    feedback = event_feedback(data, session)
    counts = Counter(row["status"] for row in feedback)
    lines = [
        f"# {data.get('run', {}).get('report_title', 'GTM Client Recette')}",
        "",
        f"Output contract: {output_contract_version(session)}",
        "",
        f"Scope: {_sidecar_text(data.get('run', {}).get('acceptance_scope'))}",
        "",
        "## Event totals",
        "",
        ", ".join(
            f"{status}: {counts.get(status, 0)}"
            for status in ("PASS", "FAIL", "BLOCKED", "REVIEW", "NOT_TESTED")
        ),
        "",
        "## Non-PASS events",
        "",
    ]
    non_pass = [row for row in feedback if row.get("status") != "PASS"]
    if not non_pass:
        lines.append("No non-PASS event.")
    else:
        lines.extend(
            f"- {row.get('plan_order')}. {row.get('event_name')} — {row.get('status')} "
            f"[{row.get('primary_outcome')}; "
            f"{_sidecar_text(row.get('anomaly_flags')) or 'no anomaly flag'}]: "
            f"{_sidecar_text(row.get('reason'))} Retest: {_sidecar_text(row.get('retest'))}"
            for row in non_pass
        )
    atomic_write_bytes(path, ("\n".join(lines) + "\n").encode("utf-8"))


def main() -> int:
    args = parse_args()
    try:
        if args.input != "-" and args.session_ledger is not None:
            recover_file_pair(Path(args.input), args.session_ledger)
        data = load_data(args.input)
        warnings = validate(data, strict=args.strict)
        refuse_unsafe_evidence(warnings)
        if args.strict and args.session_ledger is None:
            raise ReportValidationError("--strict final certification requires --session-ledger.")
        session = load_data(str(args.session_ledger)) if args.session_ledger is not None else None
        if session is not None:
            execution_errors = validate_session(
                session,
                results=data,
                final=args.strict,
            )
            if args.strict and execution_errors:
                raise ReportValidationError("\n".join(execution_errors))
            warnings.extend(execution_errors)
        if args.validate_only:
            print("Schema-v3 recette results are valid.")
            if warnings:
                print(f"Completed with {len(warnings)} validation warning(s).")
            return 0
        if not args.output:
            raise ReportValidationError(
                "An output .xlsx path is required unless --validate-only is used."
            )
        output = Path(args.output)
        if output.suffix.lower() != ".xlsx":
            raise ReportValidationError("Output path must use the .xlsx extension.")
        inputs = [
            Path(args.input).resolve() if args.input != "-" else None,
            args.session_ledger.resolve() if args.session_ledger is not None else None,
        ]
        outputs = [
            output.resolve(),
            *(
                path.resolve()
                for path in (args.defects_csv, args.defects_md, args.stakeholder_summary)
                if path is not None
            ),
        ]
        if len(set(outputs)) != len(outputs):
            raise ReportValidationError("Workbook and sidecar outputs must use distinct paths.")
        if any(path in outputs for path in inputs if path is not None):
            raise ReportValidationError("An output path cannot overwrite an input ledger.")
        temporary = output.with_name(f".{output.stem}.{uuid4().hex}.validated.xlsx")
        try:
            build_workbook(data, temporary, warnings, session)
            atomic_write_bytes(output, temporary.read_bytes())
        finally:
            temporary.unlink(missing_ok=True)
        defects = defect_rows(data, session)
        if args.defects_csv:
            write_defects_csv(args.defects_csv, defects)
        if args.defects_md:
            write_defects_markdown(args.defects_md, defects, session)
        if args.stakeholder_summary:
            write_stakeholder_summary(args.stakeholder_summary, data, session)
    except (OSError, json.JSONDecodeError, ReportValidationError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    print(f"Created {output.resolve()}")
    if warnings:
        print(f"Completed with {len(warnings)} validation warning(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
