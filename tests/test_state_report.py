from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from helpers import bundle_fixture

from state import (
    RunError,
    abandon_run,
    commit_event,
    finish_run,
    preflight_run,
    read_records,
    resolve_workbook_input,
    start_event,
    start_run,
)
from xlsx_plan import extract_workbook, validate_inspection_plan


def plan_file(path: Path) -> Path:
    workbook = Workbook()
    workbook.active["A1"] = "Synthetic GA4 plan"
    workbook.save(path)
    workbook.close()
    return path


def canonical(events: int = 1) -> dict:
    return {
        "schema_version": "3.0.0",
        "events": [
            {
                "event_name": "view_item" if index == 0 else "add_to_cart",
                "parameters": [
                    {
                        "data_layer_path": "ecommerce.currency",
                        "ga4_parameter_name": "currency",
                        "value": "EUR",
                        "value_semantics": "EXAMPLE",
                        "json_type": "string",
                        "required": None,
                        "source_refs": ["Sheet!A1"],
                    }
                ],
                "data_layer_payload": None,
                "definition": "Synthetic event.",
                "trigger_description": "Use the synthetic button.",
                "entry_url": "https://example.test/product",
                "expected_destination_id": "G-TEST",
                "source_refs": ["Sheet!A1"],
            }
            for index in range(events)
        ],
    }


def begin(root: Path, events: int = 1) -> dict:
    source = plan_file(root / "plan.xlsx")
    evidence = extract_workbook(source)
    plan = validate_inspection_plan(evidence, canonical(events))
    with patch("state._workspace_root", return_value=root):
        preflight_run(source)
        return start_run(source, plan)


def images(action: dict) -> tuple[Path, Path]:
    before = Path(action["before_image_temporary"])
    after = Path(action["after_image_temporary"])
    before.write_bytes(b"before-image")
    after.write_bytes(b"after-image")
    return before, after


def bound_bundle(action: dict, cursor: int) -> dict:
    bundle = bundle_fixture()
    bundle["action_id"] = action["action_id"]
    bundle["event_id"] = action["event_id"]
    bundle["preview_cursor"] = cursor
    return bundle


class StateAndReportTests(unittest.TestCase):
    def test_complete_one_scenario_lifecycle_and_exact_report(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            started = begin(root)
            run = Path(started["run_directory"])
            action = start_event(run, 0)
            images(action)
            result = commit_event(run, bound_bundle(action, 1))
            self.assertEqual(result["scenario_id"], "view_item")
            self.assertTrue((run / "inspection-plan.json").is_file())
            self.assertTrue((run / "evidence-E-0001.json").is_file())
            self.assertTrue((run / "image-E-0001-before.png").is_file())
            terminal = finish_run(run)
            self.assertEqual(terminal["status"], "COMPLETE")
            workbook = load_workbook(terminal["workbook_path"], read_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Event feedback"])
                self.assertEqual(workbook["Event feedback"].max_row, 6)
            finally:
                workbook.close()

    def test_blocked_event_never_stops_the_next_event(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run = Path(begin(root, 2)["run_directory"])
            first = start_event(run, 0)
            bundle = bound_bundle(first, 0)
            for surface in ("reality", "source", "gtm", "network", "behavior"):
                bundle[surface] = {
                    "complete": False,
                    "attributable": False,
                    "reason": "Unavailable.",
                }
            bundle["reality"]["expected"] = {}
            images(first)
            commit_event(run, bundle)
            second = start_event(run, 0)
            self.assertEqual(second["event_id"], "E-0002")

    def test_cursor_and_event_identity_fail_fast(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run = Path(begin(root)["run_directory"])
            action = start_event(run, 3)
            bundle = bound_bundle(action, 4)
            bundle["event_id"] = "E-9999"
            images(action)
            with self.assertRaisesRegex(RunError, "event_id"):
                commit_event(run, bundle)
            self.assertFalse((run / "evidence-E-0001.json").exists())

    def test_corrupted_stream_is_not_repaired(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run = Path(begin(root)["run_directory"])
            with (run / "events.ndjson").open("a", encoding="utf-8") as handle:
                handle.write("not-json\n")
            with self.assertRaisesRegex(RunError, "corrupted"):
                read_records(run)

    def test_evidence_marker_is_not_published_when_final_replace_fails(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run = Path(begin(root)["run_directory"])
            action = start_event(run, 0)
            images(action)
            import state

            original = state.os.replace

            def replace(source: Path | str, target: Path | str) -> None:
                if str(target).endswith("evidence-E-0001.json"):
                    raise OSError("injected evidence publication failure")
                original(source, target)

            with (
                patch("state.os.replace", side_effect=replace),
                self.assertRaisesRegex(OSError, "injected"),
            ):
                commit_event(run, bound_bundle(action, 1))
            self.assertFalse((run / "evidence-E-0001.json").exists())
            self.assertFalse((run / "image-E-0001-before.png").exists())
            self.assertFalse((run / "image-E-0001-after.png").exists())
            self.assertFalse(
                any(record["kind"] == "EVENT_COMMITTED" for record in read_records(run))
            )

    def test_artifacts_roll_back_when_commit_record_fails(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run = Path(begin(root)["run_directory"])
            action = start_event(run, 0)
            images(action)
            with (
                patch("state._append", side_effect=OSError("injected log failure")),
                self.assertRaisesRegex(OSError, "injected log failure"),
            ):
                commit_event(run, bound_bundle(action, 1))
            self.assertFalse((run / "evidence-E-0001.json").exists())
            self.assertFalse((run / "image-E-0001-before.png").exists())
            self.assertFalse((run / "image-E-0001-after.png").exists())

    def test_fatal_cleanup_removes_only_uncommitted_temporary_images(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            run = Path(begin(root)["run_directory"])
            action = start_event(run, 0)
            before, after = images(action)
            terminal = abandon_run(run, "TAG_ASSISTANT", "METHOD_UNAVAILABLE", "No chevron.")
            self.assertEqual(terminal["status"], "FATAL")
            self.assertFalse(before.exists())
            self.assertFalse(after.exists())
            self.assertEqual(read_records(run)[-1]["kind"], "RUN_ABORTED")
            with self.assertRaisesRegex(RunError, "terminal"):
                start_event(run, 0)

    def test_workbook_input_resolves_filename_and_workspace_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = plan_file(root / "named-plan.xlsx")
            with patch("state._workspace_root", return_value=root):
                resolved = resolve_workbook_input("named-plan.xlsx")
            self.assertEqual(resolved["plan_path"], str(source.resolve()))
            self.assertEqual(resolved["output_directory"], str(root.resolve()))


if __name__ == "__main__":
    unittest.main()
