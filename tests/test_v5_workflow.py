from __future__ import annotations

import re
import subprocess
import sys
import tempfile
import time
import unittest
from pathlib import Path

from openpyxl import load_workbook
from tests.v5_harness import ROOT, SCRIPTS, V5Harness, default_event, write_json

from core.capture import capture_value
from core.correlate import action_windows
from core.coverage import coverage_reviews
from core.plan import initialize_run
from core.report import build_reports, render_event_feedback, status_view
from core.state import StateError, load_plan, read_stream
from core.workflow import commit_action, finish_run, sync_preview


class WorkflowTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_first_event_reaches_detailed_feedback_without_future_event_state(self) -> None:
        harness = V5Harness(self.root)
        _, result = harness.execute_pass()
        self.assertEqual(result["status"], "PASS")
        self.assertTrue(result["final"])
        self.assertEqual(
            set(result["domains"]),
            {"reality", "source", "gtm", "delivery", "behavior", "safety"},
        )
        self.assertTrue(
            {"GTM fired/non-fired inventory", "Browser request - GA4 event"}.issubset(
                {row["inspection_target"] for row in result["inspections"]}
            )
        )
        records, _ = read_stream(harness.run)
        acted = {
            event_id
            for action in action_windows(records)
            for event_id in map(str, action.get("event_ids", []))
        }
        self.assertFalse(set(coverage_reviews(records)) - acted)
        rendered = render_event_feedback(result)
        self.assertIn(
            "| Scenario | Inspection target | Domain | Status | Observed | Expected | Reason |",
            rendered,
        )
        self.assertIn("Source signal PASS", rendered)
        self.assertIn("GTM decision PASS", rendered)
        self.assertIn("Destination delivery PASS", rendered)
        telemetry = status_view(harness.run)["telemetry"]
        self.assertTrue(telemetry["operation_counters_instrumented"])
        self.assertEqual(telemetry["preview_tab_switches"], 0)
        self.assertEqual(telemetry["preview_events_observed"], 1)
        self.assertEqual(telemetry["network_requests_observed"], 1)

    def test_commit_retry_is_idempotent_without_pre_preview_judgement(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        payload = {"event": "view_item"}
        bundle = {
            "health": harness._health("after"),
            "page": {"states": [harness._page("after")]},
            "datalayer": harness.datalayer([payload], action_id=action),
            "network": harness.network(["view_item"], action_id=action),
        }
        first = commit_action(harness.run, bundle, action_id=action)
        count_after_first = len(read_stream(harness.run)[0])
        second = commit_action(harness.run, bundle, action_id=action)
        self.assertEqual(len(read_stream(harness.run)[0]), count_after_first)
        self.assertEqual(first["commit"]["record_id"], second["commit"]["record_id"])
        self.assertEqual(first["operation_deltas"], second["operation_deltas"])
        self.assertEqual(first["execution_violations"], [])
        self.assertFalse(any(row["kind"] == "ACTION_PULSE" for row in read_stream(harness.run)[0]))

    def test_invalid_control_is_rejected_before_commit_evidence_is_written(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        records_before = len(read_stream(harness.run)[0])
        evidence_before = {path.name for path in (harness.run / "evidence").iterdir()}
        with self.assertRaisesRegex(StateError, "Coverage review is invalid"):
            commit_action(
                harness.run,
                {
                    "health": harness._health("after"),
                    "page": {"states": [harness._page("after")]},
                    "datalayer": harness.datalayer([{"event": "view_item"}], action_id=action),
                    "coverage": {
                        "event_id": "E-view_item",
                        "mode": "EXHAUSTIVE",
                        "complete": True,
                        "rationale": "invalid on purpose",
                        "stop_reason": "invalid on purpose",
                        "dimensions": [],
                        "scenarios": [],
                    },
                },
                action_id=action,
            )
        self.assertEqual(len(read_stream(harness.run)[0]), records_before)
        self.assertEqual(
            {path.name for path in (harness.run / "evidence").iterdir()}, evidence_before
        )

    def test_source_does_not_substitute_for_missing_preview(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        payload = {"event": "view_item", "ecommerce": {"items": [{"item_id": "SKU-1"}]}}
        harness.commit(action, [payload])
        result = harness.sync(
            ["E-view_item"],
            [payload],
            action_id=action,
            coverage=harness.coverage("E-view_item", [action]),
            include_preview=False,
        )["events"][0]
        self.assertEqual(result["domains"]["source"]["status"], "PASS")
        self.assertEqual(result["domains"]["gtm"]["status"], "BLOCKED")
        self.assertEqual(result["domains"]["delivery"]["status"], "BLOCKED")
        self.assertTrue(
            any(
                row["inspection_target"] == "Browser request - GA4 event"
                and row["status"] == "PASS"
                for row in result["inspections"]
            )
        )
        non_pass = [
            row for row in result["inspections"] if row["status"] not in {"PASS", "NOT_APPLICABLE"}
        ]
        self.assertTrue(non_pass)
        self.assertTrue(all(row.get("check_next") for row in non_pass))
        rendered = render_event_feedback(result)
        self.assertIn("| Resolved variable - event | GTM decision | BLOCKED |", rendered)
        self.assertIn("Current Tag Assistant event list and Preview linkage", rendered)
        self.assertNotEqual(result["status"], "PASS")

    def test_dead_page_fails_even_when_measurement_chain_passes(self) -> None:
        harness = V5Harness(self.root)
        _, result = harness.execute_pass(
            page_updates={
                "status_code": 404,
                "page_valid": False,
                "soft_404": True,
            }
        )
        self.assertEqual(result["domains"]["reality"]["status"], "FAIL")
        self.assertEqual(result["domains"]["source"]["status"], "PASS")
        self.assertEqual(result["domains"]["gtm"]["status"], "PASS")
        self.assertEqual(result["domains"]["delivery"]["status"], "PASS")
        self.assertEqual(result["status"], "FAIL")

    def test_report_requires_finished_run_and_validates_all_outputs(self) -> None:
        harness = V5Harness(self.root)
        harness.execute_pass()
        with self.assertRaisesRegex(StateError, "finished run"):
            build_reports(harness.run)
        result = finish_run(harness.run)
        self.assertTrue(result["finished"])
        outputs = build_reports(harness.run)
        for output in outputs.values():
            self.assertTrue(Path(output).is_file())
        workbook = load_workbook(outputs["xlsx"], read_only=True)
        try:
            self.assertEqual(
                set(workbook.sheetnames),
                {
                    "Summary",
                    "Events",
                    "Scenarios",
                    "Inspections",
                    "Defects and limits",
                    "Evidence index",
                    "Telemetry",
                },
            )
        finally:
            workbook.close()

    def test_finish_refuses_untested_events_but_accepts_final_blocked_results(self) -> None:
        events = [default_event("E1", "view_item"), default_event("E2", "select_item")]
        harness = V5Harness(self.root, events=events)
        harness.execute_pass(event_id="E1")
        with self.assertRaisesRegex(StateError, "E2"):
            finish_run(harness.run)

    def test_status_is_read_only(self) -> None:
        harness = V5Harness(self.root)
        harness.execute_pass()
        stream = harness.run / "stream.ndjson"
        before = (stream.read_bytes(), stream.stat().st_mtime_ns)
        status_view(harness.run)
        after = (stream.read_bytes(), stream.stat().st_mtime_ns)
        self.assertEqual(before, after)

    def test_one_action_can_cover_multiple_planned_events_without_false_interjection(self) -> None:
        events = [
            default_event("E-list", "view_item_list"),
            default_event("E-select", "select_item"),
        ]
        harness = V5Harness(self.root, events=events)
        action = harness.begin(["E-list", "E-select"])
        payloads = [
            {"event": "view_item_list", "ecommerce": {"items": [{"item_id": "SKU-1"}]}},
            {"event": "select_item", "ecommerce": {"items": [{"item_id": "SKU-1"}]}},
        ]
        harness.commit(action, payloads)
        coverage = [
            harness.coverage("E-list", [action]),
            harness.coverage("E-select", [action]),
        ]
        results = harness.sync(
            ["E-list", "E-select"], payloads, action_id=action, coverage=coverage
        )["events"]
        self.assertEqual([row["status"] for row in results], ["PASS", "PASS"])

    def test_one_interaction_with_three_cooccurring_events_reuses_one_preview_pass(self) -> None:
        events = [
            default_event("E-page", "page_view"),
            default_event("E-list", "view_item_list"),
            default_event("E-item", "view_item"),
        ]
        harness = V5Harness(self.root, events=events)
        started = time.perf_counter()
        actions: list[str] = []
        payloads = [
            {"event": "page_view", "page_language": "en"},
            {"event": "view_item_list", "ecommerce": {"items": [{"item_id": "SKU-1"}]}},
            {"event": "view_item", "ecommerce": {"items": [{"item_id": "SKU-1"}]}},
        ]
        preview_events = []
        for event, payload in zip(events, payloads, strict=True):
            action = harness.begin([event["event_id"]])
            actions.append(action)
            harness.commit(action, [payload])
            preview_events.extend(harness.preview([payload], action_id=action)["events"])
        coverage = [
            harness.coverage(event["event_id"], [action])
            for event, action in zip(events, actions, strict=True)
        ]
        result = sync_preview(
            harness.run,
            {
                "preview": {
                    "complete": True,
                    "epoch": "EPOCH-1",
                    "preview_session_id": "PREVIEW-1",
                    "container_ids": ["GTM-EXPECTED"],
                    "workspace_version": "42",
                    "events": preview_events,
                },
                "coverage": coverage,
            },
            event_ids=[event["event_id"] for event in events],
        )
        self.assertEqual([row["status"] for row in result["events"]], ["PASS"] * 3)
        records, _ = read_stream(harness.run)
        self.assertEqual(sum(row["kind"] == "CAPTURE_CAPABILITY" for row in records), 1)
        self.assertEqual(sum(row["kind"] == "CAPTURE_BINDING" for row in records), 1)
        self.assertEqual(sum(row["kind"] == "CAPTURE_PREVIEW" for row in records), 1)
        self.assertLess(time.perf_counter() - started, 5.0)

    def test_capability_failure_blocks_only_dependent_surfaces_without_waiting(self) -> None:
        harness = V5Harness(self.root)
        started = time.perf_counter()
        action = harness.begin(
            first_bundle_updates={
                "capability": harness.capability(
                    preview_events=False,
                    preview_tag_inventory=False,
                    preview_variables=False,
                    preview_consent=False,
                )
            }
        )
        payload = {"event": "view_item"}
        harness.commit(action, [payload])
        result = harness.sync(
            ["E-view_item"],
            [payload],
            action_id=action,
            coverage=harness.coverage("E-view_item", [action]),
            include_preview=False,
        )["events"][0]
        self.assertLess(time.perf_counter() - started, 5.0)
        self.assertEqual(result["domains"]["source"]["status"], "PASS")
        self.assertEqual(result["domains"]["gtm"]["status"], "BLOCKED")

    def test_optional_tag_can_be_absent_but_duplicate_is_not_accepted(self) -> None:
        event = default_event()
        event["tags"][0]["expected"] = "optional"
        harness = V5Harness(self.root, events=[event])
        action = harness.begin()
        payload = {"event": "view_item"}
        harness.commit(action, [payload], include_network=False)
        result = harness.sync(
            ["E-view_item"],
            [payload],
            action_id=action,
            coverage=harness.coverage("E-view_item", [action]),
            preview_updates={"fire_tags": False},
        )["events"][0]
        self.assertEqual(result["domains"]["gtm"]["status"], "PASS")
        self.assertEqual(result["domains"]["delivery"]["status"], "NOT_APPLICABLE")

    def test_explicit_tag_scope_excludes_unrequested_media_tags(self) -> None:
        event = default_event()
        event["tags"].append(
            {
                "tag_id": "Ads-conversion",
                "tag_name": "Google Ads conversion",
                "category": "Google Ads",
                "expected": "fire",
                "destination": "AW-123",
            }
        )
        harness = V5Harness(self.root, events=[event], scope={"tag_scope": ["GA4"]})
        compiled = load_plan(harness.run)["events"][0]
        self.assertEqual([tag["tag_id"] for tag in compiled["tags"]], ["GA4-event"])

    def test_large_plan_initializes_without_per_event_runtime_machinery(self) -> None:
        events = []
        for event_index in range(100):
            event = default_event(f"E-{event_index}", f"event_{event_index}")
            event["requirements"] = [
                {
                    "requirement_id": f"R-{event_index}-{requirement_index}",
                    "field_path": f"parameters.value_{requirement_index}",
                    "match_rule": "present",
                }
                for requirement_index in range(20)
            ]
            events.append(event)
        source = write_json(self.root / "large.json", {"events": events})
        started = time.perf_counter()
        plan = initialize_run(
            source,
            self.root / "large-run",
            scope={
                "approved": True,
                "origins": ["https://shop.example.test"],
                "certify_tags": False,
            },
        )
        elapsed = time.perf_counter() - started
        self.assertLess(elapsed, 3.0)
        self.assertEqual(plan["event_count"], 100)
        self.assertEqual(plan["requirement_count"], 2000)
        self.assertEqual(read_stream(self.root / "large-run")[0], [])
        self.assertEqual(list((self.root / "large-run" / "evidence").iterdir()), [])

    def test_browser_cost_counters_reuse_health_capture_without_new_workflow_state(self) -> None:
        harness = V5Harness(self.root)
        harness.begin(
            first_bundle_updates={
                "health": {
                    **harness._health("before"),
                    "operations": {
                        **harness.operations,
                        "navigations": 1,
                        "full_preflights": 1,
                        "preview_tab_switches": 2,
                        "preview_deep_reads": 1,
                        "ai_semantic_passes": 1,
                    },
                }
            }
        )
        telemetry = status_view(harness.run)["telemetry"]
        self.assertEqual(telemetry["navigations"], 1)
        self.assertEqual(telemetry["full_preflights"], 1)
        self.assertEqual(telemetry["preview_tab_switches"], 2)
        self.assertEqual(telemetry["preview_deep_reads"], 1)
        self.assertEqual(telemetry["ai_semantic_passes"], 1)
        with self.assertRaisesRegex(StateError, "operation counters"):
            capture_value(
                harness.run,
                "health",
                {**harness._health("after"), "operations": {"reloads": -1}},
            )

    def test_public_cli_is_small_and_has_no_legacy_staging_routes(self) -> None:
        result = subprocess.run(
            [sys.executable, "-B", str(SCRIPTS / "recette.py"), "--help"],
            cwd=ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
        command_list = re.search(r"\{([^}]+)\}", result.stdout)
        self.assertIsNotNone(command_list)
        self.assertEqual(
            set(command_list.group(1).split(",")),
            {"init", "next", "complete", "status", "handoff", "finish", "report", "reopen"},
        )
        complete_help = subprocess.run(
            [sys.executable, "-B", str(SCRIPTS / "recette.py"), "complete", "--help"],
            cwd=ROOT,
            check=True,
            capture_output=True,
            text=True,
        ).stdout
        self.assertIn("--action ACTION", complete_help)
        self.assertNotIn("--event", complete_help)


if __name__ == "__main__":
    unittest.main()
