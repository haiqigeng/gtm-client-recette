from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from tests.v5_harness import V5Harness, default_event, write_json

from client_side_rules import MISSING
from core.capture import capture_value, redact_for_persistence
from core.constants import utc_now
from core.correlate import build_model
from core.plan import normalize_plan
from core.predicates import PredicateError, compile_predicate, evaluate_predicate
from core.protocols import decode_logical_sends
from core.state import StateError, load_plan, read_stream
from core.workflow import add_handoff, begin_action
from decode_browser_requests import decode_requests


class CompilerAndEvidenceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def scope(self, **updates: object) -> dict:
        return {
            "approved": True,
            "origins": ["https://shop.example.test"],
            "expected_container": ["GTM-EXPECTED"],
            "destination": ["G-TEST123"],
            "tag_scope": ["GA4"],
            **updates,
        }

    def test_common_two_block_workbook_compiles_without_custom_normalization(self) -> None:
        workbook = Workbook()
        core = workbook.active
        core.title = "Core DataLayer"
        core.append([None, "Nature", "Page", "Trigger"])
        core.append([None, "Core DataLayer", "All pages", "Before GTM"])
        core.append([])
        core.append([None, "variables", "Type", "Status", "Summary", "Values"])
        core.append([None, "page_language", "string", "Mandatory", "Locale", "en | fr"])
        core.append([None, "product_name", "string", "Mandatory", "Product", "Example..."])
        core.append([])
        core.append([None, "CODE :"])
        core.append([None, "dataLayer.push({event: 'should_not_compile'})"])
        item = workbook.create_sheet("View Item")
        item.append([None, "Name of the event", "Page", "Trigger"])
        item.append([None, "view_item", "Product", "View product"])
        item.append([])
        item.append([None, "variables", "Type", "Status", "Summary", "Values"])
        item.append([None, "event", "string", "Mandatory", "Event", "view_item"])
        item.append([None, "item_id", "string", "Mandatory", "ID", "SKU..."])
        item.append([None, "currency", "string", "Mandatory", "Currency", "EUR"])
        item.append([None, "quantity", "number", "Mandatory", "Quantity", 2])
        item.append([None, "checkout_step", "number", "Mandatory", "Step", 3])
        source = self.root / "two-block.xlsx"
        workbook.save(source)

        plan = normalize_plan(source, scope=self.scope())
        self.assertEqual(plan["event_count"], 2)
        self.assertEqual(plan["requirement_count"], 7)
        self.assertFalse(plan["events"][0]["compile_errors"])
        self.assertEqual(plan["events"][0]["mode"], "state_only")
        self.assertIsNone(plan["events"][0]["source_event_name"])
        self.assertEqual(plan["events"][0]["delivery_event_name"], "page_view")
        core_claims = plan["events"][0]["claims"]
        self.assertTrue(
            all(
                claim["target"].get("event_name") is None
                for claim in core_claims
                if claim["target"].get("check") == "data_layer_state"
            )
        )
        self.assertTrue(
            all(
                claim["target"].get("event_name") == "page_view"
                for claim in core_claims
                if claim["target"].get("check")
                in {
                    "resolved_variable",
                    "effective_mapping",
                    "runtime_parameter",
                    "tag_inventory",
                    "tag_configuration",
                    "tag_firing",
                    "destination_request",
                    "request_parameter",
                }
            )
        )
        self.assertEqual(
            plan["events"][0]["known_dimensions"][0]["values"],
            [{"value": "en", "source": "plan"}, {"value": "fr", "source": "plan"}],
        )
        item_paths = {claim.get("target", {}).get("path") for claim in plan["events"][1]["claims"]}
        self.assertIn("ecommerce.items[].item_id", item_paths)
        self.assertNotIn("should_not_compile", json.dumps(plan))
        source_predicates = {
            claim["target"].get("path"): claim["predicate"]
            for claim in plan["events"][1]["claims"]
            if claim["domain"] == "source"
        }
        self.assertEqual(source_predicates["event"]["expected"], "view_item")
        self.assertEqual(source_predicates["ecommerce.checkout_step"]["expected"], 3)
        self.assertEqual(source_predicates["ecommerce.currency"]["operator"], "present")
        self.assertEqual(source_predicates["ecommerce.items[].quantity"]["operator"], "present")
        self.assertEqual(source_predicates["ecommerce.items[].item_id"]["operator"], "present")

    def test_json_yaml_and_xlsx_compile_to_typed_claims_with_source_coordinates(self) -> None:
        json_path = write_json(self.root / "plan.json", {"events": [default_event()]})
        json_plan = normalize_plan(json_path, scope=self.scope())
        self.assertEqual(json_plan["schema_version"], "5.0")
        self.assertTrue(json_plan["events"][0]["claims"])

        yaml_path = self.root / "plan.yaml"
        yaml_path.write_text(
            "events:\n  - event_id: E1\n    event_name: page_view\n    requirements: []\n",
            encoding="utf-8",
        )
        yaml_plan = normalize_plan(
            yaml_path, scope=self.scope(certify_tags=False, browser_send_required=False)
        )
        self.assertEqual(yaml_plan["event_count"], 1)

        xlsx_path = self.root / "plan.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tracking"
        sheet.append(["event_name", "field_path", "match_rule", "expected_value"])
        sheet.append(["view_item", "event", "equals", "view_item"])
        workbook.save(xlsx_path)
        workbook.close()
        xlsx_plan = normalize_plan(
            xlsx_path,
            scope=self.scope(certify_tags=False, browser_send_required=False),
        )
        sources = [claim["source"]["reference"] for claim in xlsx_plan["events"][0]["claims"]]
        self.assertTrue(any("Tracking!" in source for source in sources))

    def test_malformed_later_event_does_not_block_first_executable_event(self) -> None:
        events = [
            default_event("E-good", "view_item"),
            {
                "event_id": "E-bad",
                "event_name": "purchase",
                "requirements": [
                    {
                        "field_path": "event",
                        "match_rule": "invented_operator",
                        "expected_value": "purchase",
                    }
                ],
            },
        ]
        plan = normalize_plan(
            write_json(self.root / "plan.json", {"events": events}),
            scope=self.scope(),
        )
        self.assertTrue(plan["events"][0]["executable"])
        self.assertFalse(plan["events"][1]["executable"])
        self.assertIn("unsupported predicate", plan["events"][1]["compile_errors"][0])

    def test_xlsx_merged_event_identity_retains_every_requirement(self) -> None:
        xlsx_path = self.root / "merged-plan.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ecommerce"
        sheet.append(
            [
                "event_name",
                "event_id",
                "field_path",
                "match_rule",
                "allowed_values",
                "expected_type",
                "tag",
                "destination",
            ]
        )
        sheet.append(
            [
                "view_item",
                "EV-view-item",
                "event",
                "equals",
                None,
                "string",
                "GA4 - View item",
                "G-TEST123",
            ]
        )
        sheet.append([None, None, "ecommerce.currency", "one_of", "EUR|USD", "string"])
        sheet.append([None, None, "ecommerce.items", "type", None, "array"])
        sheet.merge_cells("A2:A4")
        sheet.merge_cells("B2:B4")
        workbook.save(xlsx_path)
        workbook.close()

        plan = normalize_plan(xlsx_path, scope=self.scope())
        event = plan["events"][0]
        normalization = plan["source"]["normalization"]

        self.assertEqual(plan["event_count"], 1)
        self.assertEqual(plan["requirement_count"], 3)
        self.assertEqual(normalization["requirements_compiled"], 3)
        self.assertEqual(normalization["carried_event_rows"], 2)
        self.assertTrue(event["executable"])
        self.assertEqual(event["destinations"], ["G-TEST123"])
        self.assertEqual(event["tags"][0]["destination"], "G-TEST123")
        currency = next(
            claim
            for claim in event["claims"]
            if claim["target"].get("path") == "ecommerce.currency"
        )
        items = next(
            claim for claim in event["claims"] if claim["target"].get("path") == "ecommerce.items"
        )
        self.assertEqual(currency["predicate"]["allowed_values"], ["EUR", "USD"])
        self.assertEqual(items["predicate"]["expected_type"], "array")
        self.assertEqual(currency["source"]["reference"], "Ecommerce!3")

    def test_tabular_orphan_requirement_fails_before_browser_work(self) -> None:
        xlsx_path = self.root / "orphan-plan.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["event_name", "field_path", "match_rule"])
        sheet.append([None, "ecommerce.items", "present"])
        workbook.save(xlsx_path)
        workbook.close()

        with self.assertRaisesRegex(StateError, "before an event name"):
            normalize_plan(xlsx_path, scope=self.scope())

    def test_blank_separator_prevents_event_identity_bleed(self) -> None:
        xlsx_path = self.root / "separated-plan.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["event_name", "field_path", "match_rule"])
        sheet.append(["view_item", "event", "present"])
        sheet.append([])
        sheet.append([None, "ecommerce.items", "present"])
        workbook.save(xlsx_path)
        workbook.close()

        with self.assertRaisesRegex(StateError, "before an event name"):
            normalize_plan(xlsx_path, scope=self.scope())

    def test_delimited_fill_down_is_visible_in_normalization_diagnostics(self) -> None:
        csv_path = self.root / "plan.csv"
        csv_path.write_text(
            "event_name,event_id,field_path,match_rule,allowed_values\n"
            "begin_checkout,EV-checkout,event,present,\n"
            ",,ecommerce.shipping_tier,one_of,standard|express\n",
            encoding="utf-8",
        )

        plan = normalize_plan(
            csv_path,
            scope=self.scope(certify_tags=False, browser_send_required=False),
        )
        normalization = plan["source"]["normalization"]

        self.assertEqual(plan["event_count"], 1)
        self.assertEqual(plan["requirement_count"], 2)
        self.assertEqual(normalization["requirements_compiled"], 2)
        self.assertEqual(normalization["carried_event_rows"], 1)
        self.assertEqual(normalization["rows_ignored"], 0)
        shipping = next(
            claim
            for claim in plan["events"][0]["claims"]
            if claim["target"].get("path") == "ecommerce.shipping_tier"
        )
        self.assertEqual(shipping["predicate"]["allowed_values"], ["standard", "express"])
        self.assertEqual(shipping["source"]["reference"], "line 3")

    def test_scope_rejects_container_destination_identity_confusion(self) -> None:
        path = write_json(self.root / "plan.json", {"events": [default_event()]})
        with self.assertRaisesRegex(StateError, "Destination IDs"):
            normalize_plan(path, scope=self.scope(destination=["GTM-WRONG-TYPE"]))
        with self.assertRaisesRegex(StateError, "container IDs"):
            normalize_plan(path, scope=self.scope(expected_container=["G-DESTINATION"]))

    def test_predicates_preserve_strict_json_types_and_wire_coercion_is_explicit(self) -> None:
        equals = compile_predicate(
            {"match_rule": "equals", "expected_value": 1, "expected_type": "number"}
        )
        self.assertEqual(evaluate_predicate(1, equals)["status"], "PASS")
        self.assertEqual(evaluate_predicate(True, equals)["status"], "FAIL")
        self.assertEqual(evaluate_predicate("1", equals, wire=True)["status"], "FAIL")
        coercing = compile_predicate(
            {
                "match_rule": "equals",
                "expected_value": 1,
                "wire_coercion": True,
            }
        )
        self.assertEqual(evaluate_predicate("1", coercing, wire=True)["status"], "PASS")
        self.assertEqual(evaluate_predicate(MISSING, {"operator": "absent"})["status"], "PASS")
        absent_state = compile_predicate({"operator": "state", "state": "absent"})
        null_state = compile_predicate({"operator": "state", "state": "null"})
        self.assertEqual(evaluate_predicate(MISSING, absent_state)["status"], "PASS")
        self.assertEqual(evaluate_predicate(None, null_state)["status"], "PASS")
        self.assertEqual(evaluate_predicate(None, absent_state)["status"], "FAIL")

    def test_unsafe_regex_and_invalid_count_are_compile_errors(self) -> None:
        with self.assertRaises(PredicateError):
            compile_predicate({"operator": "regex", "pattern": "(a+)+$"})
        with self.assertRaises(PredicateError):
            compile_predicate({"operator": "count", "exact": -1})

    def test_ga4_batched_request_decodes_each_logical_send(self) -> None:
        request = {
            "request_id": "R1",
            "endpoint": "https://www.google-analytics.com/g/collect",
            "query": {"tid": "G-TEST123"},
            "body": {
                "records": [
                    {"en": "view_item", "ep.item_id": "SKU-1"},
                    {"en": "add_to_cart", "ep.item_id": "SKU-1"},
                ]
            },
            "action_id": "A1",
            "tag_ids": ["GA4-event"],
        }
        sends = decode_logical_sends(request)
        self.assertEqual([row["event_name"] for row in sends], ["view_item", "add_to_cart"])
        self.assertEqual([row["destination"] for row in sends], ["G-TEST123", "G-TEST123"])
        self.assertTrue(all(row["tag_ids"] == ["GA4-event"] for row in sends))

    def test_google_ads_destination_survives_sensitive_path_redaction(self) -> None:
        decoded = decode_requests(
            {
                "complete": True,
                "requests": [
                    {
                        "request_id": "ADS-1",
                        "url": (
                            "https://www.googleadservices.com/pagead/conversion/"
                            "123456789/?label=checkout"
                        ),
                    }
                ],
            }
        )["requests"][0]
        sends = decode_logical_sends(decoded)
        self.assertEqual(len(sends), 1)
        self.assertEqual(sends[0]["protocol"], "google_ads")
        self.assertEqual(sends[0]["destination"], "AW-123456789")

    def test_ga4_wire_items_are_decoded_to_canonical_typed_paths(self) -> None:
        sends = decode_logical_sends(
            {
                "request_id": "R-items",
                "endpoint": "https://www.google-analytics.com/g/collect",
                "query": {
                    "tid": "G-TEST123",
                    "en": "view_item",
                    "cu": "EUR",
                    "epn.value": "29.9",
                    "pr1": "idSKU-1~nmProduct~pr29.9~qt2",
                },
            }
        )
        parameters = sends[0]["parameters"]
        self.assertEqual(parameters["ecommerce"]["currency"], "EUR")
        self.assertEqual(parameters["ecommerce"]["value"], 29.9)
        self.assertEqual(parameters["ecommerce"]["items"][0]["item_id"], "SKU-1")
        self.assertEqual(parameters["ecommerce"]["items"][0]["quantity"], 2)

    def test_bundle_validation_is_transactional(self) -> None:
        harness = V5Harness(self.root)
        invalid = {
            "health": harness._health("before"),
            "page": {"states": [harness._page("before")]},
            "capability": harness.capability(),
            "binding": harness.binding(),
            "preview": {"epoch": "E1", "events": [{"event_name": "view_item"}]},
        }
        with self.assertRaisesRegex(StateError, "workflow phase"):
            begin_action(harness.run, ["E-view_item"], invalid)
        self.assertEqual(read_stream(harness.run)[0], [])
        self.assertEqual(list((harness.run / "evidence").iterdir()), [])

    def test_accumulated_tag_assistant_state_cannot_be_laundered_as_source(self) -> None:
        harness = V5Harness(self.root)
        with self.assertRaisesRegex(StateError, "not direct-source evidence"):
            capture_value(
                harness.run,
                "source",
                {
                    "complete": True,
                    "signals": [
                        {
                            "signal_id": "S1",
                            "mechanism": "tag_assistant_message",
                            "event_name": "view_item",
                            "payload": {"event": "view_item"},
                        }
                    ],
                },
            )

    def test_request_lifecycle_updates_merge_without_false_identity_conflict(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        base = {
            "complete": False,
            "requests": [
                {
                    "request_id": "R1",
                    "action_id": action,
                    "timestamp": utc_now(),
                    "url": "https://www.google-analytics.com/g/collect?tid=G-TEST123&en=view_item",
                    "outcome": "initiated",
                }
            ],
        }
        capture_value(harness.run, "network", base, source_id="network-start")
        capture_value(
            harness.run,
            "network",
            {
                **base,
                "complete": True,
                "requests": [{**base["requests"][0], "outcome": "settled", "response_status": 204}],
            },
            source_id="network-end",
        )
        model = build_model(harness.run, load_plan(harness.run), read_stream(harness.run)[0])
        self.assertEqual(len(model["requests"]), 1)
        self.assertEqual(model["requests"][0]["outcome"], "settled")
        self.assertEqual(model["ambiguous"], [])

    def test_incompatible_request_identity_reuse_blocks_confidence(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        for index, host in enumerate(("www.google-analytics.com", "evil.example.test"), start=1):
            capture_value(
                harness.run,
                "network",
                {
                    "complete": True,
                    "requests": [
                        {
                            "request_id": "R1",
                            "action_id": action,
                            "timestamp": utc_now(),
                            "url": f"https://{host}/g/collect?tid=G-TEST123&en=view_item",
                        }
                    ],
                },
                source_id=f"conflict-{index}",
            )
        model = build_model(harness.run, load_plan(harness.run), read_stream(harness.run)[0])
        self.assertEqual(model["ambiguous"][0]["kind"], "request_identity_conflict")

    def test_sensitive_values_are_redacted_before_persistence_without_fingerprint(self) -> None:
        safe, findings = redact_for_persistence(
            {"email": "person@example.test", "phone": "+33123456789"}
        )
        serialized = json.dumps(safe)
        self.assertNotIn("person@example.test", serialized)
        self.assertNotIn("33123456789", serialized)
        self.assertNotIn('value_fingerprint": "', serialized)
        self.assertTrue(findings)
        self.assertTrue(all(row["value_fingerprint"] == "not-retained" for row in findings))

    def test_protected_handoff_resumes_only_exact_browser_lineage(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin(
            replay_safety="PROTECTED",
            first_bundle_updates={
                "datalayer": harness.datalayer([{"event": "gtm.js"}], action_id=None)
            },
        )
        binding = {
            "browser_context_id": "CTX-1",
            "tab_id": "TAB-SITE",
            "document_id": "DOC-1",
            "action_id": action,
        }
        pending = add_handoff(
            harness.run,
            {"gate": "CAPTCHA", "status": "PENDING", "binding": binding},
        )
        with self.assertRaisesRegex(StateError, "exact"):
            add_handoff(
                harness.run,
                {
                    "handoff_id": pending["data"]["handoff_id"],
                    "gate": "CAPTCHA",
                    "status": "RESUMED",
                    "binding": {**binding, "tab_id": "NEW-TAB"},
                },
            )
        resumed = add_handoff(
            harness.run,
            {
                "handoff_id": pending["data"]["handoff_id"],
                "gate": "CAPTCHA",
                "status": "RESUMED",
                "binding": binding,
            },
        )
        self.assertEqual(resumed["data"]["status"], "RESUMED")


if __name__ == "__main__":
    unittest.main()
