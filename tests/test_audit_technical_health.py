#!/usr/bin/env python3
"""Adversarial regressions for privacy, integrity, output, and path hardening."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
TESTS = ROOT / "tests"
sys.path[:0] = [str(SCRIPTS), str(TESTS)]

from test_pipeline import execution_fixture, fixture  # noqa: E402

from build_recette_report import (  # noqa: E402
    LEGACY_CASE_HEADERS,
    LEGACY_OUTPUT_CONTRACT_VERSION,
    LEGACY_PUSH_HEADERS,
    ReportValidationError,
    build_workbook,
    validate_workbook,
)
from decode_browser_requests import decode_requests  # noqa: E402
from evidence_integrity import build_integrity_record, catalog_digest  # noqa: E402
from safe_regex import compile_pattern  # noqa: E402


class AuditTechnicalHealthTests(unittest.TestCase):
    def test_request_decoder_redacts_url_query_and_body_before_normal_output(self) -> None:
        rows = [
            {
                "request_id": "REQ-1",
                "url": (
                    "https://user:secret@collect.example.test/path"
                    "?email=synthetic.user%40example.com&en=purchase"
                ),
                "method": "POST",
                "headers": {"Content-Type": "application/json"},
                "post_data": json.dumps(
                    {
                        "email": "synthetic.user@example.com",
                        "first_name": "Synthetic",
                        "password": "do-not-store-this",
                        "value": 29.9,
                    }
                ),
            }
        ]
        decoded = decode_requests(rows)
        request = decoded["requests"][0]
        serialized = json.dumps(request)
        self.assertNotIn("synthetic.user@example.com", serialized)
        self.assertNotIn("do-not-store-this", serialized)
        self.assertNotIn("user:secret", request["request_url"])
        self.assertEqual("purchase", request["query"]["en"])
        self.assertTrue(request["query"]["email"].startswith("<redacted:"))
        self.assertEqual(29.9, request["body"]["decoded"]["value"])
        self.assertTrue(request["body"]["decoded"]["first_name"].startswith("<redacted:"))
        with self.assertRaisesRegex(ValueError, "quarantine"):
            decode_requests(rows, retain_raw_body=True)
        quarantined = decode_requests(rows, retain_raw_body=True, quarantine=True)
        self.assertTrue(quarantined["quarantine"]["contains_raw_sensitive_values"])

    def test_every_local_evidence_artifact_is_required_and_hashed(self) -> None:
        data = fixture()
        with tempfile.TemporaryDirectory() as tempdir:
            base = Path(tempdir)
            for row in data["evidence"]:
                path = base / str(row["path_or_url"])
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text(row["evidence_id"], encoding="utf-8")
            missing = data["evidence"][-1]
            (base / str(missing["path_or_url"])).unlink()
            failed = build_integrity_record(data, base)
            self.assertEqual("FAILED", failed["status"])
            missing_row = next(
                row for row in failed["files"] if row["evidence_id"] == missing["evidence_id"]
            )
            self.assertEqual("MISSING_FILE", missing_row["status"])

            (base / str(missing["path_or_url"])).write_text(
                missing["evidence_id"], encoding="utf-8"
            )
            verified = build_integrity_record(data, base)
            self.assertEqual(2, verified["version"])
            self.assertEqual("VERIFIED", verified["status"])
            self.assertTrue(all(row.get("sha256") for row in verified["files"]))

    def test_catalog_digest_binds_runtime_and_request_metadata(self) -> None:
        data = fixture()
        before = catalog_digest(data)
        data["evidence"][0]["runtime_check_id"] = "RTC-CHANGED"
        self.assertNotEqual(before, catalog_digest(data))

    def test_dynamic_regex_rejects_ambiguous_or_nested_repetition(self) -> None:
        compile_pattern(r"^https://[^?]+$")
        for unsafe in (r"(ab+)+", r"(a|aa)+$", r"(a)\1"):
            with self.subTest(pattern=unsafe), self.assertRaises(ValueError):
                compile_pattern(unsafe)

    def test_legacy_workbook_keeps_exact_v2_column_contract(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "legacy.xlsx"
            build_workbook(data, output, [], session)
            workbook = load_workbook(output, read_only=False, data_only=False)
            try:
                case_headers = [cell.value for cell in workbook["Interaction Cases"][1]]
                push_headers = [cell.value for cell in workbook["Observed Push Stream"][1]]
                self.assertEqual(LEGACY_CASE_HEADERS, case_headers)
                self.assertEqual(LEGACY_PUSH_HEADERS, push_headers)
                self.assertEqual(
                    LEGACY_OUTPUT_CONTRACT_VERSION,
                    workbook["Client Summary"]["B4"].value,
                )
            finally:
                workbook.close()

    def test_workbook_reload_validation_checks_decision_values(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "tampered.xlsx"
            build_workbook(data, output, [], session)
            workbook = load_workbook(output, read_only=False, data_only=False)
            sheet = workbook["Layer Verdicts"]
            status_column = next(cell.column for cell in sheet[1] if cell.value == "status")
            sheet.cell(row=2, column=status_column).value = "FAIL"
            workbook.save(output)
            workbook.close()
            with self.assertRaisesRegex(ReportValidationError, "decision field 'status'"):
                validate_workbook(output, data, session)

    def test_tracking_plan_alias_is_rejected_before_input_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            source = Path(tempdir) / "plan.csv"
            source.write_text("event,value\nview_item,1\n", encoding="utf-8")
            before = source.read_bytes()
            completed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "inspect_tracking_plan.py"),
                    str(source),
                    str(source),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertNotEqual(0, completed.returncode)
            self.assertEqual(before, source.read_bytes())

    def test_supporting_artifact_cannot_alias_its_ledger_in_place(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            ledger = Path(tempdir) / "ledger.json"
            ledger.write_text(json.dumps(fixture()), encoding="utf-8")
            before = ledger.read_bytes()
            completed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "register_supporting_artifact.py"),
                    str(ledger),
                    str(ledger),
                    "--artifact-id",
                    "ART-1",
                    "--artifact-type",
                    "gtm_container_audit_facts",
                    "--source-skill",
                    "gtm-container-audit-cleanup",
                    "--source-run-id",
                    "RUN-OLD",
                    "--source-version",
                    "1",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertNotEqual(0, completed.returncode)
            self.assertEqual(before, ledger.read_bytes())

    def test_migration_output_cannot_overwrite_legacy_session(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            legacy_results = Path(tempdir) / "results.json"
            legacy_session = Path(tempdir) / "session.json"
            case_manifest = Path(tempdir) / "cases.json"
            legacy_results.write_text(json.dumps(fixture()), encoding="utf-8")
            legacy_session.write_text(json.dumps(execution_fixture(fixture())), encoding="utf-8")
            before = legacy_session.read_bytes()
            completed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "migrate_schema_v2_to_v3.py"),
                    str(legacy_results),
                    str(legacy_session),
                    "--legacy-session",
                    str(legacy_session),
                    "--case-manifest",
                    str(case_manifest),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertNotEqual(0, completed.returncode)
            self.assertEqual(before, legacy_session.read_bytes())


if __name__ == "__main__":
    unittest.main()
