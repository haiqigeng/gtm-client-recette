#!/usr/bin/env python3
"""Incident-driven regression tests for the expert operator-v2 recette contract."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from openpyxl import load_workbook  # noqa: E402
from test_pipeline import execution_fixture, fixture  # noqa: E402

from build_recette_report import build_workbook, required_sheets  # noqa: E402
from classify_datalayer_snapshot import build_review  # noqa: E402
from event_feedback import event_feedback, final_conclusion  # noqa: E402
from evidence_integrity import build_integrity_record  # noqa: E402
from execution_contract import validate_session  # noqa: E402
from gated_flow_contract import gated_flow_errors  # noqa: E402
from layer_contract import layer_applicability  # noqa: E402
from page_context_contract import acquisition_errors, handoff_errors  # noqa: E402
from recette_schema import validate  # noqa: E402
from scenario_coverage import coverage_errors  # noqa: E402
from semantic_contract import semantic_contract_errors  # noqa: E402
from stream_contract import stream_errors, stream_summary  # noqa: E402

TIMESTAMP = "2026-07-25T10:01:03+00:00"


def direct_evidence(
    evidence_id: str,
    kind: str,
    source: str,
    path: str,
    *,
    action_id: str = "ACT-001",
    runtime_check_id: str | None = None,
    runtime_phase: str | None = None,
) -> dict[str, object]:
    row: dict[str, object] = {
        "evidence_id": evidence_id,
        "kind": kind,
        "source": source,
        "capture_mode": "direct",
        "action_id": action_id,
        "path_or_url": path,
        "captured_at": TIMESTAMP,
        "description": f"Direct {kind} fixture evidence.",
    }
    if runtime_check_id:
        row["runtime_check_id"] = runtime_check_id
    if runtime_phase:
        row["runtime_phase"] = runtime_phase
    return row


def upgrade_to_v2(base_dir: Path) -> tuple[dict, dict]:
    data = fixture()
    session = execution_fixture(data)
    data["run"]["operator_contract_version_required"] = 2
    session["operator_contract_version"] = 2
    session["run_id"] = data["run"]["run_id"]
    session["browser_binding"] = {
        "browser_instance_id": "BROWSER-EXISTING-001",
        "browser_context_id": "desktop-default",
        "profile_path": session["profile_path"],
        "approved_existing_session": True,
        "registered_at": "2026-07-25T09:58:00+00:00",
    }

    case = session["cases"][0]
    case.update(
        {
            "coverage_decision_id": "COV-EVG-001",
            "scenario_class_id": "SCN-DEFAULT",
            "sample_role": "SINGLETON",
            "selection_rationale": "Only one materially distinct add-to-cart contract exists.",
            "population_member_id": "product-default",
            "dimension_values": {"DIM-CONTRACT": "default"},
            "acquisition_context": {
                "kind": "NOT_APPLICABLE",
                "method": "NOT_APPLICABLE",
                "limitations": [],
            },
            "gated_flow_kind": "NONE",
        }
    )
    session["coverage_decisions"] = [
        {
            "coverage_decision_id": "COV-EVG-001",
            "event_group_id": "EVG-001",
            "revision": 1,
            "status": "FROZEN",
            "recorded_at": "2026-07-25T09:58:00+00:00",
            "frozen_at": "2026-07-25T09:58:30+00:00",
            "discovery_sources": ["tracking_plan", "visible_website_state"],
            "population_scope": "One materially distinct add-to-cart behavior contract.",
            "population_complete": True,
            "limitations": [],
            "expansion_triggers": [
                "NEW_BEHAVIOR_SIGNATURE",
                "ANOMALY_OR_FAILURE",
                "UNSEEN_MATERIAL_DIMENSION_VALUE",
                "CONDITIONAL_RUNTIME_BRANCH",
            ],
            "dimensions": [
                {
                    "dimension_id": "DIM-CONTRACT",
                    "name": "runtime behavior contract",
                    "source": "tracking_plan",
                    "reason": "The plan and site expose one payload/tag/journey contract.",
                    "treatment": "ENUMERATE",
                    "values": ["default"],
                    "material": True,
                }
            ],
            "scenario_classes": [
                {
                    "scenario_class_id": "SCN-DEFAULT",
                    "name": "Default product add-to-cart",
                    "population_source": "visible website and tracking plan",
                    "selection_method": "Exhaust the singleton behavior class.",
                    "selection_mode": "SINGLETON",
                    "behavior_signature": {
                        "action_path": "product CTA click",
                        "page_or_component": "product detail",
                        "data_source": "dataLayer ecommerce object",
                        "payload_contract": "add_to_cart with value",
                        "tag_contract": "GA4 add_to_cart browser request",
                        "consent_context": "natural granted baseline",
                        "acquisition_context": "not applicable",
                        "journey_precondition": "empty basket",
                    },
                    "dimension_values": {"DIM-CONTRACT": "default"},
                    "population_estimate": 1,
                    "limitations": [],
                    "required_sample_roles": ["SINGLETON"],
                    "case_ids": ["CASE-001"],
                    "expansion_review": {
                        "status": "NOT_TRIGGERED",
                        "reason": "No anomaly, new signature, or unseen material branch appeared.",
                        "additional_case_ids": [],
                        "trigger_reviews": {
                            trigger: {
                                "detected": False,
                                "outcome": "NOT_TRIGGERED",
                                "reason": "No current-run evidence activated this trigger.",
                                "additional_case_ids": [],
                            }
                            for trigger in (
                                "NEW_BEHAVIOR_SIGNATURE",
                                "ANOMALY_OR_FAILURE",
                                "UNSEEN_MATERIAL_DIMENSION_VALUE",
                                "CONDITIONAL_RUNTIME_BRANCH",
                            )
                        },
                    },
                }
            ],
        }
    ]

    action = session["actions"][0]
    action["datalayer_call_index_before"] = 0
    action["datalayer_call_index_after"] = 1
    action["page_health_before"] = {"status": "PASS"}
    action["page_health_after"] = {"status": "PASS"}
    page_evidence = []
    for check in session["runtime_checks"]:
        before = check["phase"] == "before_action"
        check.update(
            {
                "browser_instance_id": "BROWSER-EXISTING-001",
                "tab_id": "TAB-SITE-001",
                "preview_session_id": "PREVIEW-001",
                "loaded_client_container_ids": ["GTM-TEST"],
                "datalayer_call_cursor": 0 if before else 1,
                "page_health": {
                    "status": "PASS",
                    "reachable": True,
                    "http_status": 200,
                    "is_error_page": False,
                    "is_soft_404": False,
                    "expected_content_present": True,
                    "action_target_present": before,
                    "reason": "Expected product page content is directly visible.",
                    "evidence_ids": ["EVD-PAGE-READY" if before else "EVD-PAGE-SETTLED"],
                },
            }
        )
        evidence_id = "EVD-PAGE-READY" if before else "EVD-PAGE-SETTLED"
        check["evidence_ids"].append(evidence_id)
        page_evidence.append(
            direct_evidence(
                evidence_id,
                "page_health",
                "Playwright",
                f"evidence/{evidence_id.lower()}.json",
                runtime_check_id=check["check_id"],
                runtime_phase=check["phase"],
            )
        )
        page_evidence[-1]["captured_at"] = check["captured_at"]
    data["evidence"].extend(page_evidence)

    push = session["business_pushes"][0]
    push.update(
        {
            "segment_id": "SEG-ACT-001",
            "preview_event_index": 11,
            "datalayer_call_index": 1,
        }
    )
    session["stream_contract"] = {
        "status": "CLOSED",
        "started_at": "2026-07-25T10:00:59+00:00",
        "closed_at": TIMESTAMP,
        "start_preview_event_index": 10,
        "start_datalayer_call_index": 0,
        "reviewed_through_preview_event_index": 12,
        "reviewed_through_datalayer_call_index": 1,
    }
    session["stream_segments"] = [
        {
            "segment_id": "SEG-INITIAL",
            "kind": "INITIAL_LOAD",
            "status": "RECONCILED",
            "connection_epoch": 1,
            "action_id": None,
            "previous_segment_id": None,
            "start_preview_event_index": 10,
            "end_preview_event_index": 10,
            "start_datalayer_call_index": 0,
            "end_datalayer_call_index": 0,
            "started_at": "2026-07-25T10:00:59+00:00",
            "ended_at": "2026-07-25T10:00:59+00:00",
            "evidence_ids": ["EVD-PAGE-READY"],
            "observed_push_ids": [],
            "datalayer_call_reviews": [],
        },
        {
            "segment_id": "SEG-ACT-001",
            "kind": "ACTION",
            "status": "RECONCILED",
            "connection_epoch": 1,
            "action_id": "ACT-001",
            "previous_segment_id": "SEG-INITIAL",
            "start_preview_event_index": 10,
            "end_preview_event_index": 12,
            "start_datalayer_call_index": 0,
            "end_datalayer_call_index": 1,
            "started_at": "2026-07-25T10:00:59+00:00",
            "ended_at": TIMESTAMP,
            "evidence_ids": ["EVD-RAW-011"],
            "observed_push_ids": ["PUSH-011"],
            "datalayer_call_reviews": [
                {
                    "call_index": 1,
                    "evidence_id": "EVD-RAW-011",
                    "reason": "The only argument is fully classified.",
                    "arguments": [
                        {
                            "argument_index": 0,
                            "event_field_present": True,
                            "event_name": "add_to_cart",
                            "disposition": "BUSINESS_EVENT",
                            "push_id": "PUSH-011",
                            "capture_complete": True,
                            "reason": "Custom event maps to the planned business push.",
                        }
                    ],
                }
            ],
        },
        {
            "segment_id": "SEG-FINAL",
            "kind": "FINAL",
            "status": "RECONCILED",
            "connection_epoch": 1,
            "action_id": None,
            "previous_segment_id": "SEG-ACT-001",
            "start_preview_event_index": 12,
            "end_preview_event_index": 12,
            "start_datalayer_call_index": 1,
            "end_datalayer_call_index": 1,
            "started_at": TIMESTAMP,
            "ended_at": TIMESTAMP,
            "evidence_ids": ["EVD-PAGE-SETTLED"],
            "observed_push_ids": [],
            "datalayer_call_reviews": [],
        },
    ]

    data["evidence"].extend(
        [
            direct_evidence(
                "EVD-JOURNEY-BEFORE",
                "journey_state",
                "Playwright",
                "evidence/journey-before.json",
            ),
            direct_evidence(
                "EVD-JOURNEY-AFTER",
                "journey_state",
                "Playwright",
                "evidence/journey-after.json",
            ),
        ]
    )
    session["journey_states"] = [
        {
            "state_id": "STATE-BEFORE",
            "event_group_id": "EVG-001",
            "action_id": "ACT-001",
            "case_id": "CASE-001",
            "phase": "BEFORE",
            "captured_at": "2026-07-25T10:00:59+00:00",
            "values": {"basket_count": 0, "basket_value": 0},
            "summary": "Basket is empty before the action.",
            "evidence_ids": ["EVD-JOURNEY-BEFORE"],
            "sensitive_scan_status": "PASS",
        },
        {
            "state_id": "STATE-AFTER",
            "event_group_id": "EVG-001",
            "action_id": "ACT-001",
            "case_id": "CASE-001",
            "phase": "AFTER",
            "captured_at": TIMESTAMP,
            "values": {"basket_count": 1, "basket_value": 29.9},
            "summary": "Basket contains the selected product after the action.",
            "evidence_ids": ["EVD-JOURNEY-AFTER"],
            "sensitive_scan_status": "PASS",
        },
    ]
    session["semantic_checks"] = [
        {
            "check_id": "SEM-PAGE-001",
            "action_id": "ACT-001",
            "case_id": "CASE-001",
            "event_group_id": "EVG-001",
            "requirement_id": None,
            "kind": "PAGE_ACTION_VALIDITY",
            "authority": "VISIBLE_PAGE",
            "comparison": "EQUAL",
            "anchor_state": "PRESENT",
            "anchor_value": True,
            "observed_value": True,
            "anchor_field_state": "value",
            "observed_field_state": "value",
            "subject": "The intended product page and CTA are valid.",
            "status": "PASS",
            "reason": "The route is healthy and the intended CTA is present.",
            "recorded_at": TIMESTAMP,
            "evidence_ids": ["EVD-PAGE-READY", "EVD-PAGE-SETTLED"],
        },
        {
            "check_id": "SEM-BUSINESS-001",
            "action_id": "ACT-001",
            "case_id": "CASE-001",
            "event_group_id": "EVG-001",
            "requirement_id": "REQ-001",
            "kind": "BUSINESS_STATE",
            "authority": "VISIBLE_PAGE",
            "comparison": "EQUAL",
            "anchor_state": "PRESENT",
            "anchor_value": 29.9,
            "observed_value": 29.9,
            "anchor_field_state": "value",
            "observed_field_state": "value",
            "subject": "Tracked basket value matches the visible basket value.",
            "status": "PASS",
            "reason": "Both the visible and tracked value are 29.9.",
            "recorded_at": TIMESTAMP,
            "evidence_ids": ["EVD-JOURNEY-AFTER", "EVD-RAW-011"],
        },
    ]
    session["protected_handoffs"] = []
    session["gated_flows"] = []
    session["event_closures"][0]["coverage_revision"] = 1

    for evidence in data["evidence"]:
        path = base_dir / str(evidence["path_or_url"])
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(f"evidence:{evidence['evidence_id']}".encode())
    session["evidence_integrity"] = build_integrity_record(data, base_dir)
    return data, session


class OperatorV2Tests(unittest.TestCase):
    def test_normal_datalayer_event_has_ten_default_mandatory_layers(self) -> None:
        rows = layer_applicability(
            [
                {
                    "scope_status": "IN_SCOPE",
                    "expectation": {"source_mechanism": "data_layer_push"},
                }
            ]
        )
        mandatory = [row["layer"] for row in rows if row["mode"] == "MANDATORY"]
        self.assertEqual(
            [
                "action_boundary",
                "raw_api_call",
                "resolved_data_layer",
                "concerned_tag_inventory",
                "gtm_variable",
                "tag_configuration",
                "tag_firing",
                "tag_parameter",
                "destination_request_when_applicable",
                "sensitive_data_scan",
            ],
            mandatory,
        )

    def test_full_v2_fixture_is_valid(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            self.assertEqual([], validate(data, strict=False))
            self.assertEqual([], validate_session(session, results=data, final=True))

    def test_dead_page_is_overall_ko_while_tag_delivery_stays_ok(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            action = session["actions"][0]
            action.update(
                {
                    "target_ready_before": False,
                    "interaction_outcome": "failed",
                    "settlement_reason": "interaction_failed",
                    "completion_signal": "HTTP 404 and error template observed.",
                }
            )
            action["layer_results"][0].update(
                {"status": "FAIL", "reason": "The intended URL returned HTTP 404."}
            )
            data["requirements"][0]["action_boundary"].update(
                {
                    "interaction_outcome": "failed",
                    "settlement_reason": "interaction_failed",
                    "completion_signal": "HTTP 404 and error template observed.",
                }
            )
            for check in session["runtime_checks"]:
                check["page_health"].update(
                    {
                        "status": "FAIL",
                        "reachable": True,
                        "http_status": 404,
                        "is_error_page": True,
                        "expected_content_present": False,
                        "action_target_present": False,
                        "reason": "HTTP 404 error template observed.",
                    }
                )
                if check["phase"] == "before_action":
                    check["target_interactive"] = False
                    check["target_uncovered"] = False
            page_check = session["semantic_checks"][0]
            page_check.update(
                {
                    "status": "FAIL",
                    "observed_value": False,
                    "reason": "The intended URL is a 404 page.",
                }
            )
            session["semantic_checks"][1].update(
                {
                    "status": "FAIL",
                    "reason": "No valid business action exists on the dead URL.",
                }
            )
            session["coverage_decisions"][0]["scenario_classes"][0]["expansion_review"].update(
                {
                    "status": "EXHAUSTED",
                    "reason": "The route defect is deterministic; another product cannot repair it.",
                    "additional_case_ids": [],
                }
            )
            session["coverage_decisions"][0]["scenario_classes"][0]["expansion_review"][
                "trigger_reviews"
            ]["ANOMALY_OR_FAILURE"].update(
                {
                    "detected": True,
                    "outcome": "EXHAUSTED",
                    "reason": "The deterministic route defect was directly confirmed.",
                }
            )
            session["evidence_integrity"] = build_integrity_record(data, Path(tempdir))
            self.assertEqual([], validate_session(session, results=data, final=True))
            feedback = event_feedback(data, session)[0]
            self.assertEqual("FAIL", feedback["status"])
            self.assertEqual("PASS", feedback["technical_status"])
            self.assertEqual("KO", feedback["status_label"])

    def test_empty_positive_anchor_cannot_pass(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            check = session["semantic_checks"][1]
            check.update(
                {
                    "anchor_value": "",
                    "observed_value": "",
                    "anchor_field_state": "empty",
                    "observed_field_state": "empty",
                }
            )
            errors = semantic_contract_errors(session, results=data, final=True)
            self.assertTrue(any("matching emptiness" in error for error in errors))

    def test_every_action_requires_explicit_business_state_judgement(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            session["semantic_checks"] = [session["semantic_checks"][0]]
            errors = semantic_contract_errors(session, results=data, final=True)
            self.assertTrue(any("missing BUSINESS_STATE" in error for error in errors))

    def test_populated_cart_with_empty_tracking_state_is_overall_ko(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            check = session["semantic_checks"][1]
            check.update(
                {
                    "status": "FAIL",
                    "comparison": "EQUAL",
                    "anchor_value": [{"item_id": "SKU-1", "quantity": 1}],
                    "observed_value": [],
                    "anchor_field_state": "value",
                    "observed_field_state": "value",
                    "subject": "Visible populated basket versus view_cart tracking state.",
                    "reason": "The basket contains a product but tracking returned no items.",
                }
            )
            session["coverage_decisions"][0]["scenario_classes"][0]["expansion_review"].update(
                {
                    "status": "EXHAUSTED",
                    "reason": "The visible/tracked contradiction is directly confirmed.",
                    "additional_case_ids": [],
                }
            )
            session["coverage_decisions"][0]["scenario_classes"][0]["expansion_review"][
                "trigger_reviews"
            ]["ANOMALY_OR_FAILURE"].update(
                {
                    "detected": True,
                    "outcome": "EXHAUSTED",
                    "reason": "The visible/tracked contradiction was directly confirmed.",
                }
            )
            self.assertEqual([], validate_session(session, results=data, final=True))
            feedback = event_feedback(data, session)[0]
            self.assertEqual("FAIL", feedback["status"])
            self.assertEqual("PASS", feedback["technical_status"])
            self.assertIn("tracking returned no items", feedback["reason"])

    def test_inter_action_custom_event_is_retained_and_classified(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            _, session = upgrade_to_v2(Path(tempdir))
            session["business_pushes"].append(
                {
                    "push_id": "PUSH-INTER-001",
                    "stream_id": "tag_assistant",
                    "connection_epoch": 1,
                    "action_id": None,
                    "case_id": None,
                    "event_group_id": "EVG-001",
                    "event_name": "mystery_cart_refresh",
                    "event_index": 13,
                    "preview_event_index": 13,
                    "datalayer_call_index": 2,
                    "segment_id": "SEG-INTER-001",
                    "captured_at": TIMESTAMP,
                    "url": "https://shop.example.test/product",
                    "page_state": "Between planned interactions.",
                    "classification": "unplanned_relevant",
                    "classification_reason": "Unexpected custom event between plan actions.",
                    "evidence_id": "EVD-RAW-011",
                    "container_id": "GTM-TEST",
                }
            )
            session["stream_contract"].update(
                {
                    "reviewed_through_preview_event_index": 13,
                    "reviewed_through_datalayer_call_index": 2,
                }
            )
            session["stream_segments"].insert(
                -1,
                {
                    "segment_id": "SEG-INTER-001",
                    "kind": "INTER_ACTION",
                    "status": "RECONCILED",
                    "connection_epoch": 1,
                    "action_id": None,
                    "previous_segment_id": "SEG-ACT-001",
                    "start_preview_event_index": 12,
                    "end_preview_event_index": 13,
                    "start_datalayer_call_index": 1,
                    "end_datalayer_call_index": 2,
                    "started_at": TIMESTAMP,
                    "ended_at": TIMESTAMP,
                    "evidence_ids": ["EVD-RAW-011"],
                    "observed_push_ids": ["PUSH-INTER-001"],
                    "datalayer_call_reviews": [
                        {
                            "call_index": 2,
                            "evidence_id": "EVD-RAW-011",
                            "reason": "The between-action call is classified.",
                            "arguments": [
                                {
                                    "argument_index": 0,
                                    "event_field_present": True,
                                    "event_name": "mystery_cart_refresh",
                                    "disposition": "BUSINESS_EVENT",
                                    "push_id": "PUSH-INTER-001",
                                    "capture_complete": True,
                                    "reason": "Custom event requires anomaly classification.",
                                }
                            ],
                        }
                    ],
                },
            )
            session["stream_segments"][-1].update(
                {
                    "previous_segment_id": "SEG-INTER-001",
                    "start_preview_event_index": 13,
                    "end_preview_event_index": 13,
                    "start_datalayer_call_index": 2,
                    "end_datalayer_call_index": 2,
                }
            )
            self.assertEqual([], stream_errors(session, final=True))
            self.assertEqual(1, stream_summary(session, "EVG-001")["inter_action_push_count"])

    def test_interaction_gap_custom_event_cannot_be_hidden(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            _, session = upgrade_to_v2(Path(tempdir))
            argument = session["stream_segments"][1]["datalayer_call_reviews"][0]["arguments"][0]
            argument["disposition"] = "STATE_UPDATE"
            argument["push_id"] = None
            errors = stream_errors(session, final=True)
            self.assertTrue(any("cannot be hidden" in error for error in errors))

    def test_gapless_call_review_rejects_missing_call(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            _, session = upgrade_to_v2(Path(tempdir))
            session["stream_segments"][1]["datalayer_call_reviews"] = []
            errors = stream_errors(session, final=True)
            self.assertTrue(any("every call index" in error for error in errors))

    def test_final_stream_requires_explicit_initial_and_final_segments(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            _, session = upgrade_to_v2(Path(tempdir))
            action_segment = session["stream_segments"][1]
            action_segment["previous_segment_id"] = None
            session["stream_segments"] = [action_segment]
            errors = stream_errors(session, final=True)
            self.assertTrue(any("leading INITIAL_LOAD" in error for error in errors))
            self.assertTrue(any("trailing FINAL" in error for error in errors))

    def test_wrong_loaded_preview_container_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            session["runtime_checks"][0]["loaded_client_container_ids"] = ["GTM-WRONG"]
            errors = validate_session(session, results=data, final=True)
            self.assertTrue(any("loaded client containers differ" in error for error in errors))

    def test_referrer_simulation_is_allowed_with_limitation(self) -> None:
        value = {
            "kind": "REFERRER",
            "method": "BROWSER_SIMULATED",
            "fresh_state": True,
            "referrer_url": "https://www.google.com/search?q=example",
            "evidence_ids": ["EVD-NAV-001"],
            "limitations": [
                "Referrer was simulated in the approved browser context, not produced by indexing."
            ],
        }
        self.assertEqual([], acquisition_errors(value))

    def test_sampling_requires_ordinary_and_contrast(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            scenario = session["coverage_decisions"][0]["scenario_classes"][0]
            scenario.update(
                {
                    "selection_mode": "SAMPLED",
                    "population_estimate": 500,
                    "required_sample_roles": ["ORDINARY"],
                }
            )
            errors = coverage_errors(session, results=data, final=True)
            self.assertTrue(any("CONTRAST" in error for error in errors))

    def test_exhaustive_coverage_cannot_claim_an_untested_finite_population(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            scenario = session["coverage_decisions"][0]["scenario_classes"][0]
            scenario.update(
                {
                    "selection_mode": "EXHAUSTIVE",
                    "population_estimate": 3,
                    "required_sample_roles": ["EXHAUSTIVE"],
                }
            )
            session["cases"][0]["sample_role"] = "EXHAUSTIVE"
            errors = coverage_errors(session, results=data, final=True)
            self.assertTrue(any("execute every population member" in error for error in errors))

    def test_captcha_handoff_must_resume_same_tab_and_preview(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            _, session = upgrade_to_v2(Path(tempdir))
            binding = {
                "browser_instance_id": "BROWSER-EXISTING-001",
                "browser_context_id": "desktop-default",
                "tab_id": "TAB-SITE-001",
                "preview_session_id": "PREVIEW-001",
            }
            session["protected_handoffs"] = [
                {
                    "handoff_id": "HANDOFF-CAPTCHA-001",
                    "gate_type": "CAPTCHA",
                    "status": "RESUMED",
                    "analyst_help_requested": True,
                    "case_id": "CASE-001",
                    "action_id": "ACT-001",
                    "requested_at": TIMESTAMP,
                    "resumed_at": TIMESTAMP,
                    "reason": "Analyst completed CAPTCHA in the approved tab.",
                    "evidence_ids": ["EVD-PAGE-READY"],
                    "before_binding": binding,
                    "after_binding": {**binding, "tab_id": "TAB-NEW"},
                }
            ]
            errors = handoff_errors(session, final=True)
            self.assertTrue(any("different tab_id" in error for error in errors))

    def test_captcha_flow_requires_handoff(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            _, session = upgrade_to_v2(Path(tempdir))
            session["cases"][0]["gated_flow_kind"] = "FORM"
            session["gated_flows"] = [
                {
                    "flow_id": "FLOW-001",
                    "case_id": "CASE-001",
                    "action_id": "ACT-001",
                    "kind": "FORM",
                    "status": "BLOCKED",
                    "safe_environment_confirmed": True,
                    "synthetic_data_used": True,
                    "consent_outcome": "ACCEPTED",
                    "captcha_outcome": "BLOCKED",
                    "states": ["DISCOVERED", "HANDOFF_REQUESTED", "BLOCKED"],
                    "recorded_at": TIMESTAMP,
                    "reason": "CAPTCHA requires an analyst handoff.",
                    "evidence_ids": ["EVD-PAGE-READY"],
                }
            ]
            errors = gated_flow_errors(session, final=True)
            self.assertTrue(any("requires a protected handoff" in error for error in errors))

    def test_recorder_snapshot_rejects_previous_run_residue(self) -> None:
        snapshot = {
            "runId": "RUN-OLD",
            "records": [{"callIndex": 1, "arguments": [{"event": "view_item"}]}],
        }
        with self.assertRaisesRegex(ValueError, "differs from the current run"):
            build_review(
                snapshot,
                segment_id="SEG-1",
                start_exclusive=0,
                end_inclusive=1,
                evidence_id="EVD-1",
                expected_run_id="RUN-NEW",
            )

    def test_recorder_review_retains_current_run_identity(self) -> None:
        review = build_review(
            {"runId": "RUN-NEW", "records": []},
            segment_id="SEG-EMPTY",
            start_exclusive=0,
            end_inclusive=0,
            evidence_id="EVD-RECORDER",
            expected_run_id="RUN-NEW",
        )
        self.assertEqual(2, review["version"])
        self.assertEqual("RUN-NEW", review["run_id"])

    def test_final_conclusion_lists_event_layers_status_and_reason(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            data, session = upgrade_to_v2(Path(tempdir))
            conclusion = final_conclusion(data, session)
            self.assertEqual("PASS", conclusion["status"])
            self.assertEqual("EVG-001", conclusion["events"][0]["event_group_id"])
            self.assertTrue(conclusion["events"][0]["layers_inspected"])
            self.assertTrue(conclusion["events"][0]["why"])

    def test_v2_workbook_contains_expert_detail_and_conclusion_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            root = Path(tempdir)
            data, session = upgrade_to_v2(root)
            output = root / "recette.xlsx"
            build_workbook(data, output, [], session)
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(required_sheets(session), workbook.sheetnames)
                self.assertEqual("Final Conclusion", workbook.sheetnames[-2])
                self.assertGreater(workbook["Semantic Checks"].max_row, 1)
                self.assertGreater(workbook["Stream Segments"].max_row, 1)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
