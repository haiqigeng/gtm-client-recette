from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from tests.v5_harness import V5Harness, default_event, write_json

from core.capture import capture_value
from core.constants import worst_status
from core.correlate import action_windows, build_model, source_event_names
from core.coverage import validate_coverage_annotation
from core.plan import normalize_plan
from core.state import StateError, read_stream
from core.workflow import commit_action, complete_action, next_action, sync_preview


class GeneralOptimizationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    @staticmethod
    def source_only_scope() -> dict:
        return {
            "approved": True,
            "origins": [],
            "certify_tags": False,
            "browser_send_required": False,
        }

    def test_sectioned_workbook_continues_after_blanks_and_reconciles_every_sheet(self) -> None:
        workbook = Workbook()
        index = workbook.active
        index.title = "Index"
        index.append(["Event Name"])
        index.append(["Core DataLayer"])
        index.append(["Missing Event"])

        core = workbook.create_sheet("Core DataLayer")
        core.append([None, "Name of the event", "Page", "Trigger"])
        core.append([None, "Message", "All pages", "Before GTM"])
        core.append([])
        core.append([None, "variables", "Type", "Status", "Summary", "Values"])
        core.append([None, "userStatus", "string", "Mandatory", "User", None])
        core.append([])
        core.append([None, "siteCountry", "string", "Mandatory", "Country", "FR | BE"])
        core.append([None, "dataLayer.push({event: 'example_only'})"])
        core.append([None, "lateField", "string", "Mandatory", "Late field", None])

        notes = workbook.create_sheet("Notes")
        notes.append(["Human notes only"])
        source = self.root / "plan.xlsx"
        workbook.save(source)
        workbook.close()

        plan = normalize_plan(source, scope=self.source_only_scope())
        core_event = plan["events"][0]
        paths = {
            claim["target"].get("path")
            for claim in core_event["claims"]
            if claim["domain"] == "source"
        }
        normalization = plan["source"]["normalization"]

        self.assertEqual(core_event["mode"], "state_only")
        self.assertTrue({"userStatus", "siteCountry", "lateField"}.issubset(paths))
        self.assertNotIn("example_only", str(plan))
        self.assertEqual(plan["scope"]["origin_mode"], "prepared_runtime")
        self.assertEqual(normalization["reconciliation"]["index_only_events"], ["Missing Event"])
        self.assertEqual(normalization["reconciliation"]["status"], "REVIEW")
        self.assertEqual(
            {row["sheet"]: row["classification"] for row in normalization["sheet_manifest"]},
            {"Index": "index", "Core DataLayer": "requirements", "Notes": "ignored"},
        )
        missing = plan["events"][1]
        self.assertFalse(missing["executable"])
        self.assertIn("no requirement sheet", missing["compile_errors"][0])

    def test_exact_machine_identifiers_keep_case_and_origin_can_come_from_runtime(self) -> None:
        event = default_event("E-custom", "customEventABC")
        event["requirements"].append({"field_path": "customParamABC", "match_rule": "present"})
        source = write_json(self.root / "plan.json", {"events": [event]})
        plan = normalize_plan(source, scope={"approved": True, "tag_scope": ["GA4"]})
        compiled = plan["events"][0]
        self.assertEqual(compiled["event_name"], "customEventABC")
        self.assertIn(
            "customParamABC",
            {claim["target"].get("path") for claim in compiled["claims"]},
        )
        self.assertEqual(plan["scope"]["origin_mode"], "plan_inferred")

    def test_runtime_contract_is_capability_based_not_exact_version_based(self) -> None:
        harness = V5Harness(self.root)
        capability = harness.capability()
        capability["runtime"].pop("mcp_version")
        capability["runtime"].pop("self_check")
        started = next_action(
            harness.run,
            harness.next_input(capability),
            event_ids=["E-view_item"],
        )
        self.assertTrue(started["action"]["data"]["action_id"].startswith("A-"))

    def test_unavailable_preview_surface_blocks_only_dependent_layers_without_retry(self) -> None:
        harness = V5Harness(self.root)
        capability = harness.capability(
            preview_events=False,
            preview_tag_inventory=False,
            preview_variables=False,
            preview_consent=False,
        )
        first_input = harness.next_input(capability)
        first_input["setup_boundary"]["preview_cursor"]["epoch"] = None
        started = next_action(
            harness.run,
            first_input,
            event_ids=["E-view_item"],
        )
        action_id = started["action"]["data"]["action_id"]
        payload = {"event": "view_item"}
        result = complete_action(
            harness.run,
            {
                "binding": harness.binding(),
                "health": harness._health("after"),
                "page": {"states": [harness._page("after")]},
                "datalayer": harness.datalayer([payload], action_id=action_id),
                "network": harness.network(["view_item"], action_id=action_id, payloads=[payload]),
                "coverage": harness.coverage("E-view_item", [action_id]),
            },
            action_id=action_id,
        )["events"][0]
        self.assertEqual(result["domains"]["source"]["status"], "PASS")
        self.assertEqual(result["domains"]["gtm"]["status"], "BLOCKED")
        self.assertEqual(
            sum(row["kind"] == "ACTION_BEGIN" for row in read_stream(harness.run)[0]), 1
        )

    def test_preview_cursor_rejects_historical_rows_and_inherits_top_level_action(self) -> None:
        harness = V5Harness(self.root)
        with self.assertRaisesRegex(StateError, "historical index"):
            capture_value(
                harness.run,
                "preview",
                {
                    "complete": True,
                    "epoch": "EPOCH-1",
                    "cursor_start": 3,
                    "events": [
                        {
                            "index": 3,
                            "fired_tags": [],
                            "not_fired_tags": [],
                        }
                    ],
                },
                source_id="historical",
            )

        action = harness.begin()
        harness.commit(action, [{"event": "view_item"}], include_source=False)
        preview = harness.preview([{"event": "view_item"}], action_id=None)
        preview["action_id"] = action
        sync_preview(
            harness.run,
            {
                "preview": preview,
                "coverage": harness.coverage("E-view_item", [action]),
            },
            event_ids=["E-view_item"],
        )
        records, _ = read_stream(harness.run)
        model = build_model(harness.run, harness.plan, records)
        self.assertEqual(model["preview_events"][0]["action_id"], action)

    def test_preview_api_fallback_keeps_other_occurrences_in_the_same_action(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        planned = {"event": "view_item"}
        harness.commit(action, [planned])
        preview = harness.preview([planned, {"event": "unexpected_event"}], action_id=action)
        sync_preview(
            harness.run,
            {
                "preview": preview,
                "coverage": harness.coverage("E-view_item", [action]),
            },
            event_ids=["E-view_item"],
        )
        records, _ = read_stream(harness.run)
        model = build_model(harness.run, harness.plan, records)
        self.assertEqual(
            source_event_names(model, action, authoritative_only=True),
            ["view_item", "unexpected_event"],
        )

    def test_preview_reconciliation_deduplicates_one_to_one_without_hiding_a_duplicate(
        self,
    ) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        planned = {"event": "view_item"}
        harness.commit(action, [planned])
        harness.sync(
            ["E-view_item"],
            [planned, planned],
            action_id=action,
            coverage=harness.coverage("E-view_item", [action]),
        )
        records, _ = read_stream(harness.run)
        model = build_model(harness.run, harness.plan, records)
        self.assertEqual(
            source_event_names(model, action, authoritative_only=True),
            ["view_item", "view_item"],
        )

    def test_evidence_defect_retest_supersedes_the_bad_action(self) -> None:
        harness = V5Harness(self.root)
        first = harness.begin()
        first_commit = harness.commit(first, [{"event": "purchase"}])
        first_result = harness.sync(
            ["E-view_item"],
            [{"event": "purchase"}],
            action_id=first,
            coverage=harness.coverage("E-view_item", [first]),
        )["events"][0]
        self.assertEqual(first_result["status"], "FAIL")

        second = harness.begin(
            retest_basis={
                "type": "EVIDENCE_DEFECT",
                "record_id": first_commit["commit"]["record_id"],
                "reason": "The first evidence window selected the wrong occurrence.",
            }
        )
        harness.commit(second, [{"event": "view_item"}])
        second_result = harness.sync(
            ["E-view_item"],
            [{"event": "view_item"}],
            action_id=second,
            coverage=harness.coverage("E-view_item", [second]),
        )["events"][0]
        actions = {row["action_id"]: row for row in action_windows(read_stream(harness.run)[0])}
        self.assertEqual(actions[first]["status"], "SUPERSEDED")
        self.assertEqual(actions[first]["superseded_by"], second)
        self.assertEqual(second_result["status"], "PASS")

    def test_evidence_defect_retest_cannot_supersede_another_event_slice(self) -> None:
        harness = V5Harness(
            self.root,
            events=[
                default_event("E-view_item", "view_item"),
                default_event("E-add_to_cart", "add_to_cart"),
            ],
        )
        first = harness.begin(["E-view_item"])
        committed = harness.commit(first, [{"event": "view_item"}])
        with self.assertRaisesRegex(StateError, "exact event slice and scenario"):
            harness.begin(
                ["E-add_to_cart"],
                retest_basis={
                    "type": "EVIDENCE_DEFECT",
                    "record_id": committed["commit"]["record_id"],
                    "reason": "The first evidence window was incomplete.",
                },
            )

    def test_high_cardinality_needs_contrast_only_for_distinct_behavior_signatures(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        harness.commit(action, [{"event": "view_item"}])
        review = harness.coverage(
            "E-view_item",
            [action],
            mode="SAMPLED",
            dimensions=[
                {
                    "name": "product",
                    "kind": "high_cardinality",
                    "material": True,
                    "values": ["SKU-1", "SKU-2", "SKU-3"],
                }
            ],
            scenarios=[
                {
                    "scenario_id": "ordinary-product",
                    "role": "ORDINARY",
                    "behavior_signature": "standard-pdp",
                    "values": {"product": "SKU-1"},
                    "action_ids": [action],
                }
            ],
        )
        self.assertEqual(
            validate_coverage_annotation(harness.plan, read_stream(harness.run)[0], review),
            [],
        )

    def test_annotation_cannot_weaken_a_compiler_known_finite_dimension(self) -> None:
        event = default_event()
        event["known_dimensions"] = [
            {
                "name": "language",
                "kind": "manageable_finite",
                "material": True,
                "values": ["en", "fr"],
            }
        ]
        harness = V5Harness(self.root, events=[event])
        action = harness.begin(scenario_values={"language": "en"})
        harness.commit(action, [{"event": "view_item"}])
        review = harness.coverage(
            "E-view_item",
            [action],
            mode="SAMPLED",
            dimensions=[
                {
                    "name": "language",
                    "kind": "high_cardinality",
                    "material": True,
                    "values": ["en"],
                }
            ],
            scenario_values={"language": "en"},
        )
        errors = validate_coverage_annotation(harness.plan, read_stream(harness.run)[0], review)
        self.assertTrue(any("untested values: fr" in error for error in errors))
        self.assertTrue(any("SAMPLED coverage" in error for error in errors))

    def test_singleton_is_valid_only_when_no_second_material_member_is_known(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        harness.commit(action, [{"event": "view_item"}])
        valid = harness.coverage("E-view_item", [action], mode="SINGLETON")
        records = read_stream(harness.run)[0]
        self.assertEqual(validate_coverage_annotation(harness.plan, records, valid), [])
        invalid = {
            **valid,
            "dimensions": [
                {
                    "name": "language",
                    "kind": "manageable_finite",
                    "material": True,
                    "values": ["en", "fr"],
                }
            ],
        }
        self.assertTrue(validate_coverage_annotation(harness.plan, records, invalid))

    def test_denied_ordinary_consent_blocks_setup_but_explicit_denial_can_prove_suppression(
        self,
    ) -> None:
        event = default_event()
        event["tags"][0]["consent_requirements"] = ["analytics_storage"]

        ordinary = V5Harness(self.root / "ordinary", events=[event])
        ordinary_action = ordinary.begin()
        payload = {"event": "view_item"}
        ordinary.commit(
            ordinary_action,
            [payload],
            include_network=False,
            lifecycle={
                "action_id": ordinary_action,
                "complete": True,
                "events": [],
                "errors": [],
                "consent_transitions": [
                    {
                        "action_id": ordinary_action,
                        "kind": "user_choice",
                        "method": "natural",
                        "state": {"analytics_storage": "denied"},
                    }
                ],
            },
        )
        ordinary_result = ordinary.sync(
            ["E-view_item"],
            [payload],
            action_id=ordinary_action,
            coverage=ordinary.coverage("E-view_item", [ordinary_action]),
            preview_updates={
                "fire_tags": False,
                "consent": {"analytics_storage": "denied"},
            },
        )["events"][0]
        self.assertEqual(ordinary_result["status"], "BLOCKED")
        self.assertTrue(
            any(
                row["reason_code"] == "consent.ordinary_context_denied"
                for row in ordinary_result["inspections"]
            )
        )

        explicit = V5Harness(self.root / "explicit", events=[event])
        explicit_action = explicit.begin(
            scenario_id="consent-denied",
            scenario_values={"analytics_storage": "denied"},
        )
        commit_action(
            explicit.run,
            {
                "health": explicit._health("after"),
                "page": {"states": [explicit._page("after")]},
                "datalayer": explicit.datalayer([payload], action_id=explicit_action),
                "network": explicit.network([], action_id=explicit_action, payloads=[]),
                "lifecycle": {
                    "action_id": explicit_action,
                    "complete": True,
                    "events": [],
                    "errors": [],
                    "consent_transitions": [
                        {
                            "action_id": explicit_action,
                            "kind": "user_choice",
                            "method": "natural",
                            "state": {"analytics_storage": "denied"},
                        }
                    ],
                },
            },
            action_id=explicit_action,
        )
        explicit_result = explicit.sync(
            ["E-view_item"],
            [payload],
            action_id=explicit_action,
            coverage=explicit.coverage(
                "E-view_item",
                [explicit_action],
                scenario_id="consent-denied",
                scenario_values={"analytics_storage": "denied"},
            ),
            preview_updates={
                "fire_tags": False,
                "consent": {"analytics_storage": "denied"},
            },
        )["events"][0]
        self.assertEqual(explicit_result["status"], "PASS")
        self.assertTrue(
            any(
                row["reason_code"] == "consent.expected_suppression"
                for row in explicit_result["inspections"]
            )
        )

    def test_browser_cookie_metadata_is_redacted_without_becoming_a_client_defect(self) -> None:
        harness = V5Harness(self.root)
        action = harness.begin()
        payload = {"event": "view_item"}
        network = harness.network(["view_item"], action_id=action, payloads=[payload])
        network["requests"][0]["headers"] = {"Cookie": "session_id=secret"}
        commit_action(
            harness.run,
            {
                "health": harness._health("after"),
                "page": {"states": [harness._page("after")]},
                "datalayer": harness.datalayer([payload], action_id=action),
                "network": network,
            },
            action_id=action,
        )
        result = harness.sync(
            ["E-view_item"],
            [payload],
            action_id=action,
            coverage=harness.coverage("E-view_item", [action]),
        )["events"][0]
        records, _ = read_stream(harness.run)
        model = build_model(harness.run, harness.plan, records)
        self.assertEqual(result["domains"]["safety"]["status"], "PASS")
        self.assertEqual(model["privacy_findings"], [])
        self.assertIn("cookie", model["requests"][0]["excluded_header_names"])

    def test_status_priority_does_not_hide_review_behind_pending_coverage(self) -> None:
        self.assertEqual(worst_status(["REVIEW", "PENDING"]), "REVIEW")
        self.assertEqual(worst_status(["FAIL", "PENDING"]), "FAIL")

    def test_scheduler_discovers_explicit_scenario_after_the_ordinary_action(self) -> None:
        event = default_event()
        event["scenarios"] = [
            {
                "scenario_id": "alternate-language",
                "label": "French",
                "values": {"page_language": "fr"},
            }
        ]
        harness = V5Harness(self.root, events=[event])
        first = next_action(
            harness.run,
            harness.next_input(),
            event_ids=["E-view_item"],
        )
        action_id = first["action"]["data"]["action_id"]
        payload = {"event": "view_item"}
        complete_action(
            harness.run,
            {
                "binding": harness.binding(),
                "health": harness._health("after"),
                "page": {"states": [harness._page("after")]},
                "network": harness.network(["view_item"], action_id=action_id, payloads=[payload]),
                "preview": harness.preview([payload], action_id=None),
            },
            action_id=action_id,
        )
        second = next_action(harness.run)
        self.assertEqual(second["action"]["data"]["scenario_id"], "alternate-language")
        self.assertEqual(second["action"]["data"]["scenario_values"], {"page_language": "fr"})


if __name__ == "__main__":
    unittest.main()
