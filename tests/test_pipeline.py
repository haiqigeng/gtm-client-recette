from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import tempfile
import tomllib
import unittest
from argparse import Namespace
from copy import deepcopy
from datetime import UTC, datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as WorkbookImage
from PIL import Image as PillowImage

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
FIXTURES = ROOT / "tests" / "fixtures"
sys.path.insert(0, str(SCRIPTS))

from build_recette_report import (  # noqa: E402
    REQUIRED_SHEETS,
    build_workbook,
    layer_verdict_rows,
    tag_rows,
)
from client_side_rules import (  # noqa: E402
    evaluate_business_rule,
    path_value,
    scan_requirement_sensitive_data,
)
from decode_browser_requests import decode_requests  # noqa: E402
from diff_recette_runs import compare as compare_runs  # noqa: E402
from event_feedback import event_feedback  # noqa: E402
from execution_contract import (  # noqa: E402
    PROTECTED_AUTHORIZATION_EXCLUSIONS,
    validate_session,
)
from import_ga4_tracking_plan_handoff import (  # noqa: E402
    HandoffError,
    interpreted_requirements,
    verify_delivery,
)
from incremental_recette import apply_event, status_rows, validate_event  # noqa: E402
from inspect_tracking_plan import inspect_xlsx  # noqa: E402
from layer_contract import (  # noqa: E402
    CANONICAL_LAYERS,
    TAG_RESULT_LAYERS,
    applicable_layers,
    declared_tag_contracts,
    inferred_tag_category,
    layer_applicability,
    tag_scope_decision,
)
from migrate_schema_v2_to_v3 import migrate_results  # noqa: E402
from preview_session_ledger import (  # noqa: E402
    revise_tag_inventory,
    scaffold_tag_results,
)
from recette_operator import _save_pair_atomic  # noqa: E402
from recette_schema import ReportValidationError, event_rollup, validate  # noqa: E402
from runtime_state_contract import runtime_snapshot_errors  # noqa: E402
from verify_release_artifact import verify_archive  # noqa: E402


def fixture(name: str = "valid_full.json") -> dict:
    return json.loads((FIXTURES / name).read_text(encoding="utf-8"))


def requirement(data: dict) -> dict:
    return data["requirements"][0]


def value_type(value: object) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return type(value).__name__


def client_side_fixture() -> dict:
    data = fixture()
    extension = fixture("client_side_extension.json")
    data["run"].update(extension["run_patch"])
    req = requirement(data)
    patch = extension["requirement_patch"]
    req["container_id"] = patch["container_id"]
    req["browser_context_id"] = patch["browser_context_id"]
    req["scenario"] = patch["scenario"]
    req["expectation"].update(patch["expectation_patch"])
    req["raw_api_call"]["payload"].update(patch["raw_payload_patch"])
    req["resolved_data_layer"]["snapshot"] = deepcopy(req["raw_api_call"]["payload"])
    req["tag"].update(patch["tag_patch"])
    for field in (
        "destination_request",
        "trigger_evaluation",
        "tag_sequence",
        "consent",
        "business_rule_results",
        "sensitive_data_scan",
        "client_checks",
        "regression",
    ):
        req[field] = deepcopy(patch[field])
    req["verdict"].update(patch["verdict_patch"])
    new_ids = [item["evidence_id"] for item in extension["evidence"]]
    req["evidence_ids"].extend(new_ids)
    data["evidence"].extend(deepcopy(extension["evidence"]))
    return data


def execution_fixture(data: dict | None = None) -> dict:
    data = data or fixture()
    req = requirement(data)
    boundary = req["action_boundary"]
    expectation = req.get("expectation", {})
    tag = req.get("tag") or {}
    destination = req.get("destination_request") or {}
    tag_name = tag.get("name") or expectation.get("tag_name") or "Synthetic concerned tag"
    tag_id = "TAG-FIXTURE-001"
    tag_container = (
        tag.get("container_id") or req.get("container_id") or data["run"]["container_id"]
    )
    vendor_family = tag.get("vendor_family") or expectation.get("vendor_family")
    tag_category = "analytics" if vendor_family in {None, "", "ga4"} else "media"
    tag_delivery = expectation.get("tag_delivery", "browser_request")
    configuration_evidence_id = tag.get("configuration_evidence_id", "EVD-TAG-CONFIG-011")
    runtime_evidence_id = tag.get("runtime_evidence_id", "EVD-TAG-RUNTIME-011")
    destination_evidence_id = destination.get("evidence_id", "EVD-NET-011")
    if destination.get("request_count", 0) == 0 and tag_delivery == "browser_request":
        destination_evidence_id = "EVD-NET-CAPTURE-001"
        if not any(
            row.get("evidence_id") == destination_evidence_id for row in data.get("evidence", [])
        ):
            data.setdefault("evidence", []).append(
                {
                    "evidence_id": destination_evidence_id,
                    "kind": "browser_network_capture",
                    "source": "Browser Network",
                    "capture_mode": "direct",
                    "action_id": "ACT-001",
                    "container_id": tag_container,
                    "path_or_url": "evidence/network-capture-001.json",
                    "captured_at": "2026-07-25T10:01:03+00:00",
                    "description": "Complete browser network capture for the action window.",
                }
            )
    consent_evidence_id = (req.get("consent") or {}).get("evidence_id", "EVD-CONSENT-BASELINE-001")
    sensitive_evidence_id = (req.get("sensitive_data_scan") or {}).get("evidence_id", "EVD-PII-011")
    tag_inventory = [
        {
            "tag_id": tag_id,
            "tag_name": tag_name,
            "container_id": tag_container,
            "tag_category": tag_category,
            "tag_delivery": tag_delivery,
            "vendor_family": vendor_family,
            "destination_id": tag.get("destination_id") or expectation.get("destination_id"),
            "template_type": tag.get("template_type", "GA4 Event"),
            "consent_required": bool(expectation.get("consent_contract")),
            "evidence_ids": [configuration_evidence_id],
        }
    ]
    tag_inventory[0]["scope_status"], tag_inventory[0]["scope_reason"] = tag_scope_decision(
        tag_inventory[0],
        data["run"]["tag_scope"],
        declared_tag_contracts(data["requirements"]),
    )
    applicability_card = layer_applicability(
        data["requirements"],
        container_count=len(data["run"]["containers"]),
        tag_inventory=tag_inventory,
    )
    applicable = [row["layer"] for row in applicability_card if row["mode"] == "MANDATORY"]
    evidence_for_layer = {
        "action_boundary": ["EVD-ACTION-001"],
        "raw_api_call": ["EVD-RAW-011"],
        "resolved_data_layer": ["EVD-DL-011"],
        "concerned_tag_inventory": [configuration_evidence_id],
        "gtm_variable": ["EVD-VAR-011"],
        "tag_configuration": [configuration_evidence_id],
        "tag_firing": [runtime_evidence_id],
        "tag_parameter": [runtime_evidence_id],
        "destination_request_when_applicable": [destination_evidence_id],
        "sensitive_data_scan": [sensitive_evidence_id],
        "business_rules_when_declared": ["EVD-RAW-011"],
        "consent_when_applicable": [consent_evidence_id],
        "source_signal_when_no_data_layer_push": [
            (req.get("source_signal") or {}).get("evidence_id", "EVD-ACTION-001")
        ],
        "trigger_logic_when_applicable": [
            (req.get("trigger_evaluation") or {}).get("evidence_id", runtime_evidence_id)
        ],
        "tag_sequence_when_applicable": [
            (req.get("tag_sequence") or {}).get("evidence_id", configuration_evidence_id)
        ],
        "client_checks_when_applicable": [
            next(
                (
                    row.get("evidence_id")
                    for row in req.get("client_checks", [])
                    if isinstance(row, dict) and row.get("evidence_id")
                ),
                "EVD-ACTION-001",
            )
        ],
        "regression_when_baseline_provided": [
            (req.get("regression") or {}).get("evidence_id", "EVD-ACTION-001")
        ],
        "container_context_when_applicable": [configuration_evidence_id],
        "conditional_scenarios_when_applicable": [
            (req.get("scenario") or {}).get("evidence_id", "EVD-ACTION-001")
        ],
    }
    layer_results = [
        {
            "layer": decision["layer"],
            "status": "PASS" if decision["mode"] == "MANDATORY" else "NOT_APPLICABLE",
            "reason": (
                f"{decision['layer']} matched the exact accepted evidence."
                if decision["mode"] == "MANDATORY"
                else f"Predicate false: {decision['predicate']}"
            ),
            "evidence_ids": evidence_for_layer.get(decision["layer"], ["EVD-ACTION-001"]),
            "semantic_ambiguity": None,
            "blocker_id": None,
            "predicate_result": None if decision["mode"] == "MANDATORY" else False,
            "recorded_at": "2026-07-25T10:01:04+00:00",
        }
        for decision in applicability_card
    ]
    resolved_value = (req.get("resolved_data_layer") or {}).get("field_value")
    variable_value = (req.get("gtm_variable") or {}).get("field_value")
    runtime_value = tag.get("runtime_value")
    configured_expected = expectation.get("expected_tag_configuration")
    configured_actual = tag.get("configured_value")
    expects_request = expectation.get("expected_request_behavior") not in {"absent", "blocked"}
    request_observed = destination.get("request_count", 0) > 0
    value_anchor = (
        {
            "source": "raw_data_layer_mapping",
            "requirement_id": req["requirement_id"],
            "path": "resolved_data_layer.field_value",
        }
        if (req.get("resolved_data_layer") or {}).get("field_state") != "absent"
        else {
            "source": "explicit_acceptance_rule",
            "requirement_id": req["requirement_id"],
            "path": "expectation.expected_firing",
            "transform": "absence_null",
        }
    )
    runtime_anchor = (
        {
            "source": "resolved_gtm_variable_contract",
            "requirement_id": req["requirement_id"],
            "path": "gtm_variable.field_value",
        }
        if (req.get("gtm_variable") or {}).get("field_state") != "absent"
        else {
            "source": "explicit_acceptance_rule",
            "requirement_id": req["requirement_id"],
            "path": "expectation.expected_firing",
            "transform": "absence_null",
        }
    )
    tag_evidence = {
        "gtm_variable": (
            ["EVD-VAR-011"],
            {
                "variables": [
                    {
                        "name": (req.get("gtm_variable") or {}).get("name") or "resolved input",
                        "expected_anchor": value_anchor,
                        "expected_value": resolved_value,
                        "expected_type": value_type(resolved_value),
                        "actual_value": variable_value,
                        "actual_type": value_type(variable_value),
                        "status": "PASS",
                    }
                ]
            },
        ),
        "tag_configuration": (
            [configuration_evidence_id],
            {
                "configuration": [
                    {
                        "name": tag.get("configuration_field") or "tag configuration",
                        "expected_anchor": {
                            "source": "tracking_plan",
                            "requirement_id": req["requirement_id"],
                            "path": "expectation.expected_tag_configuration",
                        },
                        "expected_value": configured_expected,
                        "expected_type": value_type(configured_expected),
                        "actual_value": configured_actual,
                        "actual_type": value_type(configured_actual),
                        "status": "PASS",
                    }
                ],
            },
        ),
        "tag_firing": (
            [runtime_evidence_id],
            {
                "expected_firing": tag.get("expected_firing"),
                "expected_firing_anchor": {
                    "source": "tracking_plan",
                    "requirement_id": req["requirement_id"],
                    "path": "expectation.expected_firing",
                },
                "actual_firing": tag.get("actual_firing"),
                "fire_count": tag.get("fire_count", 0),
            },
        ),
        "tag_parameter": (
            [runtime_evidence_id],
            {
                "parameters": [
                    {
                        "name": tag.get("configuration_field") or "runtime parameter",
                        "expected_anchor": runtime_anchor,
                        "expected_value": variable_value,
                        "expected_type": value_type(variable_value),
                        "actual_value": runtime_value,
                        "actual_type": value_type(runtime_value),
                        "status": "PASS",
                    }
                ],
                "runtime_state": tag.get("runtime_state"),
                "runtime_type": tag.get("runtime_type"),
            },
        ),
        "destination_request_when_applicable": (
            [destination_evidence_id],
            {
                "request_count": destination.get("request_count", 0),
                "request_ids": (
                    [destination.get("request_id")]
                    if request_observed and destination.get("request_id")
                    else []
                ),
                "request_behavior": destination.get("request_behavior"),
                "expected_request_behavior": expectation.get("expected_request_behavior"),
                "expected_request_behavior_anchor": {
                    "source": "tracking_plan",
                    "requirement_id": req["requirement_id"],
                    "path": "expectation.expected_request_behavior",
                },
                "local_only_configuration_proved": tag_delivery == "local_only",
                "parameters": [
                    {
                        "name": "matching request occurrence",
                        "expected_anchor": {
                            "source": "explicit_acceptance_rule",
                            "requirement_id": req["requirement_id"],
                            "path": "expectation.expected_request_behavior",
                            "transform": "request_expected",
                        },
                        "expected_value": expects_request,
                        "expected_type": "boolean",
                        "actual_value": request_observed,
                        "actual_type": "boolean",
                        "status": "PASS",
                    }
                ],
            },
        ),
        "consent_when_applicable": (
            [consent_evidence_id],
            {"predicate_reason": "No consent condition is configured for this tag."},
        ),
        "trigger_logic_when_applicable": (
            [runtime_evidence_id],
            {"predicate_reason": "Expected firing and count matched."},
        ),
        "tag_sequence_when_applicable": (
            [configuration_evidence_id],
            {"predicate_reason": "No GTM sequence is configured."},
        ),
    }
    tag_layer_results = []
    mandatory_tag_layers = {
        row["layer"] for row in applicability_card if row["mode"] == "MANDATORY"
    }
    for layer in TAG_RESULT_LAYERS:
        evidence_ids, details = tag_evidence[layer]
        active = layer in mandatory_tag_layers
        tag_layer_results.append(
            {
                "action_id": "ACT-001",
                "tag_id": tag_id,
                "tag_name": tag_name,
                "container_id": tag_container,
                "tag_category": tag_category,
                "tag_delivery": tag_delivery,
                "layer": layer,
                "status": "PASS" if active else "NOT_APPLICABLE",
                "reason": (
                    "Exact per-tag evidence matched."
                    if active
                    else "The event-level conditional predicate is false."
                ),
                "details": details,
                "evidence_ids": evidence_ids,
                "semantic_ambiguity": None,
                "blocker_id": None,
                "recorded_at": "2026-07-25T10:01:04+00:00",
            }
        )
    before_runtime_evidence_ids = ["EVD-READY-ACTION-001", "EVD-READY-NETWORK-001"]
    after_runtime_evidence_ids = ["EVD-SETTLE-ACTION-001", "EVD-SETTLE-NETWORK-001"]
    runtime_evidence = [
        {
            "evidence_id": "EVD-READY-ACTION-001",
            "kind": "action_boundary",
            "source": "Playwright",
            "capture_mode": "direct",
            "action_id": "ACT-001",
            "runtime_check_id": boundary["readiness_check_id"],
            "runtime_phase": "before_action",
            "path_or_url": "evidence/ready-action-001.json",
            "captured_at": "2026-07-25T10:00:59+00:00",
            "description": "Direct readiness boundary for this exact action.",
        },
        {
            "evidence_id": "EVD-READY-NETWORK-001",
            "kind": "browser_network_capture",
            "source": "Browser Network",
            "capture_mode": "direct",
            "action_id": "ACT-001",
            "runtime_check_id": boundary["readiness_check_id"],
            "runtime_phase": "before_action",
            "container_id": tag_container,
            "path_or_url": "evidence/ready-network-001.json",
            "captured_at": "2026-07-25T10:00:59+00:00",
            "description": "Active network cursor at the readiness boundary.",
        },
        {
            "evidence_id": "EVD-SETTLE-ACTION-001",
            "kind": "action_boundary",
            "source": "Playwright",
            "capture_mode": "direct",
            "action_id": "ACT-001",
            "runtime_check_id": boundary["settlement_check_id"],
            "runtime_phase": "after_action",
            "path_or_url": "evidence/settle-action-001.json",
            "captured_at": "2026-07-25T10:01:03+00:00",
            "description": "Direct settlement boundary for this exact action.",
        },
        {
            "evidence_id": "EVD-SETTLE-NETWORK-001",
            "kind": "browser_network_capture",
            "source": "Browser Network",
            "capture_mode": "direct",
            "action_id": "ACT-001",
            "runtime_check_id": boundary["settlement_check_id"],
            "runtime_phase": "after_action",
            "container_id": tag_container,
            "path_or_url": "evidence/settle-network-001.json",
            "captured_at": "2026-07-25T10:01:03+00:00",
            "description": "Complete request cursor capture for the settled action window.",
        },
    ]
    known_evidence_ids = {str(row.get("evidence_id", "")) for row in data.get("evidence", [])}
    data.setdefault("evidence", []).extend(
        row for row in runtime_evidence if row["evidence_id"] not in known_evidence_ids
    )
    return {
        "schema_version": 3,
        "operator_contract_version": 1,
        "created_at": "2026-07-25T09:55:00+00:00",
        "updated_at": "2026-07-25T10:02:00+00:00",
        "profile_path": "profiles/run-synthetic-001",
        "connection_epoch": 1,
        "approved_origins": ["https://shop.example.test"],
        "surfaces": {
            "gtm": {
                "role": "gtm_workspace",
                "url": "https://tagmanager.google.com/",
                "container_id": tag_container,
                "workspace": data["run"]["workspace"],
            },
            "preview": {
                "role": "tag_assistant",
                "url": "https://tagassistant.google.com/",
                "connected": True,
            },
            "site": {
                "role": "website",
                "url": "https://shop.example.test/product",
            },
        },
        "runtime_checks": [
            {
                "check_id": boundary["readiness_check_id"],
                "phase": "before_action",
                "action_id": "ACT-001",
                "case_id": "CASE-001",
                "event_group_id": "EVG-001",
                "captured_at": "2026-07-25T10:00:59+00:00",
                "recorded_at": "2026-07-25T10:00:59+00:00",
                "capture_source": "playwright_runtime_probe",
                "browser_context_id": "desktop-default",
                "connection_epoch": 1,
                "gtm_workspace_surface_id": "gtm",
                "tag_assistant_surface_id": "preview",
                "website_surface_id": "site",
                "containers": [
                    {
                        "container_id": tag_container,
                        "workspace": data["run"]["workspace"],
                    }
                ],
                "website_url": "https://shop.example.test/product",
                "selected_page_url": "https://shop.example.test/product",
                "preview_connected": True,
                "target_interactive": True,
                "target_uncovered": True,
                "lifecycle_observed": True,
                "stream_quiet": True,
                "network_capture_active": True,
                "preview_event_cursor": boundary["last_event_before"],
                "network_request_cursor": boundary["network_request_cursor_before"],
                "evidence_ids": before_runtime_evidence_ids,
                "consumed": True,
                "consumed_by_action_id": "ACT-001",
            },
            {
                "check_id": boundary["settlement_check_id"],
                "phase": "after_action",
                "action_id": "ACT-001",
                "case_id": "CASE-001",
                "event_group_id": "EVG-001",
                "captured_at": "2026-07-25T10:01:03+00:00",
                "recorded_at": "2026-07-25T10:01:03+00:00",
                "capture_source": "playwright_runtime_probe",
                "browser_context_id": "desktop-default",
                "connection_epoch": 1,
                "gtm_workspace_surface_id": "gtm",
                "tag_assistant_surface_id": "preview",
                "website_surface_id": "site",
                "containers": [
                    {
                        "container_id": tag_container,
                        "workspace": data["run"]["workspace"],
                    }
                ],
                "website_url": "https://shop.example.test/product",
                "selected_page_url": "https://shop.example.test/product",
                "preview_connected": True,
                "target_interactive": True,
                "target_uncovered": True,
                "lifecycle_observed": True,
                "stream_quiet": True,
                "network_capture_active": True,
                "preview_event_cursor": boundary["settled_final_event"],
                "network_request_cursor": boundary["network_request_cursor_after"],
                "first_event_after": boundary["first_event_after"],
                "observed_business_push_count": 1,
                "evidence_ids": after_runtime_evidence_ids,
                "consumed": True,
                "consumed_by_action_id": "ACT-001",
            },
        ],
        "event_closures": [
            {
                "event_group_id": "EVG-001",
                "plan_order": 1,
                "case_ids": ["CASE-001"],
                "final_action_ids": ["ACT-001"],
                "closed_at": "2026-07-25T10:02:00+00:00",
                "feedback_emitted_at": "2026-07-25T10:02:00+00:00",
            }
        ],
        "closure_history": [],
        "operator_state": {
            "status": "FINISHED",
            "current_event_group_id": None,
        },
        "authorizations": [],
        "cases": [
            {
                "case_id": "CASE-001",
                "event_group_id": "EVG-001",
                "requirement_ids": ["REQ-001"],
                "url": "https://shop.example.test/product",
                "element": "Add to cart",
                "placement": "product detail",
                "action": "click",
                "material_variant": {"quantity": 1},
                "discovered_from": "tracking_plan",
                "scope_status": "IN_SCOPE",
                "execution_status": "EXECUTED",
                "reason": None,
                "authorization_ids": [],
                "tag_scope": deepcopy(data["run"]["tag_scope"]),
                "declared_tag_contracts": declared_tag_contracts(data["requirements"]),
                "source_expectations": [deepcopy(req["expectation"])],
                "tag_inventory_status": "COMPLETE",
                "tag_inventory_reason": "Container and Preview inventory completed before action.",
                "tag_inventory_evidence_ids": [configuration_evidence_id],
                "tag_inventory": deepcopy(tag_inventory),
                "conditional_activations": {},
                "applicability_status": "FROZEN",
                "layer_applicability": deepcopy(applicability_card),
                "applicability_frozen_at": "2026-07-25T09:59:00+00:00",
                "applicable_layers": applicable,
                "container_ids": [tag_container],
                "registered_at": "2026-07-25T09:58:00+00:00",
                "final_action_id": "ACT-001",
            }
        ],
        "actions": [
            {
                "action_id": "ACT-001",
                "case_id": "CASE-001",
                "event_group_id": "EVG-001",
                "requirement_ids": ["REQ-001"],
                "url": "https://shop.example.test/product",
                "element": "Add to cart",
                "placement": "product detail",
                "material_variant": {"quantity": 1},
                "action": "click",
                "attempt_number": 1,
                "inventory_revision": 1,
                "connection_epoch": 1,
                "retry_of_action_id": None,
                "readiness_check_id": boundary["readiness_check_id"],
                "readiness_evidence_ids": before_runtime_evidence_ids,
                "preview_connected_before": True,
                "target_ready_before": True,
                "last_event_before": boundary["last_event_before"],
                "network_request_cursor_before": boundary["network_request_cursor_before"],
                "consent_state_before": "analytics_storage=granted",
                "browser_context_id": "desktop-default",
                "container_ids": [tag_container],
                "observed_url_before": "https://shop.example.test/product",
                "selected_page_url_before": "https://shop.example.test/product",
                "action_timestamp": boundary["action_timestamp"],
                "quiet_window_ms": boundary["quiet_window_ms"],
                "timeout_ms": boundary["timeout_ms"],
                "layer_results": layer_results,
                "tag_layer_results": tag_layer_results,
                "first_event_after": boundary["first_event_after"],
                "settled_final_event": boundary["settled_final_event"],
                "settlement_check_id": boundary["settlement_check_id"],
                "settlement_evidence_ids": after_runtime_evidence_ids,
                "network_request_cursor_after": boundary["network_request_cursor_after"],
                "expected_seen": True,
                "preview_connected_after": True,
                "interaction_outcome": boundary["interaction_outcome"],
                "completion_signal": boundary["completion_signal"],
                "stream_settled": boundary["stream_settled"],
                "settlement_reason": boundary["settlement_reason"],
                "observed_business_push_count": 1,
                "settled_at": "2026-07-25T10:01:03+00:00",
                "state": "SETTLED",
            }
        ],
        "business_pushes": [
            {
                "push_id": "PUSH-011",
                "stream_id": "tag_assistant",
                "action_id": "ACT-001",
                "case_id": "CASE-001",
                "event_group_id": "EVG-001",
                "event_name": "add_to_cart",
                "event_index": 11,
                "captured_at": "2026-07-25T10:01:00+00:00",
                "url": "https://shop.example.test/product",
                "page_state": "Basket count is 1",
                "classification": "expected",
                "classification_reason": "Expected once after the completed product CTA.",
                "evidence_id": "EVD-RAW-011",
                "container_id": "GTM-TEST",
            }
        ],
        "checkpoints": [],
    }


def configure_absent_event(data: dict, blocker_status: str | None = None) -> dict:
    req = requirement(data)
    req["event_observed"] = False
    req["occurrence_evidence"] = {
        "actual_count": 0,
        "event_indexes": [],
        "evidence_id": "EVD-ACTION-001",
    }
    req["raw_api_call"] = None
    req["resolved_data_layer"] = None
    req["gtm_variable"] = {
        "applicable": True,
        "name": "DLV - ecommerce.value",
    }
    req["tag"] = {
        "applicable": True,
        "relevance": "explains_non_firing",
        "container_id": "GTM-TEST",
        "vendor_family": "ga4",
        "destination_id": "G-TEST123",
        "event_name": "add_to_cart",
        "template_type": "GA4 Event",
        "name": "GA4 - Event - add_to_cart",
        "expected_firing": "fired_once",
        "actual_firing": "not_evaluated",
        "fire_count": 0,
        "configuration_field": "eventParameters.value",
        "configured_value": "{{DLV - ecommerce.value}}",
        "configuration_evidence_id": "EVD-TAG-CONFIG-011",
        "non_firing_reason": "Expected event did not occur.",
        "reason_source": "preview",
    }
    req["destination_request"] = {
        "applicable": True,
        "vendor_family": "ga4",
        "destination_id": "G-TEST123",
        "event_name": "add_to_cart",
        "request_behavior": "not_observed",
        "request_count": 0,
        "capture_source": "browser_network",
        "request_id": "NET-011",
        "parameter_path": 'query["ep.value"]',
        "field_state": "absent",
        "field_type": "absent",
        "evidence_id": "EVD-NET-011",
    }
    component = blocker_status or "BLOCKED"
    req["verdict"].update(
        {
            "event_occurrence": "FAIL" if blocker_status is None else blocker_status,
            "raw_payload": "FAIL" if blocker_status is None else blocker_status,
            "resolved_data_layer": component,
            "gtm_variable": component,
            "tag_configuration": component,
            "tag_firing": component,
            "tag_parameter": component,
            "destination_request": component,
            "destination_parameter": component,
            "overall": "FAIL" if blocker_status is None else blocker_status,
            "failure_layer": "event_occurrence",
            "mismatch": "Expected event was not observed.",
        }
    )
    req["action_boundary"]["first_event_after"] = None
    return req


def add_blocker(
    data: dict,
    blocker_type: str,
    *,
    help_requested: bool,
    settled: bool = True,
) -> dict:
    req = configure_absent_event(data, blocker_status="BLOCKED")
    req["journey"]["execution_status"] = "BLOCKED"
    req["blocker_id"] = "BLK-001"
    req["action_boundary"]["stream_settled"] = settled
    if not settled:
        req["action_boundary"]["settlement_reason"] = (
            "preview_disconnected" if blocker_type == "PREVIEW_DISCONNECTED" else "timeout"
        )
    data["blockers"] = [
        {
            "blocker_id": "BLK-001",
            "type": blocker_type,
            "checkpoint": "Protected or external checkpoint",
            "description": "Synthetic blocker",
            "requirement_ids": ["REQ-001"],
            "analyst_intervention_required": blocker_type
            in {
                "GOOGLE_SIGN_IN",
                "MFA",
                "CAPTCHA",
                "EMAIL_VERIFICATION",
                "SMS_VERIFICATION",
                "MAGIC_LINK",
                "REAL_PAYMENT",
                "EXTERNAL_APPROVAL",
                "IRREVERSIBLE_ACTION",
            },
            "analyst_help_requested": help_requested,
            "analyst_response": "Unable to complete" if help_requested else "",
            "outcome": "Blocked",
            "status": "BLOCKED",
            "evidence_ids": ["EVD-ACTION-001"],
            "notes": "",
        }
    ]
    return req


def add_consent_override(
    data: dict,
    *,
    approved: bool,
    production: bool = False,
    production_approved: bool = False,
) -> dict:
    req = requirement(data)
    data["run"]["environment_class"] = "production" if production else "preprod"
    data["blockers"] = [
        {
            "blocker_id": "BLK-CMP",
            "type": ("CMP_PRODUCTION_ENVIRONMENT" if production else "CMP_TEST_ENVIRONMENT"),
            "checkpoint": "CMP initialization",
            "description": "CMP did not initialize in preprod.",
            "requirement_ids": ["REQ-001"],
            "analyst_intervention_required": False,
            "analyst_help_requested": False,
            "analyst_response": "",
            "outcome": "Downstream test state proposed",
            "status": "BLOCKED",
            "evidence_ids": ["EVD-ACTION-001"],
            "notes": "",
        }
    ]
    req["expectation"]["expected_consent_state"] = "analytics_storage=granted"
    req["consent"] = {
        "applicable": True,
        "scenario_id": "CONSENT-OVERRIDE",
        "scenario": "Approved test-environment override",
        "source": "session_override",
        "state_at_event": {"analytics_storage": "granted"},
        "before_state": {"analytics_storage": "denied"},
        "override_approved": approved,
        "approval_evidence_id": "EVD-CMP-APPROVAL-001",
        "override_method": "Session-scoped gtag consent update",
        "override_scope": "session_only",
        "native_cmp_status": "FAIL",
        "native_cmp_acceptance_in_scope": False,
        "blocker_id": "BLK-CMP",
        "evidence_id": "EVD-CONSENT-001",
    }
    if production:
        req["consent"].update(
            {
                "production_exception_approved": production_approved,
                "production_approval_evidence_id": "EVD-CMP-PROD-APPROVAL-001",
                "restoration_confirmed": production_approved,
            }
        )
    req["evidence_ids"].append("EVD-CONSENT-001")
    req["evidence_ids"].append("EVD-CMP-APPROVAL-001")
    data["evidence"].append(
        {
            "evidence_id": "EVD-CONSENT-001",
            "kind": "consent_state",
            "source": "Tag Assistant",
            "capture_mode": "direct",
            "action_id": "ACT-001",
            "event_index": 11,
            "container_id": "GTM-TEST",
            "path_or_url": "evidence/consent-001.json",
            "captured_at": "2026-07-25T10:01:03+00:00",
            "description": "Event-level consent after the approved session override.",
        }
    )
    data["evidence"].append(
        {
            "evidence_id": "EVD-CMP-APPROVAL-001",
            "kind": "analyst_approval",
            "source": "Analyst supplied",
            "capture_mode": "analyst_supplied",
            "path_or_url": "evidence/cmp-approval-001.json",
            "captured_at": "2026-07-25T10:01:03+00:00",
            "description": "Explicit analyst decision for the proposed CMP override.",
        }
    )
    if production:
        req["evidence_ids"].append("EVD-CMP-PROD-APPROVAL-001")
        data["evidence"].append(
            {
                "evidence_id": "EVD-CMP-PROD-APPROVAL-001",
                "kind": "analyst_approval",
                "source": "Analyst supplied",
                "capture_mode": "analyst_supplied",
                "path_or_url": "evidence/cmp-prod-approval-001.json",
                "captured_at": "2026-07-25T10:01:03+00:00",
                "description": "Explicit production exception approval.",
            }
        )
    req["verdict"]["consent"] = "PASS"
    req["verdict"]["overall"] = "PASS"
    return req


class PipelineTests(unittest.TestCase):
    def assert_invalid(self, data: dict, message: str) -> None:
        with self.assertRaises(ReportValidationError) as context:
            validate(data, strict=True)
        self.assertIn(message, str(context.exception))

    def test_valid_full_schema_and_workbook(self) -> None:
        data = fixture()
        self.assertEqual([], validate(data, strict=True))
        data["evidence"][0]["path_or_url"] = "https://example.test/evidence/action"
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "recette.xlsx"
            build_workbook(data, output)
            workbook = load_workbook(output, read_only=False)
            self.assertEqual(REQUIRED_SHEETS, workbook.sheetnames)
            self.assertEqual("PASS", workbook["Client Summary"]["B3"].value)
            self.assertEqual("Output contract", workbook["Client Summary"]["A4"].value)
            self.assertEqual(2, workbook["Client Summary"]["B4"].value)
            self.assertEqual(2, workbook["Requirement Matrix"].max_row)
            event_sheet = workbook["Event Evidence"]
            event_headers = {cell.value: cell.column for cell in event_sheet[1]}
            self.assertEqual(
                "ACT-001",
                event_sheet.cell(row=2, column=event_headers["action_id"]).value,
            )
            self.assertEqual(
                "completed",
                event_sheet.cell(
                    row=2,
                    column=event_headers["interaction_outcome"],
                ).value,
            )
            self.assertEqual(
                "expected_and_quiet",
                event_sheet.cell(
                    row=2,
                    column=event_headers["settlement_reason"],
                ).value,
            )
            evidence_sheet = workbook["Evidence Catalogue"]
            evidence_headers = {cell.value: cell.column for cell in evidence_sheet[1]}
            self.assertIsNotNone(
                evidence_sheet.cell(
                    row=2,
                    column=evidence_headers["path_or_url"],
                ).hyperlink
            )
            workbook.close()

    def test_data_layer_applicability_implies_the_full_tag_chain(self) -> None:
        data = fixture("valid_limited_layers.json")
        layers = applicable_layers(data["requirements"])
        for layer in (
            "resolved_data_layer",
            "concerned_tag_inventory",
            "gtm_variable",
            "tag_configuration",
            "tag_firing",
            "tag_parameter",
            "destination_request_when_applicable",
        ):
            self.assertIn(layer, layers)
        with self.assertRaises(ReportValidationError):
            validate(data, strict=True)
        self.assertIsNone(requirement(data).get("tag"))

    def test_valid_full_client_side_extension_and_workbook(self) -> None:
        data = client_side_fixture()
        self.assertEqual([], validate(data, strict=True))
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "client-side-recette.xlsx"
            build_workbook(data, output)
            workbook = load_workbook(output, read_only=False)
            self.assertEqual(REQUIRED_SHEETS, workbook.sheetnames)
            self.assertEqual(2, workbook["Destination Evidence"].max_row)
            self.assertEqual(8, workbook["Business Rules"].max_row)
            self.assertEqual(15, workbook["Client Checks"].max_row)
            self.assertEqual(3, workbook["Container Context"].max_row)
            workbook.close()

    def test_declared_client_layer_cannot_be_omitted_from_run_metadata(self) -> None:
        data = client_side_fixture()
        data["run"]["included_layers"].remove("destination_request_when_applicable")
        self.assert_invalid(data, "included_layers omits declared client-side layers")

    def test_destination_parameter_mismatch_cannot_pass(self) -> None:
        data = client_side_fixture()
        requirement(data)["destination_request"]["field_value"] = 30
        self.assert_invalid(data, "PASS destination parameter contradicts")

    def test_destination_event_name_mismatch_cannot_pass(self) -> None:
        data = client_side_fixture()
        requirement(data)["destination_request"]["event_name"] = "Purchase"
        self.assert_invalid(data, "destination event_name differs from expectation")

    def test_destination_claim_must_match_browser_request(self) -> None:
        data = client_side_fixture()
        requirement(data)["destination_request"]["request_url"] = (
            "https://www.facebook.com/tr/?id=META-WRONG&ev=AddToCart&value=29.9"
        )
        self.assert_invalid(data, "decoded destination ID differs from browser request")

    def test_destination_value_must_match_browser_request(self) -> None:
        data = client_side_fixture()
        requirement(data)["destination_request"]["request_url"] = (
            "https://www.facebook.com/tr/?id=META-TEST-001&ev=AddToCart&value=999"
        )
        self.assert_invalid(data, "decoded destination parameter differs from browser request")

    def test_literal_vendor_request_keys_are_addressable(self) -> None:
        self.assertEqual(
            ["29.9", "EUR"],
            [
                path_value(
                    {"query": {"cd[value]": "29.9", "ep.currency": "EUR"}},
                    'query["cd[value]"]',
                ),
                path_value(
                    {"query": {"cd[value]": "29.9", "ep.currency": "EUR"}},
                    'query["ep.currency"]',
                ),
            ],
        )
        data = client_side_fixture()
        req = requirement(data)
        req["expectation"]["destination_parameter_path"] = 'query["cd[value]"]'
        req["destination_request"]["parameter_path"] = 'query["cd[value]"]'
        req["destination_request"]["request_url"] = (
            "https://www.facebook.com/tr/?id=META-TEST-001&ev=AddToCart&cd%5Bvalue%5D=29.9"
        )
        self.assertEqual([], validate(data, strict=True))

    def test_destination_verdict_cannot_be_omitted(self) -> None:
        data = client_side_fixture()
        requirement(data)["verdict"].pop("destination_request")
        self.assert_invalid(data, "destination expectation requires destination_request verdict")

    def test_vendor_helper_alone_cannot_prove_browser_send(self) -> None:
        data = client_side_fixture()
        requirement(data)["destination_request"]["capture_source"] = "vendor_helper"
        self.assert_invalid(data, "first-party browser-network evidence")

    def test_advanced_consent_v2_requires_all_four_signals(self) -> None:
        data = client_side_fixture()
        del requirement(data)["expectation"]["consent_contract"]["signals"]["ad_user_data"]
        self.assert_invalid(data, "must declare all four consent signals")

    def test_conditional_pass_requires_condition_evidence(self) -> None:
        data = client_side_fixture()
        requirement(data)["scenario"]["condition_met"] = False
        self.assert_invalid(data, "conditional PASS requires evidence")

    def test_trigger_logic_false_pass_is_rejected(self) -> None:
        data = client_side_fixture()
        requirement(data)["trigger_evaluation"]["actual_result"] = "blocked"
        self.assert_invalid(data, "PASS trigger result differs")

    def test_trigger_condition_truth_is_recomputed(self) -> None:
        data = client_side_fixture()
        requirement(data)["trigger_evaluation"]["conditions"][0]["actual"] = "wrong"
        self.assert_invalid(data, "matched differs from its expected/actual values")

    def test_matched_blocking_exception_cannot_hide_behind_pass(self) -> None:
        data = client_side_fixture()
        requirement(data)["trigger_evaluation"]["blocking_exceptions"][0]["matched"] = True
        self.assert_invalid(data, "trigger actual_result differs from condition/exception evidence")

    def test_tag_sequence_false_pass_is_rejected(self) -> None:
        data = client_side_fixture()
        requirement(data)["tag_sequence"]["actual_order"] = [
            "Meta - AddToCart",
            "Media - Setup",
            "Media - Cleanup",
        ]
        self.assert_invalid(data, "PASS tag sequence contradicts")

    def test_extra_sequence_step_requires_explicit_allowance(self) -> None:
        data = client_side_fixture()
        requirement(data)["tag_sequence"]["actual_order"].append("Unexpected - Cleanup")
        self.assert_invalid(data, "PASS tag sequence contradicts")

    def test_cross_field_business_rule_false_pass_is_rejected(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["raw_api_call"]["payload"]["ecommerce"]["items"][0]["price"] = 10
        req["resolved_data_layer"]["snapshot"] = deepcopy(req["raw_api_call"]["payload"])
        self.assert_invalid(data, "business rule result BR-VALUE contradicts")

    def test_business_rule_verdict_cannot_be_omitted(self) -> None:
        data = client_side_fixture()
        requirement(data)["verdict"].pop("business_rule")
        self.assert_invalid(data, "declared business_rules require business_rule verdict")

    def test_business_rule_equality_is_type_strict(self) -> None:
        result = evaluate_business_rule(
            {
                "rule_id": "BR-TYPE-STRICT",
                "operator": "equals_path",
                "left_path": "left",
                "right_path": "right",
            },
            {"left": True, "right": 1},
        )
        self.assertEqual("FAIL", result["status"])

    def test_all_items_equal_rejects_non_object_items(self) -> None:
        result = evaluate_business_rule(
            {
                "rule_id": "BR-ITEMS",
                "operator": "all_items_equal",
                "items_path": "items",
                "item_field": "currency",
                "expected_path": "currency",
            },
            {"items": [{"currency": "EUR"}, "EUR"], "currency": "EUR"},
        )
        self.assertEqual("FAIL", result["status"])

    def test_zero_business_tolerance_has_no_relative_slack(self) -> None:
        result = evaluate_business_rule(
            {
                "rule_id": "BR-EXACT",
                "operator": "sum_product_equals",
                "target_path": "value",
                "items_path": "items",
                "price_field": "price",
                "quantity_field": "quantity",
                "tolerance": 0,
            },
            {
                "value": 1_000_000_000_001,
                "items": [{"price": 1_000_000_000_000, "quantity": 1}],
            },
        )
        self.assertEqual("FAIL", result["status"])

    def test_business_rule_output_redacts_sensitive_primitives(self) -> None:
        email = "synthetic.user@example.com"
        result = evaluate_business_rule(
            {
                "rule_id": "BR-EMAIL",
                "operator": "equals_path",
                "left_path": "left",
                "right_path": "right",
            },
            {"left": email, "right": email},
        )
        self.assertEqual("PASS", result["status"])
        self.assertNotIn(email, json.dumps(result))

    def test_business_rule_output_redacts_keyed_names_and_phones(self) -> None:
        for path, value, payload in (
            (
                "profile.phone",
                "+33 6 12 34 56 78",
                {
                    "profile": {
                        "phone": "+33 6 12 34 56 78",
                        "first_name": "Synthetic Alice",
                    }
                },
            ),
            (
                "profile.first_name",
                "Synthetic Alice",
                {
                    "profile": {
                        "phone": "+33 6 12 34 56 78",
                        "first_name": "Synthetic Alice",
                    }
                },
            ),
            ("left", "Synthetic Alice", {"left": "Synthetic Alice"}),
        ):
            with self.subTest(path=path):
                result = evaluate_business_rule(
                    {
                        "rule_id": "BR-SENSITIVE",
                        "operator": "equals_path",
                        "left_path": path,
                        "right_path": path,
                    },
                    payload,
                )
                self.assertEqual("PASS", result["status"])
                self.assertNotIn(value, json.dumps(result))

    def test_sensitive_data_false_pass_is_rejected_and_output_is_redacted(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["raw_api_call"]["payload"]["contact_email"] = "synthetic.user@example.com"
        req["resolved_data_layer"]["snapshot"] = deepcopy(req["raw_api_call"]["payload"])
        self.assert_invalid(data, "sensitive_data_scan differs from deterministic scan")
        findings = scan_requirement_sensitive_data(req, req["expectation"]["sensitive_data_policy"])
        self.assertTrue(findings)
        self.assertNotIn("synthetic.user@example.com", json.dumps(findings))
        self.assertTrue(all("value_fingerprint" in item for item in findings))

    def test_sensitive_policy_cannot_disappear_from_active_scan_layer(self) -> None:
        data = client_side_fixture()
        requirement(data)["expectation"].pop("sensitive_data_policy")
        requirement(data).pop("sensitive_data_scan")
        requirement(data)["verdict"].pop("sensitive_data")
        self.assert_invalid(data, "sensitive_data_scan layer requires sensitive_data_policy")

    def test_invalid_custom_sensitive_pattern_is_rejected(self) -> None:
        data = client_side_fixture()
        requirement(data)["expectation"]["sensitive_data_policy"]["custom_patterns"] = [
            {
                "pattern_id": "CUSTOM-BAD",
                "pattern": "[",
                "category": "custom",
                "confidence": "confirmed",
            }
        ]
        self.assert_invalid(data, "has invalid regular expression")

    def test_destination_field_value_is_a_sensitive_scan_target(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["destination_request"]["field_value"] = "synthetic.user@example.com"
        findings = scan_requirement_sensitive_data(req, req["expectation"]["sensitive_data_policy"])
        self.assertTrue(any(item["path"] == "destination_request.field_value" for item in findings))

    def test_request_headers_are_sensitive_scan_targets(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["destination_request"]["request_headers"] = {
            "X-Test-Contact": "synthetic.user@example.com"
        }
        findings = scan_requirement_sensitive_data(req, req["expectation"]["sensitive_data_policy"])
        self.assertTrue(
            any(item["path"].startswith("destination_request.request_headers") for item in findings)
        )

    def test_stored_redaction_cannot_retain_raw_sensitive_value(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        email = "synthetic.user@example.com"
        req["raw_api_call"]["payload"]["contact_email"] = email
        req["resolved_data_layer"]["snapshot"] = deepcopy(req["raw_api_call"]["payload"])
        findings = scan_requirement_sensitive_data(req, req["expectation"]["sensitive_data_policy"])
        req["sensitive_data_scan"]["findings"] = findings
        req["sensitive_data_scan"]["status"] = "FAIL"
        req["verdict"]["sensitive_data"] = "FAIL"
        req["verdict"]["overall"] = "FAIL"
        req["sensitive_data_scan"]["findings"][0]["redacted_value"] = email
        self.assert_invalid(data, "sensitive_data_scan differs from deterministic scan")

    def test_sensitive_data_scan_includes_page_title(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["journey"]["page_title"] = "Account synthetic.user@example.com"
        findings = scan_requirement_sensitive_data(req, req["expectation"]["sensitive_data_policy"])
        self.assertTrue(any(item["path"] == "journey.page_title" for item in findings))

    def test_sensitive_data_cli_returns_redacted_failure(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["raw_api_call"]["payload"]["contact_email"] = "synthetic.user@example.com"
        req["resolved_data_layer"]["snapshot"] = deepcopy(req["raw_api_call"]["payload"])
        with tempfile.TemporaryDirectory() as tempdir:
            path = Path(tempdir) / "recette.json"
            path.write_text(json.dumps(data), encoding="utf-8")
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "scan_sensitive_data.py"),
                    str(path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(1, result.returncode)
        self.assertNotIn("synthetic.user@example.com", result.stdout)
        self.assertEqual("FAIL", json.loads(result.stdout)["status"])

    def test_workbook_refuses_unquarantined_sensitive_evidence(self) -> None:
        data = client_side_fixture()
        req = requirement(data)
        req["raw_api_call"]["payload"]["contact_email"] = "synthetic.user@example.com"
        req["resolved_data_layer"]["snapshot"] = deepcopy(req["raw_api_call"]["payload"])
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "unsafe.xlsx"
            with self.assertRaises(ReportValidationError) as context:
                build_workbook(data, output)
            self.assertIn("unsafe sensitive content", str(context.exception))
            self.assertFalse(output.exists())

    def test_client_check_false_pass_is_rejected(self) -> None:
        data = client_side_fixture()
        requirement(data)["client_checks"][6]["actual"]["push_method_intact"] = False
        self.assert_invalid(data, "client check CHECK-DATALAYER status contradicts")

    def test_invalid_client_check_regex_cannot_pass(self) -> None:
        data = client_side_fixture()
        check = requirement(data)["client_checks"][0]
        check.update(
            {
                "comparison": "regex",
                "expected": "[",
                "actual": "history_push",
                "status": "PASS",
            }
        )
        self.assert_invalid(data, "client check CHECK-SPA status contradicts")

    def test_client_check_verdict_cannot_be_omitted(self) -> None:
        data = client_side_fixture()
        requirement(data)["verdict"].pop("client_checks")
        self.assert_invalid(data, "supplied client_checks require client_checks verdict")

    def test_tag_consent_check_is_recomputed(self) -> None:
        data = client_side_fixture()
        requirement(data)["consent"]["tag_consent_checks"][0]["actual"] = "denied"
        self.assert_invalid(data, "status differs from expected/actual values")

    def test_consent_verdict_cannot_be_omitted(self) -> None:
        data = client_side_fixture()
        requirement(data)["verdict"].pop("consent")
        self.assert_invalid(data, "consent_contract requires consent verdict")

    def test_server_container_is_explicitly_out_of_scope(self) -> None:
        data = client_side_fixture()
        data["run"]["containers"][1]["container_type"] = "server"
        self.assert_invalid(data, "server-side GTM is out of scope")

    def test_regression_classification_false_pass_is_rejected(self) -> None:
        data = client_side_fixture()
        requirement(data)["regression"]["change"] = "REGRESSED"
        self.assert_invalid(data, "regression change classification is inconsistent")

    def test_run_baseline_requires_requirement_regression_evidence(self) -> None:
        data = client_side_fixture()
        requirement(data).pop("regression")
        requirement(data)["verdict"].pop("regression")
        self.assert_invalid(data, "run regression_context requires requirement regression evidence")

    def test_evidence_rows_require_provenance_metadata(self) -> None:
        data = client_side_fixture()
        data["evidence"][0].pop("source")
        self.assert_invalid(data, "missing provenance field 'source'")

    def test_evidence_source_must_match_evidence_kind(self) -> None:
        data = client_side_fixture()
        raw_evidence = next(row for row in data["evidence"] if row["evidence_id"] == "EVD-RAW-011")
        raw_evidence["source"] = "Analyst supplied"
        self.assert_invalid(data, "source is incompatible with kind 'api_call'")

    def test_evidence_provenance_cannot_contain_sensitive_content(self) -> None:
        data = client_side_fixture()
        data["evidence"][0]["description"] = "Captured synthetic.user@example.com"
        self.assert_invalid(data, "provenance contains sensitive content")

    def test_nested_evidence_ids_are_bound_to_expected_kinds(self) -> None:
        evidence_ids = (
            "EVD-ACTION-001",
            "EVD-RAW-011",
            "EVD-DL-011",
            "EVD-VAR-011",
            "EVD-META-CONFIG-001",
            "EVD-META-RUNTIME-001",
            "EVD-SCENARIO-001",
            "EVD-TRIGGER-001",
            "EVD-SEQUENCE-001",
            "EVD-BUSINESS-001",
            "EVD-SENSITIVE-001",
            "EVD-CLIENT-CHECKS-001",
            "EVD-REGRESSION-001",
        )
        for evidence_id in evidence_ids:
            with self.subTest(evidence_id=evidence_id):
                data = client_side_fixture()
                evidence = next(
                    row for row in data["evidence"] if row["evidence_id"] == evidence_id
                )
                evidence["kind"] = "screenshot"
                self.assert_invalid(data, "evidence kind must be")

    def test_consent_override_approval_has_dedicated_evidence_kind(self) -> None:
        data = fixture()
        add_consent_override(data, approved=True)
        approval = next(
            row for row in data["evidence"] if row["evidence_id"] == "EVD-CMP-APPROVAL-001"
        )
        approval["kind"] = "screenshot"
        self.assert_invalid(data, "consent.approval: evidence kind must be")

    def test_blocked_retest_is_unverified_not_a_proven_regression(self) -> None:
        baseline = client_side_fixture()
        current = deepcopy(baseline)
        requirement(current)["verdict"]["overall"] = "BLOCKED"
        result = compare_runs(baseline, current)[0]
        self.assertEqual("UNVERIFIED", result["change"])
        self.assertFalse(result["regression"])

    def test_non_datalayer_client_source_is_supported_without_fabricated_push(self) -> None:
        data = fixture()
        req = requirement(data)
        data["run"]["included_layers"].append("source_signal_when_no_data_layer_push")
        req["expectation"]["source_mechanism"] = "direct_vendor_call"
        req["expectation"]["resolved_data_layer_applicable"] = False
        req["expectation"].pop("variable_name")
        req["raw_api_call"] = None
        req["resolved_data_layer"] = None
        req["gtm_variable"] = None
        req["source_signal"] = {
            "mechanism": "direct_vendor_call",
            "event_name": "add_to_cart",
            "capture_source": "browser_console",
            "observed": True,
            "evidence_id": "EVD-SOURCE-001",
        }
        req["occurrence_evidence"]["evidence_id"] = "EVD-SOURCE-001"
        req["verdict"].update(
            {
                "source_signal": "PASS",
                "raw_payload": None,
                "resolved_data_layer": None,
                "gtm_variable": None,
                "tag_configuration": "PASS",
                "tag_parameter": "PASS",
            }
        )
        req["evidence_ids"].append("EVD-SOURCE-001")
        data["evidence"].append(
            {
                "evidence_id": "EVD-SOURCE-001",
                "kind": "direct_vendor_call",
                "source": "Browser Console",
                "capture_mode": "direct",
                "action_id": "ACT-001",
                "event_index": 11,
                "source_detail": "Observed through Playwright console instrumentation",
                "path_or_url": "evidence/source-001.json",
                "captured_at": "2026-07-25T10:01:03+00:00",
                "description": "Direct client-side vendor call with no dataLayer push.",
            }
        )
        self.assertEqual([], validate(data, strict=True))

    def test_multiple_vendors_and_destinations_use_atomic_requirements(self) -> None:
        data = client_side_fixture()
        second = deepcopy(requirement(data))
        second["requirement_id"] = "REQ-002"
        second["source"]["reference"] = "tracking-plan.xlsx / Events / row 12 / K12"
        second["source"]["plan_order"] = 2
        second["container_id"] = "GTM-TEST"
        second["expectation"]["vendor_family"] = "ga4"
        second["expectation"]["destination_id"] = "G-TEST000001"
        second["expectation"]["destination_event_name"] = "add_to_cart"
        second["expectation"]["destination_id_parameter_path"] = "query.tid"
        second["expectation"]["destination_event_parameter_path"] = "query.en"
        second["expectation"]["tag_name"] = "GA4 - Event - add_to_cart"
        second["expectation"]["expected_endpoint_pattern"] = (
            "^https://www\\.google-analytics\\.com/g/collect"
        )
        second["resolved_data_layer"]["snapshot"] = deepcopy(second["raw_api_call"]["payload"])
        second["tag"].update(
            {
                "container_id": "GTM-TEST",
                "vendor_family": "ga4",
                "destination_id": "G-TEST000001",
                "event_name": "add_to_cart",
                "name": "GA4 - Event - add_to_cart",
                "configuration_evidence_id": "EVD-GA4-CONFIG-002",
                "runtime_evidence_id": "EVD-GA4-RUNTIME-002",
            }
        )
        second["destination_request"].update(
            {
                "container_id": "GTM-TEST",
                "vendor_family": "ga4",
                "destination_id": "G-TEST000001",
                "event_name": "add_to_cart",
                "request_id": "NET-GA4-002",
                "evidence_id": "EVD-GA4-NET-002",
                "request_url": (
                    "https://www.google-analytics.com/g/collect"
                    "?tid=G-TEST000001&en=add_to_cart&value=29.9"
                ),
            }
        )
        second["evidence_ids"].extend(
            [
                "EVD-GA4-CONFIG-002",
                "EVD-GA4-RUNTIME-002",
                "EVD-GA4-NET-002",
            ]
        )
        data["evidence"].extend(
            [
                {
                    "evidence_id": "EVD-GA4-CONFIG-002",
                    "kind": "tag_configuration",
                    "tag_id": "TAG-REQ-002",
                    "source": "Tag Assistant",
                    "capture_mode": "direct",
                    "action_id": "ACT-001",
                    "event_index": 11,
                    "container_id": "GTM-TEST",
                    "tag_name": "GA4 - Event - add_to_cart",
                    "configuration_field": "eventParameters.value",
                    "path_or_url": "evidence/ga4-config-002.json",
                    "captured_at": "2026-07-25T10:01:03+00:00",
                    "description": "Exact GA4 tag configuration.",
                },
                {
                    "evidence_id": "EVD-GA4-RUNTIME-002",
                    "kind": "tag_runtime",
                    "tag_id": "TAG-REQ-002",
                    "source": "Tag Assistant",
                    "capture_mode": "direct",
                    "action_id": "ACT-001",
                    "event_index": 11,
                    "container_id": "GTM-TEST",
                    "tag_name": "GA4 - Event - add_to_cart",
                    "configuration_field": "eventParameters.value",
                    "path_or_url": "evidence/ga4-runtime-002.json",
                    "captured_at": "2026-07-25T10:01:03+00:00",
                    "description": "Exact GA4 runtime parameter.",
                },
                {
                    "evidence_id": "EVD-GA4-NET-002",
                    "kind": "browser_network_request",
                    "tag_id": "TAG-REQ-002",
                    "source": "Browser Network",
                    "capture_mode": "direct",
                    "action_id": "ACT-001",
                    "request_id": "NET-GA4-002",
                    "container_id": "GTM-TEST",
                    "path_or_url": "evidence/ga4-network-002.json",
                    "captured_at": "2026-07-25T10:01:03+00:00",
                    "description": "Exact decoded GA4 browser request.",
                },
            ]
        )
        data["requirements"].append(second)
        data["run"]["requirement_inventory"].append("REQ-002")
        self.assertEqual([], validate(data, strict=True))

    def test_business_rule_and_regression_clis(self) -> None:
        baseline = client_side_fixture()
        current = deepcopy(baseline)
        current["run"]["run_id"] = "RUN-CURRENT-001"
        requirement(current)["verdict"]["tag_firing"] = "FAIL"
        requirement(current)["verdict"]["overall"] = "FAIL"
        with tempfile.TemporaryDirectory() as tempdir:
            baseline_path = Path(tempdir) / "baseline.json"
            current_path = Path(tempdir) / "current.json"
            baseline_path.write_text(json.dumps(baseline), encoding="utf-8")
            current_path.write_text(json.dumps(current), encoding="utf-8")
            rules = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "validate_business_rules.py"),
                    str(baseline_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            diff = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "diff_recette_runs.py"),
                    str(baseline_path),
                    str(current_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(0, rules.returncode)
        self.assertEqual("PASS", json.loads(rules.stdout)["status"])
        self.assertEqual(1, diff.returncode)
        self.assertEqual("REGRESSED", json.loads(diff.stdout)["results"][0]["change"])

    def test_schema_v1_is_rejected_actionably(self) -> None:
        data = fixture()
        data["schema_version"] = 1
        self.assert_invalid(data, "must be migrated or re-normalized")

    def test_missing_acceptance_source_is_rejected(self) -> None:
        data = fixture()
        data["run"]["tracking_plan_source"] = ""
        self.assert_invalid(data, "tracking_plan_source")

    def test_run_type_modes_are_rejected(self) -> None:
        data = fixture()
        data["run"]["run_type"] = "FULL_TRACKING_PLAN_RECETTE"
        self.assert_invalid(data, "run_type is obsolete")

    def test_inventory_omission_is_rejected(self) -> None:
        data = fixture()
        data["run"]["requirement_inventory"] = []
        self.assert_invalid(data, "requirement_inventory")

    def test_placeholder_raw_payload_is_rejected(self) -> None:
        data = fixture()
        requirement(data)["raw_api_call"]["payload"]["ecommerce"]["value"] = "..."
        self.assert_invalid(data, "placeholder")

    def test_browser_interception_cannot_replace_preview_dependent_api_call(self) -> None:
        data = fixture()
        requirement(data)["raw_api_call"]["capture_source"] = "browser_interception"
        self.assert_invalid(data, "planned dataLayer acceptance requires exact Tag Assistant")

    def test_fixed_value_mismatch_cannot_hide_behind_pass(self) -> None:
        data = fixture()
        requirement(data)["raw_api_call"]["field_value"] = 30
        requirement(data)["raw_api_call"]["payload"]["ecommerce"]["value"] = 30
        self.assert_invalid(data, "PASS raw_payload contradicts")

    def test_raw_absent_and_stale_resolved_value_remain_separate(self) -> None:
        data = fixture()
        req = requirement(data)
        raw = req["raw_api_call"]
        raw["payload"]["ecommerce"] = {}
        raw["field_state"] = "absent"
        raw["field_type"] = "absent"
        raw.pop("field_value")
        req["verdict"]["raw_payload"] = "FAIL"
        req["verdict"]["resolved_data_layer"] = "FAIL"
        req["verdict"]["overall"] = "FAIL"
        req["verdict"]["failure_layer"] = "raw_payload"
        req["verdict"]["mismatch"] = "Raw field absent; resolved state retained 29.9."
        self.assertEqual([], validate(data, strict=True))

    def test_tag_firing_and_undefined_parameter_have_separate_verdicts(self) -> None:
        data = fixture()
        req = requirement(data)
        tag = req["tag"]
        tag["runtime_state"] = "undefined"
        tag["runtime_type"] = "undefined"
        tag.pop("runtime_value")
        req["verdict"]["tag_firing"] = "PASS"
        req["verdict"]["tag_parameter"] = "FAIL"
        req["verdict"]["overall"] = "FAIL"
        req["verdict"]["failure_layer"] = "tag_parameter"
        req["verdict"]["mismatch"] = "Runtime parameter was undefined."
        self.assertEqual([], validate(data, strict=True))

    def test_wrong_runtime_parameter_cannot_pass(self) -> None:
        data = fixture()
        tag = requirement(data)["tag"]
        tag["runtime_value"] = "29.9"
        tag["runtime_type"] = "string"
        self.assert_invalid(data, "PASS tag parameter differs")

    def test_wrong_tag_configuration_cannot_pass(self) -> None:
        data = fixture()
        requirement(data)["tag"]["configured_value"] = "{{Wrong Variable}}"
        self.assert_invalid(data, "PASS tag configuration differs")

    def test_correct_data_layer_cannot_mask_configuration_or_runtime_failures(
        self,
    ) -> None:
        data = fixture()
        req = requirement(data)
        req["tag"]["configured_value"] = "{{Wrong Variable}}"
        req["tag"]["runtime_value"] = "29.9"
        req["tag"]["runtime_type"] = "string"
        req["verdict"].update(
            {
                "tag_configuration": "FAIL",
                "tag_parameter": "FAIL",
                "overall": "FAIL",
                "failure_layer": "tag_configuration",
                "mismatch": "Configured source and runtime type differ from plan.",
            }
        )
        self.assertEqual([], validate(data, strict=True))
        req["verdict"]["overall"] = "PASS"
        self.assert_invalid(data, "does not equal worst applicable component")

    def test_expected_tag_configuration_cannot_be_silently_omitted(self) -> None:
        data = fixture()
        requirement(data)["verdict"].pop("tag_configuration")
        self.assert_invalid(data, "requires a configuration verdict")

    def test_required_base_layer_verdicts_cannot_be_omitted(self) -> None:
        cases = (
            ("raw_payload", "required raw API-call layer"),
            ("resolved_data_layer", "required resolved Data Layer"),
            ("gtm_variable", "expected GTM variable requires"),
            ("tag_configuration", "requires a configuration verdict"),
        )
        for verdict_field, expected_error in cases:
            with self.subTest(verdict_field=verdict_field):
                data = fixture()
                requirement(data)["verdict"].pop(verdict_field)
                self.assert_invalid(data, expected_error)

    def test_wrong_gtm_variable_mapping_cannot_pass(self) -> None:
        data = fixture()
        variable = requirement(data)["gtm_variable"]
        variable["field_value"] = 30
        self.assert_invalid(data, "PASS GTM variable differs")

    def test_wanted_nonfired_tag_requires_reason(self) -> None:
        data = fixture()
        req = requirement(data)
        req["tag"]["actual_firing"] = "not_fired"
        req["tag"]["fire_count"] = 0
        req["verdict"]["tag_firing"] = "FAIL"
        req["verdict"]["tag_parameter"] = "BLOCKED"
        req["verdict"]["overall"] = "FAIL"
        self.assert_invalid(data, "lacks non_firing_reason")

    def test_action_boundary_timestamp_and_cursor_order_are_strict(self) -> None:
        data = fixture()
        requirement(data)["action_boundary"]["action_timestamp"] = "not-a-time"
        self.assert_invalid(data, "action_timestamp must be ISO 8601 with timezone")

        data = fixture()
        requirement(data)["action_boundary"]["first_event_after"] = 9
        self.assert_invalid(data, "first_event_after must follow last_event_before")

        data = fixture()
        requirement(data)["occurrence_evidence"]["event_indexes"] = [13]
        self.assert_invalid(data, "occurrence event index exceeds settled_final_event")

    def test_completed_interaction_requires_independent_completion_signal(self) -> None:
        data = fixture()
        requirement(data)["action_boundary"].pop("completion_signal")
        self.assert_invalid(
            data,
            "completed interaction requires an independent completion_signal",
        )

    def test_failed_interaction_cannot_prove_expected_event_absence(self) -> None:
        data = fixture()
        req = configure_absent_event(data)
        req["action_boundary"]["interaction_outcome"] = "failed"
        req["action_boundary"]["completion_signal"] = "Overlay intercepted the click"
        req["action_boundary"]["settlement_reason"] = "interaction_failed"
        self.assert_invalid(
            data,
            "failed or uncertain interaction cannot prove expected-event absence",
        )

    def test_reviewed_attempt_still_requires_action_boundary(self) -> None:
        data = fixture()
        req = requirement(data)
        req["journey"]["execution_status"] = "REVIEW"
        req.pop("action_boundary")
        self.assert_invalid(data, "missing action_boundary")

    def test_journey_action_value_contract_is_required(self) -> None:
        data = fixture()
        requirement(data)["journey"].pop("action_value_source")
        self.assert_invalid(data, "journey missing 'action_value_source'")

    def test_client_container_inventory_is_required(self) -> None:
        data = fixture()
        data["run"].pop("containers")
        self.assert_invalid(data, "non-empty client-side web-container array")

    def test_expected_event_absence_with_settled_boundary_is_valid_failure(self) -> None:
        data = fixture()
        req = configure_absent_event(data)
        self.assertEqual([], validate(data, strict=True))
        self.assertEqual("FAIL", req["verdict"]["raw_payload"])
        self.assertEqual("BLOCKED", req["verdict"]["resolved_data_layer"])
        self.assertEqual("FAIL", event_rollup(data)[0]["status"])

    def test_session_rejects_missing_source_without_raw_fail_and_blocked_chain(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        action = session["actions"][0]
        action["expected_seen"] = False
        action["observed_business_push_count"] = 0
        session["business_pushes"] = []
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any(
                "absent expected dataLayer source requires raw_api_call=FAIL" in error
                for error in errors
            ),
            errors,
        )

    def test_duplicate_push_cannot_pass_an_expected_once_rule(self) -> None:
        data = fixture()
        occurrence = requirement(data)["occurrence_evidence"]
        occurrence["actual_count"] = 2
        occurrence["event_indexes"] = [11, 12]
        self.assert_invalid(
            data,
            "PASS event occurrence contradicts observed chronology/count",
        )

    def test_absent_event_requires_applicable_consent_to_be_blocked(self) -> None:
        data = client_side_fixture()
        req = configure_absent_event(data)
        for component in (
            "destination_request",
            "destination_parameter",
            "trigger_logic",
            "tag_sequence",
            "business_rule",
        ):
            req["verdict"][component] = "BLOCKED"
        req["verdict"]["consent"] = "PASS"
        self.assert_invalid(
            data,
            "absent expected event requires downstream consent=BLOCKED",
        )

    def test_preview_disconnect_is_blocked_not_implementation_fail(self) -> None:
        data = fixture()
        req = add_blocker(
            data,
            "PREVIEW_DISCONNECTED",
            help_requested=False,
            settled=False,
        )
        req["action_boundary"]["preview_connected_before"] = False
        req["action_boundary"]["target_ready_before"] = False
        self.assertEqual([], validate(data, strict=True))

    def test_protected_blocker_requires_analyst_help_request(self) -> None:
        data = fixture()
        add_blocker(data, "MFA", help_requested=False)
        self.assert_invalid(data, "analyst help must be requested")

    def test_protected_blocker_with_handoff_is_valid(self) -> None:
        data = fixture()
        add_blocker(data, "EMAIL_VERIFICATION", help_requested=True)
        self.assertEqual([], validate(data, strict=True))

    def test_ui_control_blocker_requires_complete_recovery_attempts(self) -> None:
        data = fixture()
        add_blocker(data, "UI_CONTROL_BLOCKER", help_requested=False)
        self.assert_invalid(data, "UI_CONTROL_BLOCKER requires attempted_methods")

        data = fixture()
        add_blocker(data, "UI_CONTROL_BLOCKER", help_requested=False)
        data["blockers"][0]["attempted_methods"] = [
            "scroll_into_view",
            "label_click",
            "direct_control",
            "pointer_click",
            "keyboard_toggle",
            "clean_state_retry",
        ]
        self.assertEqual([], validate(data, strict=True))

    def test_http_403_is_blocked_not_not_tested(self) -> None:
        data = fixture()
        add_blocker(data, "HTTP_403", help_requested=False)
        self.assertEqual([], validate(data, strict=True))

    def test_not_tested_cannot_hide_attempted_blocker(self) -> None:
        data = fixture()
        req = add_blocker(data, "HTTP_403", help_requested=False)
        for key in (
            "event_occurrence",
            "raw_payload",
            "resolved_data_layer",
            "gtm_variable",
            "tag_firing",
            "tag_parameter",
        ):
            req["verdict"][key] = "NOT_TESTED"
        req["verdict"]["overall"] = "NOT_TESTED"
        self.assert_invalid(data, "only valid for confirmed OUT_OF_SCOPE")

    def test_limited_layer_chronology_requires_anchor_evidence(self) -> None:
        data = fixture("valid_limited_layers.json")
        requirement(data)["occurrence_evidence"].pop("anchor_event_index")
        self.assert_invalid(data, "requires anchor_event_index")

    def test_unrelated_tag_matrix_row_is_rejected(self) -> None:
        data = fixture()
        requirement(data)["tag"]["relevance"] = "unrelated"
        self.assert_invalid(data, "tag relevance")

    def test_unapproved_cmp_override_is_rejected(self) -> None:
        data = fixture()
        add_consent_override(data, approved=False)
        self.assert_invalid(data, "lacks explicit analyst approval")

    def test_approved_nonproduction_cmp_override_is_valid(self) -> None:
        data = fixture()
        add_consent_override(data, approved=True)
        self.assertEqual([], validate(data, strict=True))

    def test_production_cmp_override_requires_explicit_exception(self) -> None:
        data = fixture()
        add_consent_override(data, approved=True, production=True)
        self.assert_invalid(data, "production_exception_approved=true")

    def test_explicit_production_cmp_override_is_valid_for_downstream_only(self) -> None:
        data = fixture()
        add_consent_override(
            data,
            approved=True,
            production=True,
            production_approved=True,
        )
        self.assertEqual([], validate(data, strict=True))

    def test_duplicate_evidence_id_is_rejected(self) -> None:
        data = fixture()
        data["evidence"][1]["evidence_id"] = data["evidence"][0]["evidence_id"]
        self.assert_invalid(data, "duplicate IDs")

    def test_tracking_plan_inspector_preserves_coordinates(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            path = Path(tempdir) / "plan.xlsx"
            assets = Path(tempdir) / "plan-assets"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Events"
            sheet["B12"] = "add_to_cart"
            sheet["B12"].hyperlink = "https://shop.example.test/product"
            sheet["B12"].comment = Comment("Open the product card.", "Analyst")
            sheet["F12"] = 29.9
            sheet.merge_cells("H2:I3")
            image_path = Path(tempdir) / "journey.png"
            PillowImage.new("RGB", (4, 4), color="white").save(image_path)
            with PillowImage.open(image_path) as image_source:
                plan_image = WorkbookImage(image_source)
                plan_image.anchor = "J4"
                sheet.add_image(plan_image)
                workbook.save(path)
            result = inspect_xlsx(path, max_rows=0, assets_dir=assets)
            cells = result["sheets"][0]["populated_rows"][0]["cells"]
            self.assertEqual(["B12", "F12"], [cell["cell"] for cell in cells])
            self.assertEqual("number", cells[1]["value_type"])
            self.assertEqual(
                "https://shop.example.test/product",
                cells[0]["hyperlink"]["target"],
            )
            self.assertEqual("Open the product card.", cells[0]["comment"]["text"])
            self.assertEqual(["H2:I3"], result["sheets"][0]["merged_ranges"])
            image = result["sheets"][0]["images"][0]
            self.assertEqual("J4", image["anchor"]["from_cell"])
            self.assertTrue(Path(image["extracted_file"]).is_file())

    def test_browser_request_decoder_preserves_repeated_and_batched_values(self) -> None:
        result = decode_requests(
            [
                {
                    "request_id": "NET-1",
                    "action_id": "ACT-1",
                    "url": (
                        "https://collect.example.test/g/collect"
                        "?id=G-TEST&ep.item=alpha&ep.item=beta"
                    ),
                    "method": "POST",
                    "headers": {
                        "Content-Type": "text/plain",
                        "Authorization": "Bearer secret",
                    },
                    "post_data": "en=one&value=1\nen=two&value=2",
                }
            ]
        )
        request = result["requests"][0]
        self.assertEqual(["alpha", "beta"], request["query"]["ep.item"])
        self.assertEqual("newline_batch", request["body"]["format"])
        self.assertEqual(2, len(request["body"]["records"]))
        self.assertNotIn("raw", request["body"])
        self.assertIn("authorization", request["excluded_header_names"])

    def test_incremental_event_validation_and_status(self) -> None:
        data = fixture()
        result = validate_event(data, "EVG-001")
        self.assertEqual("PASS", result["status"])
        self.assertEqual("PASS", status_rows(data)[0]["status"])

    def test_incremental_event_patch_preserves_layer_coherence(self) -> None:
        data = fixture()
        patch = {
            "event_group_id": "EVG-001",
            "requirements": deepcopy(data["requirements"]),
            "evidence": [],
        }
        updated, event_group_id = apply_event(data, patch)
        self.assertEqual("EVG-001", event_group_id)
        self.assertEqual("PASS", validate_event(updated, event_group_id)["status"])

    def test_synthetic_profile_uses_reserved_example_domain(self) -> None:
        result = subprocess.run(
            [
                sys.executable,
                str(SCRIPTS / "generate_synthetic_profile.py"),
                "--seed",
                "RUN-001",
                "--locale",
                "fr-FR",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        profile = json.loads(result.stdout)
        self.assertTrue(profile["synthetic"])
        self.assertTrue(profile["email"].endswith("@example.com"))
        self.assertNotIn("password", profile)

    def test_coverage_initializer_preserves_plan_inventory(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            input_path = Path(tempdir) / "requirements.json"
            output_path = Path(tempdir) / "ledger.json"
            source = fixture()["requirements"]
            input_path.write_text(
                json.dumps({"requirements": source}),
                encoding="utf-8",
            )
            subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "init_coverage_ledger.py"),
                    str(input_path),
                    str(output_path),
                    "--run-id",
                    "RUN-LEDGER",
                    "--title",
                    "Ledger test",
                    "--site-url",
                    "https://shop.example.test/",
                    "--environment",
                    "Preprod",
                    "--environment-class",
                    "preprod",
                    "--container-id",
                    "GTM-TEST",
                    "--workspace",
                    "Recette",
                    "--tracking-plan-source",
                    "tracking-plan.xlsx",
                    "--acceptance-scope",
                    "Full plan",
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            ledger = json.loads(output_path.read_text(encoding="utf-8"))
            self.assertEqual(["REQ-001"], ledger["run"]["requirement_inventory"])
            self.assertEqual("EVG-001", ledger["run"]["event_inventory"][0]["event_group_id"])
            self.assertEqual(
                "PENDING",
                ledger["requirements"][0]["journey"]["execution_status"],
            )
            self.assertEqual(
                "PENDING",
                ledger["requirements"][0]["verdict"]["overall"],
            )
            self.assertEqual("analytics_only", ledger["run"]["tag_scope"]["mode"])
            self.assertTrue(ledger["run"]["journey_authority"]["complete_ordinary_journeys"])
            self.assertTrue(ledger["run"]["journey_authority"]["ordinary_form_submissions"])

    def test_final_execution_contract_accepts_complete_case_and_push_stream(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        self.assertEqual([], validate_session(session, results=data, final=True))

    def test_analytics_scope_keeps_detected_media_visible_out_of_scope(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        case = session["cases"][0]
        media_evidence = deepcopy(
            next(row for row in data["evidence"] if row.get("evidence_id") == "EVD-TAG-CONFIG-011")
        )
        media_evidence.update(
            {
                "evidence_id": "EVD-META-CONFIG-OUT",
                "tag_id": "TAG-META-DETECTED",
                "tag_name": "Meta - AddToCart",
            }
        )
        data["evidence"].append(media_evidence)
        media_tag = {
            "tag_id": "TAG-META-DETECTED",
            "tag_name": "Meta - AddToCart",
            "container_id": "GTM-TEST",
            "tag_category": "media",
            "tag_delivery": "browser_request",
            "vendor_family": "meta",
            "destination_id": "META-OTHER",
            "template_type": "gallery_template",
            "consent_required": True,
            "evidence_ids": ["EVD-META-CONFIG-OUT"],
        }
        media_tag["scope_status"], media_tag["scope_reason"] = tag_scope_decision(
            media_tag,
            data["run"]["tag_scope"],
            case["declared_tag_contracts"],
        )
        self.assertEqual("OUT_OF_SCOPE", media_tag["scope_status"])
        case["tag_inventory"].append(media_tag)
        case["layer_applicability"] = layer_applicability(
            data["requirements"],
            container_count=1,
            tag_inventory=case["tag_inventory"],
        )
        case["applicable_layers"] = [
            row["layer"] for row in case["layer_applicability"] if row["mode"] == "MANDATORY"
        ]
        self.assertEqual([], validate_session(session, results=data, final=True))
        feedback = event_feedback(data, session)[0]
        excluded = next(
            row for row in feedback["tag_feedback"] if row["tag_id"] == "TAG-META-DETECTED"
        )
        self.assertEqual("OUT_OF_SCOPE", excluded["scope_status"])
        self.assertEqual(2, len(tag_rows(data, session)))

    def test_tracking_plan_tag_fields_do_not_control_runtime_tag_layers(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        req = requirement(data)
        for field in (
            "tag_name",
            "tag_delivery",
            "expected_firing",
            "tag_configuration_field",
            "expected_tag_configuration",
            "vendor_family",
            "destination_id",
            "destination_event_name",
            "destination_id_parameter_path",
            "destination_event_parameter_path",
            "destination_parameter_path",
            "expected_destination_value",
            "expected_destination_type",
            "expected_endpoint_pattern",
            "expected_request_behavior",
        ):
            req["expectation"].pop(field, None)
        req["tag"] = None
        req["destination_request"] = None
        session["cases"][0]["declared_tag_contracts"] = []
        session["cases"][0]["source_expectations"] = [deepcopy(req["expectation"])]
        self.assertEqual([], validate(data, strict=True))
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("expected anchor path is absent" in error for error in errors))
        self.assertFalse(any("omitted per-tag layers" in error for error in errors))

    def test_exact_plan_declared_media_is_in_scope_under_analytics_default(self) -> None:
        tag = {
            "tag_name": "Meta - Purchase",
            "tag_category": "media",
            "vendor_family": "meta",
            "destination_id": "META-123",
        }
        status, reason = tag_scope_decision(
            tag,
            {"mode": "analytics_only", "include_plan_declared_media": True},
            [
                {
                    "tag_name": "Meta - Purchase",
                    "vendor_family": "meta",
                    "destination_id": "META-123",
                }
            ],
        )
        self.assertEqual("IN_SCOPE", status)
        self.assertIn("explicitly declared", reason)

    def test_known_vendor_metadata_prevents_tag_category_scope_manipulation(self) -> None:
        self.assertEqual(
            "analytics",
            inferred_tag_category({"vendor_family": "ga4", "template_type": "GA4 Event"}),
        )
        self.assertEqual(
            "media",
            inferred_tag_category({"vendor_family": "meta", "template_type": "Gallery template"}),
        )
        data = fixture()
        session = execution_fixture(data)
        session["cases"][0]["tag_inventory"][0]["tag_category"] = "media"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any(
                "tag_category contradicts direct vendor/template metadata" in error
                for error in errors
            ),
            errors,
        )

    def test_empty_analytics_tag_inventory_requires_explicit_failure_chain(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        data["evidence"].append(
            {
                "evidence_id": "EVD-TAG-INVENTORY-EMPTY",
                "kind": "tag_inventory",
                "source": "GTM read-only",
                "capture_mode": "direct",
                "container_id": "GTM-TEST",
                "path_or_url": "evidence/tag-inventory-empty.json",
                "captured_at": "2026-07-25T10:01:03+00:00",
                "description": "Complete concerned-tag search found no in-scope analytics tag.",
            }
        )
        case = session["cases"][0]
        case["tag_inventory"] = []
        case["tag_inventory_evidence_ids"] = ["EVD-TAG-INVENTORY-EMPTY"]
        case["layer_applicability"] = layer_applicability(
            data["requirements"], container_count=1, tag_inventory=[]
        )
        case["applicable_layers"] = [
            row["layer"] for row in case["layer_applicability"] if row["mode"] == "MANDATORY"
        ]
        action = session["actions"][0]
        action["tag_layer_results"] = []
        status_by_layer = {
            "concerned_tag_inventory": "FAIL",
            "tag_configuration": "FAIL",
            "tag_firing": "FAIL",
            "gtm_variable": "BLOCKED",
            "tag_parameter": "BLOCKED",
            "destination_request_when_applicable": "BLOCKED",
        }
        for row in action["layer_results"]:
            if row["layer"] in status_by_layer:
                row["status"] = status_by_layer[row["layer"]]
                row["reason"] = "No in-scope analytics tag was identified for the event."
                row["evidence_ids"] = ["EVD-TAG-INVENTORY-EMPTY"]
                row["blocker_id"] = "NO_IN_SCOPE_TAG" if row["status"] == "BLOCKED" else None
        req = requirement(data)
        req["verdict"].update(
            {
                "gtm_variable": "BLOCKED",
                "tag_configuration": "FAIL",
                "tag_firing": "FAIL",
                "tag_parameter": "BLOCKED",
                "destination_request": "BLOCKED",
                "destination_parameter": "BLOCKED",
                "overall": "FAIL",
                "failure_layer": "tag_configuration",
                "mismatch": "No in-scope analytics tag was identified.",
            }
        )
        req["evidence_ids"].append("EVD-TAG-INVENTORY-EMPTY")
        self.assertEqual([], validate(data, strict=True))
        self.assertEqual([], validate_session(session, results=data, final=True))

    def test_available_network_capture_without_match_cannot_be_blocked(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        action = session["actions"][0]
        destination = next(
            row
            for row in action["tag_layer_results"]
            if row["layer"] == "destination_request_when_applicable"
        )
        destination.update(
            {
                "status": "BLOCKED",
                "reason": "No matching request was observed in an available capture.",
                "blocker_id": "NETWORK",
            }
        )
        destination["details"].update(
            {"request_count": 0, "request_ids": [], "capture_unavailable": False}
        )
        aggregate = next(
            row
            for row in action["layer_results"]
            if row["layer"] == "destination_request_when_applicable"
        )
        aggregate.update({"status": "BLOCKED", "blocker_id": "NETWORK"})
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("available capture with no match is FAIL" in error for error in errors),
            errors,
        )

    def test_per_tag_runtime_parameter_false_pass_is_rejected(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        parameter = next(
            row
            for row in session["actions"][0]["tag_layer_results"]
            if row["layer"] == "tag_parameter"
        )
        parameter["details"]["parameters"][0]["actual_value"] = "29.9"
        parameter["details"]["parameters"][0]["actual_type"] = "string"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("contradicts exact value/type comparison" in error for error in errors),
            errors,
        )

    def test_feedback_and_workbook_rows_expose_every_event_and_tag_layer(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        feedback = event_feedback(data, session)[0]
        self.assertEqual(len(CANONICAL_LAYERS), len(feedback["layer_feedback"]))
        self.assertEqual(len(TAG_RESULT_LAYERS), len(feedback["tag_feedback"][0]["layers"]))
        self.assertEqual(
            list(TAG_RESULT_LAYERS),
            [row["layer"] for row in feedback["tag_feedback"][0]["layers"]],
        )
        rows = layer_verdict_rows(data, session)
        self.assertEqual(len(CANONICAL_LAYERS) + len(TAG_RESULT_LAYERS), len(rows))
        self.assertEqual(
            list(CANONICAL_LAYERS),
            [row["layer"] for row in rows[: len(CANONICAL_LAYERS)]],
        )
        self.assertEqual(
            list(TAG_RESULT_LAYERS),
            [row["layer"] for row in rows[len(CANONICAL_LAYERS) :]],
        )
        self.assertTrue(all(str(row.get("status", "")).strip() for row in rows))
        self.assertTrue(all(str(row.get("reason", "")).strip() for row in rows))

    def test_false_consent_predicate_still_requires_natural_baseline_evidence(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        consent = next(
            row
            for row in session["actions"][0]["layer_results"]
            if row["layer"] == "consent_when_applicable"
        )
        consent["evidence_ids"] = ["EVD-ACTION-001"]
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any(
                "consent_when_applicable: no direct evidence of the required kind" in error
                for error in errors
            ),
            errors,
        )

    def test_run_authorization_is_reusable_but_credentials_remain_ephemeral(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["authorizations"] = [
            {
                "authorization_id": "AUTH-SAFE-FORMS",
                "scope": "ordinary_form_submission",
                "description": "Complete safe synthetic forms for this controlled run.",
                "environment_class": "preprod",
                "exact_method": None,
                "session_only": True,
                "protected_exclusions": list(PROTECTED_AUTHORIZATION_EXCLUSIONS),
                "approved_at": "2026-07-25T09:57:00+00:00",
            }
        ]
        session["cases"][0]["authorization_ids"] = ["AUTH-SAFE-FORMS"]
        self.assertEqual([], validate_session(session, results=data, final=True))
        session["cases"][0]["material_variant"]["email"] = "user@example.com"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("must remain ephemeral" in error for error in errors),
            errors,
        )

    def test_final_execution_contract_rejects_an_unexecuted_material_case(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        second = deepcopy(session["cases"][0])
        second.update(
            {
                "case_id": "CASE-002",
                "material_variant": {"quantity": 2},
                "execution_status": "PENDING",
                "final_action_id": None,
            }
        )
        session["cases"].append(second)
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("applicable case remains PENDING" in error for error in errors),
            errors,
        )

    def test_final_execution_contract_rejects_an_unreconciled_business_push(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["actions"][0]["observed_business_push_count"] = 2
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("does not match classified stream rows" in error for error in errors),
            errors,
        )

    def test_anomalous_business_push_requires_an_unexpected_finding(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["business_pushes"][0]["classification"] = "duplicate"
        session["business_pushes"][0]["classification_reason"] = (
            "Duplicate add_to_cart in the same action window."
        )
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("anomalous push is absent from unexpected" in error for error in errors),
            errors,
        )

    def test_final_execution_contract_rejects_missing_applicable_network_layer(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["cases"][0]["applicable_layers"].remove("destination_request_when_applicable")
        session["actions"][0]["layer_results"] = [
            row
            for row in session["actions"][0]["layer_results"]
            if row["layer"] != "destination_request_when_applicable"
        ]
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("omits explicit layer results" in error for error in errors),
            errors,
        )

    def test_local_only_tag_does_not_invent_a_network_layer(self) -> None:
        data = fixture()
        req = requirement(data)
        req["expectation"]["tag_delivery"] = "local_only"
        for field in (
            "vendor_family",
            "destination_id",
            "destination_event_name",
            "destination_id_parameter_path",
            "destination_event_parameter_path",
            "destination_parameter_path",
            "expected_destination_value",
            "expected_destination_type",
            "expected_endpoint_pattern",
            "expected_request_behavior",
        ):
            req["expectation"].pop(field, None)
        req["destination_request"] = None
        req["verdict"]["destination_request"] = None
        req["verdict"]["destination_parameter"] = None
        data["run"]["included_layers"].remove("destination_request_when_applicable")
        self.assertNotIn(
            "destination_request_when_applicable",
            applicable_layers(data["requirements"]),
        )
        self.assertEqual([], validate(data, strict=True))

    def test_reconstructed_evidence_cannot_claim_direct_api_call_proof(self) -> None:
        data = fixture()
        raw_evidence = next(row for row in data["evidence"] if row["evidence_id"] == "EVD-RAW-011")
        raw_evidence["capture_mode"] = "supplemental"
        self.assert_invalid(data, "not reconstructed or inferred")

    def test_review_is_reserved_for_a_precise_semantic_question(self) -> None:
        data = fixture()
        req = requirement(data)
        req["verdict"]["tag_parameter"] = "REVIEW"
        req["verdict"]["overall"] = "REVIEW"
        self.assert_invalid(data, "review_basis=semantic_ambiguity")
        req["verdict"]["review_basis"] = "semantic_ambiguity"
        req["verdict"]["review_question"] = (
            "Does the approved plan intentionally round value to one decimal?"
        )
        self.assertEqual([], validate(data, strict=True))

    def test_event_feedback_contains_layer_results_and_exact_retest_location(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        req = requirement(data)
        req["verdict"]["tag_parameter"] = "FAIL"
        req["verdict"]["overall"] = "FAIL"
        req["verdict"]["failure_layer"] = "tag_parameter"
        req["verdict"]["mismatch"] = "Runtime value is 29 instead of 29.9."
        layer = next(
            row for row in session["actions"][0]["layer_results"] if row["layer"] == "tag_parameter"
        )
        layer["status"] = "FAIL"
        layer["reason"] = "Runtime value is 29 instead of 29.9."
        feedback = event_feedback(data, session)[0]
        self.assertEqual("FAIL", feedback["status"])
        self.assertEqual("FAIL", feedback["verified_layers"]["tag_parameter"])
        self.assertIn("https://shop.example.test/product", feedback["retest"])
        self.assertIn("Add to cart", feedback["retest"])

    def test_event_feedback_uses_the_final_retry_layers(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        first_attempt = session["actions"][0]
        first_attempt["interaction_outcome"] = "failed"
        failed_layer = next(
            row for row in first_attempt["layer_results"] if row["layer"] == "tag_parameter"
        )
        failed_layer["status"] = "FAIL"
        failed_layer["reason"] = "The first click was intercepted before completion."
        final_attempt = deepcopy(first_attempt)
        final_attempt.update(
            {
                "action_id": "ACT-002",
                "attempt_number": 2,
                "retry_of_action_id": "ACT-001",
                "interaction_outcome": "completed",
            }
        )
        final_layer = next(
            row for row in final_attempt["layer_results"] if row["layer"] == "tag_parameter"
        )
        final_layer["status"] = "PASS"
        final_layer["reason"] = "The completed retry matched the planned value."
        session["actions"].append(final_attempt)
        session["cases"][0]["final_action_id"] = "ACT-002"

        feedback = event_feedback(data, session)[0]

        self.assertEqual("PASS", feedback["status"])
        self.assertEqual("PASS", feedback["verified_layers"]["tag_parameter"])
        self.assertNotIn("first click was intercepted", feedback["reason"])

    def test_workbook_includes_interaction_cases_and_observed_push_stream(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "complete-recette.xlsx"
            build_workbook(data, output, session=session)
            workbook = load_workbook(output, read_only=True)
            self.assertEqual(2, workbook["Interaction Cases"].max_row)
            self.assertEqual(2, workbook["Observed Push Stream"].max_row)
            workbook.close()

    def test_strict_workbook_cli_requires_and_accepts_the_session_ledger(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        with tempfile.TemporaryDirectory() as tempdir:
            results_path = Path(tempdir) / "results.json"
            session_path = Path(tempdir) / "session.json"
            output = Path(tempdir) / "recette.xlsx"
            results_path.write_text(json.dumps(data), encoding="utf-8")
            session_path.write_text(json.dumps(session), encoding="utf-8")
            missing = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "build_recette_report.py"),
                    str(results_path),
                    str(output),
                    "--strict",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            complete = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "build_recette_report.py"),
                    str(results_path),
                    str(output),
                    "--strict",
                    "--session-ledger",
                    str(session_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(2, missing.returncode)
        self.assertIn("requires --session-ledger", missing.stderr)
        self.assertEqual(0, complete.returncode, complete.stderr)

    def test_preview_session_ledger_supports_checkpointed_action(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            ledger = Path(tempdir) / "session.json"
            results = Path(tempdir) / "results.json"
            results.write_text(json.dumps(fixture()), encoding="utf-8")
            script = str(SCRIPTS / "preview_session_ledger.py")

            def run(*args: str) -> subprocess.CompletedProcess[str]:
                return subprocess.run(
                    [sys.executable, script, *args],
                    check=True,
                    capture_output=True,
                    text=True,
                )

            def record_runtime(
                check_id: str,
                phase: str,
                action_id: str,
                preview_cursor: int,
                network_cursor: int,
                *,
                first_event_after: int | None = None,
                push_count: int | None = None,
            ) -> None:
                captured_at = datetime.now(UTC).isoformat(timespec="seconds")
                runtime_ids = [f"EVD-{check_id}-ACTION", f"EVD-{check_id}-NETWORK"]
                current_results = json.loads(results.read_text(encoding="utf-8"))
                current_results.setdefault("evidence", []).extend(
                    [
                        {
                            "evidence_id": runtime_ids[0],
                            "kind": "action_boundary",
                            "source": "Playwright",
                            "capture_mode": "direct",
                            "action_id": action_id,
                            "runtime_check_id": check_id,
                            "runtime_phase": phase,
                            "path_or_url": f"evidence/{check_id}-action.json",
                            "captured_at": captured_at,
                            "description": "Direct runtime action boundary.",
                        },
                        {
                            "evidence_id": runtime_ids[1],
                            "kind": "browser_network_capture",
                            "source": "Browser Network",
                            "capture_mode": "direct",
                            "action_id": action_id,
                            "runtime_check_id": check_id,
                            "runtime_phase": phase,
                            "container_id": "GTM-TEST",
                            "path_or_url": f"evidence/{check_id}-network.json",
                            "captured_at": captured_at,
                            "description": "Direct runtime network boundary.",
                        },
                    ]
                )
                results.write_text(json.dumps(current_results), encoding="utf-8")
                snapshot = {
                    "check_id": check_id,
                    "captured_at": captured_at,
                    "capture_source": "playwright_runtime_probe",
                    "browser_context_id": "desktop-default",
                    "connection_epoch": 1,
                    "gtm_workspace_surface_id": "gtm_workspace",
                    "tag_assistant_surface_id": "tag_assistant",
                    "website_surface_id": "website",
                    "containers": [{"container_id": "GTM-TEST", "workspace": "Recette"}],
                    "website_url": "https://shop.example.test/product",
                    "selected_page_url": "https://shop.example.test/product",
                    "preview_connected": True,
                    "target_interactive": True,
                    "target_uncovered": True,
                    "lifecycle_observed": True,
                    "stream_quiet": True,
                    "network_capture_active": True,
                    "preview_event_cursor": preview_cursor,
                    "network_request_cursor": network_cursor,
                    "evidence_ids": runtime_ids,
                }
                if phase == "after_action":
                    snapshot["first_event_after"] = first_event_after
                    snapshot["observed_business_push_count"] = push_count
                snapshot_path = Path(tempdir) / f"{check_id}.json"
                snapshot_path.write_text(json.dumps(snapshot), encoding="utf-8")
                run(
                    "record-runtime-check",
                    str(ledger),
                    str(snapshot_path),
                    "--results",
                    str(results),
                    "--phase",
                    phase,
                    "--action-id",
                    action_id,
                    "--case-id",
                    "CASE-ADD-DESKTOP",
                )

            run(
                "init",
                str(ledger),
                "--profile-path",
                str(Path(tempdir) / "profile"),
                "--approved-origin",
                "https://shop.example.test",
            )
            run(
                "register-surface",
                str(ledger),
                "--role",
                "gtm_workspace",
                "--url",
                "https://tagmanager.google.com/",
                "--title",
                "GTM",
                "--container-id",
                "GTM-TEST",
                "--workspace",
                "Recette",
            )
            run(
                "register-surface",
                str(ledger),
                "--role",
                "tag_assistant",
                "--url",
                "https://tagassistant.google.com/",
                "--title",
                "Tag Assistant",
                "--connected",
                "true",
            )
            run(
                "register-surface",
                str(ledger),
                "--role",
                "website",
                "--url",
                "https://shop.example.test/product",
                "--title",
                "Product",
            )
            run(
                "register-case",
                str(ledger),
                "--results",
                str(results),
                "--case-id",
                "CASE-ADD-DESKTOP",
                "--event-group-id",
                "EVG-001",
                "--url",
                "https://shop.example.test/product",
                "--element",
                "Add to cart",
                "--placement",
                "product detail",
                "--action",
                "click",
                "--variant",
                "quantity=1",
                "--discovered-from",
                "tracking_plan",
            )
            run(
                "register-tag",
                str(ledger),
                "--case-id",
                "CASE-ADD-DESKTOP",
                "--tag-id",
                "TAG-GA4-ADD",
                "--tag-name",
                "GA4 - Event - add_to_cart",
                "--container-id",
                "GTM-TEST",
                "--tag-category",
                "analytics",
                "--tag-delivery",
                "browser_request",
                "--vendor-family",
                "ga4",
                "--destination-id",
                "G-TEST123",
                "--template-type",
                "GA4 Event",
                "--consent-required",
                "false",
                "--evidence-id",
                "EVD-TAG-CONFIG-011",
            )
            run(
                "complete-tag-inventory",
                str(ledger),
                "--case-id",
                "CASE-ADD-DESKTOP",
                "--reason",
                "Container and Preview tag inventory completed.",
                "--evidence-id",
                "EVD-TAG-CONFIG-011",
            )
            record_runtime("READY-ACT-001", "before_action", "ACT-001", 10, 20)
            run(
                "begin-action",
                str(ledger),
                "--action-id",
                "ACT-001",
                "--case-id",
                "CASE-ADD-DESKTOP",
                "--readiness-check-id",
                "READY-ACT-001",
                "--consent-state",
                "analytics_storage=granted",
            )
            record_runtime(
                "SETTLE-ACT-001",
                "after_action",
                "ACT-001",
                10,
                20,
                push_count=0,
            )
            run(
                "settle-action",
                str(ledger),
                "--action-id",
                "ACT-001",
                "--settlement-check-id",
                "SETTLE-ACT-001",
                "--expected-seen",
                "false",
                "--interaction-outcome",
                "failed",
                "--completion-signal",
                "Overlay intercepted the click",
                "--settlement-reason",
                "interaction_failed",
            )
            record_runtime("READY-ACT-002", "before_action", "ACT-002", 10, 20)
            run(
                "begin-action",
                str(ledger),
                "--action-id",
                "ACT-002",
                "--retry-of-action-id",
                "ACT-001",
                "--case-id",
                "CASE-ADD-DESKTOP",
                "--readiness-check-id",
                "READY-ACT-002",
                "--consent-state",
                "analytics_storage=granted",
                "--quiet-window-ms",
                "3000",
                "--timeout-ms",
                "20000",
            )
            run(
                "record-push",
                str(ledger),
                "--push-id",
                "PUSH-011",
                "--action-id",
                "ACT-002",
                "--event-index",
                "11",
                "--event-name",
                "add_to_cart",
                "--classification",
                "expected",
                "--classification-reason",
                "Expected add_to_cart for the completed product CTA case.",
                "--page-state",
                "Basket count is 1",
                "--evidence-id",
                "EVD-RAW-011",
                "--container-id",
                "GTM-TEST",
            )
            record_runtime(
                "SETTLE-ACT-002",
                "after_action",
                "ACT-002",
                12,
                21,
                first_event_after=11,
                push_count=1,
            )
            run(
                "settle-action",
                str(ledger),
                "--action-id",
                "ACT-002",
                "--settlement-check-id",
                "SETTLE-ACT-002",
                "--expected-seen",
                "true",
                "--interaction-outcome",
                "completed",
                "--completion-signal",
                "Basket count changed from 0 to 1",
                "--settlement-reason",
                "expected_and_quiet",
            )
            for layer, evidence_id in (
                ("raw_api_call", "EVD-RAW-011"),
                ("resolved_data_layer", "EVD-DL-011"),
                ("gtm_variable", "EVD-VAR-011"),
                ("tag_configuration", "EVD-TAG-CONFIG-011"),
                ("tag_firing", "EVD-TAG-RUNTIME-011"),
                ("tag_parameter", "EVD-TAG-RUNTIME-011"),
                ("destination_request_when_applicable", "EVD-NET-011"),
            ):
                run(
                    "record-layer",
                    str(ledger),
                    "--action-id",
                    "ACT-002",
                    "--layer",
                    layer,
                    "--status",
                    "PASS",
                    "--reason",
                    f"{layer} matched the tracking-plan expectation.",
                    "--evidence-id",
                    evidence_id,
                )
            state = json.loads(run("status", str(ledger)).stdout)
            self.assertEqual("SETTLED", state["actions"][0]["state"])
            self.assertEqual("failed", state["actions"][0]["interaction_outcome"])
            self.assertEqual("ACT-001", state["actions"][1]["retry_of_action_id"])
            self.assertEqual("completed", state["actions"][1]["interaction_outcome"])
            self.assertEqual(
                "Basket count changed from 0 to 1",
                state["actions"][1]["completion_signal"],
            )
            self.assertTrue(state["actions"][1]["preview_connected_after"])
            self.assertTrue(state["actions"][1]["stream_settled"])
            record_runtime("READY-ACT-003", "before_action", "ACT-003", 12, 21)
            run(
                "begin-action",
                str(ledger),
                "--action-id",
                "ACT-003",
                "--retry-of-action-id",
                "ACT-002",
                "--case-id",
                "CASE-ADD-DESKTOP",
                "--readiness-check-id",
                "READY-ACT-003",
                "--consent-state",
                "analytics_storage=granted",
            )
            record_runtime(
                "SETTLE-ACT-003",
                "after_action",
                "ACT-003",
                12,
                21,
                push_count=0,
            )
            invalid = subprocess.run(
                [
                    sys.executable,
                    script,
                    "settle-action",
                    str(ledger),
                    "--action-id",
                    "ACT-003",
                    "--settlement-check-id",
                    "SETTLE-ACT-003",
                    "--expected-seen",
                    "false",
                    "--interaction-outcome",
                    "completed",
                    "--settlement-reason",
                    "quiet_without_expected",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(0, invalid.returncode)
            self.assertIn("--completion-signal", invalid.stderr)

    def test_runtime_readiness_rejects_wrong_page_and_inactive_network(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        snapshot = deepcopy(session["runtime_checks"][0])
        snapshot["selected_page_url"] = "https://shop.example.test/cart"
        snapshot["network_capture_active"] = False
        errors = runtime_snapshot_errors(
            snapshot,
            phase="before_action",
            action_id="ACT-001",
            case=session["cases"][0],
            ledger=session,
            results=data,
            expected_connection_epoch=1,
        )
        self.assertTrue(any("selected Tag Assistant page differs" in row for row in errors))
        self.assertTrue(any("network_capture_active=true" in row for row in errors))

    def test_runtime_capture_rejects_untrusted_source_future_time_and_reused_proof(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["runtime_checks"][0]["capture_source"] = "manual_guess"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("supported browser runtime probe" in row for row in errors))

        session = execution_fixture(data)
        session["runtime_checks"][0]["captured_at"] = "2026-07-25T11:00:59+00:00"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("later than recorded_at" in row for row in errors))

        session = execution_fixture(data)
        session["runtime_checks"][0]["captured_at"] = "2026-07-25T09:50:00+00:00"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("runtime snapshot is stale" in row for row in errors))

        session = execution_fixture(data)
        session["runtime_checks"][1]["evidence_ids"] = list(
            session["runtime_checks"][0]["evidence_ids"]
        )
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("distinct action-window evidence IDs" in row for row in errors))

    def test_legacy_schema_v3_remains_readable_without_new_boundary_contract(self) -> None:
        current = fixture()
        current_session = execution_fixture(current)
        legacy = deepcopy(current)
        legacy["run"].pop("action_boundary_contract_version", None)
        for row in legacy["requirements"]:
            boundary = row.get("action_boundary", {})
            for field in (
                "readiness_check_id",
                "settlement_check_id",
                "network_request_cursor_before",
                "network_request_cursor_after",
            ):
                boundary.pop(field, None)
        self.assertEqual([], validate(legacy, strict=True))

        strict_current = deepcopy(current)
        strict_current["requirements"][0]["action_boundary"].pop("readiness_check_id")
        strict_errors = validate(strict_current, strict=False)
        self.assertTrue(any("readiness_check_id" in row for row in strict_errors))

        legacy_session = deepcopy(current_session)
        for field in (
            "operator_contract_version",
            "runtime_checks",
            "event_closures",
            "closure_history",
            "operator_state",
        ):
            legacy_session.pop(field, None)
        for action in legacy_session["actions"]:
            for field in (
                "readiness_check_id",
                "settlement_check_id",
                "readiness_evidence_ids",
                "settlement_evidence_ids",
                "network_request_cursor_before",
                "network_request_cursor_after",
            ):
                action.pop(field, None)
        self.assertEqual([], validate_session(legacy_session, results=legacy, final=True))

    def test_guided_operator_rejects_legacy_results_without_fabricating_runtime_proof(
        self,
    ) -> None:
        data = fixture()
        session = execution_fixture(data)
        data["run"].pop("action_boundary_contract_version", None)
        with tempfile.TemporaryDirectory() as tempdir:
            results_path = Path(tempdir) / "results.json"
            session_path = Path(tempdir) / "session.json"
            results_path.write_text(json.dumps(data), encoding="utf-8")
            session_path.write_text(json.dumps(session), encoding="utf-8")
            completed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "finish-run",
                    str(results_path),
                    str(session_path),
                    str(Path(tempdir) / "recette.xlsx"),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(0, completed.returncode)
            self.assertIn("missing historical fields are never fabricated", completed.stdout)

    def test_guided_operator_pause_resume_between_events_and_reopen_closed_event(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["event_closures"] = []
        session["operator_state"] = {"status": "ACTIVE", "current_event_group_id": None}
        with tempfile.TemporaryDirectory() as tempdir:
            results_path = Path(tempdir) / "results.json"
            session_path = Path(tempdir) / "session.json"
            results_path.write_text(json.dumps(data), encoding="utf-8")
            session_path.write_text(json.dumps(session), encoding="utf-8")
            paused = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "pause-run",
                    str(session_path),
                    "--label",
                    "between events",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, paused.returncode, paused.stdout + paused.stderr)
            resumed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "resume-run",
                    str(results_path),
                    str(session_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, resumed.returncode, resumed.stdout + resumed.stderr)
            self.assertIsNone(json.loads(resumed.stdout)["fresh_runtime_check_id"])

            session = execution_fixture(data)
            session_path.write_text(json.dumps(session), encoding="utf-8")
            reopened = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "reopen-event",
                    str(results_path),
                    str(session_path),
                    "--event-group-id",
                    "EVG-001",
                    "--reason",
                    "Late material footer interaction discovered.",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, reopened.returncode, reopened.stdout + reopened.stderr)
            reopened_session = json.loads(session_path.read_text(encoding="utf-8"))
            self.assertEqual([], reopened_session["event_closures"])
            self.assertEqual(1, len(reopened_session["closure_history"]))
            self.assertEqual([], validate_session(reopened_session, results=data, final=False))
            patch_path = Path(tempdir) / "no-op-patch.json"
            patch_path.write_text(
                json.dumps(
                    {
                        "event_group_id": "EVG-001",
                        "requirements": deepcopy(data["requirements"]),
                        "evidence": [],
                    }
                ),
                encoding="utf-8",
            )
            unchanged_close = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "close-event",
                    str(results_path),
                    str(session_path),
                    str(patch_path),
                    "--event-group-id",
                    "EVG-001",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(0, unchanged_close.returncode)
            self.assertIn("requires a new material case", unchanged_close.stdout)

    def test_paired_event_close_write_rolls_back_after_second_replace_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            results_path = Path(tempdir) / "results.json"
            session_path = Path(tempdir) / "session.json"
            original_results = b'{"state":"old-results"}\n'
            original_session = b'{"state":"old-session"}\n'
            results_path.write_bytes(original_results)
            session_path.write_bytes(original_session)
            path_type = type(results_path)
            real_replace = path_type.replace
            failed = False

            def fail_second_replace(source: Path, target: Path) -> Path:
                nonlocal failed
                if not failed and Path(target) == session_path and source.suffix == ".tmp":
                    failed = True
                    raise OSError("injected second replace failure")
                return real_replace(source, target)

            with (
                patch.object(path_type, "replace", new=fail_second_replace),
                self.assertRaises(OSError),
            ):
                _save_pair_atomic(
                    results_path,
                    {"state": "new-results"},
                    session_path,
                    {"state": "new-session"},
                )
            self.assertEqual(original_results, results_path.read_bytes())
            self.assertEqual(original_session, session_path.read_bytes())

    def test_operator_contract_rejects_uncaptured_cursor_and_missing_closure(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["actions"][0]["last_event_before"] -= 1
        session["event_closures"] = []
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("differs from readiness capture" in row for row in errors))
        self.assertTrue(any("one closure for every plan event" in row for row in errors))

    def test_event_feedback_computes_missing_event_and_occurrence_anomaly(self) -> None:
        missing = fixture()
        configure_absent_event(missing)
        missing_feedback = event_feedback(missing, execution_fixture(missing))[0]
        self.assertEqual("DATALAYER_EVENT_ABSENT", missing_feedback["primary_outcome"])
        self.assertIn("MISSING_EXPECTED_OCCURRENCE", missing_feedback["anomaly_flags"])

        duplicate = fixture()
        duplicate_session = execution_fixture(duplicate)
        duplicate_session["business_pushes"][0]["classification"] = "duplicate"
        duplicate_feedback = event_feedback(duplicate, duplicate_session)[0]
        self.assertIn("DUPLICATE_OCCURRENCE", duplicate_feedback["anomaly_flags"])

    def test_guided_operator_closes_event_with_feedback_then_builds_workbook(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["event_closures"] = []
        session["operator_state"] = {
            "status": "ACTIVE",
            "current_event_group_id": "EVG-001",
        }
        patch = {
            "event_group_id": "EVG-001",
            "requirements": deepcopy(data["requirements"]),
            "evidence": [],
        }
        with tempfile.TemporaryDirectory() as tempdir:
            results_path = Path(tempdir) / "results.json"
            session_path = Path(tempdir) / "session.json"
            patch_path = Path(tempdir) / "event-patch.json"
            workbook_path = Path(tempdir) / "recette.xlsx"
            results_path.write_text(json.dumps(data), encoding="utf-8")
            session_path.write_text(json.dumps(session), encoding="utf-8")
            patch_path.write_text(json.dumps(patch), encoding="utf-8")
            close = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "close-event",
                    str(results_path),
                    str(session_path),
                    str(patch_path),
                    "--event-group-id",
                    "EVG-001",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, close.returncode, close.stdout + close.stderr)
            feedback = json.loads(close.stdout)
            self.assertEqual("PASS", feedback["status"])
            self.assertTrue(feedback["layer_feedback"])
            self.assertTrue(feedback["tag_feedback"])
            finish = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recette_operator.py"),
                    "finish-run",
                    str(results_path),
                    str(session_path),
                    str(workbook_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, finish.returncode, finish.stdout + finish.stderr)
            self.assertTrue(workbook_path.exists())

    def test_ga4_tracking_plan_handoff_initializes_exact_ordered_requirements(self) -> None:
        plan = {
            "events": [
                {
                    "event_name": "generate_lead",
                    "classification": "official",
                    "journey_ids": ["lead_generation"],
                    "measurement_opportunity_ids": ["lead_success"],
                    "trigger": "Confirmed lead success.",
                    "locations": [
                        {"url_pattern": "https://example.test/quote", "component": "quote form"}
                    ],
                    "data_layer": {
                        "clear": ["event_data"],
                        "push": {
                            "event": "generate_lead",
                            "event_data": {"form_name": "quote_request"},
                        },
                    },
                    "parameters": [
                        {
                            "name": "form_name",
                            "scope": "event",
                            "type": "string",
                            "requirement": "required",
                            "data_layer_path": "event_data.form_name",
                            "destination": "ga4_event_parameter",
                            "allowed_values": ["quote_request"],
                        }
                    ],
                }
            ]
        }
        expected = {
            "events": [
                {
                    "event_name": "generate_lead",
                    "measurement_opportunity_ids": ["lead_success"],
                }
            ]
        }
        with tempfile.TemporaryDirectory() as raw:
            delivery = Path(raw)
            plan_path = delivery / "plan.json"
            expected_path = delivery / "expected-events.json"
            plan_path.write_text(json.dumps(plan), encoding="utf-8")
            expected_path.write_text(json.dumps(expected), encoding="utf-8")
            plan_hash = hashlib.sha256(plan_path.read_bytes()).hexdigest()
            expected_hash = hashlib.sha256(expected_path.read_bytes()).hexdigest()
            handoff = {
                "handoff_version": "1.0.0",
                "skill": {"name": "ga4-tracking-plan", "version": "2.5.0"},
                "approval": {"state": "reviewed"},
                "plan": {
                    "canonical_sha256": plan_hash,
                    "target_sites": ["https://example.test/"],
                },
                "artifacts": [
                    {
                        "path": "plan.json",
                        "role": "canonical_tracking_plan",
                        "sha256": plan_hash,
                    },
                    {
                        "path": "expected-events.json",
                        "role": "runtime_expected_events_contract",
                        "sha256": expected_hash,
                    },
                ],
            }
            (delivery / "handoff.json").write_text(json.dumps(handoff), encoding="utf-8")
            verified_handoff, verified_plan, verified_expected = verify_delivery(delivery)
            imported = interpreted_requirements(verified_handoff, verified_plan, verified_expected)
            self.assertEqual(len(imported["requirements"]), 2)
            self.assertEqual(
                [item["source"]["plan_order"] for item in imported["requirements"]],
                [1, 2],
            )
            self.assertEqual(
                imported["requirements"][1]["expectation"]["expected_value"],
                ["quote_request"],
            )
            expected_path.write_text("{}", encoding="utf-8")
            with self.assertRaisesRegex(HandoffError, "hash-mismatched"):
                verify_delivery(delivery)

    def test_per_tag_evidence_cannot_be_reused_for_another_tag(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        case = session["cases"][0]
        case["tag_inventory"][0]["tag_id"] = "TAG-OTHER"
        for row in session["actions"][0]["tag_layer_results"]:
            row["tag_id"] = "TAG-OTHER"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("not bound to this exact tag" in error for error in errors))

    def test_self_asserted_equal_comparison_is_rejected_by_anchor(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        row = next(
            item
            for item in session["actions"][0]["tag_layer_results"]
            if item["layer"] == "tag_configuration"
        )
        comparison = row["details"]["configuration"][0]
        comparison["expected_value"] = "{{Fake - Variable}}"
        comparison["actual_value"] = "{{Fake - Variable}}"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any(
                "expected value differs from its accepted source anchor" in error
                for error in errors
            )
        )

    def test_browser_request_ids_must_reconcile_with_direct_evidence(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        row = next(
            item
            for item in session["actions"][0]["tag_layer_results"]
            if item["layer"] == "destination_request_when_applicable"
        )
        row["details"]["request_ids"] = ["NET-FABRICATED"]
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(any("request_ids do not reconcile" in error for error in errors))

    def test_session_per_tag_sensitive_value_blocks_export_contract(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        row = next(
            item
            for item in session["actions"][0]["tag_layer_results"]
            if item["layer"] == "tag_parameter"
        )
        row["details"]["diagnostic"] = "raw-person@example.test"
        errors = validate_session(session, results=data, final=True)
        self.assertTrue(
            any("session contains unredacted sensitive content" in error for error in errors)
        )
        with (
            tempfile.TemporaryDirectory() as tempdir,
            self.assertRaisesRegex(ReportValidationError, "unredacted sensitive content"),
        ):
            build_workbook(data, Path(tempdir) / "unsafe.xlsx", session=session)

    def test_schema_v2_migration_preserves_order_but_resets_proof(self) -> None:
        legacy = fixture()
        legacy["schema_version"] = 2
        migrated = migrate_results(legacy)
        self.assertEqual(3, migrated["schema_version"])
        self.assertEqual(
            legacy["run"]["requirement_inventory"], migrated["run"]["requirement_inventory"]
        )
        self.assertEqual([], migrated["evidence"])
        self.assertEqual([], migrated["unexpected"])
        self.assertTrue(
            all(row["verdict"]["overall"] == "PENDING" for row in migrated["requirements"])
        )
        self.assertTrue(
            all(row["journey"]["execution_status"] == "PENDING" for row in migrated["requirements"])
        )

    def test_analytics_vendor_taxonomy_covers_common_non_ga4_tags(self) -> None:
        for vendor in (
            "Piano Analytics",
            "Adobe Analytics",
            "Matomo",
            "Piwik PRO",
            "Snowplow",
            "Realytics",
        ):
            with self.subTest(vendor=vendor):
                self.assertEqual(
                    "analytics",
                    inferred_tag_category(
                        {"vendor_family": vendor, "template_type": f"{vendor} event"}
                    ),
                )

    def test_scaffold_tag_results_emits_exact_eight_layer_matrix(self) -> None:
        session = execution_fixture(fixture())
        session["actions"][0]["state"] = "OPEN"
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "tag-results.json"
            scaffold_tag_results(
                session,
                Namespace(action_id="ACT-001", output=output),
            )
            rows = json.loads(output.read_text(encoding="utf-8"))["tag_layer_results"]
        self.assertEqual(list(TAG_RESULT_LAYERS), [row["layer"] for row in rows])
        self.assertTrue(all(row["status"] == "PENDING" for row in rows))
        self.assertTrue(all(row["tag_id"] == "TAG-FIXTURE-001" for row in rows))

    def test_late_tag_discovery_versions_inventory_and_forces_retest(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        data["evidence"].append(
            {
                "evidence_id": "EVD-LATE-TAG",
                "kind": "tag_configuration",
                "source": "Tag Assistant",
                "capture_mode": "direct",
                "action_id": "ACT-001",
                "event_index": 11,
                "container_id": "GTM-TEST",
                "tag_id": "TAG-LATE-PIANO",
                "tag_name": "Piano - add_to_cart",
                "configuration_field": "event_name",
                "path_or_url": "evidence/late-piano.json",
                "captured_at": "2026-07-25T10:01:03+00:00",
                "description": "Late direct tag inventory evidence.",
            }
        )
        revise_tag_inventory(
            session,
            Namespace(
                case_id="CASE-001",
                tag_id="TAG-LATE-PIANO",
                tag_name="Piano - add_to_cart",
                container_id="GTM-TEST",
                tag_category="analytics",
                tag_delivery="browser_request",
                vendor_family="piano",
                destination_id="PA-TEST",
                template_type="Piano Analytics Event",
                consent_required="false",
                evidence_id=["EVD-LATE-TAG"],
                reason="Late direct Preview discovery requires a new attempt.",
            ),
        )
        case = session["cases"][0]
        self.assertEqual(2, case["inventory_revision"])
        self.assertEqual(1, len(case["applicability_history"]))
        self.assertEqual("PENDING", case["execution_status"])
        self.assertIsNone(case["final_action_id"])
        self.assertEqual("ACT-001", case["required_retest_of_action_id"])
        self.assertEqual([], validate_session(session, results=data, final=False))

    def test_workbook_splits_oversized_structured_cells_without_truncation(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        row = next(
            item
            for item in session["actions"][0]["tag_layer_results"]
            if item["layer"] == "tag_parameter"
        )
        row["details"]["large_safe_diagnostic"] = "x" * 40000
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "large.xlsx"
            build_workbook(data, output, session=session)
            workbook = load_workbook(output, read_only=True)
            values = [
                cell.value
                for sheet_row in workbook["Layer Verdicts"].iter_rows()
                for cell in sheet_row
                if isinstance(cell.value, str)
            ]
            workbook.close()
        self.assertTrue(any(value.startswith("[part 1/2]") for value in values))
        self.assertTrue(any(value.startswith("[part 2/2]") for value in values))
        self.assertTrue(all(len(value) <= 32767 for value in values))

    def test_release_archive_manifest_is_hash_verified(self) -> None:
        version = str(
            tomllib.loads((ROOT / "pyproject.toml").read_text(encoding="utf-8"))["project"][
                "version"
            ]
        )
        with tempfile.TemporaryDirectory() as tempdir:
            archives = []
            for directory in ("first", "second"):
                archive = Path(tempdir) / directory / f"gtm-preview-recette-v{version}.zip"
                completed = subprocess.run(
                    [
                        sys.executable,
                        str(SCRIPTS / "build_skill_package.py"),
                        "--output",
                        str(archive),
                    ],
                    check=False,
                    capture_output=True,
                    text=True,
                )
                self.assertEqual(0, completed.returncode, completed.stderr)
                archives.append(archive)
            manifest = verify_archive(archives[0])
            self.assertEqual(archives[0].read_bytes(), archives[1].read_bytes())
        self.assertEqual(f"v{version}", manifest["release"])
        self.assertIn("SKILL.md", manifest["files"])
        self.assertIn("scripts/build_skill_package.py", manifest["files"])
        self.assertIn("tests/test_v220_regressions.py", manifest["files"])


if __name__ == "__main__":
    unittest.main()
