from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from helpers import action_fixture, bundle_fixture, event_fixture

from judge import judge_event
from report import HEADERS, LAYER_ORDER, feedback_rows, render_report


class ReportTests(unittest.TestCase):
    def result(self) -> dict:
        result = judge_event(event_fixture(), action_fixture(), bundle_fixture())
        result["evidence_file"] = "evidence-E-0001.json"
        return result

    def test_exact_one_sheet_four_columns_and_five_rows(self) -> None:
        result = self.result()
        plan = {
            "events": [
                {
                    "event_id": result["event_id"],
                    "event_name": result["event_name"],
                }
            ]
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "gtm-client-recette-results.xlsx"
            render_report(plan, [result], path)
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, ["Event feedback"])
                sheet = workbook["Event feedback"]
                self.assertEqual(tuple(cell.value for cell in sheet[1]), HEADERS)
                self.assertEqual(sheet.max_row, 6)
                self.assertEqual(
                    [sheet.cell(row, 2).value for row in range(2, 7)],
                    list(LAYER_ORDER),
                )
            finally:
                workbook.close()

    def test_details_use_fixed_verifiable_grammar(self) -> None:
        for row in feedback_rows(self.result()):
            self.assertTrue(row["details"].startswith("reason="))
            self.assertIn("; expected=", row["details"])
            self.assertIn("; observed=", row["details"])
            self.assertIn("; evidence=evidence-E-0001.json#", row["details"])

    def test_duplicate_or_missing_event_results_fail_before_write(self) -> None:
        result = self.result()
        plan = {
            "events": [
                {
                    "event_id": result["event_id"],
                    "event_name": result["event_name"],
                }
            ]
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            with self.assertRaisesRegex(ValueError, "exactly one result"):
                render_report(plan, [], path)
            with self.assertRaisesRegex(ValueError, "exactly one result"):
                render_report(plan, [result, result], path)
            self.assertFalse(path.exists())

    def test_validation_failure_leaves_no_final_or_temporary_workbook(self) -> None:
        result = self.result()
        plan = {"events": [{"event_id": result["event_id"], "event_name": result["event_name"]}]}
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            with (
                patch(
                    "report.load_workbook",
                    side_effect=ValueError("injected validation failure"),
                ),
                self.assertRaisesRegex(ValueError, "injected"),
            ):
                render_report(plan, [result], path)
            self.assertFalse(path.exists())
            self.assertFalse(path.with_name(path.name + ".tmp.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
