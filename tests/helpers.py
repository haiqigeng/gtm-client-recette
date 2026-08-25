from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook


def add_event_sheet(
    workbook: Workbook,
    title: str,
    event_name: str,
    fields: list[tuple[str, str, Any]],
    *,
    core: bool = False,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.cell(5, 2, "Nature" if core else "Name of the event")
    sheet.cell(5, 3, "Type page")
    sheet.cell(5, 4, "Triggers")
    sheet.cell(6, 2, "Core DataLayer" if core else event_name)
    sheet.cell(6, 3, "All pages" if core else "Synthetic page")
    sheet.cell(6, 4, "Observe prepared page" if core else "Perform the named interaction")
    for column, value in enumerate(
        ["variables", "Type", "Status", "Summary", "Values", "Develop"], 2
    ):
        sheet.cell(9, column, value)
    for row, (name, value_type, values) in enumerate(fields, 10):
        sheet.cell(row, 2, name)
        sheet.cell(row, 3, value_type)
        sheet.cell(row, 4, "mandatory")
        sheet.cell(row, 5, f"Synthetic {name}")
        sheet.cell(row, 6, values)
        sheet.cell(row, 7, "OK")
    sheet.cell(10 + len(fields) + 1, 2, "IMAGES :")
    sheet.cell(10 + len(fields) + 3, 2, "CODE :")
    sheet.cell(
        10 + len(fields) + 4,
        2,
        'window.dataLayer.push({"event":"discarded_example"});',
    )


def make_plan(path: Path, *, two_events: bool = False, finite: bool = False) -> Path:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["A1"] = "Synthetic plan"
    index = workbook.create_sheet("Events")
    index["A1"] = "Event"
    index["A2"] = "view_item"
    index["A3"] = "display-only typo"
    add_event_sheet(
        workbook,
        "View Item",
        "view_item",
        [
            ("item_name", "string", "Example product"),
            ("quantity", "number", 1),
            ("currency", "string", "EUR"),
            ("page_language", "string", "en/fr" if finite else "dynamic"),
        ],
    )
    if two_events:
        add_event_sheet(
            workbook,
            "Add To Cart",
            "add_to_cart",
            [("quantity", "number", 1), ("currency", "string", "EUR")],
        )
    workbook.save(path)
    workbook.close()
    return path


def event_fixture() -> dict[str, Any]:
    return {
        "event_id": "View-Item",
        "event_name": "view_item",
        "label": "View Item",
        "selector": {"event": "view_item"},
        "fields": [
            {
                "path": "ecommerce.items[].item_name",
                "type": "string",
                "required": True,
                "rule": "present",
            },
            {
                "path": "ecommerce.items[].quantity",
                "type": "number",
                "required": True,
                "rule": "present",
            },
            {
                "path": "ecommerce.currency",
                "type": "string",
                "required": True,
                "rule": "equals",
                "expected": "EUR",
            },
        ],
    }


def payload(quantity: int = 1, currency: str = "EUR") -> dict[str, Any]:
    return {
        "event": "view_item",
        "ecommerce": {
            "items": [{"item_name": "Synthetic product", "quantity": quantity}],
            "currency": currency,
        },
    }


def action_fixture() -> dict[str, Any]:
    return {
        "action_id": "A-0001",
        "event_id": "View-Item",
        "preview_cursor": 0,
    }


def bundle_fixture(quantity: int = 1) -> dict[str, Any]:
    value = payload(quantity)
    return {
        "observer_contract": "playwright-mcp-v8",
        "action_id": "A-0001",
        "event_id": "View-Item",
        "preview_cursor": 1,
        "scenario": {
            "id": "default",
            "signature": "product-detail-default",
            "values": {
                "ecommerce.items[].item_name": ["Synthetic product"],
                "ecommerce.items[].quantity": [quantity],
                "ecommerce.currency": "EUR",
            },
        },
        "coverage": {
            "complete": True,
            "rationale": "All material synthetic paths are covered.",
            "remaining": [],
            "unreachable": [],
        },
        "reality": {
            "complete": True,
            "attributable": True,
            "page": {
                "url": "https://example.test/product",
                "status_code": 200,
                "soft_404": False,
            },
            "outcome": True,
            "expected": {
                "ecommerce.items[].item_name": ["Synthetic product"],
                "ecommerce.items[].quantity": [quantity],
                "ecommerce.currency": "EUR",
            },
            "findings": [],
        },
        "source": {
            "complete": True,
            "attributable": True,
            "occurrence_count": 1,
            "selected": {"cursor": 1, "payload": deepcopy(value)},
            "calls": [
                {
                    "cursor": 1,
                    "event_name": "view_item",
                    "payload": deepcopy(value),
                    "business": True,
                }
            ],
        },
        "gtm": {
            "complete": True,
            "attributable": True,
            "tags": [
                {
                    "name": "GA4 - view_item",
                    "concerned": True,
                    "fired": True,
                    "firing_count": 1,
                    "mapped_paths": ["event", "ecommerce"],
                    "mappings": {"event": "view_item", "ecommerce": "Data Layer"},
                    "runtime": deepcopy(value),
                }
            ],
            "findings": [],
        },
        "network": {
            "complete": True,
            "attributable": True,
            "requests": [
                {
                    "url": "https://analytics.example.test/collect",
                    "status": 204,
                    "failed": False,
                    "duplicate": False,
                    "logical_hit_id": "hit-1",
                    "parameters": deepcopy(value),
                }
            ],
            "findings": [],
        },
        "behavior": {
            "complete": True,
            "attributable": True,
            "messages": [
                {
                    "cursor": 1,
                    "event_name": "view_item",
                    "payload": deepcopy(value),
                    "business": True,
                }
            ],
            "findings": [],
        },
    }


def blocked_bundle(action_id: str, event_id: str, cursor: int = 0) -> dict[str, Any]:
    incomplete = {"complete": False, "attributable": False, "reason": "Preview unavailable."}
    return {
        "observer_contract": "playwright-mcp-v8",
        "action_id": action_id,
        "event_id": event_id,
        "preview_cursor": cursor,
        "scenario": {"id": "blocked", "signature": "unobservable", "values": {}},
        "coverage": {
            "complete": True,
            "rationale": "Only reachable path was attempted.",
            "remaining": [],
            "unreachable": [],
        },
        "reality": {**incomplete, "expected": {}},
        "source": deepcopy(incomplete),
        "gtm": deepcopy(incomplete),
        "network": deepcopy(incomplete),
        "behavior": deepcopy(incomplete),
    }


def action_bundle(action: dict[str, Any], cursor: int) -> dict[str, Any]:
    values = {field["path"]: f"value-{index}" for index, field in enumerate(action["fields"], 1)}
    payload_value = {"event": action["event_name"], **values}
    return {
        "observer_contract": "playwright-mcp-v8",
        "action_id": action["action_id"],
        "event_id": action["event_id"],
        "preview_cursor": cursor,
        "scenario": {"id": "default", "signature": "default", "values": values},
        "coverage": {
            "complete": True,
            "rationale": "Only one synthetic behavior signature exists.",
            "remaining": [],
            "unreachable": [],
        },
        "reality": {
            "complete": True,
            "attributable": True,
            "page": {"url": "https://example.test/", "status_code": 200, "soft_404": False},
            "outcome": True,
            "expected": values,
            "findings": [],
        },
        "source": {
            "complete": True,
            "attributable": True,
            "occurrence_count": 1,
            "selected": {"cursor": cursor, "payload": deepcopy(payload_value)},
            "calls": [
                {
                    "cursor": cursor,
                    "event_name": action["event_name"],
                    "payload": deepcopy(payload_value),
                    "business": True,
                }
            ],
        },
        "gtm": {
            "complete": True,
            "attributable": True,
            "tags": [
                {
                    "name": f"Tag - {action['event_name']}",
                    "concerned": True,
                    "fired": True,
                    "firing_count": 1,
                    "mapped_paths": ["event", *values],
                    "runtime": deepcopy(payload_value),
                }
            ],
            "findings": [],
        },
        "network": {
            "complete": True,
            "attributable": True,
            "requests": [
                {
                    "url": "https://analytics.example.test/collect",
                    "status": 204,
                    "failed": False,
                    "duplicate": False,
                    "logical_hit_id": f"hit-{cursor}",
                    "parameters": deepcopy(payload_value),
                }
            ],
            "findings": [],
        },
        "behavior": {
            "complete": True,
            "attributable": True,
            "messages": [
                {
                    "cursor": cursor,
                    "event_name": action["event_name"],
                    "payload": deepcopy(payload_value),
                    "business": True,
                }
            ],
            "findings": [],
        },
    }
