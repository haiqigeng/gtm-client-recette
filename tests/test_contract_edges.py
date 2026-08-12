from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from argparse import Namespace
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
TESTS = ROOT / "tests"
sys.path[:0] = [str(SCRIPTS), str(TESTS)]

from test_pipeline import execution_fixture, fixture, requirement  # noqa: E402

from build_recette_report import (  # noqa: E402
    build_workbook,
    defect_rows,
    write_defects_csv,
    write_defects_markdown,
    write_stakeholder_summary,
)
from build_retest_manifest import build_manifest  # noqa: E402
from client_side_rules import (  # noqa: E402
    evaluate_business_rule,
    evaluate_report_business_rules,
    scan_sensitive_value,
)
from event_feedback import event_feedback  # noqa: E402
from execution_contract import validate_session  # noqa: E402
from incremental_recette import (  # noqa: E402
    _session_event_view,
    event_view,
    scaffold_event_patch,
)
from init_coverage_ledger import initialize_requirement  # noqa: E402
from layer_contract import applicable_layers  # noqa: E402
from preview_session_ledger import import_cases, record_push  # noqa: E402
from recette_schema import validate  # noqa: E402


class ContractEdgeTests(unittest.TestCase):
    def test_event_slice_keeps_only_its_own_reopen_history(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        closure = deepcopy(session["event_closures"][0])
        session["closure_history"] = [
            {
                "reopened_event_group_id": "EVG-001",
                "reopened_at": "2026-07-25T10:03:00+00:00",
                "reason": "Late material interaction.",
                "invalidated_closures": [
                    closure,
                    {
                        **closure,
                        "event_group_id": "EVG-002",
                        "plan_order": 2,
                    },
                ],
            }
        ]
        sliced = _session_event_view(session, "EVG-001")
        self.assertEqual(
            ["EVG-001"],
            [row["event_group_id"] for row in sliced["closure_history"][0]["invalidated_closures"]],
        )
        unchanged = execution_fixture(data)
        unchanged_closure = deepcopy(unchanged["event_closures"][0])
        unchanged["closure_history"] = [
            {
                "reopened_event_group_id": "EVG-001",
                "reopened_at": "2026-07-25T10:03:00+00:00",
                "reason": "Late material interaction.",
                "invalidated_closures": [unchanged_closure],
            }
        ]
        errors = validate_session(unchanged, results=data, final=True)
        self.assertTrue(any("requires a new case or final action" in row for row in errors))

    def test_workbook_forces_untrusted_formula_text_to_literal_strings(self) -> None:
        data = fixture()
        requirement(data)["notes"] = '=HYPERLINK("https://example.test","click")'
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "safe.xlsx"
            build_workbook(data, output)
            workbook = load_workbook(output, data_only=False)
            try:
                sheet = workbook["Requirement Matrix"]
                headers = {str(cell.value): cell.column for cell in sheet[1]}
                cell = sheet.cell(row=2, column=headers["notes"])
                self.assertEqual(requirement(data)["notes"], cell.value)
                self.assertEqual("s", cell.data_type)
            finally:
                workbook.close()

    def test_malformed_business_rule_paths_can_never_pass(self) -> None:
        payload = {"event": "purchase", "value": 42}
        result = evaluate_business_rule(
            {
                "rule_id": "BR-BAD-PATH",
                "operator": "equals_path",
                "left_path": "",
                "right_path": "value[",
            },
            payload,
        )
        self.assertEqual("REVIEW", result["status"])
        self.assertIn("Invalid configured path syntax", result["reason"])

    def test_encoded_sensitive_query_values_are_scanned_without_value_fingerprints(self) -> None:
        email = "person@example.com"
        for url in (
            "https://collect.example/g/collect?ep.user=person%40example.com",
            "/g/collect?ep.user=person%40example.com",
        ):
            findings = scan_sensitive_value(url)
            self.assertTrue(any(row["category"] == "email" for row in findings), findings)
            self.assertTrue(
                all(row["value_fingerprint"] == "not-retained" for row in findings),
                findings,
            )
            self.assertNotIn(email, json.dumps(findings))

    def test_vendor_user_data_query_values_distinguish_plaintext_from_hashes(self) -> None:
        raw_url = (
            "https://collect.example/pixel?em=person%40example.com&ph=%2B33123456789"
            "&fn=Alice&ln=Example&external_id=customer-1&uip=192.0.2.1&up.postal=75001"
        )
        raw_findings = scan_sensitive_value(raw_url)
        self.assertTrue(
            {"email", "phone", "person_name", "ip_address", "postal_address"}
            <= {row["category"] for row in raw_findings},
            raw_findings,
        )
        status_by_path = {row["path"]: row["status"] for row in raw_findings}
        self.assertEqual("FAIL", status_by_path["$?em"])
        self.assertEqual("FAIL", status_by_path["$?ph"])
        self.assertEqual("REVIEW", status_by_path["$?fn"])
        self.assertEqual("REVIEW", status_by_path["$?ln"])

        digest = "a" * 64
        hash_url = (
            "https://collect.example/pixel?"
            f"em=tv.1~em{digest}&ph={digest}&fn={digest}&ln={digest}&external_id={digest}"
        )
        hash_findings = scan_sensitive_value(hash_url)
        self.assertEqual(5, len(hash_findings), hash_findings)
        self.assertTrue(
            all(row["category"] == "hashed_user_data" for row in hash_findings),
            hash_findings,
        )
        self.assertTrue(all(row["status"] == "PASS" for row in hash_findings), hash_findings)
        self.assertNotIn(digest, json.dumps(hash_findings))

        decoded_findings = scan_sensitive_value(
            {"em": "person@example.com", "ph": "+33123456789", "external_id": digest}
        )
        self.assertEqual(
            {"email", "phone", "hashed_user_data"},
            {row["category"] for row in decoded_findings},
            decoded_findings,
        )
        unencoded_url_findings = scan_sensitive_value(
            "https://collect.example/pixel?em=person@example.com"
        )
        self.assertEqual(
            1,
            sum(row["category"] == "email" for row in unencoded_url_findings),
            unencoded_url_findings,
        )

        ambiguous_findings = scan_sensitive_value(
            "https://collect.example/pixel?pn=SKU-123&fn=process_checkout"
        )
        self.assertTrue(ambiguous_findings)
        self.assertEqual({"REVIEW"}, {row["status"] for row in ambiguous_findings})

    def test_unexpected_failure_requires_a_known_event_group(self) -> None:
        data = fixture()
        data["unexpected"] = [
            {
                "unexpected_id": "UNX-NOGROUP",
                "status": "FAIL",
                "classification_reason": "Confirmed wrong-context business push.",
                "evidence_ids": ["EVD-RAW-011"],
            }
        ]
        errors = validate(data, strict=False)
        self.assertTrue(any("missing event_group_id" in error for error in errors), errors)

    def test_unplanned_push_inherits_its_action_event_group(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["actions"][0]["state"] = "OPEN"
        record_push(
            session,
            Namespace(
                push_id="PUSH-UNPLANNED",
                action_id="ACT-001",
                event_index=12,
                connection_epoch=1,
                stream_id="tag_assistant",
                event_group_id=None,
                classification="unplanned_relevant",
                event_name="view_item_list",
                url="https://example.test/shop",
                page_state="homepage",
                classification_reason="Wrong-context event observed.",
                evidence_id="EVD-RAW-011",
                container_id="GTM-TEST",
            ),
        )
        self.assertEqual("EVG-001", session["business_pushes"][-1]["event_group_id"])

    def test_uniqueness_never_pools_raw_and_resolved_surfaces(self) -> None:
        rule = {
            "rule_id": "BR-UNIQUE",
            "operator": "unique_across_requirements",
            "path": "transaction_id",
        }
        data = {
            "requirements": [
                {
                    "requirement_id": "REQ-RAW",
                    "event_group_id": "EVG-RAW",
                    "expectation": {
                        "source_mechanism": "data_layer_push",
                        "business_rules": [rule],
                    },
                    "raw_api_call": {
                        "event_index": 1,
                        "payload": {"transaction_id": "T-1"},
                    },
                },
                {
                    "requirement_id": "REQ-RESOLVED",
                    "event_group_id": "EVG-RESOLVED",
                    "expectation": {
                        "source_mechanism": "data_layer_push",
                        "business_rules": [rule],
                    },
                    "raw_api_call": None,
                    "resolved_data_layer": {"snapshot": {"transaction_id": "T-1"}},
                },
            ]
        }
        results = evaluate_report_business_rules(data)
        self.assertEqual({"REVIEW"}, {row["status"] for row in results}, results)
        self.assertTrue(
            all("heterogeneous evidence surfaces" in row["reason"] for row in results),
            results,
        )

    def test_uniqueness_counts_resolved_occurrences_not_requirement_rows(self) -> None:
        rule = {
            "rule_id": "BR-UNIQUE",
            "operator": "unique_across_requirements",
            "path": "transaction_id",
        }

        def row(requirement_id: str, evidence_id: str) -> dict:
            return {
                "requirement_id": requirement_id,
                "event_group_id": "EVG-PURCHASE",
                "expectation": {
                    "source_mechanism": "data_layer_push",
                    "business_rules": [rule],
                },
                "raw_api_call": None,
                "resolved_data_layer": {
                    "snapshot": {"transaction_id": "T-1"},
                    "evidence_id": evidence_id,
                },
            }

        same_occurrence = {"requirements": [row("REQ-1", "EVD-1"), row("REQ-2", "EVD-1")]}
        self.assertEqual(
            {"PASS"},
            {item["status"] for item in evaluate_report_business_rules(same_occurrence)},
        )
        distinct_occurrences = {"requirements": [row("REQ-1", "EVD-1"), row("REQ-2", "EVD-2")]}
        self.assertEqual(
            {"FAIL"},
            {item["status"] for item in evaluate_report_business_rules(distinct_occurrences)},
        )

    def test_mixed_surface_review_never_overwrites_a_conclusive_uniqueness_failure(self) -> None:
        rule = {
            "rule_id": "BR-UNIQUE",
            "operator": "unique_across_requirements",
            "path": "transaction_id",
        }
        data = {
            "requirements": [
                {
                    "requirement_id": "REQ-RAW-1",
                    "event_group_id": "EVG-1",
                    "expectation": {"business_rules": [rule]},
                    "raw_api_call": {
                        "event_index": 1,
                        "payload": {"transaction_id": "DUPLICATE"},
                    },
                },
                {
                    "requirement_id": "REQ-RAW-2",
                    "event_group_id": "EVG-2",
                    "expectation": {"business_rules": [rule]},
                    "raw_api_call": {
                        "event_index": 2,
                        "payload": {"transaction_id": "DUPLICATE"},
                    },
                },
                {
                    "requirement_id": "REQ-RESOLVED",
                    "event_group_id": "EVG-3",
                    "expectation": {"business_rules": [rule]},
                    "raw_api_call": None,
                    "resolved_data_layer": {
                        "snapshot": {"transaction_id": "OTHER"},
                        "evidence_id": "EVD-RESOLVED",
                    },
                },
            ]
        }
        results = evaluate_report_business_rules(data)
        raw_results = [
            row for row in results if str(row.get("requirement_id", "")).startswith("REQ-RAW")
        ]
        self.assertEqual({"FAIL"}, {row["status"] for row in raw_results}, results)

    def test_capture_limitation_rejects_conclusive_layer_verdict_and_is_visible_in_xlsx(
        self,
    ) -> None:
        data = fixture()
        req = requirement(data)
        req["raw_api_call"]["field_value"] = {
            "__gtm_recette_type": "snapshot_truncated",
            "reason": "max_nodes",
        }
        errors = validate(data, strict=False)
        self.assertTrue(
            any("verdict.raw_payload must be BLOCKED" in error for error in errors),
            errors,
        )
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir) / "limited-capture.xlsx"
            build_workbook(data, output, warnings=errors)
            workbook = load_workbook(output, data_only=False)
            try:
                sheet = workbook["Requirement Matrix"]
                headers = {str(cell.value): cell.column for cell in sheet[1]}
                self.assertIn(
                    "snapshot_truncated",
                    str(sheet.cell(2, headers["raw_value"]).value),
                )
            finally:
                workbook.close()

    def test_null_event_group_uses_the_missing_group_error_branch(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["business_pushes"][0]["event_group_id"] = None
        errors = validate_session(session, results=data, final=False)
        self.assertTrue(any("classification requires event_group_id" in row for row in errors))
        self.assertFalse(any("unknown event_group_id 'None'" in row for row in errors))

    def test_explicit_action_epoch_controls_default_push_epoch(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        action = session["actions"][0]
        action.update(
            {
                "action_id": "ACT-EPOCH-3",
                "state": "OPEN",
                "connection_epoch": 3,
                "last_event_before": 100,
            }
        )
        session["business_pushes"] = []
        record_push(
            session,
            Namespace(
                push_id="PUSH-EPOCH-3",
                action_id="ACT-EPOCH-3",
                event_index=101,
                connection_epoch=None,
                stream_id="tag_assistant",
                event_group_id="EVG-001",
                classification="expected",
                event_name="add_to_cart",
                url="https://shop.example.test/product",
                page_state="product",
                classification_reason="Expected push.",
                evidence_id="EVD-RAW-011",
                container_id="GTM-TEST",
                captured_at=None,
            ),
        )
        self.assertEqual(3, session["business_pushes"][0]["connection_epoch"])

    def test_retest_manifest_and_scaffold_never_inherit_verdicts_or_evidence(self) -> None:
        data = fixture()
        req = requirement(data)
        req["verdict"]["tag_firing"] = "FAIL"
        req["verdict"]["overall"] = "FAIL"
        req["verdict"]["failure_layer"] = "tag_firing"
        req["verdict"]["mismatch"] = "Expected tag did not fire."
        session = execution_fixture(data)
        manifest = build_manifest(data, session)
        self.assertFalse(manifest["verdicts_inherited"])
        self.assertFalse(manifest["evidence_inherited"])
        self.assertEqual("PENDING", manifest["cases"][0]["execution_status"])
        self.assertNotIn("evidence_ids", manifest["cases"][0])

        patch = scaffold_event_patch(data, "EVG-001", session)
        self.assertFalse(patch["capture_context"]["verdicts_inherited"])
        self.assertEqual([], patch["evidence"])
        self.assertEqual("PENDING", patch["requirements"][0]["verdict"]["overall"])
        self.assertEqual([], patch["requirements"][0]["evidence_ids"])

        with tempfile.TemporaryDirectory() as tempdir:
            manifest_path = Path(tempdir) / "retest.json"
            results_path = Path(tempdir) / "results.json"
            manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
            results_path.write_text(json.dumps(data), encoding="utf-8")
            new_session = execution_fixture(data)
            new_session["cases"] = []
            new_session["actions"] = []
            new_session["business_pushes"] = []
            import_cases(
                new_session,
                Namespace(manifest=manifest_path, results=results_path),
            )
            self.assertEqual("PENDING", new_session["cases"][0]["execution_status"])
            self.assertEqual([], new_session["cases"][0]["authorization_ids"])

        incomplete_data = deepcopy(data)
        requirement(incomplete_data)["journey"]["url"] = ""
        empty_session = execution_fixture(incomplete_data)
        empty_session["cases"] = []
        empty_session["actions"] = []
        empty_session["business_pushes"] = []
        incomplete_manifest = build_manifest(incomplete_data, empty_session)
        self.assertEqual([], incomplete_manifest["cases"])
        self.assertEqual("EVG-001", incomplete_manifest["limitations"][0]["event_group_id"])

    def test_defect_register_contains_short_actionable_non_pass_rows(self) -> None:
        data = fixture()
        req = requirement(data)
        req["verdict"]["tag_firing"] = "FAIL"
        req["verdict"]["overall"] = "FAIL"
        req["verdict"]["failure_layer"] = "tag_firing"
        req["verdict"]["mismatch"] = "Expected GA4 tag did not fire."
        session = execution_fixture(data)
        layer_result = next(
            row for row in session["actions"][0]["layer_results"] if row["layer"] == "tag_firing"
        )
        layer_result.update({"status": "FAIL", "reason": "GA4 tag did not fire in this placement."})
        rows = defect_rows(data, session)
        self.assertEqual(2, len(rows))
        requirement_defect = next(row for row in rows if row["requirement_id"] == "REQ-001")
        case_defect = next(
            row for row in rows if row["case_id"] == "CASE-001" and row["requirement_id"] is None
        )
        self.assertEqual("tag_firing", requirement_defect["failed_layer"])
        self.assertEqual("Expected GA4 tag did not fire.", requirement_defect["concise_reason"])
        self.assertEqual("GA4 tag did not fire in this placement.", case_defect["concise_reason"])
        with tempfile.TemporaryDirectory() as tempdir:
            csv_path = Path(tempdir) / "defects.csv"
            markdown_path = Path(tempdir) / "defects.md"
            summary_path = Path(tempdir) / "summary.md"
            write_defects_csv(csv_path, rows)
            write_defects_markdown(markdown_path, rows)
            write_stakeholder_summary(summary_path, data, session)
            csv_text = csv_path.read_text(encoding="utf-8-sig")
            self.assertTrue(csv_text.startswith("output_contract_version,"))
            self.assertIn("2,DEF-", csv_text)
            self.assertIn("tag_firing", csv_text)
            self.assertIn("Expected GA4 tag did not fire", markdown_path.read_text())
            self.assertIn("Output contract: 2", markdown_path.read_text())
            self.assertIn("Output contract: 2", summary_path.read_text())
            self.assertIn("Non-PASS events", summary_path.read_text())

    def test_supporting_artifact_is_metadata_only_and_has_no_verdict_authority(self) -> None:
        data = fixture()
        artifact = {
            "contract_version": 1,
            "artifact_id": "ART-AUDIT-001",
            "artifact_type": "gtm_container_audit_facts",
            "source_skill": "gtm-container-audit-cleanup",
            "source_run_id": "AUDIT-RUN-001",
            "source_version": "1.0.0",
            "sha256": "a" * 64,
            "file_name": "audit-facts.json",
            "role": "supporting_only",
            "verdict_authority": False,
            "registered_at": "2026-08-01T10:00:00+00:00",
            "notes": None,
        }
        data["run"]["supporting_artifacts"] = [artifact]
        self.assertEqual([], validate(data, strict=True))
        artifact["verdict_authority"] = True
        errors = validate(data, strict=False)
        self.assertTrue(any("verdict_authority must be false" in row for row in errors))
        artifact["verdict_authority"] = False
        artifact["notes"] = "Contact person@example.test"
        errors = validate(data, strict=False)
        self.assertTrue(any("metadata contains sensitive content" in row for row in errors))

    def test_bulk_push_import_is_transactional_on_a_late_error(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["actions"][0]["state"] = "OPEN"
        with tempfile.TemporaryDirectory() as tempdir:
            ledger_path = Path(tempdir) / "session.json"
            pushes_path = Path(tempdir) / "pushes.json"
            original = json.dumps(session, ensure_ascii=False, indent=2) + "\n"
            ledger_path.write_text(original, encoding="utf-8")
            common = {
                "action_id": "ACT-001",
                "event_name": "add_to_cart",
                "classification": "expected",
                "classification_reason": "Expected action-window push.",
                "event_group_id": "EVG-001",
                "url": "https://shop.example.test/product",
                "page_state": "product",
                "evidence_id": "EVD-RAW-011",
                "container_id": "GTM-TEST",
            }
            pushes_path.write_text(
                json.dumps(
                    [
                        {**common, "push_id": "PUSH-IMPORT-1", "event_index": 12},
                        {**common, "push_id": "PUSH-IMPORT-2", "event_index": 12},
                    ]
                ),
                encoding="utf-8",
            )
            completed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "preview_session_ledger.py"),
                    "import-pushes",
                    str(ledger_path),
                    str(pushes_path),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertNotEqual(0, completed.returncode)
            self.assertEqual(original, ledger_path.read_text(encoding="utf-8"))

            pushes_path.write_text(
                json.dumps([{**common, "push_id": "PUSH-IMPORT-OK", "event_index": 12}]),
                encoding="utf-8",
            )
            succeeded = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "preview_session_ledger.py"),
                    "import-pushes",
                    str(ledger_path),
                    str(pushes_path),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, succeeded.returncode, succeeded.stderr)
            imported = json.loads(ledger_path.read_text(encoding="utf-8"))
            self.assertEqual("PUSH-IMPORT-OK", imported["business_pushes"][-1]["push_id"])

    def test_supporting_artifact_cli_registers_a_digest_not_artifact_content(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            ledger_path = Path(tempdir) / "results.json"
            artifact_path = Path(tempdir) / "audit-facts.json"
            ledger_path.write_text(json.dumps(fixture()), encoding="utf-8")
            artifact_path.write_text('{"synthetic_fact":"tag changed"}', encoding="utf-8")
            completed = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "register_supporting_artifact.py"),
                    str(ledger_path),
                    str(artifact_path),
                    "--artifact-id",
                    "ART-AUDIT-CLI",
                    "--artifact-type",
                    "gtm_container_audit_facts",
                    "--source-skill",
                    "gtm-container-audit-cleanup",
                    "--source-run-id",
                    "AUDIT-CLI",
                    "--source-version",
                    "1.0.0",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, completed.returncode, completed.stderr)
            updated = json.loads(ledger_path.read_text(encoding="utf-8"))
            metadata = updated["run"]["supporting_artifacts"][0]
            self.assertFalse(metadata["verdict_authority"])
            self.assertEqual(64, len(metadata["sha256"]))
            self.assertNotIn("synthetic_fact", json.dumps(updated))

    def test_out_of_scope_case_is_neutral_in_event_status_rollup(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        out_of_scope = deepcopy(session["cases"][0])
        out_of_scope.update(
            {
                "case_id": "CASE-OOS",
                "scope_status": "OUT_OF_SCOPE",
                "execution_status": "NOT_TESTED",
                "reason": "Confirmed placement exclusion.",
                "final_action_id": None,
            }
        )
        session["cases"].append(out_of_scope)
        session["event_closures"][0]["case_ids"].append("CASE-OOS")
        session["event_closures"][0]["case_ids"].reverse()
        self.assertEqual([], validate_session(session, results=data, final=True))

    def test_conditional_absence_can_pass_schema_and_session_contracts(self) -> None:
        data = fixture()
        req = requirement(data)
        req["expectation"]["expected_occurrence"] = {
            "rule": "conditional",
            "condition_id": "OUT-OF-STOCK",
            "branch_rule": "absent",
        }
        req["scenario"] = {
            "scenario_id": "SCN-OUT",
            "kind": "conditional",
            "condition": "Product is out of stock",
            "branch": "absent",
            "condition_met": True,
            "evidence_id": "EVD-SCENARIO-OUT",
        }
        data["evidence"].append(
            {
                "evidence_id": "EVD-SCENARIO-OUT",
                "kind": "scenario_branch",
                "source": "Playwright",
                "capture_mode": "direct",
                "action_id": "ACT-001",
                "event_index": 11,
                "path_or_url": "evidence/scenario-out.json",
                "captured_at": "2026-07-25T10:01:03+00:00",
                "description": "Out-of-stock branch was visibly confirmed.",
            }
        )
        req["evidence_ids"].append("EVD-SCENARIO-OUT")
        req["event_observed"] = False
        req["occurrence_evidence"] = {
            "actual_count": 0,
            "event_indexes": [],
            "evidence_id": "EVD-ACTION-001",
        }
        req["raw_api_call"] = None
        req["resolved_data_layer"].update(
            {"snapshot": {}, "field_state": "absent", "field_type": "absent"}
        )
        req["resolved_data_layer"].pop("field_value", None)
        req["gtm_variable"].update({"field_state": "absent", "field_type": "absent"})
        req["gtm_variable"].pop("field_value", None)
        req["expectation"].update(
            {
                "expected_firing": "not_fired",
                "expected_request_behavior": "absent",
                "business_rules": [
                    {
                        "rule_id": "BR-NOT-APPLICABLE-WHEN-ABSENT",
                        "operator": "range",
                        "path": "ecommerce.value",
                        "min": 0,
                    }
                ],
            }
        )
        req["tag"].update(
            {
                "expected_firing": "not_fired",
                "actual_firing": "not_fired",
                "fire_count": 0,
                "runtime_state": "absent",
                "runtime_type": "absent",
            }
        )
        req["tag"].pop("runtime_value", None)
        req["destination_request"].update({"request_behavior": "not_observed", "request_count": 0})
        for field in (
            "destination_parameter_path",
            "expected_destination_value",
            "expected_destination_type",
        ):
            req["expectation"].pop(field, None)
        for field in ("parameter_path", "field_state", "field_type", "field_value"):
            req["destination_request"].pop(field, None)
        req["verdict"]["destination_parameter"] = None

        layers = applicable_layers(data["requirements"], container_count=1)
        data["run"]["included_layers"] = layers
        session = execution_fixture(data)
        session["business_pushes"] = []
        session["actions"][0]["observed_business_push_count"] = 0
        session["actions"][0]["expected_seen"] = False
        session["runtime_checks"][1]["observed_business_push_count"] = 0
        self.assertEqual([], evaluate_report_business_rules(data))
        self.assertEqual([], validate(data, strict=True))
        self.assertEqual([], validate_session(session, results=data, final=True))

    def test_non_datalayer_business_rules_use_the_source_signal(self) -> None:
        data = {
            "requirements": [
                {
                    "requirement_id": "REQ-DIRECT",
                    "event_group_id": "EVG-DIRECT",
                    "expectation": {
                        "source_mechanism": "direct_vendor_call",
                        "business_rules": [
                            {
                                "rule_id": "BR-DIRECT",
                                "operator": "equals_path",
                                "left_path": "value",
                                "right_path": "expected",
                            }
                        ],
                    },
                    "source_signal": {
                        "event_index": 7,
                        "payload": {"value": 1, "expected": 1},
                    },
                }
            ]
        }
        results = evaluate_report_business_rules(data)
        self.assertEqual(1, len(results))
        self.assertEqual("PASS", results[0]["status"])
        self.assertEqual("source_signal.payload", results[0]["evaluation_source"])

    def test_non_datalayer_business_rules_can_use_a_scalar_source_value(self) -> None:
        data = {
            "requirements": [
                {
                    "requirement_id": "REQ-DIRECT-VALUE",
                    "event_group_id": "EVG-DIRECT",
                    "expectation": {
                        "source_mechanism": "direct_vendor_call",
                        "business_rules": [
                            {
                                "rule_id": "BR-DIRECT-VALUE",
                                "operator": "range",
                                "path": "value",
                                "min": 1,
                                "max": 2,
                            }
                        ],
                    },
                    "source_signal": {"event_index": 8, "value": 1},
                }
            ]
        }
        results = evaluate_report_business_rules(data)
        self.assertEqual(1, len(results))
        self.assertEqual("PASS", results[0]["status"])
        self.assertEqual("source_signal.value", results[0]["evaluation_source"])

    def test_reconnected_preview_indexes_are_unique_by_connection_epoch(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        session["business_pushes"][0]["connection_epoch"] = 1
        duplicate_index = deepcopy(session["business_pushes"][0])
        duplicate_index.update(
            {
                "push_id": "PUSH-EPOCH-2",
                "connection_epoch": 2,
            }
        )
        session["business_pushes"].append(duplicate_index)
        session["actions"][0]["observed_business_push_count"] = 2
        errors = validate_session(session, results=data, final=True)
        self.assertFalse(any("duplicate stream/connection-epoch" in error for error in errors))

    def test_unplanned_review_rolls_up_as_review_instead_of_fail(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        push = session["business_pushes"][0]
        push["classification"] = "unplanned_relevant"
        push["classification_reason"] = "Plan meaning requires analyst confirmation."
        data["unexpected"] = [
            {
                "unexpected_id": "UNX-001",
                "observed_push_id": push["push_id"],
                "event_group_id": "EVG-001",
                "action_id": "ACT-001",
                "case_id": "CASE-001",
                "classification": "unplanned_relevant",
                "classification_reason": push["classification_reason"],
                "status": "REVIEW",
                "review_basis": "semantic_ambiguity",
                "review_question": "Is this companion push intentionally in scope?",
                "evidence_ids": ["EVD-RAW-011"],
            }
        ]
        self.assertEqual([], validate(data, strict=True))
        self.assertEqual([], validate_session(session, results=data, final=True))
        self.assertEqual("REVIEW", event_feedback(data, session)[0]["status"])

    def test_tagless_direct_vendor_call_keeps_browser_destination_layer(self) -> None:
        layers = applicable_layers(
            [
                {
                    "scope_status": "IN_SCOPE",
                    "expectation": {
                        "source_mechanism": "direct_vendor_call",
                        "vendor_family": "meta",
                        "destination_id": "PIXEL-1",
                        "destination_event_name": "Lead",
                        "expected_endpoint_pattern": "^https://www\\.facebook\\.com/tr",
                        "expected_request_behavior": "sent_once",
                    },
                }
            ]
        )
        self.assertIn("source_signal_when_no_data_layer_push", layers)
        self.assertIn("destination_request_when_applicable", layers)
        self.assertNotIn("tag_firing", layers)

    def test_initializer_emits_no_rejected_tag_object_when_tag_is_not_applicable(self) -> None:
        source = deepcopy(requirement(fixture()))
        for field in (
            "tag_name",
            "tag_delivery",
            "expected_firing",
            "tag_configuration_field",
            "expected_tag_configuration",
        ):
            source["expectation"].pop(field, None)
        initialized = initialize_requirement(source)
        self.assertIsNone(initialized["tag"])

    def test_initializer_omits_payload_rule_layer_for_expected_absence(self) -> None:
        source = deepcopy(requirement(fixture()))
        source["expectation"]["expected_occurrence"] = "absent"
        source["expectation"]["business_rules"] = [
            {
                "rule_id": "BR-ABSENT",
                "operator": "range",
                "path": "ecommerce.value",
                "min": 0,
            }
        ]
        initialized = initialize_requirement(source)
        self.assertIsNone(initialized["verdict"]["business_rule"])
        self.assertNotIn(
            "business_rules_when_declared",
            applicable_layers([initialized]),
        )

    def test_run_wide_blocker_survives_incremental_event_projection(self) -> None:
        data = fixture()
        data["blockers"] = [
            {
                "blocker_id": "BLK-RUN",
                "type": "CMP_TEST_ENVIRONMENT",
                "status": "BLOCKED",
            }
        ]
        projected = event_view(data, "EVG-001")
        self.assertEqual(["BLK-RUN"], [row["blocker_id"] for row in projected["blockers"]])

    def test_apply_event_does_not_persist_before_session_validation(self) -> None:
        data = fixture()
        session = execution_fixture(data)
        patch = {
            "event_group_id": "EVG-001",
            "requirements": deepcopy(data["requirements"]),
            "evidence": [],
        }
        patch["requirements"][0]["notes"] = "MUST-NOT-PERSIST"
        session["actions"][0]["layer_results"][0]["status"] = "FAIL"
        session["actions"][0]["layer_results"][0]["reason"] = "Synthetic mismatch."
        with tempfile.TemporaryDirectory() as tempdir:
            ledger_path = Path(tempdir) / "ledger.json"
            patch_path = Path(tempdir) / "patch.json"
            session_path = Path(tempdir) / "session.json"
            ledger_path.write_text(json.dumps(data), encoding="utf-8")
            patch_path.write_text(json.dumps(patch), encoding="utf-8")
            session_path.write_text(json.dumps(session), encoding="utf-8")
            result = subprocess.run(
                [
                    sys.executable,
                    "-B",
                    str(SCRIPTS / "incremental_recette.py"),
                    "apply-event",
                    str(ledger_path),
                    str(patch_path),
                    "--session-ledger",
                    str(session_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(0, result.returncode)
            retained = json.loads(ledger_path.read_text(encoding="utf-8"))
            self.assertNotEqual("MUST-NOT-PERSIST", retained["requirements"][0]["notes"])


if __name__ == "__main__":
    unittest.main()
