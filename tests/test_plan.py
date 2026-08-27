from __future__ import annotations

import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from xlsx_plan import PlanError, extract_workbook, validate_inspection_plan


def interpretation(reference: str) -> dict:
    return {
        "schema_version": "3.0.0",
        "events": [
            {
                "event_name": "view_item",
                "parameters": [
                    {
                        "data_layer_path": "ecommerce.currency",
                        "ga4_parameter_name": "currency",
                        "value": "EUR",
                        "value_semantics": "EXAMPLE",
                        "json_type": "string",
                        "required": None,
                        "source_refs": [reference],
                    }
                ],
                "data_layer_payload": None,
                "definition": None,
                "trigger_description": None,
                "entry_url": None,
                "expected_destination_id": None,
                "source_refs": [reference],
            }
        ],
    }


def workbook(path: Path, value: str = "view_item") -> Path:
    book = Workbook()
    book.active["A1"] = value
    book.save(path)
    book.close()
    return path


class PlanTests(unittest.TestCase):
    def test_variable_layout_extracts_cells_formulas_links_images_and_code(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "variable.xlsx"
            stream = BytesIO()
            with Image.new("RGB", (8, 8), "blue") as image:
                image.save(stream, format="PNG")
            stream.seek(0)
            embedded = Image.open(stream)
            book = Workbook()
            sheet = book.active
            sheet.title = "Any Layout"
            sheet["C4"] = "view_item"
            sheet["D8"] = "=1+1"
            sheet["C5"] = "Site"
            sheet["C5"].hyperlink = "https://example.test/product"
            sheet["B20"] = "dataLayer.push({event: 'view_item', ecommerce: {currency: 'EUR'}});"
            sheet.add_image(WorksheetImage(embedded), "F3")
            book.save(path)
            book.close()
            embedded.close()
            evidence = extract_workbook(path)
        self.assertIn("Any Layout!C4", evidence["source_refs"])
        self.assertIn("image:Any Layout:1", evidence["source_refs"])
        self.assertEqual(evidence["code_calls"][0]["payload"]["event"], "view_item")

    def test_parameter_table_reconstructs_canonical_data_layer(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = workbook(Path(directory) / "plan.xlsx")
            plan = validate_inspection_plan(extract_workbook(path), interpretation("Sheet!A1"))
        event = plan["events"][0]
        self.assertEqual(event["data_layer_payload"]["event"], "view_item")
        self.assertEqual(event["data_layer_payload"]["ecommerce"]["currency"], "EUR")
        self.assertIn("window.dataLayer.push", event["data_layer_snippet"])
        self.assertEqual(event["scenario"], {"id": "view_item"})

    def test_complete_data_layer_reconstructs_parameter_table(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "plan.xlsx"
            book = Workbook()
            book.active["A1"] = (
                "dataLayer.push({event: 'view_item', ecommerce: {currency: 'EUR', "
                "items: [{item_id: 'SKU-1', quantity: 1}]}});"
            )
            book.save(path)
            book.close()
            evidence = extract_workbook(path)
            value = interpretation("Sheet!A1")
            value["events"][0]["parameters"] = []
            plan = validate_inspection_plan(evidence, value)
        self.assertEqual(
            {parameter["data_layer_path"] for parameter in plan["events"][0]["parameters"]},
            {"ecommerce.currency", "ecommerce.items[].item_id", "ecommerce.items[].quantity"},
        )

    def test_optional_context_is_nonfatal(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = workbook(Path(directory) / "plan.xlsx")
            event = validate_inspection_plan(extract_workbook(path), interpretation("Sheet!A1"))[
                "events"
            ][0]
        self.assertIsNone(event["entry_url"])
        self.assertEqual(
            {notice["field"] for notice in event["mapping_notices"]},
            {"definition", "trigger_description", "entry_url"},
        )

    def test_missing_mandatory_semantics_or_unknown_reference_fails(self) -> None:
        evidence = {
            "source": {"path": "x", "filename": "x", "sha256": "0" * 64},
            "source_refs": ["Sheet!A1"],
            "code_calls": [],
        }
        value = interpretation("Sheet!A1")
        value["events"][0]["parameters"] = []
        with self.assertRaisesRegex(PlanError, "parameter values"):
            validate_inspection_plan(evidence, value)
        with self.assertRaisesRegex(PlanError, "unknown source reference"):
            validate_inspection_plan(evidence, interpretation("Sheet!Z9"))

    def test_payload_and_table_contradiction_is_fatal(self) -> None:
        evidence = {
            "source": {"path": "x", "filename": "x", "sha256": "0" * 64},
            "source_refs": ["Sheet!A1"],
            "code_calls": [],
        }
        value = interpretation("Sheet!A1")
        value["events"][0]["data_layer_payload"] = {
            "event": "view_item",
            "ecommerce": {"currency": "USD"},
        }
        with self.assertRaisesRegex(PlanError, "value contradicts"):
            validate_inspection_plan(evidence, value)

    def test_dynamic_placeholder_does_not_require_literal_payload_match(self) -> None:
        evidence = {
            "source": {"path": "x", "filename": "x", "sha256": "0" * 64},
            "source_refs": ["Sheet!A1"],
            "code_calls": [],
        }
        value = interpretation("Sheet!A1")
        parameter = value["events"][0]["parameters"][0]
        parameter["value"] = "%currency%"
        parameter["value_semantics"] = "DYNAMIC"
        value["events"][0]["data_layer_payload"] = {
            "event": "view_item",
            "ecommerce": {"currency": "EUR"},
        }
        plan = validate_inspection_plan(evidence, value)
        self.assertEqual(plan["events"][0]["parameters"][0]["value_semantics"], "DYNAMIC")


if __name__ == "__main__":
    unittest.main()
